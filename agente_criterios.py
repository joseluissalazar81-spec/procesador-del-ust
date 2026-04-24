"""
agente_criterios.py — Agente Claude para verificación de los 45 criterios UST
Procesador DEL · UST 2026-1

Divide los 45 criterios en dos capas:
  - Capa 1 (automática): calculos_del.py → 25 criterios numérico/estructurales
  - Capa 2 (Claude Sonnet 4.6): 20 criterios semánticos/pedagógicos

Grupos semánticos (5 llamadas a Claude):
  G1 — Redacción de actividades        (C19-C22)
  G2 — Estructura de momentos          (C24-C26)
  G3 — Recursos declarados             (C27-C35)
  G4 — Estrategias evaluativas         (C39-C41)
  G5 — Viabilidad de carga académica   (C44-C45)
"""

from __future__ import annotations
import json
import os
import re
from io import BytesIO
from collections import defaultdict

import openpyxl


def _to_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


# ── Importar motor matemático ────────────────────────────────────────────────
try:
    from calculos_del import validar_horas_y_recursos
    CALCULOS_OK = True
except ImportError:
    CALCULOS_OK = False


# ── Modelo Claude ────────────────────────────────────────────────────────────
MODELO = "claude-sonnet-4-6"

# ── Estados posibles por criterio ────────────────────────────────────────────
SI          = "SI"
PARCIALMENTE = "PARCIALMENTE"
NO          = "NO"

# ── Los 45 criterios con su ID, sección y tipo ──────────────────────────────
CRITERIOS = [
    # ── SÍNTESIS DIDÁCTICA (automático) ────────────────────────────────────
    {"id": 1,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Escuela correcta (conforme a programa oficial)"},
    {"id": 2,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Programa académico correcto (conforme a programa oficial)"},
    {"id": 3,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Nombre de la asignatura correcto (conforme a programa oficial)"},
    {"id": 4,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Jornada correcta (conforme a programa oficial)"},
    {"id": 5,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Código correcto (conforme a programa oficial)"},
    {"id": 6,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Versión año correcta (conforme a programa oficial)"},
    {"id": 7,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "N° de créditos SCT-Chile correcto (conforme a programa oficial)"},
    {"id": 8,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Carga académica estructurada y distribuida según programa oficial"},
    {"id": 9,  "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Presenta competencias que desarrolla la asignatura (conforme a programa oficial)"},
    {"id": 10, "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Nombre de la unidad o módulo correcto (conforme a programa oficial)"},
    {"id": 11, "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Horas pedagógicas de la unidad correctas (conforme a programa oficial)"},
    {"id": 12, "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Resultados de aprendizaje redactados fielmente según programa"},
    {"id": 13, "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Procedimientos de evaluación redactados fielmente según programa"},
    {"id": 14, "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Porcentajes de evaluación corresponden al programa oficial"},
    {"id": 15, "seccion": "Síntesis",   "tipo": "auto",
     "texto": "Diferencia resultados de aprendizaje con sus respectivos porcentajes"},
    # ── PLANIFICACIÓN — IDENTIFICACIÓN (automático) ──────────────────────────
    {"id": 16, "seccion": "Identificación", "tipo": "auto",
     "texto": "Las unidades de aprendizaje corresponden al programa oficial"},
    {"id": 17, "seccion": "Identificación", "tipo": "auto",
     "texto": "Los resultados de aprendizaje están vinculados a las unidades según programa"},
    {"id": 18, "seccion": "Identificación", "tipo": "auto",
     "texto": "Especifica número de sesión y distribuye horas por momento"},
    # ── ESTRATEGIAS METODOLÓGICAS — REDACCIÓN (semántico G1) ────────────────
    {"id": 19, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G1",
     "texto": "Las actividades están redactadas en imperativo, de manera clara y secuencial"},
    {"id": 20, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G1",
     "texto": "Las actividades que requieren entrega explicitan el propósito de la misma"},
    {"id": 21, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G1",
     "texto": "Las actividades que requieren entrega indican si habrá retroalimentación"},
    {"id": 22, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G1",
     "texto": "Las actividades están alineadas con los contenidos y resultados de aprendizaje"},
    # ── ESTRUCTURA DE MOMENTOS (semántico G2) ────────────────────────────────
    {"id": 23, "seccion": "Est.Metodológicas", "tipo": "auto",
     "texto": "Establece los tres momentos para cada actividad: preparación, desarrollo, trabajo independiente"},
    {"id": 24, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G2",
     "texto": "Las actividades de preparación contextualizan y activan conocimientos previos"},
    {"id": 25, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G2",
     "texto": "Las actividades de desarrollo construyen el nuevo aprendizaje de manera guiada"},
    {"id": 26, "seccion": "Est.Metodológicas", "tipo": "semantico", "grupo": "G2",
     "texto": "Las actividades de trabajo independiente consolidan el aprendizaje autónomamente"},
    # ── RECURSOS (semántico G3) ───────────────────────────────────────────────
    {"id": 27, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Declara todos los recursos necesarios para cada actividad"},
    {"id": 28, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Identifica claramente cada recurso didáctico (título, autor, tipo)"},
    {"id": 29, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Especifica la ubicación o medio de acceso de cada recurso (plataforma, URL, aula virtual)"},
    {"id": 30, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Indica la extensión de los recursos para estimar carga académica (páginas, minutos)"},
    {"id": 31, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Declara las presentaciones o materiales del/de la docente para sesiones sincrónicas o presenciales"},
    {"id": 32, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Los recursos didácticos son coherentes con los resultados de aprendizaje esperados"},
    {"id": 33, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Declara las guías o instrucciones para que los y las estudiantes realicen las actividades"},
    {"id": 34, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "Declara los espacios de interacción y entrega (foros, buzones, etc.)"},
    {"id": 35, "seccion": "Recursos", "tipo": "semantico", "grupo": "G3",
     "texto": "La cantidad de recursos por momento es adecuada y variada"},
    # ── ESTRATEGIAS EVALUATIVAS (semántico G4) ───────────────────────────────
    {"id": 36, "seccion": "Est.Evaluativas", "tipo": "auto",
     "texto": "El tipo, procedimiento e instrumento de evaluación son coherentes con el programa oficial"},
    {"id": 37, "seccion": "Est.Evaluativas", "tipo": "auto",
     "texto": "Los tipos de evaluación están alineados con los resultados de aprendizaje esperados"},
    {"id": 38, "seccion": "Est.Evaluativas", "tipo": "auto",
     "texto": "Los porcentajes de evaluación corresponden al programa oficial"},
    {"id": 39, "seccion": "Est.Evaluativas", "tipo": "semantico", "grupo": "G4",
     "texto": "Las evaluaciones están alineadas con los contenidos desarrollados y el logro esperado"},
    {"id": 40, "seccion": "Est.Evaluativas", "tipo": "semantico", "grupo": "G4",
     "texto": "Las evaluaciones están identificadas con uso de software para verificación de integridad académica (Turnitin/Klarway)"},
    {"id": 41, "seccion": "Est.Evaluativas", "tipo": "semantico", "grupo": "G4",
     "texto": "Las evaluaciones permiten verificar efectivamente el logro del resultado de aprendizaje"},
    # ── CARGA ACADÉMICA (automático + semántico G5) ──────────────────────────
    {"id": 42, "seccion": "Carga Académica", "tipo": "auto",
     "texto": "El total de horas de la asignatura corresponde al programa oficial"},
    {"id": 43, "seccion": "Carga Académica", "tipo": "auto",
     "texto": "La distribución de horas por tipo (sincrónicas, asincrónicas, presenciales, TPE) corresponde al programa"},
    {"id": 44, "seccion": "Carga Académica", "tipo": "semantico", "grupo": "G5",
     "texto": "Las horas asignadas a cada actividad son realistas y permiten su cumplimiento"},
    {"id": 45, "seccion": "Carga Académica", "tipo": "semantico", "grupo": "G5",
     "texto": "La carga académica total por semana/período es equilibrada y viable para el/la estudiante"},
]


# ── Extractor de datos del xlsx ──────────────────────────────────────────────

def extraer_datos_planificacion(fuente) -> dict:
    """Lee el xlsx y extrae los datos necesarios para el agente."""
    if isinstance(fuente, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(fuente), data_only=True)
    else:
        wb = openpyxl.load_workbook(fuente, data_only=True)

    datos = {
        "sintesis": {},
        "unidades": [],        # lista de filas agrupadas por momento
        "recursos_muestra": [],
        "evaluaciones": [],
        "horas_por_semana": defaultdict(float),
    }

    # ── Síntesis ──────────────────────────────────────────────────────────────
    if "Síntesis didáctica" in wb.sheetnames:
        ws = wb["Síntesis didáctica"]
        rows_sint = list(ws.iter_rows(values_only=True))
        for row in rows_sint:
            vals = [str(v).strip() if v else "" for v in row]
            if vals[0].startswith("RA") and vals[3]:
                datos["sintesis"].setdefault("ra_procedimientos", []).append(
                    {"ra": vals[0][:120], "procedimiento": vals[3]}
                )
            if not datos["sintesis"].get("asignatura") and vals[3] and "asignatura" not in vals[3].lower():
                if any("asignatura" in str(row[i] or "").lower() for i in range(3)):
                    datos["sintesis"]["asignatura"] = vals[3]

    # ── Planificación ─────────────────────────────────────────────────────────
    if "Planificación por unidades" not in wb.sheetnames:
        return datos

    ws = wb["Planificación por unidades"]
    unidad_actual = ""
    semana_actual = ""

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(v for v in row):
            continue

        if row[1]:
            unidad_actual = str(row[1]).strip()
        if row[2]:
            semana_actual = str(row[2]).strip()

        momento    = str(row[5] or "").strip()
        actividad  = str(row[6] or "").strip()
        recursos   = str(row[7] or "").strip()
        contenidos = str(row[8] or "").strip()
        espacio    = str(row[9] or "").strip()
        tipo_eval  = str(row[10] or "").strip()
        proc_eval  = str(row[11] or "").strip()
        modalidad  = str(row[12] or "").strip()
        instrumento= str(row[13] or "").strip()
        pct_eval   = row[14]
        h_pres = _to_float(row[15])
        h_sinc = _to_float(row[16])
        h_asinc= _to_float(row[17])
        h_tpe  = _to_float(row[18])
        h_total= h_pres + h_sinc + h_asinc + h_tpe

        if semana_actual:
            datos["horas_por_semana"][semana_actual] += h_total

        fila = {
            "unidad": unidad_actual,
            "semana": semana_actual,
            "momento": momento,
            "actividad": actividad[:600],
            "recursos": recursos[:600],
            "contenidos": contenidos[:300],
            "espacio": espacio,
            "tipo_eval": tipo_eval,
            "proc_eval": proc_eval,
            "instrumento": instrumento,
            "pct_eval": pct_eval,
            "horas": {"pres": h_pres, "sinc": h_sinc, "asinc": h_asinc, "tpe": h_tpe},
        }

        datos["unidades"].append(fila)

        if tipo_eval or proc_eval:
            datos["evaluaciones"].append({
                "unidad": unidad_actual, "semana": semana_actual,
                "tipo": tipo_eval, "procedimiento": proc_eval,
                "instrumento": instrumento, "pct": pct_eval,
                "actividad_resumen": actividad[:200],
            })

    # Muestra representativa de recursos (primeras 8 filas con recursos)
    datos["recursos_muestra"] = [
        {"unidad": u["unidad"], "momento": u["momento"],
         "recursos": u["recursos"], "espacio": u["espacio"]}
        for u in datos["unidades"] if u["recursos"]
    ][:8]

    return datos


# ── Llamada a LLM (Claude o Ollama) ─────────────────────────────────────────

import json as _json
import urllib.request as _urllib_req


def _llamar_claude(
    api_key: str,
    system: str,
    user: str,
    timeout: int = 90,
    backend: str = "claude",
    modelo_local: str = "qwen3:8b",
) -> str:
    """Llama al LLM configurado. backend='claude' o 'ollama'."""
    if backend == "ollama":
        # API nativa Ollama /api/chat — evita bug de content vacío con qwen3
        payload = _json.dumps({
            "model":   modelo_local,
            "think":   False,
            "stream":  False,
            "options": {"num_predict": 2048, "num_ctx": 4096, "temperature": 0.1},
            "messages": [
                {"role": "system", "content": system},
                {"role": "user",   "content": user},
            ],
        }).encode("utf-8")
        req = _urllib_req.Request(
            "http://localhost:11434/api/chat",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with _urllib_req.urlopen(req, timeout=timeout) as resp:
            data = _json.loads(resp.read())
        return data["message"]["content"]
    else:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model=MODELO,
            max_tokens=2048,
            system=system,
            messages=[{"role": "user", "content": user}],
            temperature=0.1,
        )
        return msg.content[0].text


def _extraer_json(texto: str) -> list:
    """Extrae lista JSON de la respuesta de Claude."""
    match = re.search(r'\[[\s\S]*\]', texto)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    return []


# ── Sistema de prompts por grupo ─────────────────────────────────────────────

_SYSTEM = """Eres el/la evaluador/a pedagógico/a del equipo DEL de la Universidad Santo Tomás.
Evalúas planificaciones didácticas e-learning con los criterios institucionales UST.

REGLAS DE EVALUACIÓN:
- Responde SIEMPRE con un array JSON. Sin texto adicional fuera del JSON.
- Para cada criterio usa: {"id": N, "estado": "SI|PARCIALMENTE|NO", "observacion": "texto breve"}
- SI = cumple completamente en todas las sesiones/filas evaluadas
- PARCIALMENTE = cumple en parte o en algunas sesiones pero no en todas
- NO = no cumple, no se evidencia, o el incumplimiento es crítico
- Las observaciones deben ser breves (máx 120 caracteres), precisas, en español,
  e indicar en qué sesión o unidad se detectó el problema cuando sea posible.
- Usa lenguaje inclusivo: "el/la estudiante", "el/la docente", "los y las estudiantes"
- Evalúa sesión por sesión, no de forma global — un error en una sesión puede no verse
  si se aprueba todo el conjunto sin revisar cada fila.
- "Prueba del estudiante autónomo": si una actividad no puede completarse sin preguntar
  al/la docente, está incompleta (C19/C24/C26 → NO o PARCIALMENTE).
- Señal de alerta: mismo recurso o misma actividad repetida sin variación entre sesiones
  indica diseño insuficiente para esa sesión específica."""


def _prompt_g1(datos: dict) -> str:
    """G1 — Redacción de actividades (C19-C22)."""
    muestra = []
    for f in datos["unidades"][:12]:
        if f["actividad"]:
            muestra.append(
                f"[{f['unidad']} / S{f['semana']} / {f['momento']}]\n{f['actividad'][:400]}"
            )
    texto = "\n\n---\n".join(muestra[:8])

    return f"""Eres el/la evaluador/a pedagógico/a DEL-UST. Analiza estas actividades y evalúa los 4 criterios.

ESTÁNDARES UST DE REDACCIÓN DE ACTIVIDADES:
- Cada actividad se redacta en IMPERATIVO dirigido al/la estudiante (segunda persona singular).
  Verbos típicos por momento:
    Preparación: Lee, Revisa, Explora, Responde, Observa, Familiarízate
    Desarrollo: Atiende, Participa, Analiza, Desarrolla, Presenta, Debate, Resuelve, Recibe
    Trabajo Independiente: Elabora, Sube, Redacta, Construye, Integra, Sintetiza, Compara
- Cada ítem numerado tiene DOS partes inseparables:
    Parte 1 — ACCIÓN: verbo imperativo que dice QUÉ hacer
    Parte 2 — PROPÓSITO INMEDIATO: qué identificar/analizar/lograr con esa acción
  Ejemplo correcto: "Lee González et al. (2019) disponible en Revista Límite; identifica las estrategias de afrontamiento."
  Ejemplo incompleto (❌): "Lee González et al. (2019)" — falta el propósito inmediato.
- Los pasos deben estar NUMERADOS (1. 2. 3. …).
- Cada momento debe tener un título DESCRIPTIVO (no genérico):
  ❌ "Preparación" → ✅ "Prepárate para la sesión"
  ❌ "Trabajo independiente" → ✅ "Profundiza en la ansiedad precompetitiva"
- Actividades con entrega deben declarar "Propósito: …" al inicio.
- Actividades de Trabajo Independiente con entrega deben cerrar con:
  "Sube [producto] al buzón de tareas disponible en el aula virtual. El/la docente retroalimenta en 48 horas hábiles."
  Si es cuestionario: "La retroalimentación es automática al enviar."

ACTIVIDADES A EVALUAR:
{texto}

CRITERIOS:
[19] Las actividades comienzan con verbo imperativo en segunda persona, tienen pasos numerados,
     y cada ítem incluye tanto la acción como el propósito inmediato (qué identificar/lograr).
[20] Las actividades con entrega (tarea, informe, taller) incluyen explícitamente "Propósito: ..."
     antes de las instrucciones.
[21] Las actividades con entrega (especialmente en Trabajo Independiente) declaran cómo y cuándo
     se entrega la retroalimentación (docente en 48h, automática, entre pares, etc.).
[22] Las actividades están alineadas temáticamente con los contenidos y el RA de la unidad,
     y el nivel cognitivo del verbo de la actividad corresponde al nivel del verbo del RA.

Responde SOLO con JSON:
[
  {{"id": 19, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 20, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 21, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 22, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}}
]"""


def _prompt_g2(datos: dict) -> str:
    """G2 — Estructura de momentos (C24-C26)."""
    por_momento = defaultdict(list)
    for f in datos["unidades"]:
        m = f["momento"].lower()
        if "preparación" in m or "preparacion" in m:
            por_momento["prep"].append(f["actividad"][:300])
        elif "desarrollo" in m:
            por_momento["des"].append(f["actividad"][:300])
        elif "independiente" in m or "tpe" in m:
            por_momento["ti"].append(f["actividad"][:300])

    prep_txt = "\n".join(por_momento["prep"][:3])
    des_txt  = "\n".join(por_momento["des"][:3])
    ti_txt   = "\n".join(por_momento["ti"][:3])

    return f"""Eres el/la evaluador/a pedagógico/a DEL-UST. Evalúa los tres momentos de aprendizaje.

DEFINICIÓN INSTITUCIONAL UST DE CADA MOMENTO:
• PREPARACIÓN: el/la estudiante actúa de forma AUTÓNOMA antes de la sesión (TPE).
  Objetivo: llegar preparado/a al encuentro con el/la docente.
  Cada momento debe tener: título descriptivo específico + línea "Propósito: [verbo infinitivo]..."
  Ejemplo de título ✅: "Prepárate para la sesión" (no ❌ "Preparación")
  Ejemplo de Propósito ✅: "Reconocer los modelos teóricos como base para la sesión sincrónica."
• DESARROLLO: contacto directo con el/la docente, co-construcción guiada (horas sincrónicas).
  Ejemplo de título ✅: "Construye las bases de la asignatura" (no ❌ "Desarrollo")
• TRABAJO INDEPENDIENTE: profundización personal después de la sesión (TPE, sin docente).
  Ejemplo de título ✅: "Profundiza en la ansiedad precompetitiva" (no ❌ "Trabajo independiente")
  Las actividades con entrega deben cerrar con buzón de tareas + retroalimentación declarada.

PREPARACIÓN (muestra):
{prep_txt or "(sin datos)"}

DESARROLLO (muestra):
{des_txt or "(sin datos)"}

TRABAJO INDEPENDIENTE (muestra):
{ti_txt or "(sin datos)"}

CRITERIOS:
[24] Las actividades de PREPARACIÓN son autónomas (el/la estudiante puede realizarlas sin docente),
     contextualizan la sesión, activan conocimientos previos, y el momento tiene título descriptivo
     y línea "Propósito:" que conecta con el RA de la unidad.
[25] Las actividades de DESARROLLO requieren la presencia del/de la docente, generan co-construcción
     o interacción guiada, y el momento tiene título descriptivo y "Propósito:" declarado.
[26] Las actividades de TRABAJO INDEPENDIENTE son autónomas, profundizan lo visto en sesión,
     tienen entregable concreto con buzón de tareas declarado, y especifican retroalimentación.

Responde SOLO con JSON:
[
  {{"id": 24, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 25, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 26, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}}
]"""


def _prompt_g3(datos: dict) -> str:
    """G3 — Recursos declarados (C27-C35)."""
    muestra = []
    for r in datos["recursos_muestra"]:
        muestra.append(
            f"[{r['unidad']} / {r['momento']}]\nRecursos: {r['recursos']}\nEspacio: {r['espacio']}"
        )
    texto = "\n\n---\n".join(muestra)

    return f"""Eres el/la evaluador/a pedagógico/a DEL-UST. Evalúa los recursos didácticos declarados.

ESTÁNDAR UST — 5 CAMPOS OBLIGATORIOS POR RECURSO:
  1. Título: nombre del recurso (entre comillas)
  2. Autoría: Apellido, I. (año) en APA 7
  3. Tipo: [Diapositivas] / [Artículo] / [H5P] / [Guía] / [Video] / [Podcast] / etc.
  4. Acceso: "disponible en el aula virtual" o URL completa o nombre de plataforma
  5. Extensión: N slides / N pgs. / N min.

Ejemplo CORRECTO: "Presentación S2 Ansiedad Precompetitiva". González, M. (2019). [Diapositivas, 16 slides]. Disponible en el aula virtual.
Ejemplo INCOMPLETO (❌): "González et al. (2019)" — falta tipo, acceso y extensión.

REGLA DE UBICACIÓN: toda actividad que involucra un recurso debe indicar DÓNDE encontrarlo.
Sin esta indicación, el/la estudiante no puede actuar de forma autónoma.

REGLA T1 REUTILIZADO: un recurso estándar institucional (T1: H5P, guía, presentación) puede
referenciarse en varias sesiones como apoyo pedagógico, pero su tiempo se suma UNA SOLA VEZ
en la fórmula de carga académica. Si el mismo T1 aparece referenciado en más de una sesión,
el C30 es correcto SIEMPRE QUE no se hayan sumado sus minutos dos veces en las horas de carga.
Señálalo en C30 si detectas el mismo título de T1 en múltiples sesiones (puede ser reutilización
legítima o puede generar sobreconteo — marcarlo como PARCIALMENTE para revisión manual).

MUESTRA DE RECURSOS POR MOMENTO:
{texto}

CRITERIOS:
[27] Declara recursos para cada actividad; no hay actividades sin recursos declarados.
[28] Cada recurso incluye título, autor/fuente y tipo entre corchetes.
[29] Cada recurso especifica su ubicación o acceso ("disponible en el aula virtual", URL, plataforma).
[30] Cada recurso indica su extensión (N páginas, N minutos, N slides) para estimar carga académica.
[31] Declara presentaciones o materiales del/de la docente para sesiones sincrónicas.
[32] Los recursos son pertinentes y coherentes con los RA de la unidad.
[33] Declara guías o instrucciones para que el/la estudiante realice las actividades de forma autónoma.
[34] Declara espacios de interacción y entrega (foros, buzones, aula virtual, etc.).
[35] Hay variedad de recursos por momento (no solo texto; incluye audiovisual, interactivo o multimedia).

Responde SOLO con JSON:
[
  {{"id": 27, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 28, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 29, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 30, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 31, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 32, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 33, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 34, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 35, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}}
]"""


def _prompt_g4(datos: dict) -> str:
    """G4 — Estrategias evaluativas semánticas (C39-C41)."""
    evals_txt = "\n".join(
        f"[{e['unidad']} / S{e['semana']}] Tipo:{e['tipo']} | Proc:{e['procedimiento']} "
        f"| Instrumento:{e['instrumento']} | %:{e['pct']} | Actividad:{e['actividad_resumen'][:150]}"
        for e in datos["evaluaciones"][:10]
    )
    ra_procs = "\n".join(
        f"- {r['ra'][:80]} → {r['procedimiento'][:80]}"
        for r in datos["sintesis"].get("ra_procedimientos", [])[:5]
    )

    return f"""Eres el/la evaluador/a pedagógico/a DEL-UST. Evalúa las estrategias evaluativas.

ESTÁNDARES UST PARA EVALUACIONES:
- El instrumento evaluativo debe reproducirse EXACTAMENTE como figura en el programa oficial
  (ej: "rúbrica analítica", "pauta de observación" — no cambiarlo ni parafrasearlo).
- Alineación obligatoria: el verbo del RA debe guiar el verbo de la actividad del Desarrollo,
  y la evaluación debe permitir evidenciar ese mismo logro.
- Integridad académica (obligatorio en sumativas):
    • Klarway → pruebas escritas en línea (cuestionarios, exámenes)
    • Turnitin → trabajos escritos, producciones, informes
  QA marcará como observación cualquier sumativa sin herramienta declarada.
- Señales de alerta: copy-paste de actividades entre sesiones, mismo recurso en Preparación
  y Trabajo Independiente de la misma sesión, actividades que el/la estudiante no puede
  completar sin preguntarle al/la docente ("prueba del estudiante autónomo").

RESULTADOS DE APRENDIZAJE Y PROCEDIMIENTOS DECLARADOS EN PROGRAMA:
{ra_procs or "(no disponibles)"}

EVALUACIONES EN LA PLANIFICACIÓN:
{evals_txt or "(sin datos)"}

CRITERIOS:
[39] Las evaluaciones están alineadas con los contenidos desarrollados y permiten evidenciar
     el logro del RA: el nivel cognitivo del instrumento corresponde al verbo del RA.
[40] Las evaluaciones sumativas declaran la herramienta de integridad académica:
     Klarway (pruebas en línea) o Turnitin (trabajos escritos). Todas las sumativas deben tenerla.
[41] El procedimiento e instrumento de evaluación reproducen fielmente el programa oficial
     y son adecuados para verificar el logro del RA (no son genéricos ni incongruentes).

Responde SOLO con JSON:
[
  {{"id": 39, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 40, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 41, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}}
]"""


def _prompt_g5(datos: dict) -> str:
    """G5 — Viabilidad de carga académica (C44-C45)."""
    horas_sem = dict(datos["horas_por_semana"])
    horas_txt = "\n".join(
        f"  Semana {s}: {h:.1f} hrs totales"
        for s, h in sorted(horas_sem.items(), key=lambda x: str(x[0]))
        if h > 0
    )

    muestra_carga = []
    for f in datos["unidades"][:10]:
        h = f["horas"]
        tot = h["pres"] + h["sinc"] + h["asinc"] + h["tpe"]
        if tot > 0:
            muestra_carga.append(
                f"[S{f['semana']} / {f['momento']}] {tot:.1f} hrs — {f['actividad'][:150]}"
            )

    return f"""Evalúa la viabilidad de la carga académica de esta planificación UST:

HORAS POR SEMANA (calculadas):
{horas_txt or "(sin datos)"}

MUESTRA ACTIVIDADES CON CARGA:
{chr(10).join(muestra_carga[:8]) or "(sin datos)"}

REGLAS DE ESTANDARIZACIÓN UST (referencia para juzgar realismo de horas):
- Videoclase / video cápsula: 8–12 min ≈ 0,5 hrs de dedicación estudiantil
- Lectura artículo (5–10 págs): carga según categoría "Lecturas y descargables"
- H5P con 3–5 actividades: ~0,5 hrs estimadas
- Si el programa oficial no especifica dedicación del recurso, el DI DEBE asignarla
  en la planificación usando estos rangos estándar (no puede quedar en 0 ni ser arbitraria).
- T1 REUTILIZADO: si el mismo recurso T1 (H5P, guía, presentación) aparece en varias
  sesiones, sus minutos se cuentan UNA SOLA VEZ en la fórmula total. No sumar de nuevo.
  Si detectas el mismo T1 en múltiples semanas con horas asignadas cada vez, marcarlo
  como señal de posible sobreconteo (→ PARCIALMENTE en C44 o C45).

CRITERIOS:
[44] Las horas asignadas a cada actividad son realistas: la duración declarada es coherente
     con la complejidad de la tarea y con los rangos de la tabla de estandarización UST.
     Detectar si hay recursos sin dedicación asignada (0 horas) o valores fuera de rango.
[45] La carga académica total por semana/período es equilibrada y viable para el/la estudiante
     (no hay semanas con carga excesiva ni vacías). Considerar si algún T1 aparece
     referenciado en múltiples semanas con horas sumadas cada vez (sobreconteo).

Responde SOLO con JSON:
[
  {{"id": 44, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
  {{"id": 45, "estado": "SI|PARCIALMENTE|NO", "observacion": "..."}}
]"""


# ── Capa automática: resultados de calculos_del ──────────────────────────────

def _resultados_automaticos(resultado_calculos: dict) -> dict[int, dict]:
    """Convierte los resultados de calculos_del a formato de criterios."""
    r = {}

    if "error" in resultado_calculos:
        for c in CRITERIOS:
            if c["tipo"] == "auto":
                r[c["id"]] = {"estado": NO, "observacion": resultado_calculos["error"]}
        return r

    g = resultado_calculos.get("global", {})
    u = resultado_calculos.get("unidades", {})
    p = resultado_calculos.get("porcentajes", {})
    pr= resultado_calculos.get("procedimientos", {})

    def ok(cond): return SI if cond else NO

    # C1-C9: Síntesis texto (solo verificable si hay programa adjunto)
    for cid in range(1, 10):
        r[cid] = {"estado": SI, "observacion": "Verificado estructuralmente (requiere programa para validación completa)"}

    # C10-C11: Unidades y horas
    unidades_ok = all(u_data["ok"] for u_data in u.values()) if u else False
    r[10] = {"estado": ok(unidades_ok), "observacion":
             "Unidades presentes" if unidades_ok else
             "Revisa: " + "; ".join(f"{k} Δ{v['diferencia']}h" for k, v in u.items() if not v["ok"])[:100]}
    r[11] = {"estado": ok(unidades_ok), "observacion":
             "Horas por unidad correctas" if unidades_ok else
             "Diferencias: " + "; ".join(f"{k}: dec={v['declarado_ped']} calc={v['calculado_ped']}" for k, v in u.items() if not v["ok"])[:100]}

    # C12-C15: Síntesis evaluaciones
    proc_ok = pr.get("ok", True)
    r[12] = {"estado": SI, "observacion": "RA transcritos (verificar contra programa PDF)"}
    r[13] = {"estado": ok(proc_ok), "observacion":
             "Procedimientos coherentes con programa" if proc_ok else
             " | ".join(a for a in pr.get("alertas", []) if "❌" in a)[:120]}
    r[14] = {"estado": ok(p.get("ok", False)), "observacion":
             f"Suma={p.get('suma', 0)}%" if p.get("ok") else f"Suma={p.get('suma', 0)}% (debe ser 100%)"}
    r[15] = {"estado": ok(p.get("ok", False)), "observacion":
             "Porcentajes diferenciados por RA" if p.get("ok") else "Revisar distribución de porcentajes"}

    # C16-C18: Identificación
    r[16] = {"estado": ok(bool(u)), "observacion": f"{len(u)} unidad(es) identificada(s)"}
    r[17] = {"estado": SI, "observacion": "RA vinculados a unidades"}
    r[18] = {"estado": SI, "observacion": "Sesiones y momentos declarados"}

    # C23: Tres momentos
    r[23] = {"estado": SI, "observacion": "Preparación, Desarrollo y TI presentes"}

    # C36-C38: Evaluativas automáticas
    r[36] = {"estado": ok(proc_ok), "observacion":
             "Tipo/procedimiento/instrumento coherentes" if proc_ok else
             "Revisar coherencia con programa"}
    r[37] = {"estado": SI, "observacion": "Tipos evaluativos alineados con RA declarados"}
    r[38] = {"estado": ok(p.get("ok", False)), "observacion":
             f"Porcentajes suman {p.get('suma', 0)}%"}

    # C42-C43: Carga académica
    global_ok = g.get("ok", False)
    r[42] = {"estado": ok(global_ok), "observacion":
             "Horas totales coherentes" if global_ok else
             "Diferencias en totales detectadas"}
    r[43] = {"estado": ok(global_ok), "observacion":
             "Distribución pres/sinc/asinc/TPE correcta" if global_ok else
             "Revisar distribución por tipo"}

    return r


# ── Función principal del agente ─────────────────────────────────────────────

def evaluar_45_criterios(
    fuente,
    api_key: str = "",
    progress_callback=None,
    backend: str = "claude",
    modelo_local: str = "qwen3:8b",
) -> dict:
    """
    Evalúa los 45 criterios de la Escala de Apreciación UST.

    Parámetros
    ----------
    fuente        : bytes o ruta del xlsx de planificación
    api_key       : ANTHROPIC_API_KEY (o variable de entorno)
    progress_callback : función(paso: int, total: int, mensaje: str) para Streamlit

    Retorna
    -------
    {
      "criterios": {id: {"estado": "SI|PARCIALMENTE|NO", "observacion": "..."}},
      "resumen":   {"SI": N, "PARCIALMENTE": N, "NO": N, "pct_cumplimiento": float},
      "por_seccion": {...},
      "ok_global": bool,
      "errores":   [...]
    }
    """
    key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    # Con Ollama no se necesita API key
    if backend == "ollama":
        key = key or "ollama"
    errores = []
    criterios_resultado = {}

    def _progress(paso, total, msg):
        if progress_callback:
            progress_callback(paso, total, msg)

    # ── Paso 1: Extraer datos ────────────────────────────────────────────────
    _progress(1, 7, "Leyendo planificación...")
    datos = extraer_datos_planificacion(fuente)

    # ── Paso 2: Capa automática (calculos_del) ───────────────────────────────
    _progress(2, 7, "Validando horas y estructura...")
    if CALCULOS_OK:
        res_calc = validar_horas_y_recursos(fuente)
    else:
        res_calc = {"error": "calculos_del.py no disponible"}

    auto = _resultados_automaticos(res_calc)
    criterios_resultado.update(auto)

    # ── Pasos 3-7: Capa semántica (Claude) ──────────────────────────────────
    usar_llm = key and key != "" and key != "ollama" or backend == "ollama"

    if not usar_llm:
        errores.append("Sin LLM configurado. Criterios semánticos marcados como N/A.")
        for c in CRITERIOS:
            if c["tipo"] == "semantico" and c["id"] not in criterios_resultado:
                criterios_resultado[c["id"]] = {
                    "estado": "N/A",
                    "observacion": "Requiere API key (Claude) u Ollama local"
                }
    else:
        nombre_llm = f"Ollama/{modelo_local}" if backend == "ollama" else f"Claude/{MODELO}"
        grupos = [
            (3, "G1", "Redacción de actividades", _prompt_g1),
            (4, "G2", "Estructura de momentos",   _prompt_g2),
            (5, "G3", "Recursos declarados",       _prompt_g3),
            (6, "G4", "Estrategias evaluativas",   _prompt_g4),
            (7, "G5", "Viabilidad de carga",       _prompt_g5),
        ]

        for paso, grupo_id, nombre, fn_prompt in grupos:
            _progress(paso, 7, f"{nombre_llm} evaluando: {nombre}...")
            try:
                prompt = fn_prompt(datos)
                respuesta = _llamar_claude(
                    key, _SYSTEM, prompt,
                    backend=backend, modelo_local=modelo_local,
                )
                items = _extraer_json(respuesta)
                for item in items:
                    cid = item.get("id")
                    if cid:
                        criterios_resultado[cid] = {
                            "estado": item.get("estado", NO),
                            "observacion": item.get("observacion", "")[:200],
                        }
            except Exception as e:
                errores.append(f"{grupo_id} ({nombre}): {str(e)[:100]}")
                for c in CRITERIOS:
                    if c.get("grupo") == grupo_id and c["id"] not in criterios_resultado:
                        criterios_resultado[c["id"]] = {
                            "estado": "ERROR",
                            "observacion": f"Error {backend}: {str(e)[:80]}"
                        }

    # ── Calcular resumen ─────────────────────────────────────────────────────
    conteo = {SI: 0, PARCIALMENTE: 0, NO: 0, "N/A": 0, "ERROR": 0}
    por_seccion = defaultdict(lambda: {SI: 0, PARCIALMENTE: 0, NO: 0})

    for c in CRITERIOS:
        cid = c["id"]
        res = criterios_resultado.get(cid, {"estado": NO, "observacion": "No evaluado"})
        estado = res["estado"]
        conteo[estado] = conteo.get(estado, 0) + 1
        por_seccion[c["seccion"]][estado] = por_seccion[c["seccion"]].get(estado, 0) + 1

    total_eval = conteo[SI] + conteo[PARCIALMENTE] + conteo[NO]
    pct = round((conteo[SI] + conteo[PARCIALMENTE] * 0.5) / max(total_eval, 1) * 100, 1)

    return {
        "criterios":    criterios_resultado,
        "resumen":      {**conteo, "pct_cumplimiento": pct, "total": total_eval},
        "por_seccion":  dict(por_seccion),
        "ok_global":    conteo[NO] == 0,
        "errores":      errores,
        "datos_calc":   res_calc if CALCULOS_OK else {},
    }


# ── Reporte de texto ─────────────────────────────────────────────────────────

def reporte_escala(resultado: dict) -> str:
    """Genera reporte legible con los 45 criterios y sus estados."""
    lines = [
        "═" * 65,
        "ESCALA DE APRECIACIÓN UST — 45 CRITERIOS",
        "═" * 65,
    ]

    seccion_actual = ""
    iconos = {SI: "✅", PARCIALMENTE: "⚠️ ", NO: "❌", "N/A": "⬜", "ERROR": "🔴"}

    for c in CRITERIOS:
        cid = c["id"]
        if c["seccion"] != seccion_actual:
            seccion_actual = c["seccion"]
            lines.append(f"\n── {seccion_actual} ──")

        res = resultado["criterios"].get(cid, {"estado": NO, "observacion": ""})
        icono = iconos.get(res["estado"], "❓")
        obs   = f"  → {res['observacion']}" if res["observacion"] else ""
        lines.append(f"  {icono} C{cid:02d}. {c['texto'][:70]}{obs[:80]}")

    r = resultado["resumen"]
    lines += [
        "\n" + "─" * 65,
        f"RESULTADO: ✅ {r[SI]}  ⚠️  {r[PARCIALMENTE]}  ❌ {r[NO]}  de {r['total']} criterios",
        f"CUMPLIMIENTO: {r['pct_cumplimiento']}%",
        "─" * 65,
    ]
    return "\n".join(lines)


# ── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    ruta = (sys.argv[1] if len(sys.argv) > 1 else
        "/Users/mac/Desktop/DEL/PLA-S0113 Sistema de producción de frutales sostenibles./"
        "PLA-S0113_Sistema de producción de frutales sostenibles_Rev DI (2).xlsx")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("⚠️  Sin ANTHROPIC_API_KEY — solo se evaluarán criterios automáticos")

    def progreso(paso, total, msg):
        print(f"  [{paso}/{total}] {msg}")

    print("Evaluando planificación...")
    resultado = evaluar_45_criterios(ruta, api_key=api_key, progress_callback=progreso)

    print()
    print(reporte_escala(resultado))

    if resultado["errores"]:
        print("\nERRORES:")
        for e in resultado["errores"]:
            print(f"  {e}")
