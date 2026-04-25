"""
app_del.py — Procesador de Planificaciones DEL · UST 2026-1
Interfaz Streamlit para el equipo de Dirección de Educación a Distancia.

Ejecutar localmente:
    streamlit run app_del.py

Desplegar en Streamlit Cloud:
    1. Subir este archivo + revisar_planificaciones.py + requirements.txt a GitHub
    2. Conectar el repositorio en share.streamlit.io
"""

import streamlit as st
import streamlit.components.v1 as components
import sys, os, glob, shutil, tempfile, re
from io import BytesIO

import validar_planificacion as vp
import dict_ust
import apa_recursos as apa
import apa_llm

# ── Cruce con programa (independiente del resto de módulos avanzados) ─────────
try:
    from cruce_programa import (
        extraer_programa_pdf as cruce_extraer_pdf,
        cruzar_con_planificacion,
    )
    CRUCE_OK = True
except ImportError:
    CRUCE_OK = False
    def cruce_extraer_pdf(*a, **kw):
        return {"_error": "Módulo cruce_programa no disponible"}
    def cruzar_con_planificacion(*a, **kw):
        return [], []

# ── Conversor ProgramaOficial → dict (formato esperado por revisar_planificaciones) ──
def _prog_a_dict(prog) -> dict:
    """
    Convierte un objeto ProgramaOficial (cruce_programa) al formato dict
    que espera revisar_planificaciones.py (verificar_horas, verificar_contra_programa).
    Si prog ya es dict o es None, lo devuelve tal cual.
    """
    if prog is None:
        return {}
    if isinstance(prog, dict):
        return prog
    try:
        horas     = getattr(prog, "horas", {}) or {}
        unidades  = []
        for u in getattr(prog, "unidades", []):
            unidades.append({
                "numero": u.get("num", ""),
                "nombre": u.get("nombre", ""),
                "horas":  u.get("horas", 0),
            })
        creditos = getattr(prog, "creditos", None)
        try:
            creditos = int(creditos) if creditos else None
        except (ValueError, TypeError):
            pass
        return {
            "codigo":             getattr(prog, "codigo", None),
            "asignatura":         getattr(prog, "nombre_asignatura", None),
            "creditos":           creditos,
            "area":               getattr(prog, "area_ocde", None),
            "total_pedagogicas":  horas.get("total"),
            "horas_tpe":          horas.get("tpe"),
            "unidades":           unidades,
            "ponderaciones":      {},
            "pct_examen":         None,
        }
    except Exception:
        return {}

# ── Módulos DEL avanzados (escala + reescritura) ──────────────────────────────
try:
    from agente_criterios import evaluar_45_criterios, reporte_escala, CRITERIOS
    from calculos_del import validar_horas_y_recursos, reporte_texto as calculos_reporte
    from reescritura_llm import reescribir_planificacion
    ESCALA_OK = True
except ImportError as _ei:
    ESCALA_OK = False
    CRITERIOS = []

# ── Configuración de página ───────────────────────────────────────────────
st.set_page_config(
    page_title="Procesador DEL — UST",
    page_icon="📋",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Cargar script revisor ─────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
try:
    import revisar_planificaciones as rp
    SCRIPT_OK = True
except ImportError:
    SCRIPT_OK = False

# ── Estilos institucionales UST ───────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ══════════════════════════════════════════════
       VARIABLES
    ══════════════════════════════════════════════ */
    :root {
        --g0: #002b14;  --g1: #004d26;  --g2: #006633;
        --g3: #009450;  --g4: #00c46a;
        --gp: #E8F5EE;  --gm: #F0F7F4;  --gb: #C8E6D4;
        --gold:  #C8A951;  --gold-d: #9a7a2a;
        --blue:  #2E5FA3;  --blue-l: #EBF2FB;
        --purp:  #6B3FA0;  --purp-l: #F0EAFB;
        --teal:  #007A7A;  --teal-l: #E6F5F5;
        --slate: #4A6080;  --slate-l: #EDF0F4;
        --amber: #C8A030;  --amber-l: #FBF3E0;
        --indg:  #4C3AA0;  --indg-l: #EDE9FB;
        --red:   #CC3333;  --red-l:  #FDDCDC;
        --text:  #1a2332;  --text-m: #4a5568;  --text-l: #718096;
        --bg:    #F4F8F5;
        --white: #FFFFFF;
        --r-sm: 8px;  --r-md: 12px;  --r-lg: 16px;  --r-xl: 20px;
        --sh-xs: 0 1px 3px rgba(0,0,0,0.08);
        --sh-sm: 0 2px 8px rgba(0,0,0,0.10);
        --sh-md: 0 4px 16px rgba(0,0,0,0.10);
        --sh-lg: 0 8px 32px rgba(0,102,51,0.12);
        --t: 0.18s ease;
    }

    /* ══════════════════════════════════════════════
       BASE
    ══════════════════════════════════════════════ */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, sans-serif !important;
        color: var(--text);
    }
    .stApp {
        background: linear-gradient(150deg, #E8F4EE 0%, #EEF6F2 50%, #F5FAF7 100%);
    }
    .block-container {
        max-width: 880px;
        padding: 1.2rem 2.4rem 2.4rem;
        background: var(--white);
        border-radius: var(--r-xl);
        box-shadow: var(--sh-lg);
        border: 1px solid rgba(0,102,51,0.06);
    }

    /* ══════════════════════════════════════════════
       CABECERA
    ══════════════════════════════════════════════ */
    .del-header {
        background: linear-gradient(135deg, var(--g0) 0%, var(--g1) 35%, var(--g2) 65%, var(--g3) 100%);
        border-radius: var(--r-lg);
        padding: 1.5rem 1.8rem 1.3rem;
        margin-bottom: 1.2rem;
        position: relative;
        overflow: hidden;
        box-shadow: 0 6px 24px rgba(0,77,38,0.28);
    }
    .del-header::before {
        content: '';
        position: absolute; inset: 0;
        background: url("data:image/svg+xml,%3Csvg width='80' height='80' viewBox='0 0 80 80' xmlns='http://www.w3.org/2000/svg'%3E%3Cg fill='%23ffffff' fill-opacity='0.03'%3E%3Ccircle cx='40' cy='40' r='36'/%3E%3Ccircle cx='0' cy='0' r='20'/%3E%3Ccircle cx='80' cy='80' r='20'/%3E%3C/g%3E%3C/svg%3E") repeat;
        pointer-events: none;
    }
    .del-header::after {
        content: '';
        position: absolute;
        bottom: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, var(--gold) 0%, rgba(200,169,81,0) 100%);
    }
    .del-header h2 {
        color: #FFFFFF !important;
        margin: 0 0 0.15rem 0;
        font-size: 1.5rem;
        font-weight: 800;
        letter-spacing: -0.02em;
        text-shadow: 0 1px 4px rgba(0,0,0,0.25);
    }
    .del-header .sub {
        color: rgba(255,255,255,0.78);
        font-size: 0.82rem;
        margin: 0 0 0.7rem 0;
        font-weight: 400;
        letter-spacing: 0.01em;
    }
    .del-header .badges { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }
    .del-header .badge {
        display: inline-flex; align-items: center; gap: 5px;
        background: rgba(255,255,255,0.13);
        border: 1px solid rgba(255,255,255,0.28);
        border-radius: 20px;
        padding: 4px 12px;
        font-size: 0.73rem;
        font-weight: 600;
        color: #FFFFFF;
        letter-spacing: 0.04em;
        backdrop-filter: blur(6px);
    }
    .del-header .badge-gold {
        background: rgba(200,169,81,0.25);
        border-color: rgba(200,169,81,0.50);
    }

    /* ══════════════════════════════════════════════
       BOTÓN NUEVA (top-right)
    ══════════════════════════════════════════════ */
    button[data-testid="baseButton-secondary"] {
        border-radius: var(--r-md) !important;
        font-weight: 600 !important;
        font-size: 0.82rem !important;
        letter-spacing: 0.02em !important;
        transition: all var(--t) !important;
        border: 1.5px solid #C8D8DC !important;
        color: var(--slate) !important;
        background: var(--white) !important;
        padding: 0.4rem 0.9rem !important;
    }
    button[data-testid="baseButton-secondary"]:hover {
        border-color: var(--g2) !important;
        color: var(--g2) !important;
        box-shadow: var(--sh-sm) !important;
        transform: translateY(-1px) !important;
    }

    /* ══════════════════════════════════════════════
       BOTONES PRIMARIOS
    ══════════════════════════════════════════════ */
    button[data-testid="baseButton-primary"] {
        border-radius: var(--r-md) !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
        letter-spacing: 0.01em !important;
        padding: 0.65rem 1.4rem !important;
        transition: all var(--t) !important;
        box-shadow: 0 3px 10px rgba(0,0,0,0.18) !important;
        border: none !important;
        text-transform: none !important;
    }
    button[data-testid="baseButton-primary"]:not(:disabled):hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 18px rgba(0,0,0,0.22) !important;
        filter: brightness(1.06) !important;
    }
    button[data-testid="baseButton-primary"]:not(:disabled):active {
        transform: translateY(0) !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.18) !important;
    }
    button[data-testid="baseButton-primary"]:disabled {
        opacity: 0.42 !important;
        cursor: not-allowed !important;
        transform: none !important;
        box-shadow: none !important;
        filter: grayscale(30%) !important;
    }

    /* ══════════════════════════════════════════════
       TABS
    ══════════════════════════════════════════════ */
    .stTabs [data-baseweb="tab-list"] {
        gap: 5px;
        background: #EDF1F5;
        border-radius: var(--r-md);
        padding: 5px;
        border: 1px solid #D4DCE6;
        flex-wrap: wrap;
        box-shadow: inset 0 1px 3px rgba(0,0,0,0.06);
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: var(--r-sm) !important;
        padding: 0.5rem 1rem !important;
        font-weight: 500 !important;
        font-size: 0.81rem !important;
        transition: all var(--t) !important;
        border: 1px solid transparent !important;
        white-space: nowrap !important;
        min-width: 0 !important;
    }
    .stTabs [aria-selected="true"] { font-weight: 700 !important; }

    /* Tab colors */
    .stTabs [data-baseweb="tab"]:nth-child(1)                          { color:#004d26; background:#DCF0E4; border-color:#B8DBC8; }
    .stTabs [data-baseweb="tab"]:nth-child(1)[aria-selected="true"]    { background:var(--g2) !important; color:#fff !important; border-color:var(--g2) !important; box-shadow:0 2px 10px rgba(0,102,51,0.35) !important; }
    .stTabs [data-baseweb="tab"]:nth-child(1):hover:not([aria-selected="true"]) { background:#C6E8D0; border-color:var(--g2); }

    .stTabs [data-baseweb="tab"]:nth-child(2)                          { color:#1a3a6b; background:var(--blue-l); border-color:#C0D4EE; }
    .stTabs [data-baseweb="tab"]:nth-child(2)[aria-selected="true"]    { background:var(--blue) !important; color:#fff !important; border-color:var(--blue) !important; box-shadow:0 2px 10px rgba(46,95,163,0.35) !important; }
    .stTabs [data-baseweb="tab"]:nth-child(2):hover:not([aria-selected="true"]) { background:#D0E4F8; border-color:var(--blue); }

    .stTabs [data-baseweb="tab"]:nth-child(3)                          { color:#4a2680; background:var(--purp-l); border-color:#D0C0EC; }
    .stTabs [data-baseweb="tab"]:nth-child(3)[aria-selected="true"]    { background:var(--purp) !important; color:#fff !important; border-color:var(--purp) !important; box-shadow:0 2px 10px rgba(107,63,160,0.35) !important; }
    .stTabs [data-baseweb="tab"]:nth-child(3):hover:not([aria-selected="true"]) { background:#E4D8F5; border-color:var(--purp); }

    .stTabs [data-baseweb="tab"]:nth-child(4)                          { color:#005050; background:var(--teal-l); border-color:#A8D8D8; }
    .stTabs [data-baseweb="tab"]:nth-child(4)[aria-selected="true"]    { background:var(--teal) !important; color:#fff !important; border-color:var(--teal) !important; box-shadow:0 2px 10px rgba(0,122,122,0.32) !important; }
    .stTabs [data-baseweb="tab"]:nth-child(4):hover:not([aria-selected="true"]) { background:#C0E4E4; border-color:var(--teal); }

    .stTabs [data-baseweb="tab"]:nth-child(5)                          { color:#5a3a00; background:var(--amber-l); border-color:#E8D090; }
    .stTabs [data-baseweb="tab"]:nth-child(5)[aria-selected="true"]    { background:var(--amber) !important; color:#fff !important; border-color:var(--amber) !important; box-shadow:0 2px 10px rgba(200,160,48,0.32) !important; }
    .stTabs [data-baseweb="tab"]:nth-child(5):hover:not([aria-selected="true"]) { background:#F0E0B0; border-color:var(--amber); }

    .stTabs [data-baseweb="tab"]:nth-child(6)                          { color:#2a1860; background:var(--indg-l); border-color:#C0B4F0; }
    .stTabs [data-baseweb="tab"]:nth-child(6)[aria-selected="true"]    { background:var(--indg) !important; color:#fff !important; border-color:var(--indg) !important; box-shadow:0 2px 10px rgba(76,58,160,0.35) !important; }
    .stTabs [data-baseweb="tab"]:nth-child(6):hover:not([aria-selected="true"]) { background:#D8D0F8; border-color:var(--indg); }

    /* ══════════════════════════════════════════════
       MÉTRICAS
    ══════════════════════════════════════════════ */
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #F3FAF6 0%, #EBF6EF 100%);
        border: 1px solid var(--gb);
        border-top: 3px solid var(--g2);
        border-radius: var(--r-md);
        padding: 1rem 1.1rem 0.9rem;
        box-shadow: var(--sh-xs);
        transition: box-shadow var(--t);
    }
    div[data-testid="stMetric"]:hover { box-shadow: var(--sh-sm); }
    div[data-testid="stMetricLabel"] {
        color: var(--g1) !important; font-weight: 700 !important;
        font-size: 0.72rem !important; text-transform: uppercase !important;
        letter-spacing: 0.06em !important;
    }
    div[data-testid="stMetricValue"] {
        color: var(--g2) !important; font-weight: 800 !important;
        font-size: 1.9rem !important; line-height: 1.2 !important;
    }
    div[data-testid="stMetricDelta"] { font-size: 0.76rem !important; font-weight: 500 !important; }

    /* ══════════════════════════════════════════════
       EXPANDERS
    ══════════════════════════════════════════════ */
    div[data-testid="stExpander"] {
        border: 1px solid var(--gb) !important;
        border-radius: var(--r-md) !important;
        box-shadow: var(--sh-xs) !important;
        transition: box-shadow var(--t) !important;
        overflow: visible !important;
    }
    div[data-testid="stExpander"]:hover { box-shadow: var(--sh-sm) !important; }
    div[data-testid="stExpander"] summary {
        background: linear-gradient(90deg, var(--gm), #F8FCF9) !important;
        color: var(--g1) !important;
        font-weight: 600 !important;
        font-size: 0.88rem !important;
        padding: 0.7rem 1rem !important;
        border-radius: var(--r-md) !important;
        transition: background var(--t) !important;
        cursor: pointer !important;
    }
    div[data-testid="stExpander"] summary:hover { background: var(--gp) !important; }
    div[data-testid="stExpander"] summary p { font-weight: 600 !important; }

    /* ══════════════════════════════════════════════
       ALERTAS
    ══════════════════════════════════════════════ */
    div[data-testid="stAlert"] {
        border-radius: var(--r-md) !important;
        border-left-width: 4px !important;
        font-size: 0.87rem !important;
        box-shadow: var(--sh-xs) !important;
    }

    /* ══════════════════════════════════════════════
       FILE UPLOADER
    ══════════════════════════════════════════════ */
    div[data-testid="stFileUploader"] {
        border: 2px dashed var(--g3) !important;
        border-radius: var(--r-md) !important;
        padding: 0.8rem !important;
        background: linear-gradient(135deg, #F5FBF7, #F0F8F4) !important;
        transition: all var(--t) !important;
    }
    div[data-testid="stFileUploader"]:hover {
        border-color: var(--g2) !important;
        background: var(--gp) !important;
        box-shadow: var(--sh-sm) !important;
    }
    div[data-testid="stFileUploader"] label {
        font-weight: 600 !important;
        color: var(--g1) !important;
    }

    /* ══════════════════════════════════════════════
       PROGRESS BAR
    ══════════════════════════════════════════════ */
    div[data-testid="stProgressBar"] > div > div {
        background: linear-gradient(90deg, var(--g2), var(--g3)) !important;
        border-radius: 99px !important;
    }

    /* ══════════════════════════════════════════════
       INPUTS / SELECTBOX
    ══════════════════════════════════════════════ */
    div[data-baseweb="select"] > div {
        border-radius: var(--r-sm) !important;
        border-color: var(--gb) !important;
        font-size: 0.88rem !important;
        transition: border-color var(--t) !important;
    }
    div[data-baseweb="select"] > div:focus-within {
        border-color: var(--g2) !important;
        box-shadow: 0 0 0 3px rgba(0,102,51,0.12) !important;
    }
    div[data-baseweb="input"] > div {
        border-radius: var(--r-sm) !important;
        border-color: var(--gb) !important;
        font-size: 0.88rem !important;
    }
    div[data-baseweb="textarea"] {
        border-radius: var(--r-sm) !important;
        border-color: var(--gb) !important;
        font-size: 0.87rem !important;
    }

    /* ══════════════════════════════════════════════
       CHECKBOXES
    ══════════════════════════════════════════════ */
    .stCheckbox label {
        color: var(--g1) !important;
        font-weight: 500 !important;
        font-size: 0.88rem !important;
    }
    .stCheckbox label:hover { color: var(--g2) !important; }

    /* ══════════════════════════════════════════════
       DATAFRAME
    ══════════════════════════════════════════════ */
    div[data-testid="stDataFrame"] {
        border-radius: var(--r-md) !important;
        overflow: hidden !important;
        box-shadow: var(--sh-xs) !important;
        border: 1px solid var(--gb) !important;
    }

    /* ══════════════════════════════════════════════
       DIVIDERS
    ══════════════════════════════════════════════ */
    hr {
        border: none !important;
        border-top: 1px solid var(--gb) !important;
        margin: 1.4rem 0 !important;
        opacity: 0.8 !important;
    }

    /* ══════════════════════════════════════════════
       CAPTIONS / TEXTO
    ══════════════════════════════════════════════ */
    .stCaption, [data-testid="stCaptionContainer"] {
        color: var(--text-m) !important;
        font-size: 0.81rem !important;
    }

    /* ══════════════════════════════════════════════
       SECTION CARD — contenedor de sección
    ══════════════════════════════════════════════ */
    .sec-card {
        background: var(--white);
        border: 1px solid var(--gb);
        border-radius: var(--r-md);
        padding: 1.1rem 1.3rem;
        margin-bottom: 1rem;
        box-shadow: var(--sh-xs);
    }
    .sec-title {
        font-size: 0.72rem; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.08em; color: var(--text-m);
        margin: 0 0 0.7rem 0; display: flex; align-items: center; gap: 6px;
    }
    .sec-title::after {
        content: ''; flex: 1; height: 1px; background: var(--gb);
    }

    /* ══════════════════════════════════════════════
       TAGS
    ══════════════════════════════════════════════ */
    .tag-ok {
        display: inline-flex; align-items: center; gap: 4px;
        background: #D4EDDA; color: #155724;
        border-radius: 6px; padding: 3px 10px;
        font-size: 0.79rem; font-weight: 600;
        border: 1px solid #B8DFBF;
    }
    .tag-error {
        display: inline-flex; align-items: center; gap: 4px;
        background: #FDDCDC; color: #7A0000;
        border-radius: 6px; padding: 3px 10px;
        font-size: 0.79rem; font-weight: 600;
        border: 1px solid #F5B8B8;
    }
    .tag-warn {
        display: inline-flex; align-items: center; gap: 4px;
        background: #FFF3CD; color: #6B4A00;
        border-radius: 6px; padding: 3px 10px;
        font-size: 0.79rem; font-weight: 600;
        border: 1px solid #ECCC7A;
    }

    /* ══════════════════════════════════════════════
       VISION / ANÁLISIS VISUAL
    ══════════════════════════════════════════════ */
    .vision-card {
        background: linear-gradient(135deg, #EDE9FB 0%, #F3F0FD 100%);
        border: 1px solid #C0B4F0;
        border-top: 3px solid var(--indg);
        border-radius: var(--r-lg);
        padding: 1.3rem 1.5rem;
        margin-bottom: 1.2rem;
        box-shadow: var(--sh-sm);
    }
    .vision-card h4 {
        color: var(--indg); margin: 0 0 0.4rem 0;
        font-size: 1.05rem; font-weight: 700;
    }
    .vision-result {
        background: #FAFBFC;
        border: 1px solid #E4E8EE;
        border-radius: var(--r-md);
        padding: 1.1rem 1.3rem;
        font-size: 0.875rem;
        line-height: 1.7;
        color: var(--text);
        white-space: pre-wrap;
        box-shadow: inset 0 1px 4px rgba(0,0,0,0.04);
        max-height: 520px;
        overflow-y: auto;
    }
    .model-badge {
        display: inline-flex; align-items: center; gap: 6px;
        background: linear-gradient(90deg, var(--indg), #6B52C8);
        color: #fff; border-radius: 20px; padding: 5px 14px;
        font-size: 0.75rem; font-weight: 700; letter-spacing: 0.04em;
        box-shadow: 0 2px 8px rgba(76,58,160,0.28);
    }

    /* ══════════════════════════════════════════════
       SPINNER — override color
    ══════════════════════════════════════════════ */
    div[data-testid="stSpinner"] > div {
        border-top-color: var(--g2) !important;
    }

    /* ══════════════════════════════════════════════
       MARKDOWN headings dentro de tabs
    ══════════════════════════════════════════════ */
    .stMarkdown h3 {
        color: var(--g1);
        font-size: 1.05rem;
        font-weight: 700;
        margin: 1.2rem 0 0.5rem;
        padding-bottom: 0.3rem;
        border-bottom: 2px solid var(--gb);
    }
    .stMarkdown h4 {
        color: var(--text);
        font-size: 0.93rem;
        font-weight: 600;
        margin: 0.8rem 0 0.3rem;
    }
</style>
""", unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ═════════════════════════════════════════════════════════════════════════

def parsear_log(log: list[str]) -> dict:
    """Extrae métricas clave del log para mostrar en el resumen."""
    resultado = {
        "total_correcciones": 0,
        "criterios_ok": 0, "criterios_error": 0, "criterios_manual": 0,
        "discrepancias_prog": 0,
        "as_ok": 0, "as_error": 0, "as_manual": 0,
        "tiene_as": False,
        "correcciones_detalle": [],
        "discrepancias_detalle": [],
        "as_detalle": [],
        "lt_errores": 0,
        "lt_revisadas": 0,
        "lt_correcciones": 0,
        "lt_detalle": [],
        "lt_ejecutado": False,
    }
    seccion = None
    for linea in log:
        # Total correcciones
        if "TOTAL" in linea and "correcciones" in linea:
            m = re.search(r"TOTAL\s*:\s*(\d+)", linea)
            if m:
                resultado["total_correcciones"] = int(m.group(1))

        # Resultado escala UST
        m = re.search(r"Resultado:\s*(\d+)✅\s*(\d+)❌\s*(\d+)⚠️\s*manual", linea)
        if m:
            resultado["criterios_ok"]     = int(m.group(1))
            resultado["criterios_error"]  = int(m.group(2))
            resultado["criterios_manual"] = int(m.group(3))

        # Discrepancias vs programa
        m = re.search(r"Verificación vs programa:\s*(\d+)", linea)
        if m:
            resultado["discrepancias_prog"] = int(m.group(1))

        # Resultado A+S
        m = re.search(r"Resultado A\+Se:\s*(\d+)✅\s*(\d+)❌\s*(\d+)⚠️\s*manual", linea)
        if m:
            resultado["tiene_as"]  = True
            resultado["as_ok"]     = int(m.group(1))
            resultado["as_error"]  = int(m.group(2))
            resultado["as_manual"] = int(m.group(3))

        # LanguageTool — resumen (nuevo formato con correcciones)
        m = re.search(r"LanguageTool:\s*(\d+)\s*celda\(s\)\s*revisadas?,\s*(\d+)\s*error\(es?\)(?:,\s*(\d+)\s*correcci.+?\s*aplicada\(s\))?", linea)
        if m:
            resultado["lt_revisadas"] = int(m.group(1))
            resultado["lt_errores"]   = int(m.group(2))
            resultado["lt_correcciones"] = int(m.group(3)) if m.group(3) else 0
            resultado["lt_ejecutado"] = True

        # Sección actual para detalles
        if "[Planificación por unidades]" in linea or "[Síntesis didáctica]" in linea:
            seccion = "correcciones"
        elif "[Verificación contra programa" in linea:
            seccion = "programa"
        elif "[Verificación A+Se" in linea:
            seccion = "as"
        elif "LanguageTool es" in linea or "autocorrección" in linea:
            seccion = "lt"

        # Líneas de detalle
        stripped = linea.strip()
        if seccion == "correcciones" and stripped.startswith("[Plan") and "→" in stripped:
            resultado["correcciones_detalle"].append(stripped)
        elif seccion == "programa" and stripped and stripped[0] in ("✅", "❌", "⚠"):
            resultado["discrepancias_detalle"].append(stripped)
        elif seccion == "as" and stripped and stripped[0] in ("✅", "❌", "⚠"):
            resultado["as_detalle"].append(stripped)
        elif seccion == "lt" and stripped and stripped[0] in ("✅", "⚠", "ℹ"):
            resultado["lt_detalle"].append(stripped)

    return resultado


def _datos_desde_xlsx(xlsx_bytes: bytes) -> dict:
    """Lee código y nombre de asignatura desde la hoja Síntesis didáctica del Excel."""
    try:
        import openpyxl as _oxl
        wb = _oxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = next(
            (wb[s] for s in wb.sheetnames
             if 'síntesis' in s.lower() or 'sintesis' in s.lower()),
            None
        )
        if ws is None:
            return {}
        codigo     = str(ws.cell(7, 1).value or "").strip()
        asignatura = str(ws.cell(4, 4).value or "").strip()
        wb.close()
        return {"codigo": codigo or None, "asignatura": asignatura or None}
    except Exception:
        return {}


def _nombre_descarga(programa: dict | None, instancia: int,
                     xlsx_bytes: bytes | None = None) -> str:
    """
    Genera el nombre del archivo de descarga:
      {CODIGO}_{Nombre_Asignatura}_Revisado_I{N}.xlsx
    Primero lee del dict programa (PDF); si falta algo usa el Excel como fallback.
    """
    prog = _prog_a_dict(programa)
    # Fallback al Excel si el programa no tiene código o nombre
    if xlsx_bytes and (not prog.get("codigo") or not prog.get("asignatura")):
        xlsx_datos = _datos_desde_xlsx(xlsx_bytes)
        prog.setdefault("codigo",     xlsx_datos.get("codigo"))
        prog.setdefault("asignatura", xlsx_datos.get("asignatura"))

    codigo = (prog.get("codigo") or "").strip()
    nombre = (prog.get("asignatura") or "").strip()
    nombre_limpio = re.sub(r'[\\/:*?"<>|]', '', nombre).strip()
    nombre_limpio = re.sub(r'\s+', '_', nombre_limpio)
    partes = [p for p in [codigo, nombre_limpio] if p]
    base   = "_".join(partes) if partes else "Planificacion"
    return f"{base}_Revisado_I{instancia}.xlsx"


def tag(texto: str, tipo: str) -> str:
    css = {"ok": "tag-ok", "error": "tag-error", "warn": "tag-warn"}.get(tipo, "tag-warn")
    return f'<span class="{css}">{texto}</span>'


# ═════════════════════════════════════════════════════════════════════════
#  CABECERA
# ═════════════════════════════════════════════════════════════════════════

components.html("""
<script>
(function() {
  var DIAS  = ['domingo','lunes','martes','miércoles','jueves','viernes','sábado'];
  var MESES = ['enero','febrero','marzo','abril','mayo','junio',
               'julio','agosto','septiembre','octubre','noviembre','diciembre'];
  var pdoc = window.parent.document;

  function tick() {
    var n = new Date();
    var dia   = DIAS[n.getDay()];
    var fecha = dia.charAt(0).toUpperCase()+dia.slice(1)+' '+
                n.getDate()+' de '+MESES[n.getMonth()]+' '+n.getFullYear();
    var hora  = String(n.getHours()).padStart(2,'0')+':'+
                String(n.getMinutes()).padStart(2,'0')+':'+
                String(n.getSeconds()).padStart(2,'0');
    var dh = pdoc.getElementById('reloj-hora');
    var df = pdoc.getElementById('reloj-fecha');
    if (dh) dh.textContent = hora;
    if (df) df.textContent = fecha;
  }
  tick();
  setInterval(tick, 1000);
})();
</script>
""", height=0)

st.markdown("""
<div class="del-header">
  <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:1rem;">
    <div>
      <h2 style="margin:0;">Procesador de Planificaciones DEL</h2>
      <p class="sub">Universidad Santo Tomás &nbsp;·&nbsp; Dirección de Educación a Distancia</p>
      <div class="badges">
        <span class="badge badge-gold">✦ Semestre 2026-1</span>
        <span class="badge">🤖 IA local</span>
        <span class="badge">📋 UST</span>
      </div>
    </div>
    <div style="text-align:right;min-width:140px;">
      <div id="reloj-hora" style="font-size:2rem;font-weight:800;color:#fff;
           letter-spacing:0.05em;line-height:1;text-shadow:0 2px 10px rgba(0,0,0,0.30);
           font-variant-numeric:tabular-nums;">
        --:--:--
      </div>
      <div id="reloj-fecha" style="font-size:0.78rem;font-weight:500;color:rgba(212,240,228,0.90);
           margin-top:5px;line-height:1.4;">
        cargando...
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Colores y estilos de botones vía JS ──────────────────────────────────────
components.html("""
<script>
(function () {
    /* bg normal · bg hover · shadow color */
    var COLOR_MAP = [
        { label: 'Revisar planificación',             bg:'#006633', hov:'#004d26', sh:'rgba(0,102,51,0.38)' },
        { label: 'Aplicar correcciones (Instancia 2)',bg:'#2E5FA3', hov:'#1d4480', sh:'rgba(46,95,163,0.38)' },
        { label: 'Aplicar correcciones (Instancia 3)',bg:'#6B3FA0', hov:'#522f80', sh:'rgba(107,63,160,0.38)' },

        { label: 'Agregar al diccionario',            bg:'#4A6080', hov:'#354860', sh:'rgba(74,96,128,0.38)' },
        { label: 'Analizar documento',                bg:'#4C3AA0', hov:'#3a2b80', sh:'rgba(76,58,160,0.38)' },
        { label: 'Nueva',                             bg:'#52606D', hov:'#3a4450', sh:'rgba(82,96,109,0.32)' },
    ];
    var done = new WeakSet();

    function styleBtn(btn, c) {
        Object.assign(btn.style, {
            backgroundColor: c.bg,
            borderColor:     c.bg,
            color:           '#ffffff',
            borderRadius:    '10px',
            fontWeight:      '700',
            fontSize:        '0.92rem',
            letterSpacing:   '0.01em',
            padding:         '0.6rem 1.3rem',
            boxShadow:       '0 3px 10px ' + c.sh,
            transition:      'all 0.18s ease',
            outline:         'none',
        });
        btn.addEventListener('mouseenter', function () {
            if (btn.disabled) return;
            btn.style.backgroundColor = c.hov;
            btn.style.borderColor     = c.hov;
            btn.style.boxShadow       = '0 5px 16px ' + c.sh;
            btn.style.transform       = 'translateY(-2px)';
        });
        btn.addEventListener('mouseleave', function () {
            btn.style.backgroundColor = c.bg;
            btn.style.borderColor     = c.bg;
            btn.style.boxShadow       = '0 3px 10px ' + c.sh;
            btn.style.transform       = 'translateY(0)';
        });
        btn.addEventListener('mousedown',  function () { btn.style.transform = 'translateY(0)'; });
        done.add(btn);
    }

    function paint() {
        try {
            var doc = window.parent.document;
            doc.querySelectorAll('button').forEach(function (btn) {
                if (done.has(btn)) return;
                var txt = (btn.innerText || btn.textContent || '').trim();
                COLOR_MAP.forEach(function (c) {
                    if (txt.indexOf(c.label) !== -1) styleBtn(btn, c);
                });
            });
        } catch(e) { /* cross-origin — silencioso */ }
    }

    paint();
    new MutationObserver(paint).observe(
        window.parent.document.body,
        { childList: true, subtree: true }
    );
})();
</script>
""", height=0, scrolling=False)

if not SCRIPT_OK:
    st.error("No se encontró `revisar_planificaciones.py` en la misma carpeta que esta app.")
    st.stop()

# ── Botón de limpieza ─────────────────────────────────────────────────────
_col_sp, _col_btn = st.columns([6, 1])
with _col_btn:
    if st.button("🔄 Nueva", help="Limpia todos los archivos y resultados para procesar otra planificación", use_container_width=True):
        st.session_state.clear()
        st.rerun()

# ═════════════════════════════════════════════════════════════════════════
#  NAVEGACIÓN PRINCIPAL
# ═════════════════════════════════════════════════════════════════════════

# Indicador visual del flujo de trabajo
st.markdown("""
<div style="display:flex;align-items:stretch;gap:0;margin:0.4rem 0 1.4rem;
            background:linear-gradient(135deg,#F2F8F4,#EBF5EE);
            border-radius:14px;padding:0;border:1px solid #C8E6D4;
            overflow:hidden;box-shadow:0 2px 8px rgba(0,102,51,0.08);">

  <div style="flex:1;text-align:center;padding:0.85rem 0.4rem;
              border-right:1px solid #C8E6D4;position:relative;">
    <div style="font-size:1.5rem;line-height:1;margin-bottom:4px">📋</div>
    <div style="font-size:0.68rem;font-weight:800;color:#006633;letter-spacing:0.06em;text-transform:uppercase">Revisar</div>
    <div style="font-size:0.60rem;color:#5a8a6a;margin-top:2px">Sube archivos</div>
    <div style="position:absolute;top:50%;right:-8px;transform:translateY(-50%);
                width:15px;height:15px;background:#C8E6D4;border-radius:50%;
                display:flex;align-items:center;justify-content:center;
                font-size:0.55rem;color:#006633;font-weight:700;z-index:1;
                border:2px solid #EBF5EE">▶</div>
  </div>

  <div style="flex:1;text-align:center;padding:0.85rem 0.4rem;
              border-right:1px solid #C8E6D4;position:relative;">
    <div style="font-size:1.5rem;line-height:1;margin-bottom:4px">📝</div>
    <div style="font-size:0.68rem;font-weight:800;color:#2E5FA3;letter-spacing:0.06em;text-transform:uppercase">Corregir</div>
    <div style="font-size:0.60rem;color:#5a7aaa;margin-top:2px">Revisoras DEL</div>
    <div style="position:absolute;top:50%;right:-8px;transform:translateY(-50%);
                width:15px;height:15px;background:#C5D8F0;border-radius:50%;
                display:flex;align-items:center;justify-content:center;
                font-size:0.55rem;color:#2E5FA3;font-weight:700;z-index:1;
                border:2px solid #EBF5EE">▶</div>
  </div>

  <div style="flex:1;text-align:center;padding:0.85rem 0.4rem;
              border-right:1px solid #C8E6D4;position:relative;">
    <div style="font-size:1.5rem;line-height:1;margin-bottom:4px">🎨</div>
    <div style="font-size:0.68rem;font-weight:800;color:#6B3FA0;letter-spacing:0.06em;text-transform:uppercase">Diseñar</div>
    <div style="font-size:0.60rem;color:#8a6ab0;margin-top:2px">Recursos T1–T4</div>
    <div style="position:absolute;top:50%;right:-8px;transform:translateY(-50%);
                width:15px;height:15px;background:#D8C8F0;border-radius:50%;
                display:flex;align-items:center;justify-content:center;
                font-size:0.55rem;color:#6B3FA0;font-weight:700;z-index:1;
                border:2px solid #EBF5EE">▶</div>
  </div>

  <div style="flex:1;text-align:center;padding:0.85rem 0.4rem;
              border-right:1px solid #C8E6D4;position:relative;">
    <div style="font-size:1.5rem;line-height:1;margin-bottom:4px">🔍</div>
    <div style="font-size:0.68rem;font-weight:800;color:#4C3AA0;letter-spacing:0.06em;text-transform:uppercase">Visual</div>
    <div style="font-size:0.60rem;color:#7060b0;margin-top:2px">Qwen2.5-VL</div>
    <div style="position:absolute;top:50%;right:-8px;transform:translateY(-50%);
                width:15px;height:15px;background:#C9BEF5;border-radius:50%;
                display:flex;align-items:center;justify-content:center;
                font-size:0.55rem;color:#4C3AA0;font-weight:700;z-index:1;
                border:2px solid #EBF5EE">▶</div>
  </div>

  <div style="flex:1;text-align:center;padding:0.85rem 0.4rem;
              background:linear-gradient(135deg,rgba(0,102,51,0.05),rgba(0,148,80,0.08));">
    <div style="font-size:1.5rem;line-height:1;margin-bottom:4px">✅</div>
    <div style="font-size:0.68rem;font-weight:800;color:#006633;letter-spacing:0.06em;text-transform:uppercase">Aprobar</div>
    <div style="font-size:0.60rem;color:#5a8a6a;margin-top:2px">Planificación válida</div>
  </div>

</div>
""", unsafe_allow_html=True)

tab_i1, tab_del, tab_recursos, tab_escala, tab_config, tab_vision = st.tabs([
    "📋 Revisar",
    "📝 Revisoras DEL",
    "🎨 Diseño de Recursos",
    "📊 Escala de Apreciación",
    "⚙️ Herramientas",
    "🔍 Análisis Visual",
])

# Alias para compatibilidad con código existente
tab_i2 = tab_del
tab_i3 = tab_del
tab_dict = tab_config

# ═════════════════════════════════════════════════════════════════════════
#  INSTANCIA 1
# ═════════════════════════════════════════════════════════════════════════

with tab_i1:

    st.markdown("### 1 · Archivos obligatorios")

    col_pdf, col_xlsx = st.columns(2)
    with col_pdf:
        pdf_file = st.file_uploader(
            "📄 Programa de asignatura",
            type="pdf",
            key="i1_pdf",
            help="PDF del programa oficial de la asignatura.",
        )
    with col_xlsx:
        xlsx_file = st.file_uploader(
            "📊 Planificación didáctica",
            type="xlsx",
            key="i1_xlsx",
            help="Archivo Excel con la planificación por unidades.",
        )

    # Mostrar resumen del programa si ya se subió
    if pdf_file:
        with st.spinner("Leyendo programa..."):
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                    tmp_pdf.write(pdf_file.read())
                    tmp_pdf_path = tmp_pdf.name
                pdf_file.seek(0)
                programa = cruce_extraer_pdf(tmp_pdf_path)
                os.unlink(tmp_pdf_path)
                _prog_error = None
            except Exception as _ep:
                programa = None
                _prog_error = str(_ep)[:200]

        if _prog_error:
            st.warning(f"No se pudo leer el PDF: {_prog_error}", icon="⚠️")
            programa = None
        else:
            with st.expander("✅ Programa leído — verificar datos extraídos", expanded=False):
                c1, c2, c3 = st.columns(3)
                c1.metric("Código",   getattr(programa, "codigo", None) or "—")
                c2.metric("Créditos", getattr(programa, "creditos", None) or "—")
                c3.metric("Área",     getattr(programa, "area_ocde", None) or "—")
                unidades = getattr(programa, "unidades", [])
                if unidades:
                    st.markdown("**Unidades:**")
                    for u in unidades:
                        st.markdown(
                            f"- Unidad {u.get('num', '?')}: **{u.get('nombre', '')}** "
                            f"· {u.get('horas', '?')}h pedagógicas"
                        )
    else:
        programa = None

    # ── Validaciones preventivas ──────────────────────────────────────────
    if xlsx_file:
        problemas = vp.validar_xlsx(xlsx_file.getvalue())
        xlsx_file.seek(0)
        errores   = [p for p in problemas if p["nivel"] == "error"]
        advertencias = [p for p in problemas if p["nivel"] == "advertencia"]
        if errores or advertencias:
            with st.expander(
                f"{'🔴' if errores else '🟡'} Validación previa — "
                f"{len(errores)} error(es), {len(advertencias)} advertencia(s)",
                expanded=bool(errores),
            ):
                for p in errores:
                    st.error(f"**[{p['codigo']}]** {p['mensaje']}", icon="🔴")
                for p in advertencias:
                    st.warning(f"**[{p['codigo']}]** {p['mensaje']}", icon="⚠️")
                if errores:
                    st.caption(
                        "Corrige los errores marcados en rojo antes de procesar."
                    )
        else:
            st.success("Archivo validado sin problemas estructurales.", icon="✅")

    st.markdown("### 2 · Opciones")
    st.caption("Solo marcar si aplica a esta asignatura.")

    col_as, col_dec = st.columns(2)
    with col_as:
        es_as = st.checkbox(
            "📌 Lineamiento A+S",
            key="i1_as",
            help="Activa la verificación de los 11 hitos obligatorios "
                 "del lineamiento Aprendizaje + Servicio (UST 2025).",
        )
    with col_dec:
        st.checkbox(
            "📜 Decreto de actualización",
            disabled=True,
            key="i1_dec",
            help="Próximamente: permite subir un decreto que actualice "
                 "campos del programa oficial.",
        )

    usar_lt_i1 = st.checkbox(
        "🔤 Revisión ortográfica y gramatical (LanguageTool)",
        value=True,
        key="i1_lt",
        help="Revisa ortografía y gramática de las celdas de actividad "
             "usando la API gratuita de LanguageTool en español. "
             "Requiere conexión a internet. Aumenta el tiempo de procesamiento.",
    )

    autocorregir_lt_i1 = st.checkbox(
        "✍️ Aplicar correcciones automáticas (conservador)",
        value=True,
        key="i1_autocorr",
        help="Aplica automáticamente las correcciones SEGURAS detectadas por LanguageTool "
             "(solo cambios unívocos: tildes faltantes, errores ortográficos claros). "
             "Los cambios se marcan en azul en el archivo descargado.",
    )

    st.markdown("**Referencias bibliográficas (columna H)**")
    col_apa1, col_apa2 = st.columns(2)
    with col_apa1:
        usar_apa_i1 = st.checkbox(
            "📚 Validar referencias APA 7 (reglas)",
            value=True,
            key="i1_apa",
            help="Verifica que las referencias de la columna H cumplan APA 7: "
                 "formato de autor, año entre paréntesis, & entre autores, "
                 "URL sin 'Disponible en', etc.",
        )
    with col_apa2:
        autocorr_apa_i1 = st.checkbox(
            "✍️ Corregir APA automáticamente",
            value=False,
            key="i1_apa_corr",
            help="Aplica correcciones APA 7 seguras: elimina 'Disponible en', "
                 "'y' → '&' entre autores, punto tras año, etc. "
                 "Los cambios se marcan en azul.",
        )

    # ── Revisión APA con LLM ──────────────────────────────────────────────
    with st.expander("🤖 Revisión APA 7 con LLM (más precisa)", expanded=False):
        st.caption(
            "Usa un modelo de lenguaje para revisar y corregir referencias APA 7 "
            "con mayor profundidad que las reglas automáticas. "
            "Ollama es gratuito y corre localmente — usa la API nativa para mejor compatibilidad con qwen3."
        )

        col_llm1, col_llm2 = st.columns([2, 2])
        with col_llm1:
            usar_llm_i1 = st.checkbox(
                "Activar revisión LLM",
                value=False,
                key="i1_llm_activo",
            )
            backend_i1 = st.selectbox(
                "Backend",
                apa_llm.BACKENDS,
                key="i1_llm_backend",
                help="Ollama = local y gratis. Los demás requieren API key.",
            )
        with col_llm2:
            # Modelo por defecto según backend — clave incluye backend para evitar caché cruzada
            modelo_default = apa_llm.BACKEND_DEFAULTS[backend_i1]["model"]
            modelo_i1 = st.text_input(
                "Modelo",
                value=modelo_default,
                key=f"i1_llm_model_{backend_i1}",
                help=f"Para Ollama texto: {apa_llm.MODELO_TEXTO} · imágenes: {apa_llm.MODELO_VISION}. "
                     "Para otros backends: nombre del modelo de la API.",
            )
            apikey_i1 = st.text_input(
                "API Key",
                value="",
                type="password",
                key="i1_llm_key",
                placeholder="No requerida para Ollama",
                help="Pega aquí tu API key si usas Claude, OpenAI o Grok.",
            )
        autocorr_llm_i1 = st.checkbox(
            "Aplicar correcciones LLM al archivo",
            value=False,
            key="i1_llm_autocorr",
            help="Si está activo, las correcciones del LLM se escriben en la "
                 "columna H del archivo descargado (marcadas en azul).",
        )
        if backend_i1 == "ollama":
            st.info(
                f"Modelo por defecto: **{modelo_default}** (API nativa `/api/chat`, compatible con qwen3). "
                "Si no lo tienes instalado: "
                f"`ollama pull {modelo_default}`",
                icon="💡",
            )

    # ── Escala de Apreciación con IA ──────────────────────────────────────────
    if ESCALA_OK:
        with st.expander("🎯 Escala de Apreciación UST (45 criterios con IA)", expanded=False):
            st.caption(
                "Evalúa los 45 criterios institucionales con IA. "
                "Usa **Ollama local** (gratis) o **Claude API** (mejor calidad)."
            )

            col_e0, col_e1 = st.columns([1, 1])
            with col_e0:
                usar_escala_i1 = st.checkbox(
                    "Activar evaluación",
                    value=False,
                    key="i1_escala_activo",
                )
                reescribir_i1 = st.checkbox(
                    "Reescribir actividades (C19-C21)",
                    value=True,
                    key="i1_reescribir",
                    help="Corrige imperativo, propósito y retroalimentación.",
                )
            with col_e1:
                _backend_opts = ["ollama (local, gratis)", "claude (API)"]
                _backend_sel  = st.selectbox(
                    "Motor IA",
                    _backend_opts,
                    key="i1_escala_backend",
                    help="Ollama corre localmente sin costo. Claude requiere API key.",
                )
                backend_escala_i1 = "ollama" if "ollama" in _backend_sel else "claude"

            if backend_escala_i1 == "ollama":
                _modelo_ollama_default = apa_llm.MODELO_TEXTO
                modelo_escala_i1 = st.text_input(
                    "Modelo Ollama",
                    value=_modelo_ollama_default,
                    key="i1_modelo_ollama_ollama",
                    help=f"Recomendado: {apa_llm.MODELO_TEXTO} — usa API nativa /api/chat (compatible con qwen3). "
                         "No correr junto al modelo de visión (16 GB RAM). "
                         "Ejecuta `ollama list` para ver los disponibles.",
                )
                apikey_escala_i1 = ""
                if usar_escala_i1:
                    st.success(
                        f"Usando **{modelo_escala_i1}** local (API nativa) — sin costo de API.",
                        icon="🦙",
                    )
            else:
                modelo_escala_i1 = ""
                apikey_escala_i1 = st.text_input(
                    "API Key de Anthropic",
                    value=os.environ.get("ANTHROPIC_API_KEY", ""),
                    type="password",
                    key="i1_escala_key",
                    placeholder="sk-ant-api03-...",
                )
                if usar_escala_i1 and not apikey_escala_i1:
                    st.warning("Ingresa la API Key para usar Claude.", icon="🔑")
    else:
        usar_escala_i1   = False
        reescribir_i1    = False
        apikey_escala_i1 = ""
        backend_escala_i1 = "ollama"
        modelo_escala_i1  = apa_llm.MODELO_TEXTO

    st.markdown("### 3 · Revisar")

    listo_i1 = bool(pdf_file and xlsx_file)
    procesar_i1 = st.button(
        "▶ Revisar planificación",
        disabled=not listo_i1,
        type="primary",
        key="btn_i1",
        use_container_width=True,
        help="Sube el PDF y el .xlsx para activar este botón." if not listo_i1 else "",
    )

    if not listo_i1:
        st.caption("⬆ Sube el PDF del programa y la planificación .xlsx para continuar.")

    if procesar_i1 and listo_i1:
        output_bytes     = None
        output_name      = None
        log              = []
        ok               = False
        escala_resultado = None
        reesc_log        = []

        rp._LANGUAGETOOL_ACTIVO = usar_lt_i1
        rp._LANGUAGETOOL_AUTOCORREGIR = autocorregir_lt_i1

        spinner_msg = ("Procesando planificación + revisión ortográfica (puede tardar)..."
                       if usar_lt_i1 else "Procesando planificación...")
        with st.spinner(spinner_msg):
            with tempfile.TemporaryDirectory() as tmp:
                carpeta_asig = os.path.join(tmp, "asig")
                carpeta_env  = os.path.join(carpeta_asig, "Enviado a DEL")
                os.makedirs(carpeta_env)

                xlsx_path = os.path.join(carpeta_env, xlsx_file.name)
                with open(xlsx_path, "wb") as f:
                    f.write(xlsx_file.getvalue())

                if not programa:
                    try:
                        pdf_path = os.path.join(tmp, pdf_file.name)
                        with open(pdf_path, "wb") as f:
                            f.write(pdf_file.getvalue())
                        programa = cruce_extraer_pdf(pdf_path)
                    except Exception:
                        programa = None

                log, ok = rp.procesar_asignatura(
                    carpeta_asig,
                    programa=_prog_a_dict(programa),
                    es_as=es_as,
                )

                if ok:
                    salidas = glob.glob(os.path.join(carpeta_asig, "Revisado", "*.xlsx"))
                    if salidas:
                        # ── APA 7 en columna H ──────────────────────────
                        if usar_apa_i1:
                            import openpyxl as _oxl
                            _wb_apa = _oxl.load_workbook(salidas[0])
                            if "Planificación por unidades" in _wb_apa.sheetnames:
                                _ws_apa = _wb_apa["Planificación por unidades"]
                                _apa_log, _apa_probs, _apa_corr = \
                                    apa.revisar_columna_recursos(
                                        _ws_apa,
                                        autocorregir=autocorr_apa_i1,
                                    )
                                log.extend(_apa_log)
                                if autocorr_apa_i1 and _apa_corr:
                                    _wb_apa.save(salidas[0])
                            _wb_apa.close()
                        # ── APA con LLM ──────────────────────────────
                        if usar_llm_i1:
                            import openpyxl as _oxl2
                            _wb_llm = _oxl2.load_workbook(salidas[0])
                            if "Planificación por unidades" in _wb_llm.sheetnames:
                                _ws_llm = _wb_llm["Planificación por unidades"]
                                _llm_log, _llm_probs, _llm_corr = \
                                    apa_llm.revisar_columna_recursos_llm(
                                        _ws_llm,
                                        backend=backend_i1,
                                        model=modelo_i1 or None,
                                        api_key=apikey_i1,
                                        autocorregir=autocorr_llm_i1,
                                    )
                                log.extend(_llm_log)
                                if autocorr_llm_i1 and _llm_corr:
                                    _wb_llm.save(salidas[0])
                            _wb_llm.close()
                        # ────────────────────────────────────────────────
                        with open(salidas[0], "rb") as f:
                            output_bytes = f.read()
                        output_name = _nombre_descarga(programa, 1, output_bytes)

                # ── Escala de Apreciación (45 criterios) ──────────────────
                _escala_activa = (
                    ESCALA_OK and usar_escala_i1 and output_bytes and
                    (backend_escala_i1 == "ollama" or apikey_escala_i1)
                )
                if _escala_activa:
                    _spinner_llm = (f"Ollama/{modelo_escala_i1}" if backend_escala_i1 == "ollama"
                                    else "Claude Sonnet")
                    with st.spinner(f"{_spinner_llm} evaluando Escala de Apreciación (45 criterios)…"):
                        def _escala_prog(p, t, msg): pass
                        try:
                            escala_resultado = evaluar_45_criterios(
                                output_bytes,
                                api_key=apikey_escala_i1,
                                progress_callback=_escala_prog,
                                backend=backend_escala_i1,
                                modelo_local=modelo_escala_i1 or apa_llm.MODELO_TEXTO,
                            )
                            st.session_state["escala_resultado"] = escala_resultado
                            st.session_state["escala_archivo"]   = output_name
                        except Exception as _e_escala:
                            escala_resultado = None
                            log.append(f"⚠️  Escala: error — {str(_e_escala)[:100]}")

                    # ── Reescritura semántica si corresponde ───────────────
                    if reescribir_i1 and escala_resultado and output_bytes:
                        _rw_necesario = any(
                            escala_resultado.get("criterios", {}).get(cid, {}).get("estado", "NO")
                            in ("NO", "PARCIALMENTE")
                            for cid in (19, 20, 21)
                        )
                        if _rw_necesario:
                            with st.spinner(f"{_spinner_llm} reescribiendo actividades (C19-C21)…"):
                                try:
                                    output_bytes, reesc_log = reescribir_planificacion(
                                        output_bytes,
                                        escala_resultado,
                                        api_key=apikey_escala_i1,
                                        backend=backend_escala_i1,
                                        modelo_local=modelo_escala_i1 or apa_llm.MODELO_TEXTO,
                                    )
                                except Exception as _e_rw:
                                    reesc_log = [f"❌ Error reescritura: {str(_e_rw)[:100]}"]

        st.divider()
        st.markdown("### Resultados")

        if not ok or not output_bytes:
            st.error("El procesamiento falló. Revisa el log.")
        else:
            metricas = parsear_log(log)
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(
                    "Correcciones aplicadas",
                    metricas["total_correcciones"],
                    help="Celdas modificadas automáticamente (marcadas en azul).",
                )
            with col2:
                total_criterios = (metricas["criterios_ok"] + metricas["criterios_error"]
                                   + metricas["criterios_manual"])
                st.metric(
                    "Criterios UST",
                    f"{metricas['criterios_ok']}/{total_criterios}",
                    delta=f"{metricas['criterios_error']} con error"
                          if metricas["criterios_error"] else "sin errores",
                    delta_color="inverse" if metricas["criterios_error"] else "normal",
                )
            with col3:
                st.metric(
                    "Discrepancias vs programa",
                    metricas["discrepancias_prog"],
                    delta="revisar" if metricas["discrepancias_prog"] else "todo coincide",
                    delta_color="inverse" if metricas["discrepancias_prog"] else "normal",
                )

            # ── Métricas APA ────────────────────────────────────────────
            if usar_apa_i1:
                _apa_probs_log = sum(
                    1 for l in log
                    if "[APA_" in l and ("❌" in l or "⚠️" in l)
                )
                _apa_corr_log = sum(
                    1 for l in log if l.strip().startswith("✏️") and "Fila" in l
                )
                col_a1, col_a2 = st.columns(2)
                with col_a1:
                    st.metric(
                        "Problemas APA 7 (col. H)",
                        _apa_probs_log,
                        delta="revisar referencias" if _apa_probs_log else "referencias OK",
                        delta_color="inverse" if _apa_probs_log else "normal",
                        help="Errores y advertencias detectados en columna H.",
                    )
                with col_a2:
                    st.metric(
                        "Correcciones APA aplicadas",
                        _apa_corr_log,
                        help="Cambios APA 7 seguros aplicados en columna H.",
                    )
                # Detalle APA en expander
                _apa_detalle = [
                    l for l in log
                    if "APA" in l or ("Fila" in l and "H" in l and "✏️" in l)
                ]
                if _apa_detalle:
                    with st.expander(
                        f"📚 Detalle APA 7 — columna H ({_apa_probs_log} problema(s))",
                        expanded=_apa_probs_log > 0,
                    ):
                        for linea in _apa_detalle:
                            st.markdown(linea)

            # ── Métricas LLM ─────────────────────────────────────────────
            if usar_llm_i1:
                _llm_probs_log = sum(
                    1 for l in log if "APA 7 LLM" not in l
                    and "❌" in l and "LLM" not in l
                    and "[APA_" not in l
                )
                _llm_corr_log = sum(
                    1 for l in log
                    if l.strip().startswith("✅") and "corrección(es) aplicada" in l
                )
                # Leer directamente del log LLM
                for l in reversed(log):
                    m = re.search(r'APA 7 LLM: (\d+) problema', l)
                    if m:
                        _llm_probs_log = int(m.group(1))
                        break
                for l in reversed(log):
                    m = re.search(r'APA 7 LLM:.*?(\d+) corrección', l)
                    if m:
                        _llm_corr_log = int(m.group(1))
                        break

                col_l1, col_l2 = st.columns(2)
                with col_l1:
                    st.metric(
                        f"Problemas APA — LLM ({backend_i1})",
                        _llm_probs_log,
                        delta="revisar" if _llm_probs_log else "referencias OK",
                        delta_color="inverse" if _llm_probs_log else "normal",
                    )
                with col_l2:
                    st.metric("Correcciones LLM aplicadas", _llm_corr_log)

                _llm_detalle = [
                    l for l in log
                    if "LLM" in l or ("Fila" in l and "corrección(es) aplicada" in l)
                ]
                if _llm_detalle:
                    with st.expander(
                        f"🤖 Detalle APA LLM ({_llm_probs_log} problema(s))",
                        expanded=_llm_probs_log > 0,
                    ):
                        for linea in _llm_detalle:
                            st.markdown(linea)

            if metricas["tiene_as"]:
                st.markdown("**Verificación A+Se:**")
                total_as = metricas["as_ok"] + metricas["as_error"] + metricas["as_manual"]
                st.progress(
                    metricas["as_ok"] / total_as if total_as else 0,
                    text=f"{metricas['as_ok']}✅  {metricas['as_error']}❌  "
                         f"{metricas['as_manual']}⚠️ manual  (de {total_as} hitos)",
                )

            if metricas["lt_ejecutado"]:
                lt_err = metricas["lt_errores"]
                lt_rev = metricas["lt_revisadas"]
                lt_corr = metricas["lt_correcciones"]
                delta_text = f"{lt_rev} celda(s) revisadas"
                if lt_corr > 0:
                    delta_text += f", {lt_corr} corregidas"
                st.metric(
                    "Errores ortográficos/gramaticales",
                    lt_err,
                    delta=delta_text,
                    delta_color="off" if lt_corr == 0 else "normal",
                    help="LanguageTool API (es). Las correcciones automáticas se aplican solo a cambios seguros.",
                )
                if metricas["lt_detalle"]:
                    with st.expander(
                        f"🔤 Detalle ortografía/gramática ({lt_err} error(es), {lt_corr} corregidos)",
                        expanded=lt_err > 0 or lt_corr > 0,
                    ):
                        for linea in metricas["lt_detalle"]:
                            st.markdown(linea)

            if metricas["discrepancias_detalle"]:
                with st.expander("📊 Detalle — verificación contra programa", expanded=True):
                    for linea in metricas["discrepancias_detalle"]:
                        st.markdown(linea)

            if metricas["tiene_as"] and metricas["as_detalle"]:
                with st.expander("📌 Detalle — hitos A+Se",
                                 expanded=metricas["as_error"] > 0):
                    for linea in metricas["as_detalle"]:
                        st.markdown(linea)

            # ── Escala de Apreciación: resumen visual ─────────────────────
            _escala_res = st.session_state.get("escala_resultado")
            if _escala_res:
                st.divider()
                st.markdown("### 📊 Escala de Apreciación UST — 45 Criterios")
                _r = _escala_res.get("resumen", {})
                _n_si   = _r.get("SI", 0)
                _n_parc = _r.get("PARCIALMENTE", 0)
                _n_no   = _r.get("NO", 0)
                _n_na   = _r.get("N/A", 0)
                _pct    = _r.get("pct_cumplimiento", 0)
                _total  = _r.get("total", 45)

                # Métricas resumen
                _ce1, _ce2, _ce3, _ce4 = st.columns(4)
                _ce1.metric("✅ Cumple", _n_si,
                            delta=f"{round(_n_si/_total*100)}%" if _total else "—")
                _ce2.metric("⚠️ Parcial", _n_parc)
                _ce3.metric("❌ No cumple", _n_no,
                            delta="revisar" if _n_no else "sin errores",
                            delta_color="inverse" if _n_no else "normal")
                _ce4.metric("Cumplimiento", f"{_pct}%")

                st.progress(
                    (_n_si + _n_parc * 0.5) / max(_total, 1),
                    text=f"Cumplimiento ponderado: {_pct}%  "
                         f"({_n_si} SI · {_n_parc} Parcial · {_n_no} No)"
                )

                # Tabla por sección
                _por_sec = _escala_res.get("por_seccion", {})
                if _por_sec:
                    with st.expander("📋 Resumen por sección", expanded=True):
                        _SEC_COLS = st.columns(len(_por_sec))
                        for _idx, (_sec, _cnt) in enumerate(_por_sec.items()):
                            _s_si   = _cnt.get("SI", 0)
                            _s_no   = _cnt.get("NO", 0)
                            _s_parc = _cnt.get("PARCIALMENTE", 0)
                            _s_tot  = _s_si + _s_no + _s_parc
                            _bg = (
                                "#C8E6D4" if _s_no == 0 else
                                "#FFF3CD" if _s_no <= _s_tot // 2 else
                                "#FDDCDC"
                            )
                            _SEC_COLS[_idx].markdown(
                                f"<div style='background:{_bg};border-radius:8px;"
                                f"padding:0.5rem 0.7rem;text-align:center;'>"
                                f"<b style='font-size:0.75rem;color:#333'>{_sec}</b><br>"
                                f"<span style='font-size:1.1rem;font-weight:700'>"
                                f"✅{_s_si} ⚠️{_s_parc} ❌{_s_no}</span></div>",
                                unsafe_allow_html=True,
                            )

                # Detalle 45 criterios colapsado
                _crits_data = _escala_res.get("criterios", {})
                if CRITERIOS and _crits_data:
                    _sec_act = ""
                    _lineas_crit = []
                    _ICONOS = {"SI": "✅", "PARCIALMENTE": "⚠️", "NO": "❌",
                               "N/A": "⬜", "ERROR": "🔴"}
                    for _c in CRITERIOS:
                        if _c["seccion"] != _sec_act:
                            _sec_act = _c["seccion"]
                            _lineas_crit.append(f"\n**── {_sec_act} ──**")
                        _res_c = _crits_data.get(_c["id"], {})
                        _ico   = _ICONOS.get(_res_c.get("estado", "NO"), "❓")
                        _obs   = _res_c.get("observacion", "")
                        _lineas_crit.append(
                            f"{_ico} **C{_c['id']:02d}** {_c['texto'][:75]}"
                            + (f"  \n&nbsp;&nbsp;&nbsp;&nbsp;↳ *{_obs[:110]}*" if _obs else "")
                        )
                    n_err_crit = _n_no + _n_parc
                    with st.expander(
                        f"🔍 Ver 45 criterios individuales ({n_err_crit} con observación)",
                        expanded=_n_no > 3,
                    ):
                        for _l in _lineas_crit:
                            st.markdown(_l)

                # Reescritura log
                if reesc_log:
                    with st.expander(
                        f"✍️ Reescritura pedagógica — {sum(1 for l in reesc_log if '✏️' in l)} actividad(es) corregida(s)",
                        expanded=False,
                    ):
                        for _rl in reesc_log:
                            st.markdown(_rl)

                # Descargar reporte texto
                _reporte_txt = reporte_escala(_escala_res) if ESCALA_OK else ""
                if _reporte_txt:
                    st.download_button(
                        "⬇ Descargar reporte Escala (.txt)",
                        data=_reporte_txt,
                        file_name=output_name.replace(".xlsx", "_Escala.txt"),
                        mime="text/plain",
                        use_container_width=True,
                    )

            st.divider()
            st.download_button(
                label=f"⬇ Descargar {output_name}",
                data=output_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            st.caption(
                "El archivo incluye todas las correcciones marcadas en azul. "
                "Las filas Formativa tienen fondo lila y las Sumativa fondo amarillo."
                + (" Actividades reescritas pedagógicamente marcadas en azul." if reesc_log else "")
            )

            with st.expander("🔍 Ver log completo de correcciones"):
                log_texto = "\n".join(log)
                st.code(log_texto, language=None)
                st.download_button(
                    "⬇ Descargar log (.txt)",
                    data=log_texto,
                    file_name=output_name.replace(".xlsx", "_log.txt"),
                    mime="text/plain",
                )

# ═════════════════════════════════════════════════════════════════════════
#  INSTANCIA 2 e INSTANCIA 3 (lógica compartida)
# ═════════════════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════════════════
#  INSTANCIA 3
# ═════════════════════════════════════════════════════════════════════════

def _render_instancia_escala(tab, instancia_num, key_prefix):
    """Renderiza el formulario de Instancia 2 o 3 (lógica idéntica, distinto número)."""
    with tab:
        st.markdown("### 1 · Archivos")
        if instancia_num == 2:
            st.caption(
                "Sube la escala completada por la **1ª revisora DEL** "
                "y la planificación del docente."
            )
        else:
            st.caption(
                "Sube la escala completada por la **2ª revisora DEL** "
                "y el archivo `_I2_REVISADO.xlsx` generado en Instancia 2."
            )

        col_e, col_p = st.columns(2)
        with col_e:
            escala_f = st.file_uploader(
                "📋 Escala de apreciación completada",
                type="xlsx",
                key=f"{key_prefix}_escala",
                help="Archivo Excel de la escala completada por la revisora DEL.",
            )
        with col_p:
            plan_f = st.file_uploader(
                "📊 Planificación",
                type="xlsx",
                key=f"{key_prefix}_plan",
                help=("Planificación del docente o _I2_REVISADO.xlsx "
                      if instancia_num == 3 else
                      "Planificación del docente o _REVISADO de Instancia 1."),
            )

        if escala_f:
            with st.spinner("Leyendo escala..."):
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_e:
                    tmp_e.write(escala_f.getvalue())
                    tmp_e_path = tmp_e.name
                crits = rp.leer_escala_completa(tmp_e_path)
                os.unlink(tmp_e_path)

            n_ok_e   = sum(1 for c in crits if c["estado"] == "Si")
            n_parc_e = sum(1 for c in crits if c["estado"] == "Parcialmente")
            n_no_e   = sum(1 for c in crits if c["estado"] == "No")

            with st.expander(
                f"✅ Escala leída — {len(crits)} criterios: "
                f"{n_ok_e} ✅  {n_parc_e} ⚠️ parcial  {n_no_e} ❌",
                expanded=False,
            ):
                for c in crits:
                    icono = ("✅" if c["estado"] == "Si"
                             else ("⚠️" if c["estado"] == "Parcialmente" else "❌"))
                    st.markdown(f"{icono} **{c['criterio'][:80]}**")
                    if c["obs_texto"]:
                        st.caption(f"↳ {c['obs_texto'][:160]}")

        st.markdown("### 2 · Opciones")
        col_as_x, col_pdf_x = st.columns(2)
        with col_as_x:
            es_as_x = st.checkbox(
                "📌 Lineamiento A+S",
                key=f"{key_prefix}_as",
                help="Activa la verificación de los 11 hitos A+Se.",
            )
        with col_pdf_x:
            pdf_x = st.file_uploader(
                "📄 Programa (opcional)",
                type="pdf",
                key=f"{key_prefix}_pdf",
                help="PDF del programa para cruzar correcciones.",
            )

        usar_lt_x = st.checkbox(
            "🔤 Revisión ortográfica y gramatical (LanguageTool)",
            value=True,
            key=f"{key_prefix}_lt",
            help="Revisa ortografía y gramática de las actividades usando "
                 "LanguageTool API en español. Requiere conexión a internet.",
        )

        autocorregir_lt_x = st.checkbox(
            "✍️ Aplicar correcciones automáticas (conservador)",
            value=True,
            key=f"{key_prefix}_autocorr",
            help="Aplica automáticamente las correcciones SEGURAS detectadas por LanguageTool "
                 "(solo cambios unívocos: tildes faltantes, errores ortográficos claros).",
        )

        # ── Escala IA 45 criterios ────────────────────────────────────────
        if ESCALA_OK:
            with st.expander("🎯 Escala de Apreciación UST (45 criterios con IA)", expanded=False):
                st.caption(
                    "Evalúa los 45 criterios institucionales con IA. "
                    "Usa **Ollama local** (gratis) o **Claude API** (mejor calidad)."
                )
                _col_ex0, _col_ex1 = st.columns([1, 1])
                with _col_ex0:
                    usar_escala_x = st.checkbox(
                        "Activar evaluación",
                        value=False,
                        key=f"{key_prefix}_escala_activo",
                    )
                with _col_ex1:
                    _bk_opts_x = ["ollama (local, gratis)", "claude (API)"]
                    _bk_sel_x  = st.selectbox(
                        "Motor IA",
                        _bk_opts_x,
                        key=f"{key_prefix}_escala_backend",
                        help="Ollama corre localmente sin costo. Claude requiere API key.",
                    )
                    backend_escala_x = "ollama" if "ollama" in _bk_sel_x else "claude"

                if backend_escala_x == "ollama":
                    modelo_escala_x = st.text_input(
                        "Modelo Ollama",
                        value=apa_llm.MODELO_TEXTO,
                        key=f"{key_prefix}_escala_modelo_ollama",
                        help=f"Recomendado: {apa_llm.MODELO_TEXTO}. Ejecuta `ollama list` para ver disponibles.",
                    )
                    apikey_escala_x = ""
                    if usar_escala_x:
                        st.success(f"Usando **{modelo_escala_x}** local — sin costo de API.", icon="🦙")
                else:
                    modelo_escala_x = ""
                    apikey_escala_x = st.text_input(
                        "API Key de Anthropic",
                        value=os.environ.get("ANTHROPIC_API_KEY", ""),
                        type="password",
                        key=f"{key_prefix}_escala_key",
                        placeholder="sk-ant-api03-...",
                    )
                    if usar_escala_x and not apikey_escala_x:
                        st.warning("Ingresa la API Key para usar Claude.", icon="🔑")
        else:
            usar_escala_x   = False
            backend_escala_x = "ollama"
            modelo_escala_x  = apa_llm.MODELO_TEXTO
            apikey_escala_x  = ""

        programa_x = None
        if pdf_x:
            with st.spinner("Leyendo programa..."):
                try:
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_px:
                        tmp_px.write(pdf_x.read())
                        tmp_px_path = tmp_px.name
                    programa_x = cruce_extraer_pdf(tmp_px_path)
                    os.unlink(tmp_px_path)
                except Exception as _epx:
                    st.warning(f"No se pudo leer el PDF: {str(_epx)[:200]}", icon="⚠️")
                    programa_x = None

        st.markdown(f"### 3 · Aplicar correcciones")

        listo_x = bool(escala_f and plan_f)
        procesar_x = st.button(
            f"▶ Aplicar correcciones (Instancia {instancia_num})",
            disabled=not listo_x,
            type="primary",
            key=f"btn_{key_prefix}",
            use_container_width=True,
            help="Sube la escala y la planificación para activar."
                 if not listo_x else "",
        )
        if not listo_x:
            st.caption("⬆ Sube la escala completada y la planificación para continuar.")

        if procesar_x and listo_x:
            rp._LANGUAGETOOL_ACTIVO = usar_lt_x
            rp._LANGUAGETOOL_AUTOCORREGIR = autocorregir_lt_x
            spinner_x = (f"Aplicando correcciones Instancia {instancia_num} + revisión ortográfica..."
                         if usar_lt_x else
                         f"Aplicando correcciones Instancia {instancia_num}...")
            with st.spinner(spinner_x):
                log_x, ok_x, out_bytes_x, out_name_x = rp.procesar_instancia2(
                    plan_bytes=plan_f.getvalue(),
                    escala_bytes=escala_f.getvalue(),
                    plan_nombre=plan_f.name,
                    programa=_prog_a_dict(programa_x),
                    es_as=es_as_x,
                    instancia_num=instancia_num,
                )
                if ok_x:
                    out_name_x = _nombre_descarga(programa_x, instancia_num, out_bytes_x)

            st.divider()
            st.markdown("### Resultados")

            if not ok_x or not out_bytes_x:
                st.error("El procesamiento falló. Revisa el log.")
                with st.expander("Log de error"):
                    st.code("\n".join(log_x), language=None)
            else:
                metricas_x = parsear_log(log_x)

                n_anot_x = 0
                for linea in log_x:
                    m = re.search(r"Anotaciones inyectadas:\s*(\d+)", linea)
                    if m:
                        n_anot_x = int(m.group(1))
                        break

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric(
                        "Auto-correcciones",
                        metricas_x["total_correcciones"],
                        help="Correcciones estándar aplicadas.",
                    )
                with col2:
                    st.metric(
                        "Anotaciones revisora",
                        n_anot_x,
                        help="Textos de la escala inyectados en azul.",
                    )
                with col3:
                    total_crit_x = (metricas_x["criterios_ok"]
                                    + metricas_x["criterios_error"]
                                    + metricas_x["criterios_manual"])
                    st.metric(
                        "Criterios UST",
                        f"{metricas_x['criterios_ok']}/{total_crit_x}",
                        delta=f"{metricas_x['criterios_error']} con error"
                              if metricas_x["criterios_error"] else "sin errores",
                        delta_color="inverse" if metricas_x["criterios_error"] else "normal",
                    )

                if metricas_x["lt_ejecutado"]:
                    lt_err_x = metricas_x["lt_errores"]
                    lt_rev_x = metricas_x["lt_revisadas"]
                    lt_corr_x = metricas_x["lt_correcciones"]
                    delta_text_x = f"{lt_rev_x} celda(s) revisadas"
                    if lt_corr_x > 0:
                        delta_text_x += f", {lt_corr_x} corregidas"
                    st.metric(
                        "Errores ortográficos/gramaticales",
                        lt_err_x,
                        delta=delta_text_x,
                        delta_color="off" if lt_corr_x == 0 else "normal",
                        help="LanguageTool API (es). Las correcciones automáticas se aplican solo a cambios seguros.",
                    )
                    if metricas_x["lt_detalle"]:
                        with st.expander(
                            f"🔤 Detalle ortografía/gramática ({lt_err_x} error(es), {lt_corr_x} corregidos)",
                            expanded=lt_err_x > 0 or lt_corr_x > 0,
                        ):
                            for linea in metricas_x["lt_detalle"]:
                                st.markdown(linea)

                if metricas_x["discrepancias_detalle"]:
                    with st.expander("📊 Verificación contra programa", expanded=True):
                        for linea in metricas_x["discrepancias_detalle"]:
                            st.markdown(linea)

                # ── Escala IA 45 criterios ────────────────────────────────
                _escala_activa_x = (
                    ESCALA_OK and usar_escala_x and
                    (backend_escala_x == "ollama" or apikey_escala_x)
                )
                if _escala_activa_x:
                    _llm_label_x = (f"Ollama/{modelo_escala_x}"
                                    if backend_escala_x == "ollama" else "Claude Sonnet")
                    with st.spinner(f"{_llm_label_x} evaluando Escala de Apreciación (45 criterios)…"):
                        try:
                            escala_resultado_x = evaluar_45_criterios(
                                plan_f.getvalue(),
                                api_key=apikey_escala_x,
                                backend=backend_escala_x,
                                modelo_local=modelo_escala_x or apa_llm.MODELO_TEXTO,
                            )
                            st.session_state["escala_resultado"] = escala_resultado_x
                            st.session_state["escala_archivo"]   = plan_f.name
                        except Exception as _ex:
                            st.warning(f"Error en Escala IA: {str(_ex)[:120]}", icon="⚠️")
                            escala_resultado_x = None

                    if escala_resultado_x:
                        _rx = escala_resultado_x.get("resumen", {})
                        st.divider()
                        st.markdown("### 📊 Escala de Apreciación UST — 45 Criterios")
                        _cx1, _cx2, _cx3, _cx4 = st.columns(4)
                        _cx1.metric("✅ Cumple (SI)", _rx.get("SI", 0))
                        _cx2.metric("⚠️ Parcialmente", _rx.get("PARCIALMENTE", 0))
                        _cx3.metric("❌ No cumple", _rx.get("NO", 0),
                                    delta="revisar" if _rx.get("NO", 0) else "sin errores críticos",
                                    delta_color="inverse" if _rx.get("NO", 0) else "normal")
                        _cx4.metric("Cumplimiento", f"{_rx.get('pct_cumplimiento', 0)}%")
                        st.progress(
                            (_rx.get("SI", 0) + _rx.get("PARCIALMENTE", 0) * 0.5) /
                            max(_rx.get("total", 45), 1),
                            text=f"Cumplimiento ponderado: {_rx.get('pct_cumplimiento', 0)}% "
                                 f"({_rx.get('SI',0)} SI · {_rx.get('PARCIALMENTE',0)} Parcial "
                                 f"· {_rx.get('NO',0)} No)"
                        )
                        # Detalle por criterio
                        if CRITERIOS:
                            _crits_x = escala_resultado_x.get("criterios", {})
                            _ICONOS_X = {"SI": "✅", "PARCIALMENTE": "⚠️", "NO": "❌",
                                         "N/A": "⬜", "ERROR": "🔴"}
                            _por_sec_x = {}
                            for _c in CRITERIOS:
                                _por_sec_x.setdefault(_c["seccion"], []).append(_c)
                            for _sec_x, _crits_sec_x in _por_sec_x.items():
                                _no_x  = sum(1 for _c in _crits_sec_x
                                             if _crits_x.get(_c["id"], {}).get("estado") == "NO")
                                _par_x = sum(1 for _c in _crits_sec_x
                                             if _crits_x.get(_c["id"], {}).get("estado") == "PARCIALMENTE")
                                with st.expander(
                                    f"{'✅' if _no_x == 0 else '❌'} **{_sec_x}** — "
                                    f"{len(_crits_sec_x) - _no_x - _par_x} SI · {_par_x} Parcial · {_no_x} No",
                                    expanded=(_no_x > 0),
                                ):
                                    for _c in _crits_sec_x:
                                        _res_cx = _crits_x.get(_c["id"], {})
                                        _ico_x  = _ICONOS_X.get(_res_cx.get("estado", "NO"), "❓")
                                        _obs_x  = _res_cx.get("observacion", "")
                                        st.markdown(f"{_ico_x} **C{_c['id']:02d}** {_c['texto']}")
                                        if _obs_x:
                                            st.caption(f"   ↳ {_obs_x}")
                        st.info(
                            "Los resultados completos también están disponibles en el tab "
                            "**📊 Escala de Apreciación**.",
                            icon="💡",
                        )

                if instancia_num == 3 and metricas_x["criterios_error"] == 0:
                    st.success(
                        "Sin criterios con error automático. "
                        "La planificación está lista para aprobación final.",
                        icon="🎓",
                    )

                st.info(
                    "La hoja **NOTAS_CORRECCIONES_DEL** en el archivo descargado "
                    "contiene el resumen completo de observaciones de la revisora.",
                    icon="📋",
                )

                st.divider()
                st.download_button(
                    label=f"⬇ Descargar {out_name_x}",
                    data=out_bytes_x,
                    file_name=out_name_x,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
                st.caption(
                    "Azul = corrección automática o anotación de la revisora. "
                    "Lila = Formativa · Amarillo = Sumativa. "
                    "Hoja NOTAS_CORRECCIONES_DEL = resumen completo."
                )

                with st.expander("🔍 Ver log completo"):
                    log_x_texto = "\n".join(log_x)
                    st.code(log_x_texto, language=None)
                    st.download_button(
                        "⬇ Descargar log (.txt)",
                        data=log_x_texto,
                        file_name=out_name_x.replace(".xlsx", "_log.txt"),
                        mime="text/plain",
                    )


# ── Revisoras DEL: I2 e I3 en una sola pestaña ───────────────────────────────
with tab_del:
    st.markdown("### Correcciones del equipo DEL")
    st.caption("Selecciona la instancia según quién está revisando.")

    _inst_sel = st.radio(
        "¿Qué instancia aplicas?",
        ["📝 1ª revisora DEL (Instancia 2)", "✅ 2ª revisora / aprobación final (Instancia 3)"],
        horizontal=True,
        key="del_instancia_radio",
    )
    _inst_num = 2 if "2" in _inst_sel else 3

    st.markdown("---")
    _render_instancia_escala(tab_del, instancia_num=_inst_num,
                             key_prefix=f"i{_inst_num}")

# ═════════════════════════════════════════════════════════════════════════
#  TAB DISEÑO DE RECURSOS
# ═════════════════════════════════════════════════════════════════════════

try:
    import generar_recursos as gr
    RECURSOS_OK = True
except ImportError:
    RECURSOS_OK = False

try:
    import extraer_borrador as eb
    BORRADOR_OK = True
except ImportError:
    BORRADOR_OK = False

with tab_recursos:
    st.markdown("### 🎨 Diseño de Recursos Didácticos T1–T4")

    if not RECURSOS_OK:
        st.error(
            "No se encontró `generar_recursos.py`. "
            "Verifica que esté en la misma carpeta que esta app.",
            icon="⚙️",
        )
        st.stop()

    # ── Selector de tipo de recurso ────────────────────────────────────────
    st.markdown("""
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:0.6rem;margin-bottom:1.2rem">
  <div style="background:#DCF0E4;border:2px solid #006633;border-radius:10px;
              padding:0.8rem 0.5rem;text-align:center;cursor:pointer">
    <div style="font-size:1.4rem">🎬</div>
    <div style="font-size:0.75rem;font-weight:700;color:#004d26">T1</div>
    <div style="font-size:0.65rem;color:#336644">Videoclase</div>
  </div>
  <div style="background:#EBF2FB;border:2px solid #2E5FA3;border-radius:10px;
              padding:0.8rem 0.5rem;text-align:center;cursor:pointer">
    <div style="font-size:1.4rem">🖼️</div>
    <div style="font-size:0.75rem;font-weight:700;color:#1d3f80">T2</div>
    <div style="font-size:0.65rem;color:#4060a0">Genially</div>
  </div>
  <div style="background:#F0EAFB;border:2px solid #6B3FA0;border-radius:10px;
              padding:0.8rem 0.5rem;text-align:center;cursor:pointer">
    <div style="font-size:1.4rem">📄</div>
    <div style="font-size:0.75rem;font-weight:700;color:#4a2880">T3</div>
    <div style="font-size:0.65rem;color:#7050b0">Guía</div>
  </div>
  <div style="background:#E6F5F5;border:2px solid #007A7A;border-radius:10px;
              padding:0.8rem 0.5rem;text-align:center;cursor:pointer">
    <div style="font-size:1.4rem">📝</div>
    <div style="font-size:0.75rem;font-weight:700;color:#005555">T4</div>
    <div style="font-size:0.65rem;color:#007070">Foro/Quiz/Tarea</div>
  </div>
</div>
""", unsafe_allow_html=True)

    _tipo_recurso = st.radio(
        "Tipo de recurso a generar",
        ["🎬 T1 — Guión de videoclase", "🖼️ T2 — Estructura Genially",
         "📄 T3 — Guía de aprendizaje", "📝 T4 — Foro / Quiz / Tarea"],
        horizontal=True,
        label_visibility="collapsed",
        key="rec_tipo",
    )
    _es_t1 = "T1" in _tipo_recurso
    _es_t2 = "T2" in _tipo_recurso
    _es_t3 = "T3" in _tipo_recurso
    _es_t4 = "T4" in _tipo_recurso

    st.divider()

    # ── 1 · Planificación ─────────────────────────────────────────────────
    st.markdown("#### 1 · Planificación")

    _rec_xlsx = st.file_uploader(
        "📊 Planificación didáctica (.xlsx)",
        type="xlsx",
        key="rec_xlsx",
        help="Sube la planificación procesada (descargada del tab Revisar) "
             "o la planificación original del docente.",
    )

    # ── Material borrador del docente ──────────────────────────────────────
    with st.expander("📎 Material borrador del docente (opcional)", expanded=False):
        st.caption(
            "Sube el material que el/la docente entregó como base de contenido "
            "(Word, PPT, Excel, PDF, TXT). El modelo de IA lo usará para construir "
            "el recurso a partir del contenido real de la asignatura."
        )
        _borrador_file = st.file_uploader(
            "Archivo borrador",
            type=["docx", "pptx", "xlsx", "pdf", "txt", "md"],
            key="rec_borrador",
            label_visibility="collapsed",
        )

    _borrador_texto = ""
    if _borrador_file:
        if BORRADOR_OK:
            with st.spinner("Extrayendo texto del borrador…"):
                try:
                    _borrador_texto = eb.extraer_texto(
                        _borrador_file.getvalue(), _borrador_file.name
                    )
                    _borrador_file.seek(0)
                    if _borrador_texto.startswith("["):
                        st.warning(_borrador_texto, icon="⚠️")
                        _borrador_texto = ""
                    else:
                        st.success(
                            f"Borrador extraído — {len(_borrador_texto):,} caracteres "
                            f"de **{_borrador_file.name}**",
                            icon="📎",
                        )
                except Exception as _eb_err:
                    st.error(f"No se pudo leer el borrador: {_eb_err}", icon="🔴")
        else:
            st.warning(
                "No se encontró `extraer_borrador.py`. El borrador no se usará.",
                icon="⚙️",
            )

    _unidades_disp: list[tuple[int, str]] = []
    if _rec_xlsx:
        try:
            with st.spinner("Leyendo unidades..."):
                _unidades_disp = gr.listar_unidades(_rec_xlsx.getvalue())
            _rec_xlsx.seek(0)
            if _unidades_disp:
                st.success(
                    f"Planificación leída — {len(_unidades_disp)} unidad(es) encontrada(s).",
                    icon="✅",
                )
            else:
                st.warning(
                    "No se encontraron unidades en el archivo. "
                    "Verifica que el Excel tenga la hoja 'Planificación por unidades'.",
                    icon="⚠️",
                )
        except Exception as _e_xl:
            st.error(f"No se pudo leer el Excel: {_e_xl}", icon="🔴")

    if not _unidades_disp:
        st.caption("⬆ Sube el archivo .xlsx para continuar.")

    # ── Nota T1 reutilizado ────────────────────────────────────────────────
    if _es_t1 and _unidades_disp:
        st.info(
            "**T1 reutilizado** — Si este recurso ya fue contabilizado en otra sesión, "
            "no vuelvas a sumar sus minutos en la fórmula de carga académica. "
            "Solo referenciarlo pedagógicamente está bien; las horas se cuentan **una sola vez**.",
            icon="ℹ️",
        )

    # ── 2 · Selección de unidad y opciones ────────────────────────────────
    if _unidades_disp:
        st.markdown("#### 2 · Opciones")

        _col_u, _col_d = st.columns([2, 1])
        with _col_u:
            _unidad_opciones = [f"Unidad {n}: {nombre}" for n, nombre in _unidades_disp]
            _unidad_sel_idx  = st.selectbox(
                "Unidad",
                range(len(_unidades_disp)),
                format_func=lambda i: _unidad_opciones[i],
                key="rec_unidad",
            )
            _num_unidad = _unidades_disp[_unidad_sel_idx][0]

        with _col_d:
            if _es_t1:
                _duracion = st.selectbox(
                    "Duración (min)",
                    [5, 8, 10, 12, 15],
                    index=2,
                    key="rec_duracion",
                    help="Rango institucional: 8–12 min ≈ 0,5 hrs de carga estudiantil. "
                         "Si el programa no especifica duración, elige dentro de este rango.",
                )
            elif _es_t4:
                _tipo_t4 = st.selectbox(
                    "Tipo T4",
                    ["tarea", "foro", "quiz"],
                    format_func=lambda t: {
                        "tarea": "📋 Tarea / producción",
                        "foro":  "💬 Foro de participación",
                        "quiz":  "✅ Quiz de autoevaluación",
                    }[t],
                    key="rec_t4_tipo",
                )
            else:
                st.empty()

        # ── Motor IA ──────────────────────────────────────────────────────
        st.markdown("**Motor de IA**")
        _col_b, _col_m, _col_k = st.columns([1, 2, 2])
        with _col_b:
            _rec_backend = st.selectbox(
                "Backend",
                gr.BACKENDS,
                key="rec_backend",
                help="Claude = mejor calidad · Ollama = local y gratuito",
            )
        with _col_m:
            _rec_model = st.text_input(
                "Modelo",
                value=gr.BACKEND_DEFAULTS[_rec_backend]["model"],
                key=f"rec_model_{_rec_backend}",
            )
        with _col_k:
            _rec_apikey = st.text_input(
                "API Key",
                value="",
                type="password",
                key="rec_apikey",
                placeholder="No requerida para Ollama",
                help="Anthropic / OpenAI / Grok según el backend.",
            )

        if _rec_backend == "ollama":
            st.info(
                f"Ollama corre localmente. Asegúrate de que esté activo (`ollama serve`) "
                f"y que el modelo `{_rec_model}` esté descargado.",
                icon="🦙",
            )
        elif _rec_backend == "claude" and not _rec_apikey:
            _env_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if _env_key:
                _rec_apikey = _env_key
                st.success("API Key de Anthropic cargada desde variable de entorno.", icon="🔑")
            else:
                st.warning("Ingresa tu API Key de Anthropic para usar Claude.", icon="🔑")

        # ── 3 · Generar ───────────────────────────────────────────────────
        st.markdown("#### 3 · Generar")

        _listo_rec = bool(
            _rec_xlsx and _unidades_disp and
            (_rec_backend == "ollama" or _rec_apikey)
        )

        _label_btn = {
            True:  "▶ Generar guión de videoclase",
            False: "▶ Generar guión de videoclase",
        }
        if _es_t2:
            _label_btn = {True: "▶ Generar estructura Genially", False: "▶ Generar estructura Genially"}
        elif _es_t3:
            _label_btn = {True: "▶ Generar guía de aprendizaje", False: "▶ Generar guía de aprendizaje"}
        elif _es_t4:
            _label_btn = {True: "▶ Generar consigna T4", False: "▶ Generar consigna T4"}

        _btn_rec = st.button(
            _label_btn[True],
            disabled=not _listo_rec,
            type="primary",
            key="btn_rec_generar",
            use_container_width=True,
            help="Sube el Excel y configura el motor IA para activar." if not _listo_rec else "",
        )
        if not _listo_rec and not _rec_xlsx:
            st.caption("⬆ Sube el Excel de planificación para continuar.")

        if _btn_rec and _listo_rec:
            _xlsx_bytes_rec = _rec_xlsx.getvalue()
            _nombre_unidad  = _unidades_disp[_unidad_sel_idx][1]

            if _es_t1:
                _spinner_msg = (
                    f"Generando guión de videoclase {_duracion} min — "
                    f"Unidad {_num_unidad}: {_nombre_unidad}…"
                )
            elif _es_t2:
                _spinner_msg = f"Generando estructura Genially — Unidad {_num_unidad}…"
            elif _es_t3:
                _spinner_msg = f"Generando guía de aprendizaje — Unidad {_num_unidad}…"
            else:
                _spinner_msg = f"Generando consigna T4 — Unidad {_num_unidad}…"

            with st.spinner(_spinner_msg):
                if _es_t1:
                    _rec_texto, _rec_error = gr.generar_guion_t1(
                        xlsx_bytes=_xlsx_bytes_rec,
                        num_unidad=_num_unidad,
                        duracion_min=_duracion,
                        backend=_rec_backend,
                        model=_rec_model or None,
                        api_key=_rec_apikey,
                        borrador_texto=_borrador_texto,
                    )
                    _rec_ext  = "txt"
                    _rec_mime = "text/plain"
                    _rec_icono = "🎬"
                elif _es_t2:
                    _rec_texto, _rec_error = gr.generar_estructura_t2(
                        xlsx_bytes=_xlsx_bytes_rec,
                        num_unidad=_num_unidad,
                        backend=_rec_backend,
                        model=_rec_model or None,
                        api_key=_rec_apikey,
                        borrador_texto=_borrador_texto,
                    )
                    _rec_ext  = "txt"
                    _rec_mime = "text/plain"
                    _rec_icono = "🖼️"
                elif _es_t3:
                    _rec_texto, _rec_error = gr.generar_guia_t3(
                        xlsx_bytes=_xlsx_bytes_rec,
                        num_unidad=_num_unidad,
                        backend=_rec_backend,
                        model=_rec_model or None,
                        api_key=_rec_apikey,
                        borrador_texto=_borrador_texto,
                    )
                    _rec_ext  = "txt"
                    _rec_mime = "text/plain"
                    _rec_icono = "📄"
                else:
                    _rec_texto, _rec_error = gr.generar_consigna_t4(
                        xlsx_bytes=_xlsx_bytes_rec,
                        num_unidad=_num_unidad,
                        tipo_t4=_tipo_t4,
                        backend=_rec_backend,
                        model=_rec_model or None,
                        api_key=_rec_apikey,
                        borrador_texto=_borrador_texto,
                    )
                    _rec_ext  = "txt"
                    _rec_mime = "text/plain"
                    _rec_icono = "📝"

            st.divider()

            if _rec_error:
                st.error(
                    f"Error al generar el recurso: {_rec_error}\n\n"
                    "Verifica la API key o que Ollama esté activo.",
                    icon="🔴",
                )
            else:
                if _es_t1:   _tipo_label = "Guión_Videoclase"
                elif _es_t2: _tipo_label = "Estructura_Genially"
                elif _es_t3: _tipo_label = "Guia_Aprendizaje"
                else:        _tipo_label = f"Consigna_T4_{_tipo_t4.capitalize()}"

                _nombre_limpio_rec = re.sub(r'[\\/:*?"<>|\s]+', '_', _nombre_unidad)
                _fname_rec = f"U{_num_unidad}_{_tipo_label}_{_nombre_limpio_rec}.{_rec_ext}"

                # Guardar en session_state para persistencia
                st.session_state["rec_ultimo_texto"] = _rec_texto
                st.session_state["rec_ultimo_fname"] = _fname_rec
                st.session_state["rec_ultimo_icono"] = _rec_icono

            # ── Mostrar resultado guardado ─────────────────────────────────
        _rec_cache_texto = st.session_state.get("rec_ultimo_texto", "")
        _rec_cache_fname = st.session_state.get("rec_ultimo_fname", "recurso.txt")
        _rec_cache_icono = st.session_state.get("rec_ultimo_icono", "📄")

        if _rec_cache_texto:
            _n_lineas = len(_rec_cache_texto.splitlines())
            _n_palabras = len(_rec_cache_texto.split())
            col_m1, col_m2 = st.columns(2)
            col_m1.metric("Líneas generadas", _n_lineas)
            col_m2.metric("Palabras", _n_palabras)

            with st.expander(
                f"{_rec_cache_icono} Vista previa — {_rec_cache_fname}",
                expanded=True,
            ):
                st.text(_rec_cache_texto)

            st.download_button(
                label=f"⬇ Descargar {_rec_cache_fname}",
                data=_rec_cache_texto.encode("utf-8"),
                file_name=_rec_cache_fname,
                mime="text/plain",
                type="primary",
                use_container_width=True,
            )
            st.caption(
                "Revisa el contenido antes de usarlo en producción. "
                "El recurso está basado en los datos declarados en la planificación."
            )

            if st.button("🗑 Limpiar resultado", key="btn_rec_limpiar"):
                st.session_state.pop("rec_ultimo_texto", None)
                st.session_state.pop("rec_ultimo_fname", None)
                st.session_state.pop("rec_ultimo_icono", None)
                st.rerun()

# ═════════════════════════════════════════════════════════════════════════
#  TAB ESCALA DE APRECIACIÓN
# ═════════════════════════════════════════════════════════════════════════

with tab_escala:
    st.markdown("### 📊 Escala de Apreciación UST — 45 Criterios")

    _escala_cache = st.session_state.get("escala_resultado")

    if not ESCALA_OK:
        st.warning(
            "Los módulos de evaluación avanzada no están disponibles. "
            "Verifica que `agente_criterios.py`, `calculos_del.py` y `reescritura_llm.py` "
            "estén en la misma carpeta que esta app.",
            icon="⚙️",
        )
    elif not _escala_cache:
        st.info(
            "Aún no hay evaluación disponible. "
            "Procesa una planificación en **Instancia 1** con la opción "
            "\"Escala de Apreciación\" activada para ver los resultados aquí.",
            icon="💡",
        )
        st.markdown("""
**¿Qué evalúa la Escala de Apreciación?**

| Sección | Criterios | Tipo | Qué verifica |
|---------|-----------|------|-------------|
| Síntesis didáctica | C01–C15 | Cruce vs programa PDF | Datos, RA, créditos |
| Identificación | C16–C18 | Estructura automática | Unidades, sesiones, momentos |
| Est. Metodológicas | C19–C26 | Semántico (IA) | Redacción imperativa, estructura de momentos, propósito/retroalimentación |
| Recursos (G3) | C27–C35 | Semántico (IA) | 5 campos por recurso, ubicación, extensión, T1 reutilizados |
| Est. Evaluativas | C36–C41 | Automático + semántico | Klarway/Turnitin, alineación RA↔evaluación |
| Carga Académica | C42–C45 | Matemático + semántico | Horas vs rangos de estandarización, sobreconteo de T1 |

**Nuevas detecciones (2026-1):**
- **T1 reutilizado** (C30): si el mismo recurso aparece en varias sesiones, se alerta si sus minutos se suman más de una vez.
- **Estandarización de carga** (C44): verifica que la duración declarada sea coherente con los rangos institucionales (videoclase 8–12 min, H5P 0,5 hrs, etc.).
- **Prueba del estudiante autónomo**: una actividad es incompleta si el/la estudiante no puede realizarla sin preguntar al/la docente.
        """)
    else:
        _r = _escala_cache.get("resumen", {})
        _n_si   = _r.get("SI", 0)
        _n_parc = _r.get("PARCIALMENTE", 0)
        _n_no   = _r.get("NO", 0)
        _pct    = _r.get("pct_cumplimiento", 0)
        _total  = _r.get("total", 45)
        _arch   = st.session_state.get("escala_archivo", "planificación")

        st.caption(f"Última evaluación: **{_arch}**")

        # ── Resumen global ───────────────────────────────────────────────
        _c1, _c2, _c3, _c4 = st.columns(4)
        _c1.metric("✅ Cumple (SI)", _n_si,
                   delta=f"{round(_n_si / _total * 100)}% del total" if _total else "—")
        _c2.metric("⚠️ Parcialmente", _n_parc)
        _c3.metric("❌ No cumple", _n_no,
                   delta="revisar" if _n_no else "sin errores críticos",
                   delta_color="inverse" if _n_no else "normal")
        _c4.metric("Cumplimiento ponderado", f"{_pct}%")

        # Barra de progreso
        _color_pct = "#006633" if _pct >= 80 else "#C8A951" if _pct >= 50 else "#CC3333"
        st.markdown(
            f"<div style='background:#E8E8E8;border-radius:99px;height:14px;overflow:hidden;margin:0.5rem 0'>"
            f"<div style='background:{_color_pct};height:100%;width:{_pct}%;border-radius:99px;'></div>"
            f"</div><p style='text-align:center;font-size:0.82rem;color:#555;margin-top:4px'>"
            f"Cumplimiento ponderado: {_pct}% "
            f"({_n_si} SI · {_n_parc} Parcial · {_n_no} No · de {_total} criterios)</p>",
            unsafe_allow_html=True,
        )

        st.divider()

        # ── Vista por sección ────────────────────────────────────────────
        if CRITERIOS:
            _ICONOS_E = {"SI": "✅", "PARCIALMENTE": "⚠️", "NO": "❌",
                         "N/A": "⬜", "ERROR": "🔴"}
            _crits_d  = _escala_cache.get("criterios", {})
            _por_sec  = {}
            for _c in CRITERIOS:
                _por_sec.setdefault(_c["seccion"], []).append(_c)

            for _sec_nombre, _sec_crits in _por_sec.items():
                _sec_si  = sum(1 for _c in _sec_crits if _crits_d.get(_c["id"], {}).get("estado") == "SI")
                _sec_no  = sum(1 for _c in _sec_crits if _crits_d.get(_c["id"], {}).get("estado") == "NO")
                _sec_par = sum(1 for _c in _sec_crits if _crits_d.get(_c["id"], {}).get("estado") == "PARCIALMENTE")
                _sec_tot = len(_sec_crits)
                _sec_ok  = _sec_no == 0

                _header_color = (
                    "background:#C8E6D4;color:#004d26" if _sec_no == 0 and _sec_par == 0 else
                    "background:#FFF3CD;color:#7A5700" if _sec_no == 0 else
                    "background:#FDDCDC;color:#8B0000"
                )

                with st.expander(
                    f"{'✅' if _sec_ok else '❌'} **{_sec_nombre}** — "
                    f"{_sec_si} SI · {_sec_par} Parcial · {_sec_no} No  ({_sec_tot} criterios)",
                    expanded=(_sec_no > 0),
                ):
                    for _c in _sec_crits:
                        _res_c = _crits_d.get(_c["id"], {})
                        _ico   = _ICONOS_E.get(_res_c.get("estado", "NO"), "❓")
                        _obs   = _res_c.get("observacion", "")
                        _tipo_badge = (
                            "<span style='font-size:0.68rem;background:#E8F5EE;"
                            "color:#006633;border-radius:4px;padding:1px 6px;margin-left:6px'>auto</span>"
                            if _c.get("tipo") == "auto" else
                            "<span style='font-size:0.68rem;background:#EDE8F5;"
                            "color:#5B2DB0;border-radius:4px;padding:1px 6px;margin-left:6px'>IA</span>"
                        )
                        st.markdown(
                            f"{_ico} **C{_c['id']:02d}** {_c['texto']}{_tipo_badge}",
                            unsafe_allow_html=True,
                        )
                        if _obs:
                            st.caption(f"   ↳ {_obs}")

        st.divider()

        # ── Descargas ────────────────────────────────────────────────────
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            if ESCALA_OK:
                _txt_escala = reporte_escala(_escala_cache)
                st.download_button(
                    "⬇ Descargar reporte (.txt)",
                    data=_txt_escala,
                    file_name=_arch.replace(".xlsx", "_Escala.txt"),
                    mime="text/plain",
                    use_container_width=True,
                )
        with col_dl2:
            import json as _json
            st.download_button(
                "⬇ Descargar datos (.json)",
                data=_json.dumps(_escala_cache, ensure_ascii=False, indent=2),
                file_name=_arch.replace(".xlsx", "_Escala.json"),
                mime="application/json",
                use_container_width=True,
            )

        if st.button("🔄 Limpiar evaluación", key="btn_escala_limpiar"):
            st.session_state.pop("escala_resultado", None)
            st.session_state.pop("escala_archivo", None)
            st.rerun()

# ═════════════════════════════════════════════════════════════════════════
#  HISTORIAL
# ═════════════════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════════════════
#  HERRAMIENTAS (Diccionario + APA)
# ═════════════════════════════════════════════════════════════════════════

with tab_config:
    st.markdown("### Diccionario UST personalizable")
    st.caption(
        "Las entradas base (en gris) vienen del script y no se pueden borrar desde aquí. "
        "Las entradas personalizadas (en verde) se suman a las base y tienen prioridad."
    )

    completo = dict_ust.obtener_dict_completo()
    custom   = dict_ust.obtener_entradas_custom()

    # ── Ver diccionario completo ──────────────────────────────────────────
    for mapa in dict_ust.MAPAS:
        entradas = completo[mapa]
        n_custom = len(custom.get(mapa, {}))
        with st.expander(
            f"**{dict_ust.etiqueta(mapa)}** — {len(entradas)} entradas "
            f"({n_custom} personalizadas)",
            expanded=False,
        ):
            import pandas as pd
            filas = []
            for inc, corr in sorted(entradas.items()):
                es_custom = inc in custom.get(mapa, {})
                filas.append({
                    "Término incorrecto": inc,
                    "→ Corrección":       corr,
                    "Origen":             "✏️ Personalizada" if es_custom else "📋 Base",
                })
            if filas:
                st.dataframe(pd.DataFrame(filas), use_container_width=True,
                             hide_index=True)

    st.divider()

    # ── Agregar nueva entrada ─────────────────────────────────────────────
    st.markdown("#### Agregar entrada")
    col_m, col_i, col_c = st.columns([2, 3, 3])
    with col_m:
        mapa_sel = st.selectbox(
            "Mapa",
            dict_ust.MAPAS,
            format_func=dict_ust.etiqueta,
            key="dict_mapa",
        )
    with col_i:
        nuevo_inc  = st.text_input("Término incorrecto", key="dict_inc",
                                   placeholder="ej: Examen")
    with col_c:
        nuevo_corr = st.text_input("Corrección UST",     key="dict_corr",
                                   placeholder="ej: Pruebas escritas u orales")

    if st.button("➕ Agregar al diccionario", key="btn_dict_add", type="primary", use_container_width=True):
        if nuevo_inc.strip() and nuevo_corr.strip():
            dict_ust.agregar_entrada(mapa_sel, nuevo_inc, nuevo_corr)
            st.success(f"Entrada agregada: «{nuevo_inc}» → «{nuevo_corr}»")
            st.rerun()
        else:
            st.warning("Completa ambos campos antes de agregar.")

    st.divider()

    # ── Eliminar entrada personalizada ────────────────────────────────────
    st.markdown("#### Eliminar entrada personalizada")
    todas_custom = [
        (mapa, inc)
        for mapa in dict_ust.MAPAS
        for inc in custom.get(mapa, {})
    ]
    if todas_custom:
        opciones = [f"{dict_ust.etiqueta(m)} → «{i}»" for m, i in todas_custom]
        sel_idx  = st.selectbox("Selecciona entrada a eliminar", range(len(opciones)),
                                format_func=lambda i: opciones[i], key="dict_del_sel")
        if st.button("🗑 Eliminar", key="btn_dict_del", type="secondary"):
            mapa_d, inc_d = todas_custom[sel_idx]
            dict_ust.eliminar_entrada(mapa_d, inc_d)
            st.success(f"Entrada «{inc_d}» eliminada.")
            st.rerun()
    else:
        st.caption("No hay entradas personalizadas para eliminar.")

    st.divider()

    # ── Exportar / Importar ───────────────────────────────────────────────
    st.markdown("#### Exportar / Importar")
    col_exp, col_imp = st.columns(2)
    with col_exp:
        st.download_button(
            "⬇ Exportar diccionario (.json)",
            data=dict_ust.exportar_json(),
            file_name="dict_ust.json",
            mime="application/json",
        )
    with col_imp:
        json_up = st.file_uploader("⬆ Importar diccionario (.json)",
                                   type="json", key="dict_import")
        if json_up:
            n = dict_ust.importar_json(json_up.getvalue())
            st.success(f"{n} entrada(s) importadas correctamente.")
            st.rerun()

    st.divider()

    # ── Generador de referencia APA 7 desde URL / DOI ─────────────────────
    st.markdown("#### Generar referencia APA 7 desde URL o DOI")
    st.caption(
        "Ingresa una URL o DOI para generar automáticamente la referencia APA 7. "
        "Requiere conexión a internet. Revisa el resultado antes de usarlo."
    )

    url_input = st.text_input(
        "URL o DOI",
        key="apa_url_input",
        placeholder="https://doi.org/10.1016/... o https://www.sitio.cl/articulo",
    )

    if st.button("🔍 Generar referencia", key="btn_apa_gen"):
        if url_input.strip():
            with st.spinner("Consultando metadatos..."):
                ref_gen, estado = apa.generar_desde_url(url_input.strip())
            if ref_gen:
                st.success(f"Referencia generada ({estado}):")
                st.code(ref_gen, language=None)
                st.caption(
                    "Copia esta referencia y pégala en la columna H del Excel. "
                    "Verifica que el año, nombre de autores y tipo de recurso sean correctos."
                )
            else:
                st.warning(f"No se pudo generar la referencia: {estado}")
        else:
            st.warning("Ingresa una URL o DOI para continuar.")

# ═════════════════════════════════════════════════════════════════════════
#  TAB ANÁLISIS VISUAL — Qwen2.5-VL
# ═════════════════════════════════════════════════════════════════════════

with tab_vision:
    st.markdown("""
<div class="vision-card">
  <h4>🔍 Análisis Visual con Qwen2.5-VL</h4>
  <p style="margin:0;font-size:0.86rem;color:#4a3a8a;">
    Sube una imagen de un documento (planificación, programa, Genially, captura de pantalla)
    y el modelo multimodal analizará su contenido de forma inteligente.
  </p>
  <br>
  <span class="model-badge">🦙 Qwen2.5-VL · Local · Sin costo de API</span>
  &nbsp;
  <span style="font-size:0.74rem;color:#7a5c00;background:#FFF3CD;border-radius:8px;
               padding:3px 10px;border:1px solid #EDD89A;">
    ⚠️ 16 GB RAM — no correr junto al modelo de texto
  </span>
</div>
""", unsafe_allow_html=True)

    col_v1, col_v2 = st.columns([1.2, 1])
    with col_v1:
        imagen_up = st.file_uploader(
            "📷 Imagen del documento",
            type=["png", "jpg", "jpeg", "webp"],
            key="vision_img",
            help="Captura de pantalla, foto o exportación PNG/JPG de la planificación, "
                 "programa u otro recurso UST.",
        )
    with col_v2:
        tipo_analisis = st.selectbox(
            "Tipo de análisis",
            ["planificacion", "programa", "genially", "libre"],
            format_func=lambda t: {
                "planificacion": "📋 Planificación didáctica",
                "programa":      "📄 Programa de asignatura",
                "genially":      "🎨 Recurso Genially / infografía",
                "libre":         "💬 Análisis libre",
            }[t],
            key="vision_tipo",
        )
        modelo_vision = st.text_input(
            "Modelo Ollama",
            value=apa_llm.MODELO_VISION,
            key="vision_model",
            help="Modelo multimodal instalado en Ollama. "
                 "Ejecuta `ollama list` para ver los disponibles.",
        )

    prompt_extra = st.text_area(
        "Instrucción adicional (opcional)",
        key="vision_prompt",
        placeholder="Ej: Enfócate especialmente en los recursos T1 y T2 de la Unidad 2.",
        height=80,
    )

    # Vista previa de la imagen
    if imagen_up:
        with st.expander("👁 Vista previa", expanded=True):
            st.image(imagen_up, use_container_width=True)
        imagen_up.seek(0)

    listo_v = bool(imagen_up)
    if st.button(
        "🔍 Analizar documento",
        disabled=not listo_v,
        type="primary",
        key="btn_vision",
        use_container_width=True,
        help="Sube una imagen para activar el análisis." if not listo_v else "",
    ):
        imagen_bytes = imagen_up.read()
        with st.spinner(f"Qwen2.5-VL analizando imagen ({tipo_analisis})…"):
            resultado = apa_llm.analizar_imagen_llm(
                imagen_bytes=imagen_bytes,
                tipo_analisis=tipo_analisis,
                prompt_extra=prompt_extra,
                model=modelo_vision,
            )

        st.divider()
        if resultado["error"]:
            st.error(
                f"Error al conectar con Ollama: {resultado['error']}\n\n"
                "Verifica que Ollama esté corriendo (`ollama serve`) "
                f"y que el modelo `{modelo_vision}` esté instalado.",
                icon="🔴",
            )
        else:
            st.markdown("### Resultado del análisis")
            st.markdown(
                f'<div class="vision-result">{resultado["resultado"]}</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                "⬇ Descargar análisis (.txt)",
                data=resultado["resultado"].encode("utf-8"),
                file_name=f"analisis_visual_{tipo_analisis}.txt",
                mime="text/plain",
                key="btn_vision_download",
            )

    if not listo_v:
        st.caption("⬆ Sube una imagen para activar el análisis.")

    st.divider()
    st.markdown("""
**¿Cómo usar esta función?**

| Caso de uso | Tipo de análisis |
|-------------|-----------------|
| Revisar una planificación escaneada o capturada | 📋 Planificación |
| Extraer datos de un programa en PDF como imagen | 📄 Programa |
| Verificar cumplimiento visual de un Genially | 🎨 Genially |
| Cualquier documento educativo | 💬 Análisis libre |

El modelo corre **100% local** en tu equipo — sin enviar datos a servidores externos.
""")
