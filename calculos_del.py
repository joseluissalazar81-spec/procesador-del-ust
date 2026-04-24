"""
calculos_del.py — Motor matemático de validación de horas y recursos T1-T4
Procesador DEL · UST 2026-1

Qué verifica:
  1. Horas declaradas en Síntesis vs horas reales en Planificación (GLOBAL y POR UNIDAD)
  2. Distribución T1/T2/T3/T4 según estándar V8
  3. Cantidad de recursos por tipo coherente con las horas asignadas
  4. Porcentajes de evaluación suman 100%

Columnas de la planificación:
  Col 2  → Unidad / módulo
  Col 6  → Momento (Preparación / Desarrollo / Trabajo independiente)
  Col 16 → Horas presenciales
  Col 17 → Horas sincrónicas
  Col 18 → Horas asincrónicas
  Col 19 → Horas TPE
  Col 15 → % evaluación (cuando aplica)
"""

from __future__ import annotations
from collections import defaultdict
from io import BytesIO
import openpyxl


# ── Estándares V8: horas promedio por recurso ────────────────────────────────
# Tiempo estimado promedio (horas) que consume 1 recurso de cada tipo
HORAS_POR_RECURSO = {
    "T1": 0.473,   # video narrado promedio (ref. col "Tiempo de deicación estimado promedio")
    "T2": 1.292,   # tarea/trabajo práctico promedio
    "T3": 0.813,   # presentación / guía sincrónica promedio
    "T4": 0.667,   # video/guía TPE promedio
}

# Distribución esperada de horas por tipo (fracción del total por modalidad)
DISTRIBUCION = {
    "T1": {"base": "asinc",    "fraccion": 0.20},
    "T2": {"base": "asinc",    "fraccion": 0.80},
    "T3": {"base": "pres_sinc","fraccion": 0.40},  # 40% de (pres + sinc)
    "T4": {"base": "tpe",      "fraccion": 0.20},
}

# Tolerancia aceptable en horas (± N horas antes de marcar error)
TOLERANCIA_HORAS = 1.0


# ── Lectura de Síntesis ──────────────────────────────────────────────────────

def leer_sintesis(ws_sint) -> dict:
    """
    Extrae de la hoja 'Síntesis didáctica':
      - horas globales (presenciales, sincrónicas, asincrónicas, TPE)
      - horas pedagógicas declaradas por unidad
      - procedimientos e instrumentos de evaluación declarados en el programa
    """
    resultado = {
        "global": {"presencial": 0, "sinc": 0, "asinc": 0, "tpe": 0},
        "unidades": {},
        "porcentajes_eval": [],
        "procedimientos_programa": [],  # [{"ra": ..., "procedimiento": ..., "instrumento": ...}]
    }

    rows = list(ws_sint.iter_rows(values_only=True))
    global_leido = False

    for i, row in enumerate(rows):
        vals = [str(v).strip() if v is not None else "" for v in row]

        # ── Horas globales: fila con EXACTAMENTE 4 números en cols 3-6 ──
        # Patrón reconocible: cols 0-2 vacías, cols 3-6 son horas (números > 0)
        if (not global_leido
                and not vals[0] and not vals[1] and not vals[2]
                and _es_numero(row[3]) and _es_numero(row[4]) and _es_numero(row[5])
                and _num(row[3]) > 0 and _num(row[3]) < 200):
            resultado["global"]["presencial"] = _num(row[3])
            resultado["global"]["sinc"]       = _num(row[4])
            resultado["global"]["asinc"]      = _num(row[5])
            if len(row) > 6 and row[6]:
                resultado["global"]["tpe"]    = _num(row[6])
            global_leido = True

        # ── Horas por unidad ──
        # Patrón: col[0] = dígito, col[1] = nombre unidad, col[6] = horas
        if vals[0].isdigit() and vals[1] and _num(row[6]) > 0:
            num    = vals[0]
            nombre = vals[1][:60]
            horas  = _num(row[6])
            clave  = f"Unidad {_romano(num)}"
            if clave not in resultado["unidades"]:
                resultado["unidades"][clave] = {"nombre": nombre, "horas_ped": horas}
            else:
                resultado["unidades"][clave]["horas_ped"] += horas

        # ── Procedimientos e instrumentos del programa ──
        # Fila con texto en col[0] (RA) y col[3] (procedimiento de evaluación)
        if vals[0].startswith("RA") and vals[3]:
            proc_raw  = vals[3]
            instrumento, procedimiento = _separar_instrumento(proc_raw)
            resultado["procedimientos_programa"].append({
                "ra":           vals[0][:80],
                "texto_raw":    proc_raw,
                "procedimiento":procedimiento,
                "instrumento":  instrumento,
            })

        # ── Examen final (fila sin RA pero con procedimiento) ──
        if any(kw in vals[3].lower() for kw in ("examen", "prueba final", "evaluación final")) \
                and not vals[0].startswith("RA"):
            proc_raw = vals[3]
            instrumento, procedimiento = _separar_instrumento(proc_raw)
            resultado["procedimientos_programa"].append({
                "ra":           "Examen final",
                "texto_raw":    proc_raw,
                "procedimiento":procedimiento,
                "instrumento":  instrumento,
            })

    return resultado


def _es_numero(v) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def _separar_instrumento(texto: str) -> tuple[str, str]:
    """
    Separa 'Prueba escrita individual con pauta de corrección'
    en instrumento='Pauta de corrección' y procedimiento='Prueba escrita individual'.
    """
    import re
    texto = texto.strip()
    instrumentos_conocidos = [
        "rúbrica", "rubrica", "pauta de corrección", "pauta de correccion",
        "lista de cotejo", "escala de valoración", "escala de valoracion",
        "escala de apreciación", "escala de apreciacion",
        "pauta de observación", "pauta de observacion",
    ]
    instrumento = ""
    for instr in instrumentos_conocidos:
        if instr in texto.lower():
            instrumento = instr.title()
            # Limpiar "con X", "con rúbrica", etc.
            procedimiento = re.sub(
                r'\s+con\s+' + re.escape(instr), '', texto, flags=re.IGNORECASE
            ).strip()
            return instrumento, procedimiento
    return "", texto


# ── Lectura de Planificación ─────────────────────────────────────────────────

def leer_planificacion(ws_plan) -> dict:
    """
    Recorre 'Planificación por unidades' y acumula:
      - horas reales por unidad (presencial, sinc, asinc, TPE)
      - conteo de recursos por tipo T1-T4 (estimado por keywords en col 8)
      - porcentajes de evaluación declarados (col 15)
    """
    horas   = defaultdict(lambda: {"presencial": 0, "sinc": 0, "asinc": 0, "tpe": 0})
    recursos = defaultdict(lambda: {"T1": 0, "T2": 0, "T3": 0, "T4": 0})
    pcts_eval = []
    unidad_actual = None

    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        # Detectar cambio de unidad
        if row[1] and str(row[1]).strip():
            unidad_actual = _normalizar_unidad(str(row[1]).strip())
        if not unidad_actual:
            continue

        # Horas (cols 16-19 → índices 15-18)
        h_pres = _num(row[15])
        h_sinc = _num(row[16])
        h_asinc= _num(row[17])
        h_tpe  = _num(row[18])

        horas[unidad_actual]["presencial"] += h_pres
        horas[unidad_actual]["sinc"]       += h_sinc
        horas[unidad_actual]["asinc"]      += h_asinc
        horas[unidad_actual]["tpe"]        += h_tpe

        # Porcentaje de evaluación (col 15 → índice 14)
        pct = _num(row[14])
        if 5 <= pct <= 100:
            pcts_eval.append(pct)

        # Conteo de recursos por tipo (col 8 → índice 7: texto de recursos)
        texto_rec = str(row[7]).lower() if row[7] else ""
        momento   = str(row[5]).lower() if row[5] else ""
        recursos[unidad_actual] = _contar_recursos(
            texto_rec, momento, recursos[unidad_actual]
        )

    # Guardar filas raw para validación de procedimientos
    filas_raw = []
    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        if any(v is not None for v in row):
            filas_raw.append(row)

    return {
        "horas":     dict(horas),
        "recursos":  dict(recursos),
        "pcts_eval": pcts_eval,
        "_filas_raw": filas_raw,
    }


# ── Motor de validación ──────────────────────────────────────────────────────

def validar_horas_y_recursos(fuente) -> dict:
    """
    Función principal. Recibe bytes o ruta del xlsx.
    Devuelve dict con:
      {
        "global":   {declarado, calculado, diferencia, ok},
        "unidades": {nombre: {declarado, calculado, diferencia, ok, detalle_horas}},
        "tipos":    {T1..T4: {horas_esperadas, horas_calculadas, recursos_esperados,
                               recursos_encontrados, ok}},
        "porcentajes": {suma, ok, valores},
        "alertas":  [lista de strings],
        "ok_global": bool,
      }
    """
    if isinstance(fuente, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(fuente), data_only=True)
    else:
        wb = openpyxl.load_workbook(fuente, data_only=True)

    if "Síntesis didáctica" not in wb.sheetnames:
        return {"error": "Falta la hoja 'Síntesis didáctica'"}
    if "Planificación por unidades" not in wb.sheetnames:
        return {"error": "Falta la hoja 'Planificación por unidades'"}

    sint = leer_sintesis(wb["Síntesis didáctica"])
    plan = leer_planificacion(wb["Planificación por unidades"])

    alertas = []
    resultado = {}

    # ── 1. Validación global de horas ────────────────────────────────────────
    g_dec  = sint["global"]
    g_calc = _sumar_todas(plan["horas"])

    global_ok = True
    for tipo in ("presencial", "sinc", "asinc", "tpe"):
        diff = abs(g_dec[tipo] - g_calc[tipo])
        if diff > TOLERANCIA_HORAS:
            alertas.append(
                f"⚠️  GLOBAL — Horas {tipo}: declaradas={g_dec[tipo]} "
                f"calculadas={g_calc[tipo]} (Δ={diff:.1f})"
            )
            global_ok = False

    resultado["global"] = {
        "declarado":  g_dec,
        "calculado":  g_calc,
        "ok":         global_ok,
    }

    # ── 2. Validación POR UNIDAD ─────────────────────────────────────────────
    resultado["unidades"] = {}
    unidades_ok = True

    for clave, info in sint["unidades"].items():
        dec_ped = info["horas_ped"]
        h_real  = plan["horas"].get(clave, {"presencial":0,"sinc":0,"asinc":0,"tpe":0})
        calc_ped = h_real["presencial"] + h_real["sinc"] + h_real["asinc"]
        diff     = abs(dec_ped - calc_ped)
        ok_u     = diff <= TOLERANCIA_HORAS

        if not ok_u:
            alertas.append(
                f"❌  {clave} «{info['nombre'][:40]}»: "
                f"declaradas={dec_ped} hrs pedagógicas, "
                f"calculadas={calc_ped:.1f} hrs (Δ={diff:.1f})"
            )
            unidades_ok = False
        else:
            alertas.append(
                f"✅  {clave}: {dec_ped} hrs declaradas = {calc_ped:.1f} calculadas"
            )

        resultado["unidades"][clave] = {
            "nombre":         info["nombre"],
            "declarado_ped":  dec_ped,
            "calculado_ped":  round(calc_ped, 2),
            "diferencia":     round(diff, 2),
            "ok":             ok_u,
            "detalle": {
                "presencial": h_real["presencial"],
                "sinc":       h_real["sinc"],
                "asinc":      h_real["asinc"],
                "tpe":        h_real["tpe"],
            },
        }

    # ── 3. Validación distribución T1-T4 ─────────────────────────────────────
    h_total_asinc = g_calc["asinc"]
    h_total_sinc  = g_calc["sinc"]
    h_total_pres  = g_calc["presencial"]
    h_total_tpe   = g_calc["tpe"]
    h_pres_sinc   = h_total_pres + h_total_sinc

    esperado = {
        "T1": round(h_total_asinc * 0.20, 2),
        "T2": round(h_total_asinc * 0.80, 2),
        "T3": round(h_pres_sinc   * 0.40, 2),
        "T4": round(h_total_tpe   * 0.20, 2),
    }

    rec_total = defaultdict(int)
    for u_rec in plan["recursos"].values():
        for t in ("T1","T2","T3","T4"):
            rec_total[t] += u_rec[t]

    resultado["tipos"] = {}
    for t in ("T1","T2","T3","T4"):
        h_esp = esperado[t]
        rec_esp  = round(h_esp / HORAS_POR_RECURSO[t], 1) if h_esp > 0 else 0
        rec_real = rec_total[t]
        ok_t = abs(rec_real - rec_esp) <= max(1, rec_esp * 0.25)  # ±25% o ±1

        if not ok_t and h_esp > 0:
            alertas.append(
                f"⚠️  {t}: esperados ≈{rec_esp:.0f} recursos "
                f"(para {h_esp} hrs), encontrados={rec_real}"
            )

        resultado["tipos"][t] = {
            "horas_esperadas":     h_esp,
            "recursos_esperados":  round(rec_esp),
            "recursos_encontrados":rec_real,
            "ok":                  ok_t,
        }

    # ── 4. Porcentajes de evaluación suman 100% ──────────────────────────────
    pcts = plan["pcts_eval"]
    suma_pct = sum(pcts)
    pct_ok = abs(suma_pct - 100) <= 1

    if not pct_ok:
        alertas.append(
            f"❌  Porcentajes de evaluación suman {suma_pct}% (debe ser 100%)"
        )
    else:
        alertas.append(f"✅  Porcentajes de evaluación: {suma_pct}%")

    resultado["porcentajes"] = {
        "valores": pcts,
        "suma":    suma_pct,
        "ok":      pct_ok,
    }

    # ── 5. Validación procedimientos e instrumentos vs programa ─────────────
    resultado["procedimientos"] = _validar_procedimientos(
        sint["procedimientos_programa"], plan
    )
    for obs in resultado["procedimientos"].get("alertas", []):
        alertas.append(obs)

    proc_ok = resultado["procedimientos"].get("ok", True)

    resultado["alertas"]   = alertas
    resultado["ok_global"] = global_ok and unidades_ok and pct_ok and proc_ok

    return resultado


def _validar_procedimientos(declarados: list, plan: dict) -> dict:
    """
    Verifica que cada procedimiento/instrumento declarado en el programa
    aparezca al menos una vez en la planificación (col 12 y col 14).
    """
    if not declarados:
        return {"ok": True, "alertas": [], "detalle": []}

    # Recopilar todos los procedimientos e instrumentos de la planificación
    proc_plan  = set()
    instr_plan = set()
    for row in plan.get("_filas_raw", []):
        if row[11]:  # col 12 → índice 11
            proc_plan.add(str(row[11]).strip().lower())
        if row[13]:  # col 14 → índice 13
            instr_plan.add(str(row[13]).strip().lower())

    alertas = []
    detalle = []
    ok = True

    for item in declarados:
        proc_dec  = item["procedimiento"].lower()
        instr_dec = item["instrumento"].lower()
        raw       = item["texto_raw"]

        # Búsqueda flexible (la planificación puede abreviar)
        proc_encontrado  = any(_similar(proc_dec, p) for p in proc_plan)
        instr_encontrado = not instr_dec or any(_similar(instr_dec, p) for p in instr_plan)

        estado = proc_encontrado and instr_encontrado
        if not estado:
            ok = False
            if not proc_encontrado:
                alertas.append(
                    f"❌  Procedimiento del programa NO encontrado en planificación: "
                    f"«{item['procedimiento']}»"
                )
            if instr_dec and not instr_encontrado:
                alertas.append(
                    f"❌  Instrumento del programa NO encontrado en planificación: "
                    f"«{item['instrumento']}»"
                )
        else:
            alertas.append(
                f"✅  Procedimiento verificado: «{item['procedimiento']}»"
                + (f" + «{item['instrumento']}»" if item['instrumento'] else "")
            )

        detalle.append({
            "ra":          item["ra"],
            "declarado":   raw,
            "proc_ok":     proc_encontrado,
            "instr_ok":    instr_encontrado,
        })

    return {"ok": ok, "alertas": alertas, "detalle": detalle}


def _similar(a: str, b: str) -> bool:
    """Verifica si a está contenido en b o viceversa (fuzzy básico)."""
    a, b = a.strip().lower(), b.strip().lower()
    if not a or not b:
        return False
    # Coincidencia exacta o parcial (ej. "prueba escrita" en "prueba escrita individual")
    return a in b or b in a or _tokens_comunes(a, b) >= 0.6


def _tokens_comunes(a: str, b: str) -> float:
    """Fracción de tokens de `a` presentes en `b`."""
    ta = set(a.split())
    tb = set(b.split())
    if not ta:
        return 0.0
    return len(ta & tb) / len(ta)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _num(v) -> float:
    try:
        return float(v) if v else 0.0
    except (TypeError, ValueError):
        return 0.0


def _romano(n: str) -> str:
    mapa = {"1":"I","2":"II","3":"III","4":"IV","5":"V",
            "6":"VI","7":"VII","8":"VIII","9":"IX","10":"X"}
    return mapa.get(n, n)


def _normalizar_unidad(texto: str) -> str:
    """Normaliza 'Unidad I', 'Unidad 1', 'Módulo I', etc. a forma canónica."""
    import re
    texto = texto.strip()
    # Ya en forma "Unidad X"
    m = re.match(r'(Unidad|Módulo|Modulo)\s+([IVXivx\d]+)', texto, re.IGNORECASE)
    if m:
        prefijo = "Unidad"
        num = m.group(2).upper()
        # Convertir dígito a romano si es necesario
        if num.isdigit():
            num = _romano(num)
        return f"{prefijo} {num}"
    return texto


def _sumar_todas(horas_por_unidad: dict) -> dict:
    total = {"presencial": 0.0, "sinc": 0.0, "asinc": 0.0, "tpe": 0.0}
    for h in horas_por_unidad.values():
        for k in total:
            total[k] += h.get(k, 0)
    return {k: round(v, 2) for k, v in total.items()}


def _contar_recursos(texto_rec: str, momento: str, acum: dict) -> dict:
    """
    Estima el tipo de recurso por keywords en el texto de la celda.
    Incrementa el contador del tipo correspondiente.
    """
    result = dict(acum)

    keywords_t1 = ["video", "cápsula", "capsula", "screencast", "podcast",
                   "interactiva", "hotspot", "línea de tiempo", "h5p"]
    keywords_t2 = ["tarea", "trabajo", "foro", "quiz", "cuestionario",
                   "infografía", "glosario", "mapa conceptual", "descargable"]
    keywords_t3 = ["presentación", "presentacion", "diapositiva", "guía de facilitación",
                   "guia de facilitacion", "dinámica", "dinamica", "plantilla"]
    keywords_t4 = ["guía de estudio", "guia de estudio", "actividad preparatoria",
                   "video tpe", "cuestionario tpe", "trabajo independiente"]

    # Asignar por momento si hay ambigüedad
    es_tpe = "trabajo independiente" in momento or "tpe" in momento
    es_sinc_pres = momento in ("desarrollo", "presencial", "sincrónico", "sincronica")

    encontrado = False
    for kw in keywords_t1:
        if kw in texto_rec:
            result["T1"] = result.get("T1", 0) + 1
            encontrado = True
            break

    if not encontrado:
        for kw in keywords_t2:
            if kw in texto_rec:
                result["T2"] = result.get("T2", 0) + 1
                encontrado = True
                break

    if not encontrado:
        for kw in keywords_t3:
            if kw in texto_rec:
                result["T3"] = result.get("T3", 0) + 1
                encontrado = True
                break

    if not encontrado and es_tpe:
        result["T4"] = result.get("T4", 0) + 1

    return result


# ── Reporte legible ──────────────────────────────────────────────────────────

def reporte_texto(resultado: dict) -> str:
    """Genera un reporte de texto plano para mostrar en Streamlit."""
    if "error" in resultado:
        return f"ERROR: {resultado['error']}"

    lines = ["═" * 60, "VALIDACIÓN DE HORAS Y RECURSOS — UST DEL", "═" * 60]

    # Global
    g = resultado.get("global", {})
    lines.append("\n📊 HORAS GLOBALES")
    dec = g.get("declarado", {})
    cal = g.get("calculado", {})
    for t in ("presencial", "sinc", "asinc", "tpe"):
        icono = "✅" if abs(dec.get(t,0) - cal.get(t,0)) <= TOLERANCIA_HORAS else "❌"
        lines.append(f"  {icono} {t:12s}: declaradas={dec.get(t,0):5.1f}  calculadas={cal.get(t,0):5.1f}")

    # Por unidad
    lines.append("\n📚 HORAS POR UNIDAD (lectivas = presencial + sinc + asínc)")
    for clave, u in resultado.get("unidades", {}).items():
        icono = "✅" if u["ok"] else "❌"
        d = u["detalle"]
        lines.append(
            f"  {icono} {clave}: declaradas={u['declarado_ped']} | "
            f"calculadas={u['calculado_ped']} "
            f"(pres={d['presencial']} sinc={d['sinc']} asinc={d['asinc']} TPE={d['tpe']})"
        )
        if not u["ok"]:
            lines.append(f"       ↳ Diferencia: {u['diferencia']:.1f} hrs — REVISAR")

    # Tipos T1-T4
    lines.append("\n🎯 DISTRIBUCIÓN DE RECURSOS T1-T4 (estándar V8)")
    for t, info in resultado.get("tipos", {}).items():
        icono = "✅" if info["ok"] else "⚠️ "
        lines.append(
            f"  {icono} {t}: {info['horas_esperadas']} hrs esperadas → "
            f"~{info['recursos_esperados']} recursos | encontrados={info['recursos_encontrados']}"
        )

    # Porcentajes
    pct = resultado.get("porcentajes", {})
    icono = "✅" if pct.get("ok") else "❌"
    lines.append(f"\n📋 PORCENTAJES DE EVALUACIÓN")
    lines.append(f"  {icono} Suma: {pct.get('suma',0)}%  (valores: {pct.get('valores',[])})")

    # Resumen
    lines.append("\n" + "─" * 60)
    estado = "✅ PLANIFICACIÓN VÁLIDA" if resultado.get("ok_global") else "❌ HAY DIFERENCIAS — REVISAR"
    lines.append(f"RESULTADO GENERAL: {estado}")

    return "\n".join(lines)


# ── CLI rápido para pruebas ──────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    ruta = sys.argv[1] if len(sys.argv) > 1 else \
        "/Users/mac/Desktop/DEL/PLA-S0113 Sistema de producción de frutales sostenibles./PLA-S0113_Sistema de producción de frutales sostenibles_Rev DI (2).xlsx"
    res = validar_horas_y_recursos(ruta)
    print(reporte_texto(res))
    print("\n── Alertas detalladas ──")
    for a in res.get("alertas", []):
        print(" ", a)
