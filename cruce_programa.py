"""
cruce_programa.py — Cruce automático PDF programa vs planificación xlsx
Procesador DEL · UST 2026-1

Extrae del programa oficial (PDF):
  - Datos de identificación (escuela, código, créditos, horas, jornada)
  - Competencias (genéricas y específicas)
  - Unidades con nombres y horas
  - Resultados de aprendizaje (RA)
  - Procedimientos de evaluación con porcentajes
  - Bibliografía básica y complementaria

Cruza contra la síntesis y planificación del xlsx y emite veredicto
Si / Parcialmente / No para los criterios C1-C15 de la Escala de Apreciación.
"""

from __future__ import annotations
import re
from io import BytesIO
from dataclasses import dataclass, field

import pdfplumber
import openpyxl


# ── Estructura del programa ──────────────────────────────────────────────────

@dataclass
class ProgramaOficial:
    escuela:           str = ""
    programa:          str = ""
    nombre_asignatura: str = ""
    codigo:            str = ""
    version_anio:      str = ""
    creditos:          str = ""
    jornada:           str = ""
    area_ocde:         str = ""
    requisitos:        str = ""
    modalidad:         str = ""
    horas: dict        = field(default_factory=dict)   # pres, sinc, asinc, tpe, total
    competencias:      list = field(default_factory=list)
    unidades:          list = field(default_factory=list)  # [{num, nombre, horas, contenidos}]
    ra:                list = field(default_factory=list)  # [{num, texto}]
    evaluaciones:      list = field(default_factory=list)  # [{proc, pct}]
    bibliografia_basica: list = field(default_factory=list)
    bibliografia_comp:   list = field(default_factory=list)


# ── Extractor de PDF ─────────────────────────────────────────────────────────

def extraer_programa_pdf(fuente) -> ProgramaOficial:
    """
    Lee el PDF del programa oficial y extrae los campos clave.
    fuente: bytes o ruta de archivo.
    """
    if isinstance(fuente, (bytes, bytearray)):
        pdf_file = BytesIO(fuente)
    else:
        pdf_file = fuente

    prog = ProgramaOficial()

    with pdfplumber.open(pdf_file) as pdf:
        texto_completo = "\n".join(
            p.extract_text() or "" for p in pdf.pages
        )

    lineas = [l.strip() for l in texto_completo.splitlines() if l.strip()]
    texto  = texto_completo

    # ── Identificación ───────────────────────────────────────────────────────
    prog.escuela           = _extraer_campo(texto, r"ESCUELA O UNIDAD\s+(.+)")
    prog.programa          = _extraer_campo(texto, r"PROGRAMA\s+(.+)")
    prog.nombre_asignatura = _extraer_campo(texto, r"NOMBRE DE LA ASIGNATURA\s+(.+)")
    prog.codigo            = _extraer_campo(texto, r"CÓDIGO\s+([A-Z]{2,4}-[A-Z]\d{3,5})")
    prog.version_anio      = _extraer_campo(texto, r"VERSIÓN AÑO\s+(\d{4})")
    prog.creditos          = _extraer_campo(texto, r"CRÉDITOS\s*ACADÉMICOS\s+(\d+)")
    prog.jornada           = _extraer_campo(texto, r"MODALIDAD\s+(.+?)(?:\n|CARÁCTER)")
    prog.area_ocde         = _extraer_campo(texto, r"ÁREA DE CONOCIMIENTO\s+(.+)")
    prog.requisitos        = _extraer_campo(texto, r"PRE-REQUISITOS\s+(.+)")

    # ── Horas ────────────────────────────────────────────────────────────────
    # Patrón: tabla de horas (cronológicas / lectivas / distribución)
    prog.horas = _extraer_horas(texto)

    # ── Competencias ─────────────────────────────────────────────────────────
    prog.competencias = _extraer_competencias(texto)

    # ── Unidades ─────────────────────────────────────────────────────────────
    prog.unidades = _extraer_unidades(texto)

    # ── Resultados de Aprendizaje ─────────────────────────────────────────────
    prog.ra = _extraer_ra(texto)

    # ── Evaluaciones ─────────────────────────────────────────────────────────
    prog.evaluaciones = _extraer_evaluaciones(texto)

    # ── Bibliografía ─────────────────────────────────────────────────────────
    prog.bibliografia_basica, prog.bibliografia_comp = _extraer_bibliografia(texto)

    return prog


# ── Comparador xlsx vs programa ──────────────────────────────────────────────

def cruzar_con_planificacion(
    prog: ProgramaOficial,
    fuente_xlsx,
) -> dict:
    """
    Compara el programa oficial con el xlsx de planificación.
    Retorna dict con criterios C1-C15 verificados.
    """
    if isinstance(fuente_xlsx, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(fuente_xlsx), data_only=True)
    else:
        wb = openpyxl.load_workbook(fuente_xlsx, data_only=True)

    sint = _leer_sintesis_xlsx(wb)
    plan = _leer_planificacion_xlsx(wb)

    resultados = {}
    alertas    = []

    def _crit(cid, estado, obs):
        resultados[cid] = {"estado": estado, "observacion": obs[:150]}
        if estado != "SI":
            alertas.append(f"C{cid:02d} {estado}: {obs[:100]}")

    # C1 — Escuela
    _crit(1, _comparar(sint.get("escuela",""), prog.escuela),
          f"Plan: «{sint.get('escuela','')}» | Prog: «{prog.escuela}»")

    # C2 — Programa académico
    _crit(2, _comparar(sint.get("programa",""), prog.programa),
          f"Plan: «{sint.get('programa','')}» | Prog: «{prog.programa}»")

    # C3 — Nombre asignatura
    _crit(3, _comparar(sint.get("nombre_asignatura",""), prog.nombre_asignatura),
          f"Plan: «{sint.get('nombre_asignatura','')}» | Prog: «{prog.nombre_asignatura}»")

    # C4 — Jornada/Modalidad
    _crit(4, _comparar(sint.get("jornada",""), prog.jornada),
          f"Plan: «{sint.get('jornada','')}» | Prog: «{prog.jornada}»")

    # C5 — Código
    _crit(5, _comparar(sint.get("codigo",""), prog.codigo),
          f"Plan: «{sint.get('codigo','')}» | Prog: «{prog.codigo}»")

    # C6 — Versión año
    _crit(6, _comparar(str(sint.get("version","")), prog.version_anio),
          f"Plan: «{sint.get('version','')}» | Prog: «{prog.version_anio}»")

    # C7 — Créditos
    _crit(7, _comparar(str(sint.get("creditos","")), prog.creditos),
          f"Plan: «{sint.get('creditos','')}» | Prog: «{prog.creditos}»")

    # C8 — Carga académica (horas)
    h_plan = sint.get("horas", {})
    h_prog = prog.horas
    horas_ok = _comparar_horas(h_plan, h_prog)
    _crit(8, horas_ok["estado"],
          f"Pres:{h_plan.get('presencial',0)} vs {h_prog.get('presencial',0)} | "
          f"Sinc:{h_plan.get('sinc',0)} vs {h_prog.get('sinc',0)} | "
          f"Asinc:{h_plan.get('asinc',0)} vs {h_prog.get('asinc',0)} | "
          f"TPE:{h_plan.get('tpe',0)} vs {h_prog.get('tpe',0)}")

    # C9 — Competencias
    comp_ok = bool(sint.get("competencias")) and bool(prog.competencias)
    _crit(9, "SI" if comp_ok else "PARCIALMENTE",
          f"{len(prog.competencias)} competencia(s) en programa, "
          f"{'presentes' if comp_ok else 'revisar'} en planificación")

    # C10 — Nombres de unidades
    u_plan = {u["nombre"].lower(): u for u in plan.get("unidades", [])}
    u_prog = prog.unidades
    nombres_ok = all(
        any(_similar(u["nombre"], k) for k in u_plan)
        for u in u_prog
    )
    _crit(10, "SI" if nombres_ok else "PARCIALMENTE",
          f"{len(u_prog)} unidad(es) en programa | "
          + (f"nombres coinciden" if nombres_ok else
             f"revisar: {', '.join(u['nombre'][:30] for u in u_prog[:3])}"))

    # C11 — Horas por unidad
    horas_u_ok = True
    horas_u_obs = []
    for u_p in u_prog:
        u_match = next(
            (u for u in plan.get("unidades", [])
             if _similar(u["nombre"], u_p["nombre"])),
            None
        )
        if u_match:
            diff = abs(float(u_p.get("horas", 0)) - float(u_match.get("horas_calc", 0)))
            if diff > 1.0:
                horas_u_ok = False
                horas_u_obs.append(
                    f"{u_p['nombre'][:25]}: prog={u_p.get('horas',0)} plan={u_match.get('horas_calc',0):.1f}"
                )
    _crit(11, "SI" if horas_u_ok else "NO",
          "Horas por unidad correctas" if horas_u_ok else
          " | ".join(horas_u_obs[:3]))

    # C12 — RA redactados fielmente
    ra_ok = _verificar_ra(prog.ra, plan.get("ra_planificacion", []))
    _crit(12, ra_ok["estado"], ra_ok["obs"])

    # C13 — Procedimientos de evaluación
    proc_ok = _verificar_procedimientos(prog.evaluaciones, plan.get("evaluaciones", []))
    _crit(13, proc_ok["estado"], proc_ok["obs"])

    # C14 — Porcentajes
    pct_prog  = sum(float(e.get("pct", 0)) for e in prog.evaluaciones if e.get("pct"))
    pct_plan  = plan.get("suma_pct", 0)
    pct_match = abs(pct_prog - pct_plan) <= 2
    _crit(14, "SI" if pct_match else "NO",
          f"Programa: {pct_prog}% | Planificación: {pct_plan}%")

    # C15 — RA diferenciados con porcentajes
    ra_con_pct = sum(1 for e in prog.evaluaciones if e.get("pct") and e.get("proc"))
    _crit(15, "SI" if ra_con_pct >= len(prog.ra) else "PARCIALMENTE",
          f"{ra_con_pct} procedimiento(s) con porcentaje para {len(prog.ra)} RA(s)")

    return {
        "criterios": resultados,
        "alertas":   alertas,
        "programa":  prog,
        "ok_global": all(v["estado"] == "SI" for v in resultados.values()),
    }


# ── Helpers de extracción ────────────────────────────────────────────────────

def _extraer_campo(texto: str, patron: str, default: str = "") -> str:
    m = re.search(patron, texto, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return default


def _extraer_horas(texto: str) -> dict:
    """Extrae horas de la tabla de carga académica del PDF."""
    horas = {}

    # Buscar patrones numéricos en contexto de horas
    m_pres   = re.search(r"Cátedra\s+[A-Za-z/]+\s+Terreno[^\d]*(\d+)\s+[^\d]*(\d+)", texto)
    m_sinc   = re.search(r"Sincrónicas?\s+(\d+)", texto, re.IGNORECASE)
    m_asinc  = re.search(r"Asincrónicas?\s+(\d+)", texto, re.IGNORECASE)
    m_tpe    = re.search(r"TPE[^\d]*(\d+)", texto, re.IGNORECASE)
    m_cat    = re.search(r"Cátedra\s+(\d+)", texto, re.IGNORECASE)
    m_terr   = re.search(r"Terreno\s+(\d+)", texto, re.IGNORECASE)

    catedral = int(m_cat.group(1)) if m_cat else 0
    terreno  = int(m_terr.group(1)) if m_terr else 0

    horas["presencial"] = catedral + terreno
    horas["sinc"]       = int(m_sinc.group(1))  if m_sinc  else 0
    horas["asinc"]      = int(m_asinc.group(1)) if m_asinc else 0
    horas["tpe"]        = int(m_tpe.group(1))   if m_tpe   else 0
    horas["total_lec"]  = horas["presencial"] + horas["sinc"] + horas["asinc"]

    return horas


def _extraer_competencias(texto: str) -> list:
    comps = []
    # Buscar competencias genéricas y específicas
    for patron in [
        r"COMPETENCIAS GENÉRICAS\s+(.+?)(?=COMPETENCIAS ESPECÍFICAS|RESULTADOS)",
        r"COMPETENCIAS ESPECÍFICAS\s+(.+?)(?=RESULTADOS|4\.)",
    ]:
        m = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
        if m:
            comps.append(m.group(1).strip()[:300])
    return comps


def _extraer_unidades(texto: str) -> list:
    """Extrae unidades con nombre y horas del PDF."""
    unidades = []
    patron = r"(?:Nombre de la )?Unidad\s+([IVX\d]+)[^\n]*\n(.+?)\((\d+)\s*horas?\)"
    for m in re.finditer(patron, texto, re.IGNORECASE):
        unidades.append({
            "num":    m.group(1).strip(),
            "nombre": m.group(2).strip()[:80],
            "horas":  int(m.group(3)),
        })

    # Fallback: buscar líneas "Unidad N ... (N horas)"
    if not unidades:
        for m in re.finditer(r"Unidad\s+([IVX\d]+)\s+(.+?)\s+\((\d+)\s*horas?\)", texto):
            unidades.append({
                "num":    m.group(1).strip(),
                "nombre": m.group(2).strip()[:80],
                "horas":  int(m.group(3)),
            })
    return unidades


def _extraer_ra(texto: str) -> list:
    """Extrae resultados de aprendizaje numerados del PDF."""
    ra_list = []

    # Buscar RA numerados tipo "1 Analizar..." o "RA1 ..."
    bloque = re.search(
        r"RESULTADOS DE APRENDIZAJE\s+(.*?)(?=ESTRATEGIA|4\.|5\.|\Z)",
        texto, re.IGNORECASE | re.DOTALL
    )
    if bloque:
        texto_ra = bloque.group(1)
        for m in re.finditer(r"(\d)\s+([A-ZÁÉÍÓÚÑ][^0-9]{20,})", texto_ra):
            ra_list.append({
                "num":   m.group(1),
                "texto": m.group(2).strip()[:300],
            })

    return ra_list


def _extraer_evaluaciones(texto: str) -> list:
    """Extrae procedimientos de evaluación con porcentajes del PDF."""
    evals = []

    # Buscar tabla de evaluaciones
    bloque = re.search(
        r"PROCEDIMIENTOS EVALUATIVOS\s*(.*?)(?=ESTRATEGIAS METODOLÓGICAS|6\.|RECURSOS|\Z)",
        texto, re.IGNORECASE | re.DOTALL
    )
    if bloque:
        texto_eval = bloque.group(1)
        # Buscar líneas con % (ej. "Prueba escrita ... 40%")
        for m in re.finditer(
            r"([A-ZÁÉÍÓÚ][^%\n]{10,60}?)\s+(\d{1,3})\s*%", texto_eval
        ):
            evals.append({
                "proc": m.group(1).strip(),
                "pct":  float(m.group(2)),
            })

    return evals


def _extraer_bibliografia(texto: str) -> tuple[list, list]:
    """Extrae bibliografía básica y complementaria del PDF."""
    basica = []
    compl  = []

    bloque_b = re.search(
        r"BIBLIOGRAFÍA BÁSICA\s*(.*?)(?=BIBLIOGRAFÍA COMPLEMENTARIA|INDAP|\Z)",
        texto, re.IGNORECASE | re.DOTALL
    )
    bloque_c = re.search(
        r"BIBLIOGRAFÍA COMPLEMENTARIA\s*(.*?)(?=INDAP|ODEPA|\Z)",
        texto, re.IGNORECASE | re.DOTALL
    )

    if bloque_b:
        for linea in bloque_b.group(1).splitlines():
            l = linea.strip()
            if l and len(l) > 15:
                basica.append(l[:200])

    if bloque_c:
        for linea in bloque_c.group(1).splitlines():
            l = linea.strip()
            if l and len(l) > 15:
                compl.append(l[:200])

    return basica[:10], compl[:10]


# ── Helpers de comparación ────────────────────────────────────────────────────

def _comparar(a: str, b: str) -> str:
    if not a or not b:
        return "PARCIALMENTE"
    a_n = _norm(a)
    b_n = _norm(b)
    if a_n == b_n:
        return "SI"
    if _similar(a_n, b_n):
        return "PARCIALMENTE"
    return "NO"


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _similar(a: str, b: str) -> bool:
    a, b = _norm(a), _norm(b)
    if not a or not b:
        return False
    if a in b or b in a:
        return True
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    if not tokens_a:
        return False
    return len(tokens_a & tokens_b) / len(tokens_a) >= 0.6


def _comparar_horas(h_plan: dict, h_prog: dict) -> dict:
    diffs = []
    for tipo in ("presencial", "sinc", "asinc", "tpe"):
        p = float(h_plan.get(tipo, 0))
        r = float(h_prog.get(tipo, 0))
        if r > 0 and abs(p - r) > 1:
            diffs.append(f"{tipo}:{p}vs{r}")
    if not diffs:
        return {"estado": "SI"}
    return {"estado": "PARCIALMENTE" if len(diffs) <= 2 else "NO"}


def _verificar_ra(ra_prog: list, ra_plan: list) -> dict:
    if not ra_prog:
        return {"estado": "PARCIALMENTE", "obs": "RA no extraíbles del PDF (verificar manualmente)"}
    if not ra_plan:
        return {"estado": "NO", "obs": "RA no encontrados en planificación"}

    ok_count = 0
    for ra_p in ra_prog:
        texto_p = _norm(ra_p["texto"])
        if any(_similar(texto_p, _norm(r)) for r in ra_plan):
            ok_count += 1

    if ok_count == len(ra_prog):
        return {"estado": "SI",     "obs": f"Los {len(ra_prog)} RA coinciden con el programa"}
    if ok_count >= len(ra_prog) // 2:
        return {"estado": "PARCIALMENTE", "obs": f"{ok_count}/{len(ra_prog)} RA coinciden"}
    return {"estado": "NO", "obs": f"Solo {ok_count}/{len(ra_prog)} RA coinciden con programa"}


def _verificar_procedimientos(evals_prog: list, evals_plan: list) -> dict:
    if not evals_prog:
        return {"estado": "PARCIALMENTE", "obs": "Procedimientos no extraíbles del PDF"}
    if not evals_plan:
        return {"estado": "NO", "obs": "Sin evaluaciones en planificación"}

    procs_plan = [_norm(e.get("proc", "") or e.get("procedimiento", "")) for e in evals_plan]
    ok_count = 0
    faltantes = []
    for ev in evals_prog:
        proc_p = _norm(ev["proc"])
        if any(_similar(proc_p, p) for p in procs_plan):
            ok_count += 1
        else:
            faltantes.append(ev["proc"][:40])

    if ok_count == len(evals_prog):
        return {"estado": "SI", "obs": f"Los {len(evals_prog)} procedimientos coinciden"}
    return {
        "estado": "PARCIALMENTE" if ok_count > 0 else "NO",
        "obs": f"{ok_count}/{len(evals_prog)} procedimientos OK. Falta: {', '.join(faltantes)}"
    }


# ── Lectura síntesis xlsx ────────────────────────────────────────────────────

def _leer_sintesis_xlsx(wb) -> dict:
    if "Síntesis didáctica" not in wb.sheetnames:
        return {}
    ws = wb["Síntesis didáctica"]
    d  = {}
    rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(rows):
        v = [str(c).strip() if c else "" for c in row]

        # Escuela
        if v[0].lower() == "agronomía" or (i > 0 and "escuela" in str(rows[i-1][0] or "").lower()):
            if not d.get("escuela") and v[0]:
                d["escuela"] = v[0]

        # Programa, nombre, jornada (fila con 4 datos)
        if v[0] and v[3] and len(v[0]) > 3 and len(v[3]) > 3:
            if not d.get("nombre_asignatura") and len(v[3]) > 10:
                d["nombre_asignatura"] = v[3]
            if not d.get("escuela"):
                d["escuela"] = v[0]

        # Código, versión, créditos
        codigo_m = re.search(r"[A-Z]{2,4}-[A-Z]\d{3,5}", str(row[0] or ""))
        if codigo_m:
            d["codigo"] = codigo_m.group(0)

        if v[1] and re.match(r"^\d{4}$", v[1]):
            d["version"] = v[1]
        if v[2] and re.match(r"^\d$", v[2]):
            d["creditos"] = v[2]

        # Horas globales: fila con 4 números en cols 3-6
        if (not d.get("horas") and not v[0] and not v[1] and not v[2]
                and _is_num(row[3]) and _is_num(row[4]) and _is_num(row[5])
                and float(row[3] or 0) > 0):
            d["horas"] = {
                "presencial": float(row[3] or 0),
                "sinc":       float(row[4] or 0),
                "asinc":      float(row[5] or 0),
                "tpe":        float(row[6] or 0) if row[6] else 0,
            }

        # RA en síntesis
        if v[0].startswith("RA") and v[3]:
            d.setdefault("ra_procedimientos", []).append({
                "ra": v[0][:120], "procedimiento": v[3]
            })

    # Porcentaje total
    pct_total = 0.0
    for row in rows:
        for v in row:
            if isinstance(v, (int, float)) and 5 <= v <= 100:
                pct_total += v
    d["suma_pct"] = pct_total

    return d


def _leer_planificacion_xlsx(wb) -> dict:
    if "Planificación por unidades" not in wb.sheetnames:
        return {}
    ws  = wb["Planificación por unidades"]
    plan = {
        "unidades":        [],
        "ra_planificacion":[],
        "evaluaciones":    [],
        "suma_pct":        0.0,
    }
    unidad_actual = ""
    horas_u       = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(v for v in row):
            continue
        if row[1]:
            if unidad_actual and unidad_actual in horas_u:
                plan["unidades"].append({
                    "nombre":     unidad_actual,
                    "horas_calc": horas_u[unidad_actual],
                })
            unidad_actual = str(row[1]).strip()
            horas_u.setdefault(unidad_actual, 0)

        # Horas lectivas
        h = sum(_to_f(row[i]) for i in (15, 16, 17))
        if unidad_actual:
            horas_u[unidad_actual] = horas_u.get(unidad_actual, 0) + h

        # RA en columna 1
        if row[0] and str(row[0]).startswith("RA"):
            txt = str(row[0]).strip()
            if txt not in plan["ra_planificacion"]:
                plan["ra_planificacion"].append(txt)

        # Evaluaciones
        tipo_ev = str(row[10] or "").strip()
        proc_ev = str(row[11] or "").strip()
        instr   = str(row[13] or "").strip()
        pct     = _to_f(row[14])
        if tipo_ev or proc_ev:
            plan["evaluaciones"].append({
                "tipo": tipo_ev, "procedimiento": proc_ev,
                "instrumento": instr, "pct": pct,
            })
        if 5 <= pct <= 100:
            plan["suma_pct"] += pct

    # Añadir última unidad
    if unidad_actual and unidad_actual in horas_u:
        plan["unidades"].append({
            "nombre":     unidad_actual,
            "horas_calc": horas_u[unidad_actual],
        })

    return plan


def _is_num(v) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False

def _to_f(v) -> float:
    try:
        return float(v) if v else 0.0
    except (TypeError, ValueError):
        return 0.0


# ── Reporte ──────────────────────────────────────────────────────────────────

def reporte_cruce(resultado: dict) -> str:
    prog = resultado.get("programa")
    lines = [
        "═" * 60,
        "CRUCE PROGRAMA OFICIAL vs PLANIFICACIÓN",
        "═" * 60,
        f"Asignatura: {prog.nombre_asignatura}" if prog else "",
        f"Código:     {prog.codigo}" if prog else "",
        "",
    ]
    iconos = {"SI": "✅", "PARCIALMENTE": "⚠️ ", "NO": "❌"}
    for cid, res in sorted(resultado["criterios"].items()):
        icono = iconos.get(res["estado"], "❓")
        lines.append(f"  {icono} C{cid:02d}: {res['observacion']}")

    ok = sum(1 for r in resultado["criterios"].values() if r["estado"] == "SI")
    total = len(resultado["criterios"])
    lines += [
        "",
        f"RESULTADO C1-C15: {ok}/{total} criterios SI",
        "═" * 60,
    ]
    return "\n".join(lines)


# ── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    pdf_path  = (sys.argv[1] if len(sys.argv) > 1 else
        "/Users/mac/Desktop/DEL/PLA-S0113 Sistema de producción de frutales sostenibles./"
        "PLA-S0113_Sistema de producción de frutales sostenibles (4).pdf")
    xlsx_path = (sys.argv[2] if len(sys.argv) > 2 else
        "/Users/mac/Desktop/DEL/PLA-S0113 Sistema de producción de frutales sostenibles./"
        "PLA-S0113_Sistema de producción de frutales sostenibles_Rev DI (2).xlsx")

    print("Extrayendo programa oficial...")
    prog = extraer_programa_pdf(pdf_path)
    print(f"  Asignatura: {prog.nombre_asignatura}")
    print(f"  Código:     {prog.codigo}")
    print(f"  Horas:      {prog.horas}")
    print(f"  Unidades:   {[u['nombre'][:30] for u in prog.unidades]}")
    print(f"  RA:         {len(prog.ra)} resultados")
    print(f"  Eval:       {prog.evaluaciones}")

    print("\nCruzando con planificación...")
    resultado = cruzar_con_planificacion(prog, xlsx_path)
    print(reporte_cruce(resultado))
