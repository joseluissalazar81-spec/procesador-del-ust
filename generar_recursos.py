"""
generar_recursos.py — Generación de recursos didácticos T1-T4
Procesador DEL · UST 2026-1

T1: Guión de videoclase (~8-15 min)
T2: Estructura Genially / interactivo
T3: Guía de aprendizaje
T4: Foro / Quiz / Tarea
"""
from __future__ import annotations

import json
import re
import urllib.request
from io import BytesIO

import openpyxl


# ── Mapeo de columnas (sincronizado con revisar_planificaciones.py) ───────────
COL = {
    'RA': 1, 'UNIDAD': 2, 'SEMANA': 3, 'NOMBRE': 4,
    'MODALIDAD': 5, 'MOMENTO': 6, 'ACTIVIDAD': 7,
    'RECURSOS': 8, 'CONTENIDOS': 9, 'MEDIO': 10,
    'TIPO': 11, 'PROC': 12, 'INDIV': 13, 'INSTR': 14, 'PCT': 15,
}

BACKENDS = ["claude", "ollama", "openai", "grok"]

BACKEND_DEFAULTS: dict[str, dict] = {
    "claude": {
        "url":        "https://api.anthropic.com/v1/messages",
        "model":      "claude-sonnet-4-6",
        "max_tokens": 4096,
    },
    "ollama": {
        "url":        "http://localhost:11434/api/chat",   # endpoint nativo
        "url_oai":    "http://localhost:11434/v1/chat/completions",
        "model":      "qwen3.5:latest",
        "max_tokens": 4096,
    },
    "openai": {
        "url":        "https://api.openai.com/v1/chat/completions",
        "model":      "gpt-4o-mini",
        "max_tokens": 4096,
    },
    "grok": {
        "url":        "https://api.x.ai/v1/chat/completions",
        "model":      "grok-3-mini",
        "max_tokens": 4096,
    },
}

# ── Contexto institucional compartido por todos los prompts ───────────────────
_CONTEXTO_UST = """\
CONTEXTO INSTITUCIONAL UST — REGLAS OBLIGATORIAS:

━━ TRES MOMENTOS DE APRENDIZAJE ━━
• Preparación: actividades autónomas ANTES de la sesión. Tiempo: TPE. Modalidad: asincrónica.
• Desarrollo: contacto con el/la docente. Tiempo: horas sincrónicas. Incluye evaluaciones.
• Trabajo Independiente: profundización personal DESPUÉS de la sesión. Tiempo: TPE. Sin docente.

━━ PROPÓSITO POR MOMENTO (obligatorio) ━━
Cada momento debe iniciar con una línea "Propósito:" que:
  - Comienza con verbo en infinitivo
  - Conecta las actividades con el RA correspondiente
  Ejemplo: "Reconocer los principales componentes del rendimiento humano como base para la sesión sincrónica."

━━ REDACCIÓN DE ACTIVIDADES (crítica) ━━
SIEMPRE en imperativo dirigido al/la estudiante, numeradas (1. 2. 3. …):
  ✅ "Lee el artículo X e identifica los tres conceptos clave"
  ✅ "Atiende la presentación de la docente sobre…"
  ✅ "Elabora un resumen y súbelo al buzón de tareas"
  ❌ "El docente presenta los conceptos" (nunca voz del docente)
  ❌ "Se revisa el artículo" (nunca descriptivo)
  ❌ "Reflexión sobre los aprendizajes" (nunca nominal sin verbo)

━━ FORMATO DE RECURSOS (5 campos obligatorios) ━━
Cada recurso declarado debe incluir:
  1. Título: nombre del recurso entre comillas
  2. Autoría: Apellido, I. (año) en APA 7
  3. Tipo: Diapositivas / Artículo / H5P / Guía / Video / etc.
  4. Acceso: "Disponible en el aula virtual" o URL completa
  5. Extensión: N slides / N pgs. / N min.
  Ejemplo: "Presentación S2 Ansiedad Precompetitiva". González, M. (2019). [Diapositivas, 16 slides]. Disponible en el aula virtual.

━━ RETROALIMENTACIÓN (obligatoria en entregas) ━━
En el Trabajo Independiente, cuando hay entrega al buzón de tareas:
  - Declarar siempre: "El/la docente retroalimenta en 48 horas hábiles."
  - Si es cuestionario: "La retroalimentación es automática al enviar."
  - Si es foro entre pares: indicarlo explícitamente.

━━ ALINEACIÓN RA → ACTIVIDAD → EVALUACIÓN ━━
Verificar por cada sesión:
  1. El verbo de la actividad de Desarrollo apunta al mismo nivel cognitivo que el verbo del RA
  2. La evaluación permite demostrar el logro del RA
  3. El instrumento (rúbrica analítica, pauta de observación, etc.) está declarado
     y se reproduce EXACTAMENTE como figura en el programa — NO inventar ni cambiar su nombre.

━━ TÍTULOS DE MOMENTO (descriptivos, no genéricos) ━━
Cada momento debe tener un título ESPECÍFICO que anticipe la acción del/la estudiante:
  ❌ "Preparación" → ✅ "Prepárate para la sesión"
  ❌ "Desarrollo"  → ✅ "Construye las bases de la asignatura"
  ❌ "Trabajo independiente" → ✅ "Profundiza en la ansiedad precompetitiva"

━━ INTEGRIDAD ACADÉMICA (obligatorio en sumativas) ━━
Toda evaluación sumativa debe declarar la herramienta de integridad:
  • Klarway → pruebas escritas en línea
  • Turnitin → trabajos escritos / producciones
  Declararlo en la Síntesis Didáctica y en la sesión correspondiente.

━━ ESTRUCTURA DE CADA ACTIVIDAD NUMERADA (dos partes inseparables) ━━
  Parte 1 — ACCIÓN: verbo imperativo que dice QUÉ hacer
  Parte 2 — PROPÓSITO INMEDIATO: qué identificar, analizar o lograr con esa acción
  Ejemplo completo: "Lee González et al. (2019) disponible en Revista Límite;
    identifica las estrategias de afrontamiento ante la ansiedad precompetitiva."
  Sin la segunda parte, la actividad está incompleta.

━━ VERBOS TÍPICOS POR MOMENTO ━━
  Preparación:          Lee, Revisa, Explora, Responde, Observa, Familiarízate
  Desarrollo:           Atiende, Participa, Analiza, Desarrolla, Presenta, Debate, Resuelve, Recibe
  Trabajo Independiente: Elabora, Sube, Redacta, Construye, Integra, Sintetiza, Compara

━━ UBICACIÓN DEL RECURSO (obligatorio) ━━
Cada actividad con recurso debe indicar dónde encontrarlo:
  "disponible en el aula virtual" / "disponible en [URL]" / "disponible en SOCHMEDEP"
  Sin esta indicación el/la estudiante no puede actuar de forma autónoma.

━━ CIERRE DE ACTIVIDADES CON ENTREGA (Trabajo Independiente) ━━
Toda actividad de TI con entrega debe cerrar con AMBAS frases:
  "Sube [el producto] al buzón de tareas disponible en el aula virtual.
   El/la docente retroalimenta en 48 horas hábiles."

━━ EJEMPLO MODELO DE TRABAJO INDEPENDIENTE ━━
  Título: Profundiza en la ansiedad precompetitiva
  Propósito: Profundizar en las estrategias de afrontamiento psicológico ante la ansiedad
    precompetitiva como base conceptual de la Unidad I.
  1. Lee González et al. (2019) disponible en Revista Límite; identifica las estrategias
     de afrontamiento ante la ansiedad precompetitiva en deportistas.
  2. Elabora un resumen de los 3 factores psicológicos más relevantes identificados.
  3. Sube el resumen al buzón de tareas disponible en el aula virtual.
     El/la docente retroalimenta en 48 horas hábiles.

━━ LENGUAJE ━━
Inclusivo siempre: "el o la docente", "los y las estudiantes", "el/la estudiante".
"""


# ═════════════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN DE DATOS DEL EXCEL
# ═════════════════════════════════════════════════════════════════════════════

def extraer_datos_planificacion(xlsx_bytes: bytes) -> dict:
    """
    Lee la planificación didáctica y retorna datos estructurados por unidad.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)

    # Síntesis didáctica
    asignatura, codigo = "", ""
    ws_sint = next(
        (wb[s] for s in wb.sheetnames
         if "síntesis" in s.lower() or "sintesis" in s.lower()), None
    )
    if ws_sint:
        codigo     = str(ws_sint.cell(7, 1).value or "").strip()
        asignatura = str(ws_sint.cell(4, 4).value or "").strip()

    # Planificación por unidades
    ws_plan = next(
        (wb[s] for s in wb.sheetnames
         if "planificaci" in s.lower()), None
    )
    unidades: dict[int, dict] = {}

    if ws_plan:
        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            def _v(idx): return str(row[idx - 1] or "").strip()

            ra        = _v(COL["RA"])
            unidad_n  = _v(COL["UNIDAD"])
            semana    = row[COL["SEMANA"] - 1]
            nombre    = _v(COL["NOMBRE"])
            momento   = _v(COL["MOMENTO"])
            actividad = _v(COL["ACTIVIDAD"])
            recursos  = _v(COL["RECURSOS"])
            contenido = _v(COL["CONTENIDOS"])
            tipo_eval = _v(COL["TIPO"])
            proc      = _v(COL["PROC"])
            instr     = _v(COL["INSTR"])
            pct_raw   = row[COL["PCT"] - 1]

            if not any([ra, unidad_n, nombre, actividad]):
                continue

            m = re.search(r"(\d+)", unidad_n)
            if not m:
                continue
            num = int(m.group(1))

            if num not in unidades:
                unidades[num] = {
                    "numero":    num,
                    "nombre":    unidad_n,
                    "ras":       [],
                    "semanas":   [],
                    "contenidos": [],
                    "recursos":  [],
                    "momentos": {
                        "Preparación":           [],
                        "Desarrollo":            [],
                        "Trabajo Independiente": [],
                    },
                    "evaluaciones": [],
                }
            u = unidades[num]

            if ra and ra not in u["ras"]:
                u["ras"].append(ra)

            try:
                s = int(semana)
                if s not in u["semanas"]:
                    u["semanas"].append(s)
            except (TypeError, ValueError):
                pass

            momento_key = "Desarrollo"
            for mk in ("Preparación", "Trabajo Independiente"):
                if mk.lower() in momento.lower():
                    momento_key = mk
                    break

            if actividad or nombre:
                u["momentos"][momento_key].append({
                    "nombre":    nombre,
                    "actividad": actividad,
                    "tipo_eval": tipo_eval,
                    "proc":      proc,
                    "instr":     instr,
                    "pct":       pct_raw,
                })

            if contenido and contenido not in u["contenidos"]:
                u["contenidos"].append(contenido)
            if recursos and recursos not in u["recursos"]:
                u["recursos"].append(recursos)

            if tipo_eval and proc:
                u["evaluaciones"].append({
                    "tipo":  tipo_eval,
                    "proc":  proc,
                    "instr": instr,
                    "pct":   pct_raw,
                })

    wb.close()
    return {"asignatura": asignatura, "codigo": codigo, "unidades": unidades}


def listar_unidades(xlsx_bytes: bytes) -> list[tuple[int, str]]:
    datos = extraer_datos_planificacion(xlsx_bytes)
    return [(num, u["nombre"]) for num, u in sorted(datos["unidades"].items())]


def _contexto_unidad(datos: dict, num_unidad: int) -> str:
    """Construye el bloque de texto con el contenido de una unidad para el LLM."""
    u = datos["unidades"].get(num_unidad)
    if not u:
        return ""

    lineas: list[str] = [
        f"ASIGNATURA: {datos['asignatura']} (código {datos['codigo']})",
        f"UNIDAD {u['numero']}: {u['nombre']}",
        "",
    ]

    if u["ras"]:
        lineas.append("RESULTADOS DE APRENDIZAJE (reproducir fielmente):")
        for ra in u["ras"]:
            lineas.append(f"  • {ra}")
        lineas.append("")

    if u["contenidos"]:
        lineas.append("CONTENIDOS DE LA UNIDAD:")
        for c in u["contenidos"]:
            lineas.append(f"  • {c}")
        lineas.append("")

    for momento_key in ("Preparación", "Desarrollo", "Trabajo Independiente"):
        acts = u["momentos"].get(momento_key, [])
        if not acts:
            continue
        lineas.append(f"ACTIVIDADES — {momento_key.upper()}:")
        for act in acts:
            if act["nombre"]:
                lineas.append(f"  [{act['nombre']}]")
            if act["actividad"]:
                lineas.append(f"    {act['actividad'][:500]}")
            if act["tipo_eval"] and act["pct"]:
                instr_str = f" · Instrumento: {act['instr']}" if act["instr"] else ""
                lineas.append(
                    f"    → Evaluación {act['tipo_eval']}{instr_str} · {act['pct']}%"
                )
        lineas.append("")

    # Evaluaciones sumativas — reproducir instrumento exacto
    sums = [e for e in u.get("evaluaciones", []) if "sumativa" in (e.get("tipo") or "").lower()]
    if sums:
        lineas.append("EVALUACIONES SUMATIVAS (instrumento EXACTO del programa):")
        for e in sums:
            lineas.append(
                f"  • {e['proc']}"
                + (f" — instrumento: {e['instr']}" if e.get("instr") else "")
                + (f" ({e['pct']}%)" if e.get("pct") else "")
            )
        lineas.append("")

    if u["recursos"]:
        lineas.append("RECURSOS BIBLIOGRÁFICOS (APA 7 — usar tal como están):")
        for r in u["recursos"][:6]:
            lineas.append(f"  • {r[:250]}")
        lineas.append("")

    return "\n".join(lineas)


# ═════════════════════════════════════════════════════════════════════════════
#  LLAMADA LLM
# ═════════════════════════════════════════════════════════════════════════════

def _llamar_llm(
    system: str,
    user_msg: str,
    backend: str = "claude",
    model: str | None = None,
    api_key: str = "",
    timeout: int = 600,   # 10 min — qwen3.5 local puede tardar para textos largos
    max_tokens: int = 4096,
) -> str:
    """Llama al LLM indicado y retorna el texto de la respuesta."""
    cfg   = BACKEND_DEFAULTS.get(backend, BACKEND_DEFAULTS["claude"])
    model = model or cfg["model"]
    key   = api_key or ""

    if backend == "claude":
        payload = json.dumps({
            "model":      model,
            "max_tokens": max_tokens,
            "system":     system,
            "messages":   [{"role": "user", "content": user_msg}],
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type":      "application/json",
                "x-api-key":         key,
                "anthropic-version": "2023-06-01",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read())
        return data["content"][0]["text"]

    elif backend == "ollama":
        # Usa endpoint nativo /api/chat que soporta think:false correctamente
        payload = json.dumps({
            "model":   model,
            "think":   False,   # desactiva thinking en Qwen3
            "stream":  False,
            "options": {"num_predict": max_tokens, "temperature": 0.7},
            "messages": [
                {"role": "system", "content": system},
                {"role": "user",   "content": user_msg},
            ],
        }).encode("utf-8")
        req = urllib.request.Request(
            "http://localhost:11434/api/chat",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read())
        return data["message"]["content"]

    else:
        # OpenAI / Grok
        payload = json.dumps({
            "model":       model,
            "messages":    [
                {"role": "system", "content": system},
                {"role": "user",   "content": user_msg},
            ],
            "temperature": 0.7,
            "max_tokens":  max_tokens,
        }).encode("utf-8")
        req = urllib.request.Request(
            cfg["url"],
            data=payload,
            headers={
                "Content-Type":  "application/json",
                "Authorization": f"Bearer {key}",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read())
        return data["choices"][0]["message"]["content"]


# ═════════════════════════════════════════════════════════════════════════════
#  T1 — GUIÓN DE VIDEOCLASE
# ═════════════════════════════════════════════════════════════════════════════

_SYSTEM_T1 = _CONTEXTO_UST + """
ROL: Eres un/a diseñador/a instruccional del equipo DEL de la Universidad Santo Tomás.
TAREA: Generar guiones profesionales para videoclases e-learning dirigidas a estudiantes.

ESTRUCTURA OBLIGATORIA DEL GUIÓN:
╔══════════════════════════════════════════════════╗
║ ENCABEZADO: título, duración, audiencia, tono    ║
╠══════════════════════════════════════════════════╣
║ INSTRUCCIONES DE PRODUCCIÓN (HeyGen / Loom)      ║
╠══════════════════════════════════════════════════╣
║ ESCENAS (5-8 según duración):                    ║
║   ══ ESCENA N — TÍTULO [MM:SS – MM:SS] ══        ║
║   [VISUAL EN PANTALLA]                           ║
║   descripción de lo que se muestra               ║
║   [TEXTO AVATAR]                                 ║
║   texto exacto que dice el presentador           ║
╠══════════════════════════════════════════════════╣
║ RESUMEN DE ESCENAS Y TIEMPOS                     ║
╠══════════════════════════════════════════════════╣
║ NOTAS DE PRODUCCIÓN                              ║
╚══════════════════════════════════════════════════╝

REGLAS ADICIONALES PARA VIDEOCLASE:
- Escena 1 siempre: INTRODUCCIÓN — presentar el RA de la unidad en voz del avatar
- Escena final siempre: CIERRE — síntesis + indicar próximos pasos al/la estudiante
- Escenas intermedias: un bloque temático por escena (no mezclar contenidos)
- Ritmo: ~120-130 palabras por minuto para el texto avatar
- Incluir 1-2 preguntas reflexivas activadoras durante el desarrollo (no evaluativas)
- Los recursos bibliográficos se muestran en pantalla — el avatar NO los verbaliza completos
- Los tiempos deben ser coherentes y sumar exactamente la duración indicada
"""


def generar_guion_t1(
    xlsx_bytes: bytes,
    num_unidad: int,
    duracion_min: int = 10,
    backend: str = "claude",
    model: str | None = None,
    api_key: str = "",
    timeout: int = 600,
    borrador_texto: str = "",
) -> tuple[str, str | None]:
    """Genera el guión de videoclase. Returns (texto, error)."""
    datos = extraer_datos_planificacion(xlsx_bytes)
    if not datos["unidades"]:
        return "", "No se encontraron unidades en la planificación."
    if num_unidad not in datos["unidades"]:
        return "", f"Unidad {num_unidad} no encontrada. Disponibles: {sorted(datos['unidades'])}"

    contexto = _contexto_unidad(datos, num_unidad)
    borrador_bloque = (
        f"\nMATERIAL BORRADOR DEL DOCENTE (usar como base de contenido — no inventar):\n"
        f"{borrador_texto[:4000]}\n"
        if borrador_texto.strip() else ""
    )
    user_msg = f"""\
Genera el guión completo de una videoclase de EXACTAMENTE {duracion_min} minutos.
Los tiempos de todas las escenas deben sumar {duracion_min} minutos.

CONTENIDO DE LA PLANIFICACIÓN:
{contexto}
{borrador_bloque}
INSTRUCCIONES:
- Si hay borrador del docente, ese contenido es la BASE — usarlo y estructurarlo, no ignorarlo
- Cubrir los resultados de aprendizaje indicados
- Estructurar el desarrollo por los contenidos declarados
- Las actividades de Preparación pueden mencionarse como punto de partida
- Las actividades de Trabajo Independiente se mencionan en el cierre como tarea
- Citar recursos bibliográficos solo en pantalla (no leerlos completos en voz)
- Genera el guión completo ahora, listo para producción audiovisual.\
"""
    try:
        return _llamar_llm(_SYSTEM_T1, user_msg, backend, model, api_key, timeout, 4096), None
    except Exception as e:
        return "", str(e)


# ═════════════════════════════════════════════════════════════════════════════
#  T2 — ESTRUCTURA GENIALLY
# ═════════════════════════════════════════════════════════════════════════════

_SYSTEM_T2 = _CONTEXTO_UST + """
ROL: Eres un/a diseñador/a instruccional e-learning de la Universidad Santo Tomás.
TAREA: Generar la estructura de contenido para un recurso interactivo en Genially.

REGLAS DE GENIALLY UST:
- Máximo 60 palabras por pantalla (Genially es visual, no textual)
- Una idea principal por pantalla
- Los tres momentos de aprendizaje deben estar presentes si aplica

ESTRUCTURA DEL RECURSO:
1. PORTADA: título de la unidad + asignatura + RA
2. MAPA DE CONTENIDOS: índice visual de secciones
3. SECCIONES (una por bloque temático o por RA):
   PANTALLA N: [título]
   TEXTO PRINCIPAL: [máx. 60 palabras, imperativo al/la estudiante]
   ELEMENTO INTERACTIVO: [botón / tooltip / pop-up / pregunta]
   TEXTO INTERACTIVO: [contenido del elemento]
   COLOR FONDO: [sugerencia paleta UST: verde #006633 / blanco / gris claro]
4. ACTIVIDAD INTEGRADORA: pregunta o ejercicio de síntesis
5. CIERRE: mensaje motivacional + recursos complementarios (APA 7)
"""


def generar_estructura_t2(
    xlsx_bytes: bytes,
    num_unidad: int,
    backend: str = "claude",
    model: str | None = None,
    api_key: str = "",
    timeout: int = 600,
    borrador_texto: str = "",
) -> tuple[str, str | None]:
    datos = extraer_datos_planificacion(xlsx_bytes)
    if num_unidad not in datos["unidades"]:
        return "", f"Unidad {num_unidad} no encontrada."
    contexto = _contexto_unidad(datos, num_unidad)
    borrador_bloque = (
        f"\nMATERIAL BORRADOR DEL DOCENTE (usar como base de contenido):\n{borrador_texto[:4000]}\n"
        if borrador_texto.strip() else ""
    )
    user_msg = f"""\
Genera la estructura completa de un recurso Genially para la siguiente unidad.
Incluye entre 8 y 14 pantallas. Cada pantalla debe estar lista para diseñar.
Si hay borrador del docente, ese contenido es la BASE.

CONTENIDO:
{contexto}
{borrador_bloque}
Genera la estructura completa ahora.\
"""
    try:
        return _llamar_llm(_SYSTEM_T2, user_msg, backend, model, api_key, timeout, 3500), None
    except Exception as e:
        return "", str(e)


# ═════════════════════════════════════════════════════════════════════════════
#  T3 — GUÍA DE APRENDIZAJE
# ═════════════════════════════════════════════════════════════════════════════

_SYSTEM_T3 = _CONTEXTO_UST + """
ROL: Eres un/a diseñador/a instruccional e-learning de la Universidad Santo Tomás.
TAREA: Generar guías de aprendizaje completas siguiendo la plantilla institucional UST.

ESTRUCTURA EXACTA DE LA GUÍA (reproducir este orden y estos títulos):

════════════════════════════════════════
[TÍTULO DE LA GUÍA]
Asignatura | Unidad N | Semana(s)
════════════════════════════════════════

Resultado de aprendizaje
[Reproducir el RA en infinitivo EXACTAMENTE como figura en el programa]

────────────────────────────────────────
[Tema 1: nombre específico del bloque]
────────────────────────────────────────
[Desarrollo del contenido: explicación clara, ejemplos, tablas si aplica.
 Si el docente entregó borrador, usar ese contenido como base.]

[Tabla / Figura si corresponde]
Fuente: [citar en APA 7]

────────────────────────────────────────
[Tema 2: nombre específico del bloque]  (si hay más temas)
────────────────────────────────────────
[Desarrollo]

────────────────────────────────────────
Recursos complementarios
────────────────────────────────────────
Para profundizar en los contenidos trabajados, se sugieren los siguientes recursos:
• [Recurso 1: tipo (lectura / video / sitio web), título y referencia APA 7 o URL]
• [Recurso 2: ídem]

────────────────────────────────────────
Bibliografía
────────────────────────────────────────
[Listar todas las fuentes citadas, en APA 7. Reproducir EXACTAMENTE las declaradas en la planificación.]

REGLAS ADICIONALES:
- Si el docente entregó borrador, ese contenido es la BASE — no inventar contenido nuevo
- Los recursos bibliográficos declarados en la planificación se reproducen exactamente
- El instrumento evaluativo (si aplica) se copia EXACTAMENTE del programa
- Tono: claro, académico, directo al/la estudiante
"""


def generar_guia_t3(
    xlsx_bytes: bytes,
    num_unidad: int,
    backend: str = "claude",
    model: str | None = None,
    api_key: str = "",
    timeout: int = 600,
    borrador_texto: str = "",
) -> tuple[str, str | None]:
    datos = extraer_datos_planificacion(xlsx_bytes)
    if num_unidad not in datos["unidades"]:
        return "", f"Unidad {num_unidad} no encontrada."
    contexto = _contexto_unidad(datos, num_unidad)
    borrador_bloque = (
        f"\nMATERIAL BORRADOR DEL DOCENTE (esta es la BASE del contenido — estructurar y enriquecer):\n"
        f"{borrador_texto[:5000]}\n"
        if borrador_texto.strip() else ""
    )
    user_msg = f"""\
Genera la guía de aprendizaje completa para la siguiente unidad siguiendo la plantilla institucional UST.
El instrumento evaluativo debe copiarse EXACTAMENTE del programa, sin modificarlo.
Si hay borrador del docente, ese contenido es la BASE — no inventar temas ni contenidos nuevos.

CONTENIDO DE LA PLANIFICACIÓN:
{contexto}
{borrador_bloque}
Genera la guía completa ahora.\
"""
    try:
        return _llamar_llm(_SYSTEM_T3, user_msg, backend, model, api_key, timeout, 3500), None
    except Exception as e:
        return "", str(e)


# ═════════════════════════════════════════════════════════════════════════════
#  T4 — FORO / QUIZ / TAREA
# ═════════════════════════════════════════════════════════════════════════════

_SYSTEM_T4_FORO = _CONTEXTO_UST + """
ROL: Eres un/a diseñador/a instruccional e-learning de la Universidad Santo Tomás.
TAREA: Generar la consigna de un foro siguiendo la plantilla institucional UST.

ESTRUCTURA EXACTA DEL FORO (reproducir este orden y estos títulos):

════════════════════════════════════════
[NOMBRE DEL FORO según planificación]
════════════════════════════════════════

Estimado y estimada estudiante,

[CONSIGNA: presentar en 2-4 oraciones el tema, problema o situación que enmarca el foro.
 Usar el borrador del docente si fue entregado. Tono académico, imperativo.]

Para orientar tu participación, te proponemos las siguientes preguntas:

1. [Pregunta detonante 1 — abierta, conectada al RA]
2. [Pregunta detonante 2 — que invite a relacionar con la práctica]
3. [Pregunta detonante 3 — que promueva el debate entre pares]

Recuerda comentar al menos una entrada de tus compañeras y compañeros.
El intercambio enriquece el aprendizaje colectivo.
Para publicar tu entrada, haz clic en el botón "Añadir entrada" dentro del foro.

────────────────────────────────────────
Datos de la actividad
────────────────────────────────────────
Nombre del foro: [completar según planificación]
Tipo de evaluación: [Formativo / Sumativo / Diagnóstico]
Forma de trabajo: Individual
Ponderación: [Indicar % si es sumativo, o "Sin ponderación en nota final" si es formativo]
Interacción mínima: Publicar una entrada y comentar al menos una participación de un/a compañero/a
Disponibilidad: Las fechas serán informadas por el o la docente

────────────────────────────────────────
Objetivo de la actividad
────────────────────────────────────────
[Redactar en infinitivo. Debe reflejar el RA de la unidad. Ej: "Reflexionar sobre..."]

────────────────────────────────────────
Instrucciones
────────────────────────────────────────
1. Lee con atención la consigna y las preguntas guía antes de redactar tu respuesta.
2. Redacta tu reflexión de forma personal y fundamentada, incorporando los conceptos trabajados en la unidad.
3. Publica tu entrada antes de la fecha límite indicada.
4. Una vez publicada tu entrada, comenta la participación de al menos una compañera o compañero con un aporte sustantivo.
"""

_SYSTEM_T4_TAREA = _CONTEXTO_UST + """
ROL: Eres un/a diseñador/a instruccional e-learning de la Universidad Santo Tomás.
TAREA: Generar la consigna de una tarea/trabajo escrito siguiendo la plantilla institucional UST.

ESTRUCTURA EXACTA DE LA TAREA (reproducir este orden y estos títulos):

════════════════════════════════════════
[NOMBRE DE LA TAREA según planificación]
════════════════════════════════════════

Estimado y estimada estudiante,

[CONSIGNA: presentar en 2-4 oraciones el tema o problema central de la tarea.
 Si el docente entregó borrador, usar ese contenido como base.
 Tono académico, imperativo, claro.]

Para el desarrollo de este trabajo, considera las siguientes instrucciones:

────────────────────────────────────────
Datos de la actividad
────────────────────────────────────────
Nombre de la tarea: [según planificación]
Tipo de evaluación: [Formativa / Sumativa]
Forma de trabajo: [Individual / Grupal]
Ponderación: [% si es sumativa, o "Sin ponderación en nota final" si es formativa]
Entrega: Buzón de tareas disponible en el aula virtual / Formato PDF o Word
Disponibilidad: Las fechas serán informadas por el o la docente
[Si es sumativa — agregar:] Verificación de integridad académica: Turnitin

────────────────────────────────────────
Objetivo de la actividad
────────────────────────────────────────
[Redactar en infinitivo el objetivo alineado al RA de la unidad]

────────────────────────────────────────
Instrucciones
────────────────────────────────────────
1. Lee con atención el enunciado completo antes de comenzar a desarrollar tu trabajo.
2. Desarrolla tu trabajo de forma [personal / grupal], con argumentos fundamentados en los contenidos de la unidad.
3. Cuida la redacción, la ortografía y la coherencia de tu texto.
4. Guarda tu archivo con el siguiente nombre: Apellido_Nombre_[Nombre de la tarea].pdf
5. Sube tu trabajo al buzón de tareas disponible en el aula virtual, antes de la fecha y hora límite indicada.
6. Revisa la retroalimentación entregada por el o la docente una vez que esta esté disponible.

[Si el docente entregó preguntas o indicaciones específicas, incluirlas aquí en formato numerado]

────────────────────────────────────────
[RÚBRICA / INSTRUMENTO DE EVALUACIÓN]
────────────────────────────────────────
[Reproducir EXACTAMENTE el instrumento declarado en el programa — no cambiar su nombre.
 Si es rúbrica analítica: tabla con 4 criterios × 4 niveles (Logrado / Medianamente logrado / En desarrollo / No logrado).
 Si es pauta: listar los indicadores con su puntaje.]
"""

_SYSTEM_T4_QUIZ = _CONTEXTO_UST + """
ROL: Eres un/a diseñador/a instruccional e-learning de la Universidad Santo Tomás.
TAREA: Generar un quiz/cuestionario de autoevaluación siguiendo la plantilla institucional UST.

ESTRUCTURA EXACTA DEL QUIZ:

════════════════════════════════════════
[NOMBRE DEL QUIZ según planificación]
════════════════════════════════════════

Estimado y estimada estudiante,
Este cuestionario te permitirá verificar tu comprensión de los contenidos de [tema/unidad].
La retroalimentación es automática al enviar tus respuestas.

────────────────────────────────────────
Datos de la actividad
────────────────────────────────────────
Tipo de evaluación: Formativa (autoevaluación)
Forma de trabajo: Individual
Intentos: [1 / ilimitados]
Disponibilidad: Las fechas serán informadas por el o la docente

────────────────────────────────────────
Preguntas
────────────────────────────────────────
[Generar 8 preguntas de opción múltiple con 4 alternativas (A-D) cada una]

Para cada pregunta incluir:
  Pregunta N: [enunciado claro, conectado al RA]
  A) [alternativa]
  B) [alternativa]
  C) [alternativa — puede ser "Todas las anteriores" si aplica]
  D) [alternativa]
  ✅ Respuesta correcta: [letra]
  Retroalimentación: [1-2 oraciones que expliquen por qué esa es la correcta]

REGLAS PARA LAS PREGUNTAS:
- Distribuir entre niveles: recordar (2), comprender (3), aplicar (3)
- No usar "Todas las anteriores" más de una vez
- Las alternativas incorrectas deben ser plausibles (no absurdas)
- Basarse en el contenido declarado en la planificación
"""

# Mantener _SYSTEM_T4 como alias para compatibilidad
_SYSTEM_T4 = _SYSTEM_T4_TAREA


def generar_consigna_t4(
    xlsx_bytes: bytes,
    num_unidad: int,
    tipo_t4: str = "tarea",
    backend: str = "claude",
    model: str | None = None,
    api_key: str = "",
    timeout: int = 600,
    borrador_texto: str = "",
) -> tuple[str, str | None]:
    datos = extraer_datos_planificacion(xlsx_bytes)
    if num_unidad not in datos["unidades"]:
        return "", f"Unidad {num_unidad} no encontrada."

    contexto = _contexto_unidad(datos, num_unidad)
    borrador_bloque = (
        f"\nMATERIAL BORRADOR DEL DOCENTE (usar como base — no ignorar):\n{borrador_texto[:4000]}\n"
        if borrador_texto.strip() else ""
    )

    # Seleccionar prompt y nombre según tipo
    system_map = {
        "foro":  _SYSTEM_T4_FORO,
        "quiz":  _SYSTEM_T4_QUIZ,
        "tarea": _SYSTEM_T4_TAREA,
    }
    tipo_nombre = {
        "foro":  "Foro de participación",
        "quiz":  "Quiz de autoevaluación",
        "tarea": "Tarea / producción del estudiante",
    }.get(tipo_t4, "Tarea")
    system = system_map.get(tipo_t4, _SYSTEM_T4_TAREA)

    user_msg = f"""\
Genera {tipo_nombre} completa para la siguiente unidad siguiendo la plantilla institucional UST.
El instrumento evaluativo debe copiarse EXACTAMENTE como figura en la planificación.
Si hay borrador del docente, ese contenido es la BASE — úsalo, no lo ignores.

CONTENIDO DE LA PLANIFICACIÓN:
{contexto}
{borrador_bloque}
Genera el recurso completo ahora.\
"""
    try:
        return _llamar_llm(system, user_msg, backend, model, api_key, timeout, 3000), None
    except Exception as e:
        return "", str(e)
