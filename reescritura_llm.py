"""
reescritura_llm.py — Reescritura pedagógica con Claude Sonnet 4.6
Procesador DEL · UST 2026-1

Corrige automáticamente actividades que no cumplen:
  C19 — Verbo imperativo 2ª persona (Revisa, Analiza, Elabora…)
  C20 — Propósito explícito en actividades con entrega
  C21 — Retroalimentación indicada en actividades con entrega
  + Lenguaje inclusivo en toda la planificación

Marca cambios en azul en el xlsx resultante.
"""
from __future__ import annotations
import os
import re
from io import BytesIO
from typing import Callable

import openpyxl
from openpyxl.styles import Font


# ── Lenguaje inclusivo (sin API) ─────────────────────────────────────────────

_INCLUSIVO = [
    (r"\bel docente\b",       "el o la docente"),
    (r"\bla docente\b",       "el o la docente"),
    (r"\blos docentes\b",     "los y las docentes"),
    (r"\bel estudiante\b",    "el/la estudiante"),
    (r"\bla estudiante\b",    "el/la estudiante"),
    (r"\blos estudiantes\b",  "los y las estudiantes"),
    (r"\blas estudiantes\b",  "los y las estudiantes"),
    (r"\blos alumnos\b",      "los y las estudiantes"),
    (r"\blas alumnas\b",      "los y las estudiantes"),
    (r"\bel alumno\b",        "el/la estudiante"),
    (r"\bla alumna\b",        "el/la estudiante"),
]


def aplicar_lenguaje_inclusivo(texto: str) -> tuple[str, bool]:
    """Aplica lenguaje inclusivo. Retorna (texto_corregido, hubo_cambio)."""
    original = texto
    for patron, reemplazo in _INCLUSIVO:
        texto = re.sub(patron, reemplazo, texto, flags=re.IGNORECASE)
    return texto, texto != original


# ── Detección de necesidades en actividades ───────────────────────────────────

_VERBOS_IMP = {
    "accede", "analiza", "aplica", "argumenta", "busca", "calcula",
    "caracteriza", "clasifica", "compara", "completa", "consulta",
    "construye", "contrasta", "crea", "debate", "define", "describe",
    "desarrolla", "diferencia", "diseña", "elabora", "evalúa", "evalua",
    "examina", "explica", "extrae", "formula", "identifica", "indica",
    "infiere", "ingresa", "integra", "interpreta", "investiga", "lee",
    "lista", "observa", "organiza", "participa", "presenta", "propone",
    "realiza", "reflexiona", "relaciona", "responde", "resume", "revisa",
    "selecciona", "sintetiza", "sistematiza", "socializa", "sube", "trabaja",
    "ubica", "utiliza", "valida", "verifica", "descarga", "escribe",
    "documenta", "registra", "entrega", "prepara", "comparte",
}

_KW_ENTREGA = {
    "tarea", "informe", "trabajo", "entrega", "sube", "portafolio",
    "proyecto", "producto", "reporte", "actividad evaluativa",
}

_KW_RETRO = {
    "retroalimentación", "retroalimentacion", "feedback",
    "el o la docente revisará", "el/la docente revisará",
    "corrección del", "comentarios del",
}


def _empieza_imperativo(texto: str) -> bool:
    if not texto:
        return False
    primer = re.split(r"[\s,.:;(]", texto.strip())[0].lower()
    # remover tildes del primer token para comparación simple
    primer_sin = (primer.replace("á","a").replace("é","e")
                        .replace("í","i").replace("ó","o").replace("ú","u"))
    return primer in _VERBOS_IMP or primer_sin in _VERBOS_IMP


def _tiene_proposito(texto: str) -> bool:
    return bool(re.search(r"propósito\s*:", texto, re.IGNORECASE))


def _tiene_retro(texto: str) -> bool:
    t = texto.lower()
    return any(p in t for p in _KW_RETRO)


def _requiere_entrega(texto: str) -> bool:
    t = texto.lower()
    return any(p in t for p in _KW_ENTREGA)


# ── Llamada a Claude Sonnet ───────────────────────────────────────────────────

_SYSTEM = """Eres un/a diseñador/a instruccional del equipo DEL de la Universidad Santo Tomás.
Tu tarea es reescribir actividades de planificaciones didácticas e-learning según los estándares UST.

REGLAS OBLIGATORIAS:
1. Inicia SIEMPRE con un verbo imperativo en segunda persona singular (Revisa, Analiza, Elabora…).
2. Usa lenguaje inclusivo: "el o la docente", "los y las estudiantes", "el/la estudiante".
3. Conserva el contenido pedagógico y los recursos declarados — solo mejora la redacción.
4. Si la actividad tiene entrega, incluye al inicio: "Propósito: [para qué sirve]".
5. Si la actividad tiene entrega, incluye al final: "Retroalimentación: el o la docente entregará [tipo] en [plazo]".
6. Escribe en español, forma clara, directa, con pasos numerados si aplica.
7. Responde SOLO con el texto reescrito, sin explicaciones ni comillas adicionales."""


OLLAMA_URL = "http://localhost:11434/v1/chat/completions"


def _llm_reescribir(
    texto: str,
    instrucciones: list[str],
    api_key: str,
    backend: str = "claude",
    modelo_local: str = "qwen3:8b",
) -> str:
    instruccion_txt = " | ".join(instrucciones)
    prompt_user = (
        f"CORRECCIONES REQUERIDAS: {instruccion_txt}\n\n"
        f"ACTIVIDAD ORIGINAL:\n{texto}\n\n"
        "Reescribe la actividad:"
    )

    if backend == "ollama":
        import requests, re as _re
        prompt_final = prompt_user + " /no_think" if "qwen3" in modelo_local else prompt_user
        resp = requests.post(
            OLLAMA_URL,
            json={
                "model": modelo_local,
                "messages": [
                    {"role": "system", "content": _SYSTEM},
                    {"role": "user",   "content": prompt_final},
                ],
                "temperature": 0.15,
                "stream": False,
                "options": {"num_predict": 1024, "num_ctx": 4096},
            },
            timeout=90,
        )
        resp.raise_for_status()
        contenido = resp.json()["choices"][0]["message"]["content"]
        contenido = _re.sub(r"<think>[\s\S]*?</think>", "", contenido).strip()
        return contenido
    else:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            system=_SYSTEM,
            messages=[{"role": "user", "content": prompt_user}],
            temperature=0.15,
        )
        return msg.content[0].text.strip()


def reescribir_actividad(
    texto: str,
    api_key: str,
    c19_falla: bool = True,
    c20_falla: bool = True,
    c21_falla: bool = True,
    backend: str = "claude",
    modelo_local: str = "qwen3:8b",
) -> tuple[str, list[str]]:
    """
    Reescribe una actividad para los criterios indicados.
    Retorna (texto_nuevo, cambios_aplicados).
    """
    if not texto or not api_key:
        return texto, []

    instrucciones = []
    cambios = []

    if c19_falla and not _empieza_imperativo(texto):
        instrucciones.append("Inicia con verbo imperativo en segunda persona (Revisa, Analiza, Elabora…)")
        cambios.append("C19")

    if c20_falla and _requiere_entrega(texto) and not _tiene_proposito(texto):
        instrucciones.append('Agrega al inicio: "Propósito: [objetivo de la entrega]"')
        cambios.append("C20")

    if c21_falla and _requiere_entrega(texto) and not _tiene_retro(texto):
        instrucciones.append('Agrega al final: "Retroalimentación: el o la docente entregará [tipo] en [plazo]"')
        cambios.append("C21")

    if not instrucciones:
        return texto, []

    try:
        nuevo = _llm_reescribir(texto, instrucciones, api_key, backend, modelo_local)
        return nuevo, cambios
    except Exception as e:
        return texto, [f"ERROR:{str(e)[:50]}"]


# ── Función principal ─────────────────────────────────────────────────────────

def reescribir_planificacion(
    fuente,
    criterios_resultado: dict | None = None,
    api_key: str = "",
    progress_callback: Callable | None = None,
    backend: str = "claude",
    modelo_local: str = "qwen3:8b",
) -> tuple[bytes, list[str]]:
    """
    Aplica correcciones lingüísticas y pedagógicas al xlsx de planificación.

    Parámetros
    ----------
    fuente             : bytes o ruta del xlsx
    criterios_resultado: resultado de evaluar_45_criterios() (opcional)
    api_key            : ANTHROPIC_API_KEY (o variable de entorno)
    progress_callback  : función(paso, total, mensaje)

    Retorna
    -------
    (xlsx_bytes, log)
    """
    key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    log: list[str] = []

    def _prog(paso, total, msg):
        if progress_callback:
            progress_callback(paso, total, msg)

    # ── Cargar workbook ───────────────────────────────────────────────────────
    if isinstance(fuente, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(fuente), data_only=True)
    else:
        wb = openpyxl.load_workbook(fuente, data_only=True)

    if "Planificación por unidades" not in wb.sheetnames:
        return _wb_bytes(wb), ["❌ Hoja 'Planificación por unidades' no encontrada"]

    ws = wb["Planificación por unidades"]

    # ── Determinar criterios fallidos ─────────────────────────────────────────
    crits = (criterios_resultado or {}).get("criterios", {})

    def _falla(cid: int) -> bool:
        return crits.get(cid, {}).get("estado", "NO") in ("NO", "PARCIALMENTE")

    c19 = _falla(19)
    c20 = _falla(20)
    c21 = _falla(21)

    # ── Paso 1: Lenguaje inclusivo (todas las hojas) ──────────────────────────
    _prog(1, 4, "Aplicando lenguaje inclusivo...")
    total_li = 0
    for sname in wb.sheetnames:
        for row in wb[sname].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    nuevo, cambio = aplicar_lenguaje_inclusivo(cell.value)
                    if cambio:
                        cell.value = nuevo
                        cell.font  = _font_azul(cell.font)
                        total_li  += 1

    log.append(
        f"✅ Lenguaje inclusivo: {total_li} celda(s) corregidas"
        if total_li else "ℹ️  Lenguaje inclusivo: sin cambios necesarios"
    )

    # ── Paso 2: Reescritura semántica (requiere API key) ──────────────────────
    _prog(2, 4, "Evaluando actividades para reescritura...")

    usar_llm = (backend == "ollama") or bool(key)
    if not usar_llm:
        log.append("⬜ Reescritura semántica: sin LLM configurado (C19-C21 omitidos)")
        _prog(4, 4, "Guardando...")
        return _wb_bytes(wb), log

    if not (c19 or c20 or c21):
        log.append("✅ Criterios C19-C21: no requieren reescritura")
    else:
        fallando = [f"C{c}" for c, b in zip([19, 20, 21], [c19, c20, c21]) if b]
        log.append(f"⚙️  Reescribiendo actividades para: {', '.join(fallando)}")

        celdas_a_reescribir = []
        for row in ws.iter_rows(min_row=4, values_only=False):
            celda = row[6]  # columna G — actividades
            txt   = str(celda.value or "").strip()
            if len(txt) < 25:
                continue
            necesita = (
                (c19 and not _empieza_imperativo(txt)) or
                (c20 and _requiere_entrega(txt) and not _tiene_proposito(txt)) or
                (c21 and _requiere_entrega(txt) and not _tiene_retro(txt))
            )
            if necesita:
                celdas_a_reescribir.append((celda, txt))

        log.append(f"  → {len(celdas_a_reescribir)} actividad(es) identificadas")
        MAX_REESCRITURAS = 25  # límite para no agotar tiempo/tokens

        total_rw = 0
        for i, (celda, texto) in enumerate(celdas_a_reescribir[:MAX_REESCRITURAS]):
            _prog(2, 4, f"Reescribiendo {i+1}/{min(len(celdas_a_reescribir), MAX_REESCRITURAS)}…")
            nuevo, cambios = reescribir_actividad(
                texto, key, c19, c20, c21,
                backend=backend, modelo_local=modelo_local,
            )
            if cambios and nuevo != texto:
                celda.value = nuevo
                celda.font  = _font_azul(celda.font)
                total_rw   += 1
                log.append(f"  ✏️  Fila {celda.row}: {', '.join(cambios)}")

        log.append(f"✅ Reescritura: {total_rw} actividad(es) modificadas")

    _prog(4, 4, "Guardando archivo...")
    return _wb_bytes(wb), log


# ── Helpers ───────────────────────────────────────────────────────────────────

def _wb_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _font_azul(font_orig) -> Font:
    """Retorna Font azul conservando propiedades existentes."""
    if font_orig is None:
        return Font(color="0000CC")
    return Font(
        name    = font_orig.name,
        size    = font_orig.size,
        bold    = font_orig.bold,
        italic  = font_orig.italic,
        color   = "0000CC",
    )


# ── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    ruta = sys.argv[1] if len(sys.argv) > 1 else (
        "/Users/mac/Desktop/DEL/PLA-S0113 Sistema de producción de frutales sostenibles./"
        "PLA-S0113_Sistema de producción de frutales sostenibles_Rev DI (2).xlsx"
    )
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    print(f"Procesando: {ruta}")
    print(f"API key: {'OK' if api_key else 'no configurada'}")

    # Simular criterios fallidos para prueba
    crits_test = {
        "criterios": {
            19: {"estado": "NO", "observacion": "test"},
            20: {"estado": "PARCIALMENTE", "observacion": "test"},
            21: {"estado": "NO", "observacion": "test"},
        }
    }

    def prog(p, t, m):
        print(f"  [{p}/{t}] {m}")

    with open(ruta, "rb") as f:
        raw = f.read()

    out_bytes, log = reescribir_planificacion(raw, crits_test, api_key=api_key, progress_callback=prog)

    salida = ruta.replace(".xlsx", "_REESCRITO.xlsx")
    with open(salida, "wb") as f:
        f.write(out_bytes)

    print(f"\nGuardado: {salida}")
    print("\n── Log ──")
    for l in log:
        print(f"  {l}")
