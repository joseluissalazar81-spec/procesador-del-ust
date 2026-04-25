#!/usr/bin/env python3
"""
revisar_planificaciones.py
Revisa y corrige automáticamente las planificaciones didácticas UST
en todas las carpetas de asignaturas dentro de ~/Desktop/DEL.

Estructura esperada por asignatura:
  ~/Desktop/DEL/<Asignatura>/
    Correcciones/   → planificación ya revisada por DEL (*.xlsx)  [opcional]
    Enviado a DEL/  → planificación original + escala (Escala*.xlsx)
    Revisado/       → carpeta de salida (se crea si no existe)

Lógica de fuente:
  1. Si existe un .xlsx en Correcciones/  → lo usa como base (ya fue revisado por DEL)
  2. Si no hay Correcciones/ o está vacía → busca el .xlsx en Enviado a DEL/
     que NO sea la escala de apreciación, y lo trata como planificación original.

Correcciones aplicadas automáticamente:
  [Síntesis didáctica]
    - Elimina bloques DIAGNÓSTICA y FORMATIVA de la columna
      "Procedimientos de evaluación", dejando solo SUMATIVA.

  [Planificación por unidades]
    Alta prioridad — alineación al programa oficial:
      - Procedimiento 'Cuestionario' → 'Pruebas escritas u orales'
      - Procedimiento 'Tarea'        → 'Producciones del estudiante'
      - Procedimiento 'Prueba'       → 'Pruebas escritas u orales'
      - Instrumento 'Pauta' (exacto) → 'Pauta de observación'
      - Instrumento 'Rúbrica'        → 'Pauta de observación'
        (solo cuando el procedimiento es Producciones/Actuación)
      - Tipo 'Formativo'             → 'Formativa'
      - Tipo 'Diagnóstico'           → 'Diagnóstica'  (normaliza mayúsculas)
      - Datos evaluativos huérfanos en filas sin Tipo de evaluación → eliminados

    Media prioridad — columna J (Medio de entrega):
      - Desarrollo + Sincrónico con evaluación  → 'No aplica'
      - Trabajo independiente / No presencial   → 'Buzón de tareas'
      - Preparación con Diagnóstico             → 'Cuestionario'

    Media prioridad — columna M (Individual/Grupal):
      - Procedimiento contiene 'equipo' o 'grupo'  → 'Grupal'
      - Unidades donde las sumativas son 'Grupal'  → Grupal en toda la unidad
      - Resto                                      → 'Individual'

    Baja prioridad:
      - % evaluación = None en filas de evaluación → 0
"""

import os
import unicodedata
import re
import glob
import shutil
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter
import urllib.request
import urllib.parse
import json
import time

def _norm_hoja(s: str) -> str:
    """Minúsculas + quitar tildes + colapsar espacios (para comparar nombres de hoja)."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().split())


def _hoja(wb, nombre: str):
    """
    Devuelve la hoja cuyo nombre coincide con nombre de forma
    tolerante a tildes, mayúsculas y espacios extra.
    Lanza KeyError si no existe.
    """
    norm = _norm_hoja(nombre)
    for sname in wb.sheetnames:
        if _norm_hoja(sname) == norm:
            return wb[sname]
    raise KeyError(f"Hoja no encontrada: '{nombre}' (hojas disponibles: {wb.sheetnames})")


def _tiene_hoja(wb, nombre: str) -> bool:
    norm = _norm_hoja(nombre)
    return any(_norm_hoja(s) == norm for s in wb.sheetnames)



# Flag global — se desactiva con --no-languagetool
_LANGUAGETOOL_ACTIVO = True


# ══════════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN Y VERIFICACIÓN DEL PROGRAMA OFICIAL (PDF)
# ══════════════════════════════════════════════════════════════════════════

def extraer_programa_pdf(pdf_path):
    """
    Extrae datos clave del programa de asignatura UST desde un PDF.
    Devuelve un dict con: codigo, asignatura, carrera, creditos, area,
    horas_tpe, unidades (lista), ponderaciones (dict unidad→%), pct_examen.
    """
    try:
        import pdfplumber
    except ImportError:
        return {'_error': 'pdfplumber no instalado'}

    programa = {}
    try:
        texto = ''
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or '') + '\n'

        # Código
        m = re.search(r'CÓDIGO\s*:\s*([A-Z]+-[A-Z0-9]+)', texto, re.IGNORECASE)
        programa['codigo'] = m.group(1).strip() if m else None

        # Asignatura (entre ASIGNATURA: y CÓDIGO:)
        m = re.search(r'ASIGNATURA\s*:\s*(.+?)(?=CÓDIGO\s*:)', texto,
                      re.IGNORECASE | re.DOTALL)
        if m:
            programa['asignatura'] = re.sub(r'\s+', ' ', m.group(1)).strip()

        # Carrera (entre CARRERA: y ASIGNATURA:)
        m = re.search(r'CARRERA\s*:\s*(.+?)(?=ASIGNATURA\s*:)', texto,
                      re.IGNORECASE | re.DOTALL)
        if m:
            programa['carrera'] = re.sub(r'\s+', ' ', m.group(1)).strip()

        # Créditos
        m = re.search(r'(\d+)\s+CRÉDITOS', texto, re.IGNORECASE)
        programa['creditos'] = int(m.group(1)) if m else None

        # Área de conocimiento
        m = re.search(r'AREA[A]?\s+DE\s+CONOCIMIENTO\s*:\s*(.+?)(?:\n|$)',
                      texto, re.IGNORECASE)
        programa['area'] = m.group(1).strip() if m else None

        # Horas TPE
        m = re.search(r'HORAS\s+TPE\*?\s*:\s*(\d+)', texto, re.IGNORECASE)
        programa['horas_tpe'] = int(m.group(1)) if m else None

        # Unidades — método primario: extracción por tablas (pdfplumber)
        total_tabla, unidades_tabla = _extraer_horas_pdf_tabla(pdf_path)
        programa['total_pedagogicas'] = total_tabla

        if unidades_tabla:
            programa['unidades'] = unidades_tabla
        else:
            # Fallback: texto + regex (estructura antigua de programas)
            m_sec = re.search(
                r'V\s+UNIDADES\s+DE\s+APRENDIZAJE\s*\n(.*?)(?:VI\s+DESGLOSE|VII\s+METODOLOGÍA)',
                texto, re.IGNORECASE | re.DOTALL
            )
            bloque_unidades = m_sec.group(1) if m_sec else texto
            unidades_raw = re.findall(
                r'UNIDAD\s+([IVX]+)\s+(.*?)\s+(\d+)\s+HORAS\s+PEDAGÓGICAS',
                bloque_unidades, re.IGNORECASE
            )
            vistos = set()
            programa['unidades'] = []
            for n, nom, h in unidades_raw:
                key = n.upper()
                if key not in vistos:
                    vistos.add(key)
                    programa['unidades'].append({
                        'numero': key,
                        'nombre': re.sub(r'\s+', ' ', nom).strip(),
                        'horas': int(h)
                    })

        # Ponderaciones por unidad (sección PONDERACIÓN)
        ponderaciones = re.findall(r'[Uu]nidad\s+([IVX]+)\s+(\d+)%', texto)
        programa['ponderaciones'] = {n.upper(): int(p) for n, p in ponderaciones}

        # Examen LGA
        m = re.search(r'[Ee]xamen[^%\n]{0,60}?(\d+)%', texto)
        programa['pct_examen'] = int(m.group(1)) if m else None

    except Exception as e:
        programa['_error'] = str(e)

    return programa


def _extraer_horas_pdf_tabla(pdf_path):
    """
    Extrae horas del programa de asignatura usando tablas pdfplumber.
    Método más fiable que el regex sobre texto plano para PDFs UST actuales.

    Retorna (total_pedagogicas, unidades) donde:
      total_pedagogicas — int o None
      unidades          — lista de dicts {'numero': 'I', 'nombre': '...', 'horas': N}
    """
    ROMANO     = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII']
    pat_horas  = re.compile(r'\((\d+)(?:\s*horas?)?\)', re.IGNORECASE)
    pat_examen = re.compile(r'[Ee]xamen[^0-9]*(\d+)\s*horas?', re.IGNORECASE)

    total_ped = None
    unidades  = []

    try:
        import pdfplumber  # importación local: módulo opcional
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    flat = ' '.join(str(c or '') for row in table for c in row)

                    # ── Tabla de distribución de horas ────────────────────
                    if 'TPE' in flat and 'Cátedra' in flat and total_ped is None:
                        header_ok = False
                        for row in table:
                            cells  = [str(c or '').strip() for c in row]
                            joined = ' '.join(cells)
                            if 'Totales' in cells and 'Cátedra' in joined:
                                header_ok = True
                                continue
                            if header_ok:
                                nums = [int(str(c).strip()) for c in row
                                        if str(c or '').strip().isdigit()]
                                if len(nums) >= 3:
                                    total_ped = nums[-1]   # último = Totales pedagógicas
                                    break

                    # ── Tabla de contenidos por unidad ────────────────────
                    if 'Nombre de la unidad' in flat and not unidades:
                        unit_idx = 0
                        for row in table:
                            cell0 = str(row[0] or '').strip()
                            m = pat_horas.search(cell0)
                            if m:
                                nombre = re.sub(r'\s*\(\d+.*', '', cell0) \
                                              .replace('\n', ' ').strip()
                                romano = ROMANO[unit_idx] if unit_idx < len(ROMANO) \
                                         else str(unit_idx + 1)
                                unidades.append({
                                    'numero': romano,
                                    'nombre': re.sub(r'\s+', ' ', nombre).strip(),
                                    'horas': int(m.group(1)),
                                })
                                unit_idx += 1
                                continue
                            # Detectar examen en cualquier celda de la fila
                            for cell in row:
                                me = pat_examen.search(str(cell or ''))
                                if me:
                                    unidades.append({
                                        'numero': 'EXAMEN',
                                        'nombre': 'Evaluación final',
                                        'horas': int(me.group(1)),
                                    })
                                    break
    except Exception:
        pass   # pdfplumber no disponible o PDF corrupto — caller usa fallback

    return total_ped, unidades


def verificar_contra_programa(wb, programa, log):
    """
    Cruza la planificación con datos extraídos del programa oficial.
    Verifica: código, créditos, área, nombre asignatura y ponderaciones.
    Devuelve número de discrepancias.
    """
    if not programa or programa.get('_error'):
        log.append('\n  [Verificación contra programa] — datos no disponibles (omitida)')
        return 0

    errores = []
    oks = []

    ws_sint = _hoja(wb, 'Síntesis didáctica') if _tiene_hoja(wb, 'Síntesis didáctica') else None
    ws_plan = _hoja(wb, 'Planificación por unidades') if _tiene_hoja(wb, 'Planificación por unidades') else None

    def sv(ws, r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v else ''

    def norm(s):
        return re.sub(r'\s+', ' ', str(s or '')).strip().upper()

    STOPWORDS = {'Y', 'E', 'DE', 'DEL', 'LA', 'LAS', 'LOS', 'EL', 'CON',
                 'EN', 'A', 'O', 'U', 'AL'}

    if ws_sint:
        # Código
        if programa.get('codigo'):
            val = sv(ws_sint, 7, 1)
            if norm(val) == norm(programa['codigo']):
                oks.append(f'Código "{programa["codigo"]}"')
            elif val:
                errores.append(
                    f'Código: planificación="{val}" vs programa="{programa["codigo"]}"')

        # Créditos
        if programa.get('creditos') is not None:
            val = ws_sint.cell(7, 3).value
            try:
                if int(val) == int(programa['creditos']):
                    oks.append(f'Créditos ({programa["creditos"]})')
                else:
                    errores.append(
                        f'Créditos: planificación={val} vs programa={programa["creditos"]}')
            except (TypeError, ValueError):
                if val:
                    errores.append(f'Créditos: valor en planificación="{val}" no es numérico')

        # Área de conocimiento
        if programa.get('area'):
            val = sv(ws_sint, 11, 1)
            if val and (norm(programa['area']) in norm(val) or
                        norm(val) in norm(programa['area'])):
                oks.append(f'Área "{programa["area"]}"')
            elif val:
                errores.append(
                    f'Área: planificación="{val}" vs programa="{programa["area"]}"')

        # Nombre asignatura (coincidencia por palabras clave)
        if programa.get('asignatura'):
            val = sv(ws_sint, 4, 4)
            palabras_prog = set(norm(programa['asignatura']).split()) - STOPWORDS
            palabras_plan = set(norm(val).split()) - STOPWORDS
            if palabras_prog:
                ratio = len(palabras_prog & palabras_plan) / len(palabras_prog)
                if ratio >= 0.7:
                    oks.append(f'Nombre asignatura (similitud {ratio:.0%})')
                elif val:
                    errores.append(
                        f'Nombre: planificación="{val}" vs programa="{programa["asignatura"]}"')

    # Ponderaciones sumativas por unidad
    if ws_plan and programa.get('ponderaciones'):
        # Mapa para convertir número arábigo → romano (Unidad 1 → I, etc.)
        ARABE_A_ROMANO = {'1': 'I', '2': 'II', '3': 'III', '4': 'IV',
                          '5': 'V', '6': 'VI', '7': 'VII', '8': 'VIII'}

        def num_unidad(texto):
            """Extrae número de unidad como romano desde texto con arábigo o romano."""
            # Romano: "UNIDAD II", "Unidad III"
            m = re.search(r'\bUNIDAD\s+([IVX]+)\b', str(texto or ''), re.IGNORECASE)
            if m:
                return m.group(1).upper()
            # Arábigo: "Unidad 1", "Unidad 2 – Nombre"
            m = re.search(r'\bUNIDAD\s+(\d+)\b', str(texto or ''), re.IGNORECASE)
            if m:
                return ARABE_A_ROMANO.get(m.group(1))
            return None

        pcts_plan = {}
        unidad_actual = None
        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            unidad = row[COL['UNIDAD'] - 1]
            tipo   = row[COL['TIPO']   - 1]
            pct    = row[COL['PCT']    - 1]
            if unidad:
                unidad_actual = str(unidad)
            if (tipo and 'sumativa' in str(tipo).lower()
                    and isinstance(pct, (int, float)) and pct > 0):
                num = num_unidad(unidad_actual)
                if num:
                    pcts_plan[num] = pcts_plan.get(num, 0) + pct

        for num, pct_prog in programa['ponderaciones'].items():
            pct_p = pcts_plan.get(num)
            if pct_p is None:
                errores.append(
                    f'Unidad {num}: ponderación {pct_prog}% no encontrada en planificación '
                    f'(¿falta la etiqueta "Unidad {num}" en col. Unidad/módulo?)')
            else:
                if pct_p < 1:
                    pct_p = round(pct_p * 100)
                if abs(pct_p - pct_prog) <= 1:
                    oks.append(f'Unidad {num}: ponderación {pct_prog}%')
                else:
                    errores.append(
                        f'Unidad {num}: planificación={pct_p}% vs programa={pct_prog}%')

    log.append('\n  [Verificación contra programa oficial]')
    for o in oks:
        log.append(f'    ✅ {o}')
    for e in errores:
        log.append(f'    ❌ {e}')
    if not oks and not errores:
        log.append('    ⚠️  Sin campos verificables (revisa el PDF extraído)')

    return len(errores)

# ── Constantes de columnas (1-based en openpyxl) ──────────────────────────
COL = {
    'RA':        1,
    'UNIDAD':    2,
    'SEMANA':    3,
    'NOMBRE':    4,
    'MODALIDAD': 5,
    'MOMENTO':   6,
    'ACTIVIDAD': 7,
    'RECURSOS':  8,
    'CONTENIDOS':9,
    'MEDIO':    10,   # J — Medio de entrega
    'TIPO':     11,   # K — Tipo de evaluación
    'PROC':     12,   # L — Procedimiento
    'INDIV':    13,   # M — Individual/Grupal
    'INSTR':    14,   # N — Instrumento
    'PCT':      15,   # O — % evaluación
}

# ── Mapas de corrección ────────────────────────────────────────────────────
PROC_MAP = {
    'Cuestionario': 'Pruebas escritas u orales',
    'Tarea':        'Producciones del estudiante',
    'Prueba':       'Pruebas escritas u orales',
}

INSTR_MAP = {
    'Pauta':   'Pauta de observación',   # solo coincidencia exacta
    'Rúbrica': 'Pauta de observación',   # cuando procedimiento es Producciones/Actuación
}

TIPO_MAP = {
    'Formativo':   'Formativa',
    'Diagnóstico': 'Diagnóstica',
}

PROC_INSTR_RÚBRICA_CONTEXTOS = {
    'Producciones del estudiante',
    'Actuación del estudiante',
    'Informe de análisis de casos',
}

# ── Escala de apreciación estándar UST (41 criterios) ─────────────────────
# Fuente: Escala institucional DEL — idéntica para todas las asignaturas.
# Cada ítem: (sección, N° dentro de sección, criterio, modo de verificación)
#   modo: 'auto' = verificable por el script
#         'manual' = requiere revisión humana
ESCALA_ESTANDAR = [
    # ── Sección 1: Síntesis Didáctica (15 ítems) ──────────────────────────
    ('Síntesis Didáctica', 1,
     'Establece correctamente la escuela (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 2,
     'Menciona correctamente el programa académico (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 3,
     'Nombre de asignatura correcto (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 4,
     'Código correcto (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 5,
     'Versión año correcto (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 6,
     'N° de créditos correcto (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 7,
     'Área del conocimiento OCDE correcta (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 8,
     'Requisitos correctos (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 9,
     'Carga académica bien estructurada (Conforme a Programa Oficial)',
     'auto'),
    ('Síntesis Didáctica', 10,
     'Presenta competencias específicas conforme a programa oficial',
     'auto'),
    ('Síntesis Didáctica', 11,
     'Estructura módulos o unidades con horas pedagógicas correspondientes',
     'auto'),
    ('Síntesis Didáctica', 12,
     'Presenta títulos de módulos o unidades académicas',
     'auto'),
    ('Síntesis Didáctica', 13,
     'Presenta resultados de aprendizaje esperados',
     'auto'),
    ('Síntesis Didáctica', 14,
     'Presenta procedimientos de evaluación acorde a resultados esperados '
     '(solo SUMATIVA en la síntesis)',
     'auto'),
    ('Síntesis Didáctica', 15,
     'Diferencia resultados de aprendizaje con porcentajes según programa oficial',
     'auto'),

    # ── Sección 2: Planificación por unidades (3 ítems) ───────────────────
    ('Planificación por unidades', 1,
     'Presenta las unidades de aprendizaje diferenciadas correctamente '
     'para cada módulo, conforme a programa oficial',
     'auto'),
    ('Planificación por unidades', 2,
     'Presenta RA correctamente vinculados a los módulos, conforme a programa oficial',
     'auto'),
    ('Planificación por unidades', 3,
     'Las distintas sesiones se encuentran bien diferenciadas y con '
     'establecimiento de horas para cada momento',
     'auto'),

    # ── Sección 3: Estrategias metodológicas (11 ítems) ───────────────────
    ('Estrategias metodológicas', 1,
     'Se presentan los contenidos de actividades',
     'auto'),
    ('Estrategias metodológicas', 2,
     'Se presentan los contenidos de las evaluaciones',
     'auto'),
    ('Estrategias metodológicas', 3,
     'Las actividades son coherentes con los contenidos',
     'manual'),
    ('Estrategias metodológicas', 4,
     'Las evaluaciones son coherentes con los contenidos',
     'manual'),
    ('Estrategias metodológicas', 5,
     'Las actividades son consistentes con los Resultados de Aprendizaje',
     'manual'),
    ('Estrategias metodológicas', 6,
     'El trabajo independiente refleja autonomía del estudiante '
     '(tiempos asignados coherentes con la actividad)',
     'auto'),
    ('Estrategias metodológicas', 7,
     'Se establece momento de Preparación para cada actividad o evaluación',
     'auto'),
    ('Estrategias metodológicas', 8,
     'Se establece momento de Desarrollo para cada actividad o evaluación',
     'auto'),
    ('Estrategias metodológicas', 9,
     'Se establece momento de Trabajo Independiente para cada actividad o evaluación',
     'auto'),
    ('Estrategias metodológicas', 10,
     'Se presentan los recursos didácticos para cada actividad o evaluación',
     'auto'),
    ('Estrategias metodológicas', 11,
     'Los recursos didácticos son consistentes con los resultados de aprendizaje',
     'manual'),

    # ── Sección 4: Estrategias evaluativas (3 ítems) ──────────────────────
    ('Estrategias evaluativas', 1,
     'El tipo, procedimiento e instrumento de evaluación son coherentes '
     'y consistentes con el programa oficial',
     'auto'),
    ('Estrategias evaluativas', 2,
     'Existe consistencia entre los tipos de evaluaciones y los RA '
     'conforme al programa oficial (diagnóstica, formativa y sumativa por unidad)',
     'auto'),
    ('Estrategias evaluativas', 3,
     'Porcentajes de evaluaciones son consistentes conforme al programa oficial',
     'auto'),

    # ── Sección 5: Carga académica (9 ítems) ──────────────────────────────
    ('Carga académica', 1,
     'Distribución de horas presenciales conforme a programa oficial',
     'auto'),
    ('Carga académica', 2,
     'Total de horas presenciales conforme a programa oficial',
     'auto'),
    ('Carga académica', 3,
     'Distribución de horas virtuales conforme a programa oficial',
     'auto'),
    ('Carga académica', 4,
     'Total de horas virtuales conforme a programa oficial',
     'auto'),
    ('Carga académica', 5,
     'Diferenciación correcta entre horas sincrónicas y asincrónicas',
     'auto'),
    ('Carga académica', 6,
     'Asignación de horas de preparación consistentes con la carga del estudiante',
     'manual'),
    ('Carga académica', 7,
     'Distribución de horas TPE conforme a programa oficial',
     'auto'),
    ('Carga académica', 8,
     'Total horas TPE conforme a programa oficial',
     'auto'),
    ('Carga académica', 9,
     'Asignación de horas de TPE consistentes con la carga del estudiante',
     'manual'),
]


# ══════════════════════════════════════════════════════════════════════════
#  TAXONOMÍA DE BLOOM — VERBOS PARA RESULTADOS DE APRENDIZAJE
# ══════════════════════════════════════════════════════════════════════════
# Fuente: Anderson & Krathwohl (2001), adaptado al español
# Cada nivel incluye verbos en infinitivo que indican el tipo de aprendizaje

BLOOM_VERBOS = {
    # Nivel 1: Recordar (conocimiento factual y conceptual básico)
    'recordar': {
        'nivel': 1,
        'descripcion': 'Recordar — recuperación de hechos y conceptos',
        'verbos': {
            'definir', 'describir', 'identificar', 'listar', 'reconocer',
            'recordar', 'repetir', 'señalar', 'enumerar', 'nombrar',
            'rotular', 'relatar', 'recitar', 'mencionar', 'indicar',
            'localizar', 'encontrar', 'obtener', 'reproducir', 'adquirir',
        },
    },
    # Nivel 2: Comprender (construcción de significado)
    'comprender': {
        'nivel': 2,
        'descripcion': 'Comprender — interpretación y explicación',
        'verbos': {
            'comprender', 'explicar', 'interpretar', 'resumir', 'clasificar',
            'comparar', 'contrastar', 'ejemplificar', 'inferir', 'parafrasear',
            'traducir', 'representar', 'ilustrar', 'discutir', 'describir',
            'predecir', 'asociar', 'distinguir', 'contextualizar', 'relacionar',
        },
    },
    # Nivel 3: Aplicar (uso de procedimientos en situaciones dadas)
    'aplicar': {
        'nivel': 3,
        'descripcion': 'Aplicar — ejecución y implementación',
        'verbos': {
            'aplicar', 'demostrar', 'ejecutar', 'implementar', 'utilizar',
            'calcular', 'completar', 'construir', 'emplear', 'llevar a cabo',
            'manipular', 'operar', 'practicar', 'producir', 'resolver',
            'usar', 'desarrollar', 'realizar', 'efectuar', 'hacer',
            'ejecutar', 'interpretar', 'ilustrar', 'modificar', 'preparar',
        },
    },
    # Nivel 4: Analizar (descomposición en partes y relaciones)
    'analizar': {
        'nivel': 4,
        'descripcion': 'Analizar — diferenciación y organización',
        'verbos': {
            'analizar', 'diferenciar', 'examinar', 'organizar', 'categorizar',
            'contrastar', 'estructurar', 'integrar', 'separar', 'desglosar',
            'investigar', 'correlacionar', 'ordenar', 'descomponer',
            'comparar', 'criticar', 'debatir', 'examinar', 'sintetizar',
        },
    },
    # Nivel 5: Evaluar (juicio basado en criterios)
    'evaluar': {
        'nivel': 5,
        'descripcion': 'Evaluar — verificación y crítica',
        'verbos': {
            'evaluar', 'juzgar', 'criticar', 'justificar', 'defender',
            'argumentar', 'valorar', 'verificar', 'validar', 'revisar',
            'seleccionar', 'elegir', 'decidir', 'apreciar', 'estimar',
            'contrastar', 'fundamentar', 'sustentar', 'cuestionar',
            'determinar', 'concluir', 'priorizar', 'diagnosticar',
        },
    },
    # Nivel 6: Crear (ensamblaje de elementos en un todo nuevo)
    'crear': {
        'nivel': 6,
        'descripcion': 'Crear — generación y planificación',
        'verbos': {
            'crear', 'diseñar', 'construir', 'desarrollar', 'planificar',
            'elaborar', 'formular', 'generar', 'inventar', 'producir',
            'proponer', 'idear', 'concebir', 'establecer', 'organizar',
            'componer', 'ensamblar', 'formalizar', 'implementar', 'sistematizar',
            'programar', 'proyectar', 'diseñar', 'elaborar', 'rediseñar',
        },
    },
}

# Diccionario inverso para búsqueda rápida: verbo -> nivel
_VERBO_A_NIVEL = {}
for cat, datos in BLOOM_VERBOS.items():
    for v in datos['verbos']:
        _VERBO_A_NIVEL[v.lower()] = cat

# Verbos débiles o genéricos que NO indican aprendizaje medible
VERBOS_DEBILES = {
    'saber', 'conocer', 'entender', 'comprender bien', 'aprender',
    'estudiar', 'revisar', 'ver', 'leer', 'escuchar', 'observar',
    'conocer sobre', 'tener conocimiento', 'estar familiarizado',
    'tener idea', 'tener nociones', 'pensar sobre',
}


def extraer_verbo_inicial(texto: str) -> str | None:
    """
    Extrae el verbo en infinitivo al inicio de un RA.
    Normaliza acentos y mayúsculas.
    """
    if not texto:
        return None
    texto = str(texto).strip().lower()
    # Eliminar numeración tipo "RA1:", "1.", "a)", etc.
    texto = re.sub(r'^(?:ra\s*\d+\s*:?\s*|\d+\.\s*|[a-z]\)\s*)', '', texto, flags=re.I)
    texto = texto.strip()
    # Buscar verbo al inicio (hasta espacio o preposición)
    match = re.match(r'^([a-záéíóúüñ]+)(?:\s|$|,)', texto)
    if match:
        return match.group(1)
    return None


def validar_verbos_bloom_ra(texto_ra: str) -> dict:
    """
    Valida si un Resultado de Aprendizaje inicia con verbo de Bloom.

    Retorna:
    {
        'verbo': str | None,           # verbo detectado
        'nivel': str | None,            # nivel Bloom ('recordar'..'crear')
        'nivel_num': int | None,        # 1-6
        'valido': bool,                 # tiene verbo Bloom
        'debil': bool,                  # es verbo débil/genérico
        'mensaje': str,                 # descripción del resultado
    }
    """
    verbo = extraer_verbo_inicial(texto_ra)
    resultado = {
        'verbo': verbo,
        'nivel': None,
        'nivel_num': None,
        'valido': False,
        'debil': False,
        'mensaje': '',
    }

    if not verbo:
        resultado['mensaje'] = 'No se detectó verbo al inicio'
        return resultado

    if verbo in VERBOS_DEBILES:
        resultado['debil'] = True
        resultado['mensaje'] = f'«{verbo}» es verbo débil — no indica aprendizaje medible'
        return resultado

    if verbo in _VERBO_A_NIVEL:
        nivel = _VERBO_A_NIVEL[verbo]
        datos = BLOOM_VERBOS[nivel]
        resultado['nivel'] = nivel
        resultado['nivel_num'] = datos['nivel']
        resultado['valido'] = True
        resultado['mensaje'] = f'«{verbo}» — Nivel {datos["nivel"]}: {datos["descripcion"]}'
        return resultado

    # Verbo no reconocido (podría ser válido pero no está en el diccionario)
    resultado['mensaje'] = f'«{verbo}» no está en taxonomía Bloom (verificar manualmente)'
    return resultado


def verificar_verbos_bloom_planificacion(wb, log: list) -> tuple[int, int, list]:
    """
    Verifica los verbos de Bloom en todos los RA de la planificación.

    Busca RA en:
    - Síntesis didáctica: filas 25-32, columna A
    - Planificación por unidades: columna A (COL['RA'])

    Parámetros
    ----------
    wb : Workbook
        Libro de planificación
    log : list
        Lista para agregar mensajes de verificación

    Retorna
    ------
    (n_validos, n_problemas, detalles)
    donde detalles es lista de dict con info de cada RA
    """
    n_validos = 0
    n_problemas = 0
    detalles = []

    log.append('\n  [Verificación Taxonomía de Bloom — Resultados de Aprendizaje]')

    # 1. RA en Síntesis didáctica
    ws_sint = _hoja(wb, 'Síntesis didáctica') if _tiene_hoja(wb, 'Síntesis didáctica') else None
    if ws_sint:
        log.append('    Síntesis didáctica:')
        for fila in range(25, 33):
            ra_texto = ws_sint.cell(fila, 1).value
            if not ra_texto:
                continue

            resultado = validar_verbos_bloom_ra(ra_texto)
            detalles.append({
                'hoja': 'Síntesis didáctica',
                'fila': fila,
                'texto': str(ra_texto)[:80],
                **resultado
            })

            icono = '✅' if resultado['valido'] else ('⚠️' if resultado['debil'] else '❓')
            log.append(f'      {icono} Fila {fila}: {resultado["mensaje"]}')
            if resultado['valido']:
                n_validos += 1
            elif resultado['debil']:
                n_problemas += 1

    # 2. RA en Planificación por unidades (columna A)
    ws_plan = _hoja(wb, 'Planificación por unidades') if _tiene_hoja(wb, 'Planificación por unidades') else None
    if ws_plan:
        log.append('    Planificación por unidades:')
        ra_vistos = set()  # evitar duplicados

        for row in ws_plan.iter_rows(min_row=4):
            ra_texto = row[COL['RA'] - 1].value if hasattr(row[COL['RA'] - 1], 'value') else row[COL['RA'] - 1]
            if not ra_texto:
                continue

            ra_str = str(ra_texto).strip()
            if ra_str in ra_vistos:
                continue
            ra_vistos.add(ra_str)

            resultado = validar_verbos_bloom_ra(ra_texto)
            detalles.append({
                'hoja': 'Planificación por unidades',
                'fila': row[COL['RA'] - 1].row if hasattr(row[COL['RA'] - 1], 'row') else 0,
                'texto': ra_str[:80],
                **resultado
            })

            icono = '✅' if resultado['valido'] else ('⚠️' if resultado['debil'] else '❓')
            fila_num = row[COL['RA'] - 1].row if hasattr(row[COL['RA'] - 1], 'row') else '?'
            log.append(f'      {icono} Fila {fila_num}: {resultado["mensaje"]}')
            if resultado['valido']:
                n_validos += 1
            elif resultado['debil']:
                n_problemas += 1

    # Resumen
    total_ra = len(detalles)
    if total_ra == 0:
        log.append('    ⚠️ No se encontraron RA para verificar')
    else:
        log.append(
            f'\n    Bloom: {n_validos}/{total_ra} RA con verbo válido, '
            f'{n_problemas} con verbo débil, '
            f'{total_ra - n_validos - n_problemas} sin verbo reconocido'
        )

    return n_validos, n_problemas, detalles


# ── Colores ───────────────────────────────────────────────────────────────
_AZUL_FUENTE    = '0070C0'
_FILL_FORMATIVA = PatternFill(fill_type='solid', fgColor='E8D5F5')  # lila pastel
_FILL_SUMATIVA  = PatternFill(fill_type='solid', fgColor='FFF2CC')  # amarillo pastel
_FILL_NONE      = PatternFill(fill_type=None)

def _reg(registro, ws, cell, tipo, original, corregido):
    """Añade un registro de cambio a la lista de seguimiento."""
    if registro is None:
        return
    col_letra = get_column_letter(cell.column)
    registro.append({
        'hoja':      ws.title,
        'celda':     f'{col_letra}{cell.row}',
        'tipo':      tipo,
        'original':  str(original)[:300] if original is not None else '(vacío)',
        'corregido': str(corregido)[:300],
    })


def escribir_registro_cambios(wb, registro):
    """
    Escribe/reemplaza la hoja 'Registro de cambios' con todos los cambios aplicados.
    Cada fila indica: N°, Hoja, Celda, Tipo de corrección, Valor original, Valor corregido.
    """
    HOJA = 'Registro de cambios'
    if HOJA in wb.sheetnames:
        del wb[HOJA]
    ws = wb.create_sheet(HOJA)

    # ── Encabezado ─────────────────────────────────────────────────────────
    CABECERAS = ['N°', 'Hoja', 'Celda', 'Tipo de corrección',
                 'Valor original', 'Valor corregido']
    fill_cab = PatternFill(fill_type='solid', fgColor='006633')   # verde UST
    font_cab = Font(bold=True, color='FFFFFF', size=10)

    for col, h in enumerate(CABECERAS, 1):
        c = ws.cell(1, col, h)
        c.font  = font_cab
        c.fill  = fill_cab
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Filas de datos ─────────────────────────────────────────────────────
    fill_par  = PatternFill(fill_type='solid', fgColor='F0F7F4')  # verde pálido (par)
    fill_impar = PatternFill(fill_type=None)
    font_norm = Font(size=9)

    for i, cambio in enumerate(registro, 1):
        fill = fill_par if i % 2 == 0 else fill_impar
        valores = [
            i,
            cambio.get('hoja', ''),
            cambio.get('celda', ''),
            cambio.get('tipo', ''),
            cambio.get('original', ''),
            cambio.get('corregido', ''),
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(i + 1, col, val)
            c.font      = font_norm
            c.fill      = fill
            c.alignment = Alignment(vertical='top', wrap_text=(col >= 4))

    # ── Anchos de columna ──────────────────────────────────────────────────
    anchos = {'A': 5, 'B': 26, 'C': 7, 'D': 38, 'E': 52, 'F': 52}
    for col_letra, ancho in anchos.items():
        ws.column_dimensions[col_letra].width = ancho

    ws.row_dimensions[1].height = 22

    # ── Resumen al final ───────────────────────────────────────────────────
    if registro:
        fila_res = len(registro) + 3
        ws.cell(fila_res, 1, f'Total: {len(registro)} cambios aplicados').font = Font(bold=True, size=9)

    return len(registro)


def aplicar_azul(cell):
    """Aplica color de fuente azul (#0070C0) sin tocar el fondo de la celda."""
    f = cell.font
    cell.font = Font(name=f.name, size=f.size, bold=f.bold,
                     italic=f.italic, underline=f.underline, color=_AZUL_FUENTE)


def aplicar_azul_diff(cell, texto_original: str, texto_nuevo: str):
    """
    Para celdas multi-línea (col G): aplica fuente azul SOLO a las líneas
    que cambiaron entre texto_original y texto_nuevo.
    Las líneas sin cambios conservan su color original.
    Requiere openpyxl >= 3.1 (CellRichText). Si no está disponible, colorea toda la celda.
    """
    import difflib
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        aplicar_azul(cell)
        return

    f = cell.font
    try:
        orig_color = (f.color.rgb if f.color and f.color.type == 'rgb' else '000000')
    except Exception:
        orig_color = '000000'

    lineas_orig = texto_original.split('\n')
    lineas_new  = texto_nuevo.split('\n')

    # Detectar qué líneas cambiaron
    sm = difflib.SequenceMatcher(None, lineas_orig, lineas_new, autojunk=False)
    cambiadas = set()
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag != 'equal':
            for idx in range(j1, j2):
                cambiadas.add(idx)

    # Construir rich text: líneas cambiadas → azul, resto → color original
    rt = CellRichText()
    for i, linea in enumerate(lineas_new):
        segmento = linea + ('\n' if i < len(lineas_new) - 1 else '')
        color    = _AZUL_FUENTE if i in cambiadas else orig_color
        rt.append(TextBlock(
            InlineFont(rFont=f.name or 'Calibri', b=f.bold, i=f.italic, color=color),
            segmento,
        ))

    cell.value = rt


def aplicar_colores_evaluacion(ws):
    """
    Colorea el fondo de las celdas de Estrategias evaluativas (columnas K–O)
    según tipo de evaluación:
      Formativa / Formativo → lila pastel  (#E8D5F5)
      Sumativa              → amarillo pastel (#FFF2CC)
    Solo se aplica a columnas 11–15 (K a O) para no afectar el resto de la fila.
    Las celdas fusionadas (MergedCell) se omiten para evitar errores.
    """
    from openpyxl.cell.cell import MergedCell
    COL_DESDE = 11   # K
    COL_HASTA = 15   # O
    for r in range(4, (ws.max_row or 4) + 1):
        tipo = ws.cell(r, COL['TIPO']).value
        if not tipo:
            continue
        t = str(tipo).lower()
        if 'formativ' in t:
            fill = _FILL_FORMATIVA
        elif 'sumativ' in t:
            fill = _FILL_SUMATIVA
        else:
            continue
        for col_idx in range(COL_DESDE, COL_HASTA + 1):
            cell = ws.cell(r, col_idx)
            if not isinstance(cell, MergedCell):
                cell.fill = fill


# ══════════════════════════════════════════════════════════════════════════
#  FUNCIONES AUXILIARES
# ══════════════════════════════════════════════════════════════════════════

def encontrar_archivo(carpeta, patron):
    """Busca el primer archivo que coincida con el patrón glob en una carpeta."""
    resultados = glob.glob(os.path.join(carpeta, patron))
    return resultados[0] if resultados else None


def encontrar_planificacion_enviada(carpeta_enviado):
    """
    En 'Enviado a DEL/', devuelve el .xlsx que NO es la escala de apreciación.
    Devuelve (path, escala_path) o (None, None).
    Si hay más de un xlsx que no es escala, usa el más reciente y advierte en stderr.
    """
    xlsx_files = glob.glob(os.path.join(carpeta_enviado, '*.xlsx'))
    planes = []
    escala_path = None
    for f in xlsx_files:
        nombre = os.path.basename(f).lower()
        if nombre.startswith('escala'):
            escala_path = f
        else:
            planes.append(f)
    if len(planes) > 1:
        planes.sort(key=os.path.getmtime, reverse=True)
        import sys
        print(
            f'  ⚠️  Múltiples planificaciones en "{carpeta_enviado}" — '
            f'se usa la más reciente: {os.path.basename(planes[0])}',
            file=sys.stderr
        )
    plan_path = planes[0] if planes else None
    return plan_path, escala_path


def leer_observaciones_escala(escala_path):
    """
    Lee la escala de apreciación y extrae las observaciones de mejora
    (columna F / 'Solicitud de Mejora') y comentarios de 2ª validación que tienen texto.
    Devuelve lista de (criterio, observacion).
    """
    if not escala_path:
        return []
    try:
        wb = openpyxl.load_workbook(escala_path, data_only=True)
        ws = wb.active
        observaciones = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            criterio   = row[0]  if len(row) > 0  else None
            obs_1      = row[5]  if len(row) > 5  else None   # col F — 1ª validación
            obs_2      = row[9]  if len(row) > 9  else None   # col J — 2ª validación
            resultado  = row[2]  if len(row) > 2  else None   # col C — Si/X
            if not criterio:
                continue
            # Recoger observaciones reales (no encabezados de columna)
            obs_texto = []
            if obs_1 and str(obs_1).strip() and str(obs_1).strip().lower() not in (
                    'observaciones', 'solicitud de mejora/ observaciones',
                    'comentarios de postgrado post-1° validación'):
                obs_texto.append(str(obs_1).strip())
            if obs_2 and str(obs_2).strip() and str(obs_2).strip().lower() not in (
                    'comentarios 2° validación', 'ok'):
                obs_texto.append(f'[2ª val.] {str(obs_2).strip()}')
            if obs_texto:
                observaciones.append((str(criterio).strip(), ' | '.join(obs_texto)))
        return observaciones
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════
#  INSTANCIA 2 — APLICAR CORRECCIONES DE ESCALA DEL
# ══════════════════════════════════════════════════════════════════════════

_OBS_IGNORAR = {
    'si', 'parcialmente', 'no', 'observaciones',
    'solicitud de mejora/ observaciones',
    'comentarios de postgrado post-1° validación',
    'comentarios de postgrado post-1Â° validaciÃ³n',
    'comentarios 2° validación', 'comentarios 2Â° validaciÃ³n',
    'ok', 'identificación', 'resultados', 'validación inicial',
    'sanción validador(a) inicial',
}

_SECCIONES_ESCALA = (
    'síntesis didáctica', 'planificación por unidades',
    'estrategias metodológicas', 'estrategias evaluativas',
    'carga académica',
)


def leer_escala_completa(escala_path):
    """
    Lee la escala de apreciación completa.
    Devuelve lista de dicts:
      { seccion, criterio, estado ('Si'|'Parcialmente'|'No'), obs_texto }
    Solo incluye filas con marca de estado (X/x en col C, D o E).
    """
    if not escala_path:
        return []
    try:
        wb_e = openpyxl.load_workbook(escala_path, data_only=True)
        ws_e = wb_e.active
    except Exception:
        return []

    criterios = []
    seccion_actual = None

    for row in ws_e.iter_rows(min_row=4, values_only=True):
        criterio = row[0] if len(row) > 0 else None
        si_val   = row[2] if len(row) > 2 else None   # col C — Sí
        parc_val = row[3] if len(row) > 3 else None   # col D — Parcialmente
        no_val   = row[4] if len(row) > 4 else None   # col E — No
        obs_f    = row[5] if len(row) > 5 else None   # col F — Solicitud de Mejora
        obs_i    = row[8] if len(row) > 8 else None   # col I — Post-1ª validación
        obs_j    = row[9] if len(row) > 9 else None   # col J — 2ª validación

        if not criterio:
            continue
        c_lower = str(criterio).lower().strip()

        # Detectar encabezado de sección
        for sec in _SECCIONES_ESCALA:
            if c_lower.startswith(sec):
                seccion_actual = str(criterio).strip()
                break

        # Saltar filas sin marca de estado
        si_ok   = bool(si_val   and str(si_val).strip().upper()   == 'X')
        parc_ok = bool(parc_val and str(parc_val).strip().upper() == 'X')
        no_ok   = bool(no_val   and str(no_val).strip().upper()   == 'X')
        if not (si_ok or parc_ok or no_ok):
            continue

        if c_lower in _OBS_IGNORAR:
            continue

        estado = 'No' if no_ok else ('Parcialmente' if parc_ok else 'Si')

        # Construir texto de observación consolidado
        partes = []
        for o in (obs_f, obs_i, obs_j):
            if o:
                t = str(o).strip()
                if t and t.lower() not in _OBS_IGNORAR:
                    partes.append(t)
        obs_texto = ' | '.join(partes) if partes else None

        criterios.append({
            'seccion':   seccion_actual,
            'criterio':  str(criterio).strip(),
            'estado':    estado,
            'obs_texto': obs_texto,
        })

    return criterios


def _col_para_criterio(criterio_texto):
    """
    Mapea el texto de un criterio a la(s) columna(s) relevantes de la planificación.
    Devuelve lista de claves de COL (pueden ser varias).
    """
    t = str(criterio_texto).lower()
    cols = []
    if 'instrumento'                               in t: cols.append('INSTR')
    if 'procedimiento'                             in t: cols.append('PROC')
    if 'tipo de evaluación' in t or 'tipos de ev' in t: cols.append('TIPO')
    if 'recurso'                                   in t: cols.append('RECURSOS')
    if 'resultado' in t and 'aprendizaje'          in t: cols.append('RA')
    if 'contenido'                                 in t: cols.append('CONTENIDOS')
    if 'medio de entrega'                          in t: cols.append('MEDIO')
    if 'porcentaje'                                in t: cols.append('PCT')
    if 'individual' in t or 'grupal'               in t: cols.append('INDIV')
    if ('redacción' in t or 'momentos' in t
            or ('actividad' in t and 'estrategia' not in t)):
                                                         cols.append('ACTIVIDAD')
    return cols


def _escribir_notas_i2(wb, fallidos):
    """Escribe/reemplaza la hoja NOTAS_CORRECCIONES_DEL con el resumen de I2."""
    HOJA = 'NOTAS_CORRECCIONES_DEL'
    if HOJA in wb.sheetnames:
        del wb[HOJA]
    ws_n = wb.create_sheet(HOJA)

    ws_n.cell(1, 1, 'OBSERVACIONES REVISORA DEL').font = Font(bold=True, size=12)
    ws_n.cell(2, 1, 'Sección').font   = Font(bold=True)
    ws_n.cell(2, 2, 'Criterio').font  = Font(bold=True)
    ws_n.cell(2, 3, 'Estado').font    = Font(bold=True)
    ws_n.cell(2, 4, 'Observación').font = Font(bold=True)

    fill_no   = PatternFill(fill_type='solid', fgColor='F8D7DA')
    fill_parc = PatternFill(fill_type='solid', fgColor='FFF3CD')

    for i, item in enumerate(fallidos, start=3):
        ws_n.cell(i, 1, item['seccion']  or '—')
        ws_n.cell(i, 2, item['criterio'])
        ws_n.cell(i, 3, item['estado'])
        ws_n.cell(i, 4, item['obs_texto'] or '—')
        fill = fill_no if item['estado'] == 'No' else fill_parc
        for c in range(1, 5):
            ws_n.cell(i, c).fill = fill

    ws_n.column_dimensions['B'].width = 60
    ws_n.column_dimensions['D'].width = 80


def aplicar_obs_escala_i2(wb, criterios, log, registro=None):
    """
    Para cada criterio ❌ o Parcialmente con texto de observación:
    - Inyecta el texto en azul en las celdas vacías de la(s) columna(s) mapeadas.
    - Escribe resumen en hoja NOTAS_CORRECCIONES_DEL.
    Devuelve número de anotaciones aplicadas.
    """
    from openpyxl.cell.cell import MergedCell

    ws_plan = _hoja(wb, 'Planificación por unidades') if _tiene_hoja(wb, 'Planificación por unidades') else None
    if not ws_plan:
        return 0

    fallidos = [c for c in criterios
                if c['estado'] in ('No', 'Parcialmente') and c['obs_texto']]

    if not fallidos:
        log.append('  [I2] Sin observaciones de la revisora con texto.')
        _escribir_notas_i2(wb, [])
        return 0

    log.append(f'\n  [Instancia 2 — Observaciones revisora DEL ({len(fallidos)} criterios)]')
    total_anotaciones = 0

    for item in fallidos:
        criterio = item['criterio']
        obs      = item['obs_texto']
        estado   = item['estado']
        cols     = _col_para_criterio(criterio)
        icono    = '❌' if estado == 'No' else '⚠️'

        log.append(f'  {icono} {criterio[:75]}')
        if obs:
            log.append(f'      → {obs[:140]}')

        if not cols:
            log.append('      [sin mapeo automático — ver hoja NOTAS_CORRECCIONES_DEL]')
            continue

        obs_corta = obs[:100] + ('…' if len(obs) > 100 else '')

        for col_key in cols:
            col_num = COL.get(col_key)
            if not col_num:
                continue

            for row in ws_plan.iter_rows(min_row=4):
                r = row[0].row
                # Solo filas con actividad o evaluación (no filas vacías)
                tiene_contenido = (ws_plan.cell(r, COL['ACTIVIDAD']).value
                                   or ws_plan.cell(r, COL['TIPO']).value)
                if not tiene_contenido:
                    continue

                target = ws_plan.cell(r, col_num)
                if isinstance(target, MergedCell):
                    continue

                if not target.value:
                    texto_obs = f'[DEL: {obs_corta}]'
                    target.value = texto_obs
                    aplicar_azul(target)
                    _reg(registro, ws_plan, target,
                         f'Observación revisora DEL ({estado}): {criterio[:60]}',
                         None, texto_obs)
                    total_anotaciones += 1

    _escribir_notas_i2(wb, fallidos)
    log.append(f'  Anotaciones inyectadas: {total_anotaciones}  '
               f'| Resumen completo en hoja NOTAS_CORRECCIONES_DEL')
    return total_anotaciones


def procesar_instancia2(plan_bytes, escala_bytes,
                        plan_nombre='planificacion.xlsx',
                        programa=None, es_as=False,
                        instancia_num=2):
    """
    Flujo Instancia 2 / 3: aplica correcciones automáticas + observaciones de la
    revisora DEL (escala completada) sobre la planificación del docente.

    Parámetros
    ----------
    plan_bytes    : bytes del .xlsx del docente
    escala_bytes  : bytes de la escala completada por la revisora
    plan_nombre   : nombre original del archivo (para el nombre de salida)
    programa      : dict extraído por extraer_programa_pdf (opcional)
    es_as         : bool — activa verificación hitos A+Se
    instancia_num : 2 ó 3 — controla encabezado y nombre del archivo de salida

    Devuelve
    --------
    (log: list[str], ok: bool, output_bytes: bytes, output_nombre: str)
    """
    from io import BytesIO

    log = [
        f'\n{"═"*70}',
        f'  INSTANCIA {instancia_num} — Aplicar correcciones de escala DEL',
        f'{"═"*70}',
        f'  Plan   : {plan_nombre}',
    ]

    # ── Leer escala ────────────────────────────────────────────────────────
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_e:
            tmp_e.write(escala_bytes)
            tmp_escala = tmp_e.name
        criterios = leer_escala_completa(tmp_escala)
        os.unlink(tmp_escala)
    except Exception as exc:
        log.append(f'  ❌ Error leyendo escala: {exc}')
        return log, False, b'', ''

    n_total   = len(criterios)
    n_fallidos = sum(1 for c in criterios if c['estado'] in ('No', 'Parcialmente'))
    n_ok       = n_total - n_fallidos
    log.append(f'  Escala  : {n_total} criterios leídos | '
               f'{n_ok} ✅ | {n_fallidos} con observaciones')

    # ── Cargar planificación ───────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(BytesIO(plan_bytes))
    except Exception as exc:
        log.append(f'  ❌ Error cargando planificación: {exc}')
        return log, False, b'', ''

    hojas = wb.sheetnames
    log.append(f'  Hojas   : {hojas}')

    # ── Correcciones automáticas (misma lógica I1) ─────────────────────────
    registro = []   # Registro estructurado de cambios para la hoja final
    total_sint = 0
    if _tiene_hoja(wb, 'Síntesis didáctica'):
        log.append('\n  [Síntesis didáctica — correcciones automáticas]')
        total_sint = corregir_sintesis(_hoja(wb, 'Síntesis didáctica'), log, registro)
        if total_sint == 0:
            log.append('    Sin cambios necesarios.')

    total_plan = 0
    if _tiene_hoja(wb, 'Planificación por unidades'):
        log.append('\n  [Planificación por unidades — correcciones automáticas]')
        a, mj, mm, b = corregir_planificacion(_hoja(wb, 'Planificación por unidades'), log, registro)
        total_leng   = corregir_lenguaje_actividades(
            _hoja(wb, 'Planificación por unidades'), log, registro)
        total_plan   = a + mj + mm + b + total_leng
        if total_plan == 0:
            log.append('    Sin cambios necesarios.')

    # ── Observaciones de la revisora ───────────────────────────────────────
    n_obs = aplicar_obs_escala_i2(wb, criterios, log, registro)

    # ── Colores de fila (Formativa/Sumativa) ──────────────────────────────
    if _tiene_hoja(wb, 'Planificación por unidades'):
        aplicar_colores_evaluacion(_hoja(wb, 'Planificación por unidades'))

    # ── Verificación 41 criterios ──────────────────────────────────────────
    resultados_escala = verificar_escala(wb)
    log.extend(formatear_verificacion(resultados_escala, 'archivo'))

    # ── Verificación de momentos (Manual Diseño Instruccional UST) ──────────
    if _tiene_hoja(wb, 'Planificación por unidades'):
        verificar_momentos(_hoja(wb, 'Planificación por unidades'), log)

    # ── Revisión ortográfica/gramatical con LanguageTool ─────────────────
    if _tiene_hoja(wb, 'Planificación por unidades') and _LANGUAGETOOL_ACTIVO:
        verificar_lenguaje_momentos(
            _hoja(wb, 'Planificación por unidades'), log,
            autocorregir=_LANGUAGETOOL_AUTOCORREGIR
        )

    # ── Verificación de horas ─────────────────────────────────────────────
    _ws_sint = _hoja(wb, 'Síntesis didáctica') if _tiene_hoja(wb, 'Síntesis didáctica') else None
    _ws_plan = _hoja(wb, 'Planificación por unidades') if _tiene_hoja(wb, 'Planificación por unidades') else None
    verificar_horas(_ws_plan, _ws_sint, programa, log)

    # ── Verificación vs programa ───────────────────────────────────────────
    verificar_contra_programa(wb, programa, log)

    # ── Verificación A+Se ──────────────────────────────────────────────────
    if es_as and _tiene_hoja(wb, 'Planificación por unidades'):
        verificar_as(_hoja(wb, 'Planificación por unidades'), log)

    # ── Guardar ────────────────────────────────────────────────────────────
    total = total_sint + total_plan + n_obs
    log.append(f'\n  RESUMEN I2:')
    log.append(f'    Auto-correcciones I1     : {total_sint + total_plan}')
    log.append(f'    Anotaciones revisora DEL : {n_obs}')
    log.append(f'    TOTAL                    : {total}')
    log.append(f'    🔵 Texto azul = corrección automática o anotación revisora')
    log.append(f'    📋 Hoja NOTAS_CORRECCIONES_DEL = resumen completo de observaciones')
    log.append(f'    📋 Hoja "Registro de cambios" = {len(registro)} cambio(s) documentado(s)')

    escribir_registro_cambios(wb, registro)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    output_bytes = buf.read()

    output_nombre = nombre_instancia(plan_nombre, instancia_num)

    return log, True, output_bytes, output_nombre


def verificar_escala(wb):
    """
    Verifica automáticamente los criterios marcados como 'auto' en ESCALA_ESTANDAR
    contra el contenido del workbook.
    Devuelve lista de (seccion, n, criterio, estado, detalle)
    donde estado es: '✅', '❌', '⚠️ manual'
    """
    resultados = []

    ws_sint = _hoja(wb, 'Síntesis didáctica') if _tiene_hoja(wb, 'Síntesis didáctica') else None
    ws_plan = _hoja(wb, 'Planificación por unidades') if _tiene_hoja(wb, 'Planificación por unidades') else None

    # Leer síntesis en dict para acceso fácil
    sint = {}
    if ws_sint:
        # Leer valores de celdas clave por posición
        def sv(r, c): return ws_sint.cell(r, c).value

        sint = {
            'escuela':      sv(4, 1),
            'programa':     sv(4, 2),
            'asignatura':   sv(4, 4),
            'codigo':       sv(7, 1),
            'version':      sv(7, 2),
            'creditos':     sv(7, 3),
            'area':         sv(11, 1),
            'requisitos':   sv(11, 2),
            'hrs_pres':     sv(9, 4),
            'hrs_sinc':     sv(9, 5),
            'hrs_asinc':    sv(9, 6),
            'hrs_tpe':      sv(9, 7),
        }
        # Competencias (filas 13-15 col B)
        sint['competencias'] = [ws_sint.cell(r, 2).value for r in range(13, 16)
                                if ws_sint.cell(r, 2).value]
        # Unidades (filas 18-21 col B y G)
        sint['unidades'] = [(ws_sint.cell(r, 2).value, ws_sint.cell(r, 7).value)
                            for r in range(18, 22)
                            if ws_sint.cell(r, 2).value]
        # RA + procedimientos (filas 25-29 col A y D)
        sint['ra_procs'] = [(ws_sint.cell(r, 1).value,
                             ws_sint.cell(r, 4).value,
                             ws_sint.cell(r, 7).value)
                            for r in range(25, 32)
                            if ws_sint.cell(r, 1).value]

    def ok(val, nombre):
        if val:
            return ('✅', f'{nombre}: presente')
        return ('❌', f'{nombre}: vacío o ausente')

    def check(seccion, n, criterio, modo, estado, detalle):
        resultados.append((seccion, n, criterio, estado, detalle))

    # ── Sección 1: Síntesis Didáctica ─────────────────────────────────────
    if ws_sint:
        e, d = ok(sint.get('escuela'), 'Escuela')
        check('Síntesis Didáctica', 1, ESCALA_ESTANDAR[0][2], 'auto', e, d)

        e, d = ok(sint.get('programa'), 'Programa')
        check('Síntesis Didáctica', 2, ESCALA_ESTANDAR[1][2], 'auto', e, d)

        e, d = ok(sint.get('asignatura'), 'Nombre asignatura')
        check('Síntesis Didáctica', 3, ESCALA_ESTANDAR[2][2], 'auto', e, d)

        e, d = ok(sint.get('codigo'), 'Código')
        check('Síntesis Didáctica', 4, ESCALA_ESTANDAR[3][2], 'auto', e, d)

        e, d = ok(sint.get('version'), 'Versión año')
        check('Síntesis Didáctica', 5, ESCALA_ESTANDAR[4][2], 'auto', e, d)

        e, d = ok(sint.get('creditos'), 'N° créditos')
        check('Síntesis Didáctica', 6, ESCALA_ESTANDAR[5][2], 'auto', e, d)

        e, d = ok(sint.get('area'), 'Área OCDE')
        check('Síntesis Didáctica', 7, ESCALA_ESTANDAR[6][2], 'auto', e, d)

        e, d = ok(sint.get('requisitos'), 'Requisitos')
        check('Síntesis Didáctica', 8, ESCALA_ESTANDAR[7][2], 'auto', e, d)

        horas = [sint.get('hrs_pres'), sint.get('hrs_sinc'),
                 sint.get('hrs_asinc'), sint.get('hrs_tpe')]
        if all(h is not None for h in horas):
            e, d = '✅', f'Horas: pres={horas[0]} sinc={horas[1]} asinc={horas[2]} TPE={horas[3]}'
        else:
            e, d = '❌', f'Faltan horas: {horas}'
        check('Síntesis Didáctica', 9, ESCALA_ESTANDAR[8][2], 'auto', e, d)

        n_comp = len(sint.get('competencias', []))
        e = '✅' if n_comp > 0 else '❌'
        check('Síntesis Didáctica', 10, ESCALA_ESTANDAR[9][2], 'auto',
              e, f'{n_comp} competencia(s) declarada(s)')

        unidades = sint.get('unidades', [])
        n_u = len(unidades)
        todas_con_horas = all(h for _, h in unidades)
        e = '✅' if n_u > 0 and todas_con_horas else ('⚠️' if n_u > 0 else '❌')
        check('Síntesis Didáctica', 11, ESCALA_ESTANDAR[10][2], 'auto',
              e, f'{n_u} unidad(es); todas con horas={todas_con_horas}')

        e = '✅' if n_u > 0 else '❌'
        check('Síntesis Didáctica', 12, ESCALA_ESTANDAR[11][2], 'auto',
              e, f'{n_u} título(s) de unidad presentes')

        ra_procs = sint.get('ra_procs', [])
        n_ra = len(ra_procs)
        e = '✅' if n_ra > 0 else '❌'
        check('Síntesis Didáctica', 13, ESCALA_ESTANDAR[12][2], 'auto',
              e, f'{n_ra} RA declarado(s)')

        # Criterio 14: procedimientos solo SUMATIVA
        procs_malos = [(r, p) for r, p, _ in ra_procs
                       if p and ('DIAGNÓSTICA' in str(p) or 'FORMATIVA' in str(p))]
        e = '✅' if not procs_malos else '❌'
        d = 'Solo SUMATIVA en procedimientos' if not procs_malos \
            else f'Contiene DIAGNÓSTICA/FORMATIVA en {len(procs_malos)} celda(s)'
        check('Síntesis Didáctica', 14, ESCALA_ESTANDAR[13][2], 'auto', e, d)

        # Criterio 15: porcentajes
        pcts = [p for _, _, p in ra_procs if isinstance(p, (int, float))]
        total_pct = sum(pcts)
        e = '✅' if abs(total_pct - 1.0) < 0.01 or abs(total_pct - 100) < 1 else '❌'
        check('Síntesis Didáctica', 15, ESCALA_ESTANDAR[14][2], 'auto',
              e, f'Suma porcentajes = {total_pct}')
    else:
        for i in range(15):
            check('Síntesis Didáctica', i+1, ESCALA_ESTANDAR[i][2], 'auto',
                  '❌', 'Hoja no encontrada')

    # ── Sección 2: Planificación por unidades ─────────────────────────────
    if ws_plan:
        unidades_plan = set()
        ra_plan = set()
        momentos = set()
        filas_eval = 0
        filas_sin_horas = 0
        sesiones_con_prep = sesiones_con_des = sesiones_con_ti = 0
        semana_actual = None

        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            unidad  = row[COL['UNIDAD']  - 1]
            ra      = row[COL['RA']      - 1]
            semana  = row[COL['SEMANA']  - 1]
            momento = row[COL['MOMENTO'] - 1]
            tipo    = row[COL['TIPO']    - 1]
            hrs_p   = row[COL['MODALIDAD'] - 1]  # col E modalidad

            if unidad: unidades_plan.add(str(unidad))
            if ra:     ra_plan.add(str(ra)[:40])
            if semana and semana != 'Total carga académica del estudiante':
                semana_actual = semana
            if momento:
                m = str(momento).lower()
                if 'preparación' in m:   sesiones_con_prep += 1
                if 'desarrollo' in m:    sesiones_con_des  += 1
                if 'independiente' in m: sesiones_con_ti   += 1
            if tipo:
                filas_eval += 1

        n_u = len(unidades_plan) - 1 if '' in unidades_plan else len(unidades_plan)
        e = '✅' if n_u > 0 else '❌'
        check('Planificación por unidades', 1, ESCALA_ESTANDAR[15][2], 'auto',
              e, f'{n_u} unidad(es) identificada(s)')

        e = '✅' if len(ra_plan) > 0 else '❌'
        check('Planificación por unidades', 2, ESCALA_ESTANDAR[16][2], 'auto',
              e, f'{len(ra_plan)} RA vinculado(s) a unidades')

        todos = sesiones_con_prep > 0 and sesiones_con_des > 0 and sesiones_con_ti > 0
        e = '✅' if todos else '❌'
        check('Planificación por unidades', 3, ESCALA_ESTANDAR[17][2], 'auto',
              e, f'Prep={sesiones_con_prep} Des={sesiones_con_des} TI={sesiones_con_ti}')
    else:
        for i in range(3):
            check('Planificación por unidades', i+1, ESCALA_ESTANDAR[15+i][2], 'auto',
                  '❌', 'Hoja no encontrada')

    # ── Sección 3: Estrategias metodológicas ──────────────────────────────
    idx = 18  # índice en ESCALA_ESTANDAR
    if ws_plan:
        filas_actividad = filas_recursos = filas_contenido_eval = 0
        filas_eval_sin_contenido = 0
        prep_rows = des_rows = ti_rows = 0
        filas_eval_total = 0

        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            momento   = str(row[COL['MOMENTO']   - 1] or '').lower()
            actividad = row[COL['ACTIVIDAD'] - 1]
            recursos  = row[COL['RECURSOS']  - 1]
            contenido = row[COL['CONTENIDOS']- 1]
            tipo      = row[COL['TIPO']      - 1]

            if actividad: filas_actividad += 1
            if recursos:  filas_recursos  += 1
            if tipo:
                filas_eval_total += 1
                if contenido: filas_contenido_eval += 1
            if 'preparación'   in momento: prep_rows += 1
            if 'desarrollo'    in momento: des_rows  += 1
            if 'independiente' in momento: ti_rows   += 1

        e = '✅' if filas_actividad > 0 else '❌'
        check('Estrategias metodológicas', 1, ESCALA_ESTANDAR[idx][2], 'auto',
              e, f'{filas_actividad} filas con instrucciones de actividad')
        idx += 1

        # Criterio 2: contenidos de evaluaciones
        # Las filas de evaluación deben tener instrucciones en col G (Actividad)
        filas_eval_con_instruc = 0
        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            tipo_r     = row[COL['TIPO']     - 1]
            actividad_r= row[COL['ACTIVIDAD']- 1]
            if tipo_r and actividad_r:
                filas_eval_con_instruc += 1
        if filas_eval_total > 0:
            e = '✅' if filas_eval_con_instruc >= filas_eval_total * 0.8 else '⚠️'
            d = f'{filas_eval_con_instruc}/{filas_eval_total} filas de evaluación con instrucciones'
        else:
            e, d = '⚠️', 'No se detectaron filas de evaluación'
        check('Estrategias metodológicas', 2, ESCALA_ESTANDAR[idx][2], 'auto', e, d)
        idx += 1

        # Criterios 3, 4, 5 — manual
        for i in range(3):
            check('Estrategias metodológicas', 3+i, ESCALA_ESTANDAR[idx][2], 'manual',
                  '⚠️ manual', 'Requiere revisión humana')
            idx += 1

        # Criterio 6: TPE autonomía — verificable estructuralmente
        e = '✅' if ti_rows > 0 else '❌'
        check('Estrategias metodológicas', 6, ESCALA_ESTANDAR[idx][2], 'auto',
              e, f'{ti_rows} filas con Trabajo Independiente')
        idx += 1

        # Criterios 7, 8, 9: momentos
        for label, count, n in [('Preparación', prep_rows, 7),
                                  ('Desarrollo',  des_rows,  8),
                                  ('T. Independiente', ti_rows, 9)]:
            e = '✅' if count > 0 else '❌'
            check('Estrategias metodológicas', n, ESCALA_ESTANDAR[idx][2], 'auto',
                  e, f'{count} filas con momento {label}')
            idx += 1

        # Criterio 10: recursos didácticos
        e = '✅' if filas_recursos > 0 else '❌'
        check('Estrategias metodológicas', 10, ESCALA_ESTANDAR[idx][2], 'auto',
              e, f'{filas_recursos} filas con recursos didácticos')
        idx += 1

        # Criterio 11 — manual
        check('Estrategias metodológicas', 11, ESCALA_ESTANDAR[idx][2], 'manual',
              '⚠️ manual', 'Requiere revisión humana')
        idx += 1

    # ── Sección 4: Estrategias evaluativas ────────────────────────────────
    if ws_plan:
        tipos_por_unidad = {}
        procs_invalidos  = []
        instrs_invalidos = []
        pct_total = 0

        unidad_actual = None
        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            unidad = row[COL['UNIDAD'] - 1]
            tipo   = row[COL['TIPO']   - 1]
            proc   = row[COL['PROC']   - 1]
            instr  = row[COL['INSTR']  - 1]
            pct    = row[COL['PCT']    - 1]

            if unidad: unidad_actual = str(unidad)
            if not tipo: continue

            t = str(tipo).lower()
            if unidad_actual not in tipos_por_unidad:
                tipos_por_unidad[unidad_actual] = set()
            tipos_por_unidad[unidad_actual].add(t)

            # Cuestionario es válido para Diagnóstica; solo flag si es Sumativa/Formativa
            _tipo_v = str(tipo or '').lower()
            if proc in PROC_MAP and not (proc == 'Cuestionario' and 'diagnóst' in _tipo_v):
                procs_invalidos.append(f'F? proc="{proc}"')
            if instr in ('Pauta', 'Rúbrica'):
                instrs_invalidos.append(f'F? instr="{instr}"')
            if isinstance(pct, (int, float)) and pct > 0:
                pct_total += pct

        # Criterio 1: tipo/proc/instr coherentes
        errores = len(procs_invalidos) + len(instrs_invalidos)
        e = '✅' if errores == 0 else '❌'
        d = 'Sin errores de procedimiento/instrumento' if errores == 0 \
            else f'{errores} error(es): {procs_invalidos + instrs_invalidos}'
        check('Estrategias evaluativas', 1, ESCALA_ESTANDAR[idx][2], 'auto', e, d)
        idx += 1

        # Criterio 2: estructura evaluativa según Manual de Diseño Instruccional UST
        # - Diagnóstica: 1 por asignatura (no por unidad) → solo se verifica global
        # - Formativa: ≥ 1 por unidad temática
        # - Sumativa: exactamente 1 por unidad temática
        # Fuente: Manual Diseño Instruccional E-Learning, Res. N°481/24, p. 18
        EXCLUIR_UNIDADES = {'evaluación final', 'lga', 'examen'}
        faltantes = []

        # Verificación global: ¿hay al menos 1 Diagnóstica en toda la planificación?
        tiene_diagnostica_global = any(
            'diagnóstica' in t
            for tipos in tipos_por_unidad.values()
            for t in tipos
        )
        if not tiene_diagnostica_global:
            faltantes.append('Asignatura: falta evaluación Diagnóstica (1 por asignatura)')

        # Verificación por unidad: Formativa ≥1 y Sumativa = 1
        for u, tipos in tipos_por_unidad.items():
            if any(ex in str(u).lower() for ex in EXCLUIR_UNIDADES):
                continue
            if not any('formativa' in t for t in tipos):
                faltantes.append(f'{u[:35]}: falta Formativa (≥1 por unidad)')
            if not any('sumativa' in t for t in tipos):
                faltantes.append(f'{u[:35]}: falta Sumativa (1 por unidad)')

        e = '✅' if not faltantes else '❌'
        d = ('Diagnóstica (global) + Formativa y Sumativa presentes por unidad'
             if not faltantes else '; '.join(faltantes))
        check('Estrategias evaluativas', 2, ESCALA_ESTANDAR[idx][2], 'auto', e, d)
        idx += 1

        # Criterio 3: porcentajes sumativos de unidades suman 100%
        # Excluir el examen final (que es adicional al 100%)
        pct_unidades = 0
        _unidad_actual_pct = None
        for row in ws_plan.iter_rows(min_row=4, values_only=True):
            tipo_r   = row[COL['TIPO']   - 1]
            pct_r    = row[COL['PCT']    - 1]
            unidad_r = row[COL['UNIDAD'] - 1]
            if unidad_r:
                _unidad_actual_pct = str(unidad_r)
            es_excluida = any(ex in str(_unidad_actual_pct or '').lower()
                              for ex in EXCLUIR_UNIDADES)
            if (tipo_r and 'sumativa' in str(tipo_r).lower()
                    and isinstance(pct_r, (int, float)) and pct_r > 0
                    and not es_excluida):
                pct_unidades += pct_r
        e = '✅' if abs(pct_unidades - 100) < 1 or abs(pct_unidades - 1.0) < 0.01 else '⚠️'
        check('Estrategias evaluativas', 3, ESCALA_ESTANDAR[idx][2], 'auto',
              e, f'Suma % sumativos (sin examen) = {pct_unidades}')
        idx += 1

    # ── Sección 5: Carga académica ─────────────────────────────────────────
    if ws_sint:
        def hv(key): return sint.get(key)

        # Criterios 1-5: presencia de valores de horas
        for n, key, label in [
            (1, 'hrs_pres',  'Horas presenciales'),
            (2, 'hrs_pres',  'Total horas presenciales'),
            (3, 'hrs_sinc',  'Horas virtuales sincrónicas'),
            (4, 'hrs_asinc', 'Horas virtuales asincrónicas'),
            (5, None,        'Diferenciación sincrónicas/asincrónicas'),
        ]:
            if key:
                v = hv(key)
                e = '✅' if v is not None else '❌'
                d = f'{label}: {v}'
            else:
                s, a = hv('hrs_sinc'), hv('hrs_asinc')
                e = '✅' if (s is not None and a is not None) else '❌'
                d = f'Sincrónicas={s} Asincrónicas={a}'
            check('Carga académica', n, ESCALA_ESTANDAR[idx][2], 'auto', e, d)
            idx += 1

        # Criterio 6 — manual
        check('Carga académica', 6, ESCALA_ESTANDAR[idx][2], 'manual',
              '⚠️ manual', 'Requiere revisión humana')
        idx += 1

        # Criterios 7-8: horas TPE
        tpe = hv('hrs_tpe')
        for n, label in [(7, 'Distribución horas TPE'), (8, 'Total horas TPE')]:
            e = '✅' if tpe is not None else '❌'
            check('Carga académica', n, ESCALA_ESTANDAR[idx][2], 'auto',
                  e, f'{label}: {tpe}')
            idx += 1

        # Criterio 9 — manual
        check('Carga académica', 9, ESCALA_ESTANDAR[idx][2], 'manual',
              '⚠️ manual', 'Requiere revisión humana')

    return resultados


def formatear_verificacion(resultados, fuente_escala):
    """
    Formatea los resultados de verificación para el reporte.
    fuente_escala: 'archivo' | 'estandar'
    """
    lineas = []
    lineas.append('\n  [Verificación — Escala Estándar UST (41 criterios — meta: 100%)]')
    if fuente_escala == 'estandar':
        lineas.append('  Fuente criterios: estándar institucional integrado (sin archivo de escala)')
    else:
        lineas.append('  Fuente criterios: estándar institucional integrado + archivo de escala disponible')

    seccion_actual = None
    total = len(resultados)
    aprobados = sum(1 for _, _, _, e, _ in resultados if e == '✅')
    manuales  = sum(1 for _, _, _, e, _ in resultados if 'manual' in str(e))
    fallidos  = total - aprobados - manuales

    for seccion, n, criterio, estado, detalle in resultados:
        if seccion != seccion_actual:
            lineas.append(f'\n    ── {seccion} ──')
            seccion_actual = seccion
        texto_criterio = criterio[:65] + '…' if len(criterio) > 65 else criterio
        lineas.append(f'    {estado} {n:02d}. {texto_criterio}')
        if estado != '✅':
            lineas.append(f'         → {detalle}')

    lineas.append(f'\n    Resultado: {aprobados}✅  {fallidos}❌  {manuales}⚠️ manual  '
                  f'(de {total} criterios)')
    return lineas


def _base_limpia(nombre):
    """
    Elimina todos los sufijos de instancia del nombre de archivo (iterativo).
    Quita cualquier combinación de: _I1, _I2, _I3, _REVISADO, _FINAL.
    """
    base, ext = os.path.splitext(nombre)
    patron = re.compile(r'(_I[123]|_REVISADO|_FINAL)+$', re.IGNORECASE)
    prev = None
    while prev != base:
        prev = base
        base = patron.sub('', base)
    return base, ext


def nombre_instancia(nombre_original, n):
    """
    Genera nombre limpio con sufijo _I1 / _I2 / _I3.
    Elimina cualquier sufijo de instancia previo antes de añadir el nuevo.
    Ejemplo: 'plan_I1_REVISADO_I2.xlsx' → 'plan_I3.xlsx'
    """
    base, ext = _base_limpia(nombre_original)
    return f'{base}_I{n}{ext}'


def nombre_revisado(nombre_original):
    """Alias de compatibilidad — genera nombre _I1."""
    return nombre_instancia(nombre_original, 1)


def solo_sumativa(texto):
    """Extrae únicamente el bloque SUMATIVA de un texto de procedimientos."""
    match = re.search(r'(SUMATIVA\s*\([^)]+\):.*)', texto, re.DOTALL)
    return match.group(1).strip() if match else texto


def tiene_bloques_extra(texto):
    """Devuelve True si el texto contiene DIAGNÓSTICA o FORMATIVA."""
    return 'DIAGNÓSTICA' in texto or 'FORMATIVA' in texto


def inferir_medio(modalidad, momento, tipo):
    """Infiere el Medio de entrega según modalidad, momento y tipo de evaluación."""
    if not tipo:
        return None
    m = str(momento or '').lower()
    mod = str(modalidad or '').lower()
    t = str(tipo or '').lower()
    if 'diagnós' in t and 'preparación' in m:
        return 'Cuestionario'
    if 'desarrollo' in m and ('sincrónico' in mod or 'presencial' in mod):
        return 'No aplica'
    if 'independiente' in m or 'no presencial' in mod:
        return 'Buzón de tareas'
    return 'No aplica'


def inferir_indiv_grupal(procedimiento, unidad_es_grupal):
    """Infiere Individual/Grupal según procedimiento y contexto de la unidad."""
    proc = str(procedimiento or '').lower()
    if 'equipo' in proc or 'grupo' in proc or 'colaborativo' in proc:
        return 'Grupal'
    if unidad_es_grupal:
        return 'Grupal'
    return 'Individual'


def detectar_unidades_grupales(ws):
    """
    Detecta qué unidades (col B) tienen al menos una sumativa con M='Grupal'.
    Devuelve un set de valores de col B.
    """
    grupales = set()
    for row in ws.iter_rows(min_row=4, values_only=False):
        tipo  = row[COL['TIPO']  - 1].value
        indiv = row[COL['INDIV'] - 1].value
        unidad = row[COL['UNIDAD'] - 1].value
        if tipo and 'sumativa' in str(tipo).lower() and indiv == 'Grupal' and unidad:
            grupales.add(unidad)
    return grupales


# ══════════════════════════════════════════════════════════════════════════
#  CORRECCIÓN SÍNTESIS DIDÁCTICA
# ══════════════════════════════════════════════════════════════════════════

def corregir_sintesis(ws, log, registro=None):
    """
    Correcciones en Síntesis didáctica:
    1. Elimina DIAGNÓSTICA y FORMATIVA de col D (procedimientos).
    2. Normaliza porcentajes en col G a número entero (30% / 0.30 → 30).
    3. Colorea filas de la tabla RA/Procedimientos:
       - Sumativa → amarillo pastel
       - Formativa / Diagnóstica → lila pastel
    """
    from openpyxl.cell.cell import MergedCell

    cambios = 0

    # ── Detectar rango de la tabla RA ─────────────────────────────────────
    # Buscar la fila del encabezado "Resultados de Aprendizaje" (col A)
    fila_header_ra = None
    fila_total_ra  = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if (cell.column == 1 and cell.value and
                    isinstance(cell.value, str) and
                    'resultado' in cell.value.lower() and
                    'aprendizaje' in cell.value.lower()):
                fila_header_ra = cell.row
            if (cell.column == 6 and cell.value and
                    isinstance(cell.value, str) and
                    cell.value.strip().upper() == 'TOTAL' and
                    fila_header_ra and cell.row > fila_header_ra):
                fila_total_ra = cell.row
                break
        if fila_total_ra:
            break

    COL_MAX_COLOR = 7   # A–G

    # ── 1 & 2 & 3: iterar filas de datos ─────────────────────────────────
    for row in ws.iter_rows(min_row=1, values_only=False):
        r = row[0].row

        for cell in row:
            # Corrección 1: eliminar DIAGNÓSTICA/FORMATIVA en col D
            if cell.column == 4 and cell.value and isinstance(cell.value, str):
                if tiene_bloques_extra(cell.value):
                    original = cell.value
                    nuevo = solo_sumativa(cell.value)
                    log.append(f'    [Síntesis F{r}] Eliminados bloques DIAGNÓSTICA/FORMATIVA')
                    cell.value = nuevo
                    aplicar_azul_diff(cell, original, nuevo)
                    _reg(registro, ws, cell,
                         'Síntesis: eliminar DIAGNÓSTICA/FORMATIVA de procedimientos',
                         original, nuevo)
                    cambios += 1

            # Corrección 2: normalizar porcentajes en col G a número entero
            if cell.column == 7 and cell.value is not None:
                val = cell.value
                num = None
                if isinstance(val, str):
                    val_s = val.strip().rstrip('%')
                    try:
                        num = float(val_s)
                        if num <= 1.0:   # era decimal 0.30 → 30
                            num = round(num * 100)
                        else:
                            num = round(num)
                    except ValueError:
                        pass
                elif isinstance(val, float):
                    if val <= 1.0:
                        num = round(val * 100)
                    else:
                        num = round(val)
                elif isinstance(val, int):
                    num = val

                if num is not None and cell.value != num:
                    original_pct = cell.value
                    cell.value = num
                    aplicar_azul_diff(cell, str(original_pct), str(num))
                    log.append(f'    [Síntesis F{r}] Porcentaje normalizado: {original_pct!r} → {num}')
                    cambios += 1

        # Corrección 3: colorear filas de la tabla RA
        if fila_header_ra and fila_total_ra:
            if fila_header_ra < r < fila_total_ra:
                proc_val = ws.cell(r, 4).value
                if proc_val and isinstance(proc_val, str):
                    pt = proc_val.lower()
                    if 'sumativ' in pt:
                        fill = _FILL_SUMATIVA
                    elif 'formativ' in pt or 'diagnóst' in pt or 'diagnost' in pt:
                        fill = _FILL_FORMATIVA
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, COL_MAX_COLOR + 1):
                            c = ws.cell(r, col_idx)
                            if not isinstance(c, MergedCell):
                                c.fill = fill

    return cambios


# ══════════════════════════════════════════════════════════════════════════
#  CORRECCIÓN PLANIFICACIÓN POR UNIDADES
# ══════════════════════════════════════════════════════════════════════════

def corregir_planificacion(ws, log, registro=None):
    """Aplica todas las correcciones a la hoja Planificación por unidades."""
    cambios_alta = 0
    cambios_media_j = 0
    cambios_media_m = 0
    cambios_baja = 0

    # Pre-scan: detectar unidades grupales (para inferir M)
    unidades_grupales = detectar_unidades_grupales(ws)

    # Rastrear unidad actual (col B puede estar vacía por merge)
    unidad_actual = None

    for row in ws.iter_rows(min_row=4, values_only=False):
        r = row[0].row

        # Leer valores clave
        unidad_cell   = ws.cell(r, COL['UNIDAD'])
        modalidad     = ws.cell(r, COL['MODALIDAD']).value
        momento       = ws.cell(r, COL['MOMENTO']).value
        tipo_cell     = ws.cell(r, COL['TIPO'])
        proc_cell     = ws.cell(r, COL['PROC'])
        indiv_cell    = ws.cell(r, COL['INDIV'])
        instr_cell    = ws.cell(r, COL['INSTR'])
        pct_cell      = ws.cell(r, COL['PCT'])
        medio_cell    = ws.cell(r, COL['MEDIO'])

        if unidad_cell.value:
            unidad_actual = unidad_cell.value

        tipo  = tipo_cell.value
        proc  = proc_cell.value
        instr = instr_cell.value

        tiene_eval = bool(tipo)

        # ── ALTA: datos huérfanos (proc/instr sin tipo) ────────────────────
        if not tipo and (proc or instr):
            actividad = ws.cell(r, COL['ACTIVIDAD']).value
            if not actividad:
                # Fila sin actividad ni tipo: limpiar proc/instr huérfanos
                if proc:
                    _reg(registro, ws, proc_cell, 'Dato huérfano: procedimiento sin tipo', proc, None)
                    proc_cell.value = None
                if instr:
                    _reg(registro, ws, instr_cell, 'Dato huérfano: instrumento sin tipo', instr, None)
                    instr_cell.value = None
                log.append(f'    [Plan F{r}] Datos huérfanos eliminados (L y N sin tipo evaluación)')
                cambios_alta += 1
            # En ambos casos, si no hay tipo no hay correcciones evaluativas que aplicar
            continue

        if not tiene_eval:
            continue

        # ── ALTA: normalizar Tipo ──────────────────────────────────────────
        if tipo in TIPO_MAP:
            nuevo_tipo = TIPO_MAP[tipo]
            log.append(f'    [Plan F{r}] Tipo: "{tipo}" → "{nuevo_tipo}"')
            tipo_cell.value = nuevo_tipo
            aplicar_azul(tipo_cell)
            _reg(registro, ws, tipo_cell, 'Normalizar tipo de evaluación', tipo, nuevo_tipo)
            tipo = nuevo_tipo
            cambios_alta += 1

        # ── ALTA: normalizar Procedimiento ────────────────────────────────
        # Excepción: Cuestionario es procedimiento válido para evaluación Diagnóstica
        # (y puede usarse en Formativas internas). Solo corregir en Sumativas.
        # Fuente: Manual Diseño Instruccional UST, p. 18
        _tipo_lower = str(tipo or '').lower()
        _es_diagnostica = 'diagnóst' in _tipo_lower
        _skip_cuestionario = (proc == 'Cuestionario' and _es_diagnostica)

        if proc in PROC_MAP and not _skip_cuestionario:
            nuevo_proc = PROC_MAP[proc]
            log.append(f'    [Plan F{r}] Procedimiento: "{proc}" → "{nuevo_proc}"')
            proc_cell.value = nuevo_proc
            aplicar_azul(proc_cell)
            _reg(registro, ws, proc_cell, 'Normalizar procedimiento de evaluación', proc, nuevo_proc)
            proc = nuevo_proc
            cambios_alta += 1
        elif _skip_cuestionario:
            log.append(f'    [Plan F{r}] Procedimiento "Cuestionario" conservado (Diagnóstica — válido)')

        # ── ALTA: normalizar Instrumento ──────────────────────────────────
        if instr == 'Pauta':
            log.append(f'    [Plan F{r}] Instrumento: "Pauta" → "Pauta de observación"')
            instr_cell.value = 'Pauta de observación'
            aplicar_azul(instr_cell)
            _reg(registro, ws, instr_cell, 'Normalizar instrumento de evaluación', instr, 'Pauta de observación')
            instr = 'Pauta de observación'
            cambios_alta += 1
        elif instr == 'Rúbrica' and proc in PROC_INSTR_RÚBRICA_CONTEXTOS:
            log.append(f'    [Plan F{r}] Instrumento: "Rúbrica" → "Pauta de observación" (proc={proc})')
            instr_cell.value = 'Pauta de observación'
            aplicar_azul(instr_cell)
            _reg(registro, ws, instr_cell, 'Normalizar instrumento (Rúbrica→Pauta en Producciones)', instr, 'Pauta de observación')
            instr = 'Pauta de observación'
            cambios_alta += 1

        # ── MEDIA: Medio de entrega (J) ────────────────────────────────────
        if medio_cell.value is None:
            nuevo_medio = inferir_medio(modalidad, momento, tipo)
            if nuevo_medio:
                medio_cell.value = nuevo_medio
                aplicar_azul(medio_cell)
                log.append(f'    [Plan F{r}] Medio de entrega: None → "{nuevo_medio}"')
                _reg(registro, ws, medio_cell, 'Inferir medio de entrega', None, nuevo_medio)
                cambios_media_j += 1

        # ── MEDIA: Individual/Grupal (M) ───────────────────────────────────
        if indiv_cell.value is None:
            es_grupal = unidad_actual in unidades_grupales
            nuevo_indiv = inferir_indiv_grupal(proc, es_grupal)
            indiv_cell.value = nuevo_indiv
            aplicar_azul(indiv_cell)
            log.append(f'    [Plan F{r}] Individual/Grupal: None → "{nuevo_indiv}"')
            _reg(registro, ws, indiv_cell, 'Inferir modalidad individual/grupal', None, nuevo_indiv)
            cambios_media_m += 1

        # ── BAJA: % = None → 0 ────────────────────────────────────────────
        if pct_cell.value is None:
            pct_cell.value = 0
            aplicar_azul(pct_cell)
            log.append(f'    [Plan F{r}] %: None → 0')
            _reg(registro, ws, pct_cell, '% vacío → 0', None, 0)
            cambios_baja += 1

    # ── Colorear filas por tipo de evaluación (al final, sin pisar fuente azul)
    aplicar_colores_evaluacion(ws)

    return cambios_alta, cambios_media_j, cambios_media_m, cambios_baja


# ══════════════════════════════════════════════════════════════════════════
#  CORRECCIÓN DE LENGUAJE EN ACTIVIDADES
# ══════════════════════════════════════════════════════════════════════════

# Verbos en 3ª persona plural al inicio de ítems → 2ª persona singular (imperativo)
_VERBOS_PLURAL_SINGULAR = {
    'presenten': 'presenta',      'elaboren': 'elabora',       'construyan': 'construye',
    'identifiquen': 'identifica', 'analicen': 'analiza',       'revisen': 'revisa',
    'discutan': 'discute',        'compartan': 'comparte',     'entreguen': 'entrega',
    'suban': 'sube',              'respondan': 'responde',     'completen': 'completa',
    'diseñen': 'diseña',          'apliquen': 'aplica',        'realicen': 'realiza',
    'observen': 'observa',        'seleccionen': 'selecciona', 'documenten': 'documenta',
    'redacten': 'redacta',        'propongan': 'propone',      'formulen': 'formula',
    'calculen': 'calcula',        'interpreten': 'interpreta', 'reflexionen': 'reflexiona',
    'investiguen': 'investiga',   'describan': 'describe',     'expliquen': 'explica',
}

# Ítems (numerados O sin número) cuyo sujeto es el o la docente
# (?:\d+\.\s+)? hace el número inicial OPCIONAL → funciona con y sin numeración
_PAT_ACCION_DOCENTE = re.compile(
    r'^(?:\d+\.\s+)?(?:la docente|el o la docente|el docente|la\/el docente)\s+\w',
    re.IGNORECASE
)

# Ítems donde el estudiante recibe pasivamente (docente es agente activo)
# Funciona con y sin numeración: "Recibe retroalimentación del o la docente…"
_PAT_RECEPCION_PASIVA = re.compile(
    r'^(?:\d+\.\s+)?(?:recibe|observa la|escucha la)\s+'
    r'.{0,60}(?:del\s+o\s+la\s+docente|del\s+docente|de\s+la\s+docente)',
    re.IGNORECASE
)

# Ítems (numerados O sin número) que son frases nominales sin verbo imperativo
_PAT_FRASE_NOMINAL = re.compile(
    r'^(?:\d+\.\s+)?'
    r'(puesta en|revisión de|análisis de|presentación de|trabajo en|'
    r'discusión sobre|reflexión sobre|síntesis de|manejo de|introducción a|'
    r'bienvenida|presentación del|presentación de la|exposición del|exposición de la|'
    r'pódcast|podcast|video|vídeo|lectura|actividad de|taller de|evaluación de)',
    re.IGNORECASE
)

# Notas editoriales entre paréntesis destinadas al docente (no al estudiante)
_PAT_NOTA_EDITORIAL = re.compile(
    r'\([^)]{5,120}(?:'
    r'mejorar|incorporar|agregar|pendiente|revisar la|revisar el|'
    r'modificar|instruccion|instrucción|presentacion docente|'
    r'ver con|falta|corregir esto|ajustar'
    r')[^)]{0,80}\)',
    re.IGNORECASE
)

# Línea de TÍTULO del momento: solo letras, tildes, ñ y espacios (sin puntuación)
# Ejemplos: "Familiarízate con la asignatura…", "Activa tus conocimientos"
_PAT_TITULO_MOMENTO = re.compile(
    r'^[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñü\s,]+$'
)

# Verbos imperativos 2ª persona válidos para iniciar ítems
# ── Stopwords para extracción de palabras clave del Propósito ────────────
_STOPWORDS_COHERENCIA = {
    'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas', 'de', 'del', 'al',
    'en', 'con', 'por', 'para', 'que', 'se', 'su', 'sus', 'y', 'o', 'a', 'e',
    'es', 'son', 'ser', 'sobre', 'como', 'este', 'esta', 'estos', 'estas',
    'cada', 'todo', 'toda', 'todos', 'todas', 'través', 'partir', 'hacia',
    'desde', 'durante', 'según', 'entre', 'sin', 'más', 'menos', 'muy',
    'bien', 'manera', 'forma', 'proceso', 'nivel', 'tipo', 'caso', 'casos',
    'parte', 'través', 'también', 'además', 'través', 'través',
    # verbos copulativos y auxiliares
    'está', 'están', 'fue', 'fueron', 'han', 'hay', 'tiene', 'tienen',
    # preposiciones compuestas
    'través', 'través', 'través',
}

# ── Frases genéricas sin especificidad de contenido ──────────────────────
_FRASES_GENERICAS = re.compile(
    r'^(?:\d+\.\s+)?(?:'
    r'participa\s+activamente|presta\s+atención|está\s+atento|'
    r'mantiene\s+(?:el\s+)?orden|escucha\s+con\s+atención|'
    r'sigue\s+las\s+instrucciones|realiza\s+la\s+actividad|'
    r'completa\s+la\s+actividad|desarrolla\s+la\s+actividad|'
    r'toma\s+apuntes|anota\s+lo\s+importante|copia\s+(?:los\s+)?contenidos|'
    r'lee\s+el\s+texto\s*$|observa\s+el\s+video\s*$|'
    r'participa\s+en\s+la\s+actividad|trabaja\s+en\s+la\s+actividad'
    r')',
    re.IGNORECASE
)

# ── Verbos/frases que indican entrega o producción concreta ──────────────
_VERBOS_ENTREGA = {
    'entrega', 'sube', 'publica', 'envía', 'envia', 'presenta',
    'elabora', 'redacta', 'escribe', 'construye', 'diseña', 'diseña',
}

# ── Verbos que no corresponden a Preparación (activación de saberes) ─────
_VERBOS_INCOMPATIBLES_PREP = re.compile(
    r'\b(?:entrega|sube al foro|sube la tarea|publica|envía su|'
    r'presenta su informe|rinde|da la prueba)\b',
    re.IGNORECASE
)


def _analizar_coherencia_bloque(titulo, proposito_texto, items, momento_col,
                                 fila, advertencias):
    """
    Verifica coherencia semántica y calidad de redacción dentro de un bloque
    de actividades (Preparación, Desarrollo o Trabajo Independiente).

    Detecta:
    1. Baja alineación entre Propósito e ítems (palabras clave no aparecen)
    2. Ítems genéricos sin especificidad de contenido
    3. Ítems demasiado breves (< 4 palabras reales)
    4. Verbos iniciales que se repiten ≥ 3 veces (monotonía cognitiva)
    5. Ítems de entrega/producción en Preparación (incoherencia de momento)
    6. Trabajo Independiente sin ninguna actividad de producción/entrega
    """
    momento_lower = str(momento_col or '').lower()
    tag = f'[Plan F{fila} {str(momento_col or "")[:12]}]'

    if not items:
        return

    # ── 1. Alineación Propósito ↔ ítems ──────────────────────────────────
    if proposito_texto:
        palabras_prop = {
            w.lower().strip('.,;:()¿?¡!"')
            for w in proposito_texto.split()
            if len(w) > 4 and w.lower().strip('.,;:()') not in _STOPWORDS_COHERENCIA
        }
        texto_items = ' '.join(items).lower()
        encontradas = sum(1 for w in palabras_prop if w in texto_items)
        total_kw = len(palabras_prop)

        if total_kw >= 3 and encontradas == 0:
            advertencias.append(
                f'    {tag} ⚠️  Incoherencia Propósito↔ítems: el Propósito declara '
                f'"{proposito_texto[:80]}" pero ningún ítem aborda esas palabras clave '
                f'— revisar alineación entre propósito y actividades'
            )
        elif total_kw >= 4 and encontradas <= 1:
            advertencias.append(
                f'    {tag} ⚠️  Débil alineación Propósito↔ítems: solo {encontradas}/'
                f'{total_kw} palabras clave del Propósito aparecen en las actividades '
                f'("{proposito_texto[:60]}") — revisar coherencia'
            )

    # ── 2. Ítems genéricos ────────────────────────────────────────────────
    for item in items:
        if _FRASES_GENERICAS.match(item):
            advertencias.append(
                f'    {tag} ⚠️  Ítem genérico sin especificidad de contenido '
                f'(indicar qué hace el estudiante y sobre qué tema/material): '
                f'"{item[:90]}"'
            )

    # ── 3. Ítems demasiado breves ─────────────────────────────────────────
    for item in items:
        # Descontar el número si lo tiene
        texto_item = re.sub(r'^\d+\.\s*', '', item).strip()
        palabras = [w for w in texto_item.split() if w]
        if 0 < len(palabras) < 4:
            advertencias.append(
                f'    {tag} ⚠️  Ítem muy breve ({len(palabras)} palabras) — '
                f'especificar la actividad completa con objeto y contexto: '
                f'"{item[:90]}"'
            )

    # ── 4. Verbos repetidos (monotonía cognitiva) ─────────────────────────
    verbos_inicio = []
    for item in items:
        m = re.match(r'^(?:\d+\.\s+)?([a-záéíóúüñ]+)', item.lower())
        if m:
            verbos_inicio.append(m.group(1))
    if len(verbos_inicio) >= 3:
        conteo = Counter(verbos_inicio)
        for verbo, cnt in conteo.items():
            if cnt >= 3:
                advertencias.append(
                    f'    {tag} ⚠️  Verbo "{verbo}" se repite {cnt} veces en el bloque '
                    f'— variar los verbos para mostrar progresión cognitiva '
                    f'(Bloom: recordar → comprender → aplicar → analizar…)'
                )

    # ── 5. Entrega/producción en Preparación (incoherente) ───────────────
    if 'preparac' in momento_lower:
        texto_items_lower = ' '.join(items).lower()
        if _VERBOS_INCOMPATIBLES_PREP.search(texto_items_lower):
            advertencias.append(
                f'    {tag} ⚠️  Ítem de entrega o producción en Preparación — '
                f'las entregas deben ir en Desarrollo o Trabajo Independiente '
                f'(Manual UST p.18)'
            )

    # ── 6. Trabajo Independiente sin producción concreta ─────────────────
    if 'independ' in momento_lower:
        texto_items_lower = ' '.join(items).lower()
        tiene_produccion = any(v in texto_items_lower for v in _VERBOS_ENTREGA)
        if not tiene_produccion:
            advertencias.append(
                f'    {tag} ⚠️  Trabajo Independiente sin actividad de entrega/producción '
                f'— debe incluir al menos una acción concreta del estudiante '
                f'(elabora, redacta, sube, publica…) (Manual UST p.20)'
            )


# Fuente base + Manual Diseño Instruccional UST (Res. N°481/24), Tabla 1 (p.19)
_VERBOS_IMPERATIVO = {
    # A
    'accede','activa','adapta','agrega','amplía','analiza','anota','aplica',
    'argumenta','asiste','atiende','ayuda',
    # B-C
    'busca','calcula','categoriza','clasifica','colabora','comenta','compara',
    'comparte','completa','comprende','conecta','construye','consulta','contesta',
    'contrasta','contribuye','coopera','corrige','crea','cuestiona',
    # D-E
    'debate','deduce','define','describe','desarrolla','destaca','determina',
    'diferencia','diseña','documenta','elabora','emplea','enumera','escribe',
    'escucha','esquematiza','evalúa','examina','explica','explora','expone',
    # F-I
    'familiarízate','formula','grafica','identifica','implementa','infiere',
    'ingresa','interpreta','investiga',
    # J-M
    'justifica','lee','lista','localiza','mapea','marca','mide','moviliza',
    # N-P
    'nota','observa','optimiza','organiza','participa','planifica','practica',
    'prepara','presenta','prioriza','profundiza','propone','publica',
    # R
    'realiza','recibe','recuerda','recoge','redacta','reflexiona','registra',
    'relaciona','repasa','responde','retoma','revisa','resume',
    # S-Z
    'selecciona','señala','sintetiza','sistematiza','socializa','soluciona',
    'subraya','sube','tabula','toma','trabaja','ubica','usa','utiliza',
    'valida','valora','verifica','visualiza',
}

# ── Inicios incorrectos: preposición, locución o gerundio en lugar de imperativo ──
_PAT_INICIO_NO_IMPERATIVO = re.compile(
    r'^(?:\d+\.\s+)?'
    r'(?:'
    # Locuciones prepositivas
    r'a\s+partir\s+de|en\s+base\s+a|con\s+base\s+en|'
    r'a\s+trav[eé]s\s+de|en\s+funci[oó]n\s+de|'
    r'teniendo\s+en\s+cuenta|tomando\s+en\s+cuenta|'
    r'considerando\s+(?:que\s+)?|teniendo\s+presente|'
    r'de\s+acuerdo\s+(?:a|con)|'
    # Frases de modalidad grupal que preceden al verbo
    r'en\s+equipo[s]?\s*,|en\s+grupo[s]?\s*,|'
    r'de\s+manera\s+(?:grupal|individual|colaborativa)\s*,|'
    r'de\s+forma\s+(?:grupal|individual|colaborativa)\s*,|'
    r'junto\s+(?:a|con)\s+tu[s]?\s+\w+\s*,|'
    # Conectores textuales
    r'luego\s+de|despu[eé]s\s+de|una\s+vez\s+que|'
    r'con\s+el\s+fin\s+de|con\s+el\s+objetivo\s+de|'
    # Infinitivos (terminaciones -ar -er -ir seguidas de espacio o fin)
    r'[a-záéíóúüñ]{3,}(?:ar|er|ir)\s|'
    # 3ª persona singular/plural
    r'el\s+estudiante|los\s+estudiantes|'
    r'el\s+alumno|la\s+alumna|los\s+alumnos|'
    # Sustantivo abstracto sin verbo
    r'reflexi[oó]n\s+sobre|an[aá]lisis\s+de|revisi[oó]n\s+de|'
    r'lectura\s+de|elaboraci[oó]n\s+de|redacci[oó]n\s+de'
    r')',
    re.IGNORECASE,
)

# Patrón para reordenar "En equipos, verbo..." → "Verbo... en equipos..."
_PAT_MODALIDAD_PREFIJA = re.compile(
    r'^(\d+\.\s+)?'
    r'(en\s+equipo[s]?|en\s+grupo[s]?|'
    r'de\s+manera\s+(?:grupal|individual|colaborativa)|'
    r'de\s+forma\s+(?:grupal|individual|colaborativa))'
    r'\s*,\s*(.+)',
    re.IGNORECASE,
)


def verificar_imperativo_momentos(ws_plan, log):
    """
    Verifica que CADA instrucción de actividad en los 3 momentos de aprendizaje
    comience con un verbo imperativo en segunda persona singular (tú).

    Aplica a: Preparación, Desarrollo, Trabajo Independiente.
    Revisa tanto ítems numerados como consignas de línea única.

    Devuelve (n_ok, n_advertencias).
    """
    if not ws_plan:
        return 0, 0

    ok  = 0
    adv = []

    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        momento   = str(row[COL['MOMENTO']   - 1] or '').strip()
        actividad = str(row[COL['ACTIVIDAD'] - 1] or '').strip()
        unidad    = str(row[COL['UNIDAD']    - 1] or '').strip()
        semana    = str(row[COL['SEMANA']    - 1] or '').strip()

        if not momento or not actividad or len(actividad) < 10:
            continue
        m = momento.lower()
        if not any(k in m for k in ('preparaci', 'desarrollo', 'independiente')):
            continue

        ref = f'{momento[:20]} {("U"+unidad) if unidad else ""}{(" S"+semana) if semana else ""}'.strip()

        # Obtener todas las líneas de instrucción (excluir título y Propósito)
        lineas = [l.strip() for l in actividad.split('\n') if l.strip()]
        instrucciones = []
        for l in lineas:
            es_titulo    = bool(_PAT_TITULO_MOMENTO.match(l)) and len(l) < 60
            es_proposito = bool(re.match(r'Prop[oó]sito\s*:', l, re.IGNORECASE))
            if not es_titulo and not es_proposito:
                instrucciones.append(l)

        if not instrucciones:
            continue

        problemas_bloque = []
        for linea in instrucciones:
            # Extraer primera palabra (sin número de ítem)
            m_ini = re.match(r'^(?:\d+\.\s+)?([A-ZÁÉÍÓÚa-záéíóúüñÑ][a-záéíóúüñÑ]*)', linea)
            if not m_ini:
                continue

            primer_palabra = m_ini.group(1).lower()

            # Caso 1: inicio con locución/preposición/infinitivo/3ª persona
            if _PAT_INICIO_NO_IMPERATIVO.match(linea):
                problemas_bloque.append(
                    f'"{linea[:80]}" — '
                    f'comienza con "{m_ini.group(1)}", no con verbo imperativo '
                    f'(usar: Redacta, Analiza, Reflexiona…)'
                )

            # Caso 2: primera palabra no está en el listado de imperativos
            elif primer_palabra not in _VERBOS_IMPERATIVO:
                # Ignorar líneas muy cortas que sean títulos no detectados
                if len(linea) < 8:
                    continue
                # Ignorar si es acción del docente (ya reportada por corregir_lenguaje_actividades)
                if _PAT_ACCION_DOCENTE.match(linea):
                    continue
                problemas_bloque.append(
                    f'"{linea[:80]}" — '
                    f'comienza con "{m_ini.group(1)}", no reconocido como imperativo 2ª persona '
                    f'(verificar: ¿es infinitivo, 3ª persona o locución?)'
                )

        if problemas_bloque:
            adv.append(f'    ⚠️  {ref}:')
            for p in problemas_bloque:
                adv.append(f'        → {p}')
        else:
            ok += 1

    log.append('\n  [Verificación: verbo imperativo 2ª persona en actividades]')
    if adv:
        for a in adv:
            log.append(a)
        log.append(
            f'    Referencia: Manual UST — las instrucciones al estudiante '
            f'deben iniciar con verbo imperativo (Analiza, Redacta, Reflexiona…)'
        )
    elif ok > 0:
        log.append(f'    ✅ {ok} bloque(s) con instrucciones correctamente en imperativo 2ª persona')
    else:
        log.append('    —  No se encontraron actividades para verificar')

    n_adv = len([a for a in adv if a.strip().startswith('⚠️')])
    return ok, n_adv


# Imperativas con acento incorrecto → forma correcta
# Las formas agudas de tipo "clasifíca" son incorrecto — el imperativo es grave: "clasifica"
_ACENTOS_IMPERATIVO = {
    'clasifíca': 'clasifica', 'identifíca': 'identifica',
    'analíza':   'analiza',   'organíza':   'organiza',
    'sintetíza': 'sintetiza', 'especifíca': 'especifica',
    'calífca':   'califica',  'justifíca':  'justifica',
    'diferéncia':'diferencia','categoríza': 'categoriza',
    'sistematíza':'sistematiza','priorizá': 'prioriza',
    'visualíza': 'visualiza', 'resumé':     'resume',
}

# ── Correcciones ortográficas generales (errores frecuentes en planificaciones)
# Fuente: revisión de planificaciones DEL UST 2025-2 / 2026-1
# Solo corregibles automáticamente con alta confianza (1:1 sustitución)
_ORTOGRAFIA_GENERAL = {
    # Tildes faltantes — adverbios y conjunciones
    'asi':           'así',
    'tambien':       'también',
    'ademas':        'además',
    'segun':         'según',
    'traves':        'través',
    'a traves':      'a través',
    'razon':         'razón',
    'mas ':          'más ',        # adverbio "más" (no "pero")
    'aun ':          'aún ',        # "aún" = todavía
    # Tildes faltantes — sustantivos terminados en -ción (muy frecuentes)
    'informacion':   'información',
    'evaluacion':    'evaluación',
    'presentacion':  'presentación',
    'participacion': 'participación',
    'elaboracion':   'elaboración',
    'aplicacion':    'aplicación',
    'utilizacion':   'utilización',
    'produccion':    'producción',
    'realizacion':   'realización',
    'orientacion':   'orientación',
    'comprension':   'comprensión',
    'reflexion':     'reflexión',
    'integracion':   'integración',
    'colaboracion':  'colaboración',
    'comunicacion':  'comunicación',
    'autoevaluacion':'autoevaluación',
    'sistematizacion':'sistematización',
    'categorizacion':'categorización',
    'planificacion': 'planificación',
    'instruccion':   'instrucción',
    'introduccion':  'introducción',
    'revision':      'revisión',
    'conclusion':    'conclusión',
    'redaccion':     'redacción',
    'descripcion':   'descripción',
    'definicion':    'definición',
    'preparacion':   'preparación',
    'construccion':  'construcción',
    'interaccion':   'interacción',
    'transicion':    'transición',
    'distribucion':  'distribución',
    'investigacion': 'investigación',
    'implementacion':'implementación',
    'retroalimentacion':'retroalimentación',
    'retroalimentacion':'retroalimentación',
    'observacion':   'observación',
    'resolucion':    'resolución',
    'solucion':      'solución',
    'situacion':     'situación',
    'intervencion':  'intervención',
    'exposicion':    'exposición',
    'composicion':   'composición',
    'relacion':      'relación',
    'comparacion':   'comparación',
    'seleccion':     'selección',
    'publicacion':   'publicación',
    'administracion':'administración',
    # Tildes faltantes — otras palabras comunes
    'diagnostico':   'diagnóstico',
    'metodologia':   'metodología',
    'tecnologia':    'tecnología',
    'pedagogia':     'pedagogía',
    'didactica':     'didáctica',
    'sistematico':   'sistemático',
    'practico':      'práctico',
    'teorico':       'teórico',
    'critico':       'crítico',
    'especifico':    'específico',
    'academico':     'académico',
    'basico':        'básico',
    'etico':         'ético',
    'logico':        'lógico',
    'dinamico':      'dinámico',
    'autonomia':     'autonomía',
    'competencia':   'competencia',   # no requiere tilde, confirma
    # Errores de transposición frecuentes
    'asigantura':    'asignatura',
    'evalaucion':    'evaluación',
    'aprendizaje':   'aprendizaje',   # no requiere tilde, confirma
    'activdad':      'actividad',
    'resolucion':    'resolución',
    'asignatrua':    'asignatura',
    # Tildes faltantes — palabras académicas adicionales
    'analisis':      'análisis',
    'sintesis':      'síntesis',
    'hipotesis':     'hipótesis',
    'enfasis':       'énfasis',
    'parentesis':    'paréntesis',
    'tesis':         'tesis',          # no requiere tilde, confirma
    'areas':         'áreas',
    'area':          'área',
    'indice':        'índice',
    'ambito':        'ámbito',
    'caracter':      'carácter',
    'modelo':        'modelo',         # no requiere tilde, confirma
    'numero':        'número',
    'calculo':       'cálculo',
    'grafico':       'gráfico',
    'simbolo':       'símbolo',
    'exito':         'éxito',
    'metodo':        'método',
    'proceso':       'proceso',        # no requiere tilde
    'objetivo':      'objetivo',       # no requiere tilde
    'diagnostico':   'diagnóstico',
    'pronostico':    'pronóstico',
    'indice':        'índice',
    'practicas':     'prácticas',
    'practica':      'práctica',
    'tecnica':       'técnica',
    'tecnicas':      'técnicas',
    'estatica':      'estática',
    'dinamica':      'dinámica',
    'economico':     'económico',
    'economica':     'económica',
    'historico':     'histórico',
    'historica':     'histórica',
    'geografico':    'geográfico',
    'bibliografico': 'bibliográfico',
    'bibliografica': 'bibliográfica',
    'matematico':    'matemático',
    'matematica':    'matemática',
    'fisico':        'físico',
    'quimico':       'químico',
    'biologico':     'biológico',
    'ecologico':     'ecológico',
    'ecologia':      'ecología',
    'taxonomia':     'taxonomía',
    'filosofia':     'filosofía',
    'sociologia':    'sociología',
    'psicologia':    'psicología',
    'antropologia':  'antropología',
    # Confusiones ortográficas — no corregir automáticamente
    'haber':         None,   # depende de contexto (hay / haber)
    'a ver':         None,   # podría ser correcto
    # Palabras con b/v
    'havlar':        'hablar',
    'havitual':      'habitual',
    'tambien':       'también',
    'tambien ':      'también ',
    'ademas':        'además',
    'segun':         'según',
    'despues':       'después',
    'antes ':        'antes ',         # no requiere tilde, confirma
    'solo ':         None,             # "solo" (adverbio) sin tilde es válido modernamente
    'aun ':          'aún ',           # ya definido arriba (duplicado eliminado implícito)
}


def corregir_lenguaje_actividades(ws, log, registro=None):
    """
    Corrige incoherencias de lenguaje en la columna Actividad (G):
      - "la docente" / "el docente" → "el o la docente" (lenguaje inclusivo UST)
      - Espacio faltante tras punto numerado: "2.Verbo" → "2. Verbo"
      - Verbos 3ª persona plural → 2ª persona singular (imperativo)
      - Correcciones ortográficas de alta confianza
    Además registra advertencias para revisión manual:
      - Ítems que describen acción del docente (no instrucción al estudiante)
      - Ítems que son frases nominales sin verbo imperativo
      - Bloques con ítems numerados sin "Propósito:" (Manual UST p.19)
    Devuelve el número de celdas modificadas.
    """
    celdas_modificadas = 0
    advertencias = []

    for row in ws.iter_rows(min_row=4, values_only=False):
        r = row[0].row
        momento = ws.cell(r, COL['MOMENTO']).value
        if not momento:
            continue  # filas sin momento (p.ej. filas de totales al final)

        cell = ws.cell(r, COL['ACTIVIDAD'])
        texto = cell.value
        if not isinstance(texto, str):
            continue

        texto_nuevo = texto
        cambios_celda = []

        # ── 1. Espacio faltante: "2.Verbo" → "2. Verbo" ──────────────────
        t2 = re.sub(r'(\d+)\.([^\s\d\n])', r'\1. \2', texto_nuevo)
        if t2 != texto_nuevo:
            cambios_celda.append('espacio faltante tras punto numerado')
            texto_nuevo = t2

        # ── 2a. "de/del el/la (o el/la)* docente" → "del o la docente" ─────
        # Cubre: "de el docente", "de la docente", "de el o la docente",
        #        "de el o el o la docente", "del o el o la docente", etc.
        t2 = re.sub(
            r'\bde(?:l)?\s+(?:el|la)(?:\s+o\s+(?:el|la))*\s+docente\b',
            'del o la docente',
            texto_nuevo, flags=re.IGNORECASE
        )
        if t2 != texto_nuevo:
            cambios_celda.append('"de el o… docente" → "del o la docente"')
            texto_nuevo = t2

        # ── 2b. "el/la docente" sueltos (no precedidos por "o ") ──────────
        # Lookbehind fijo: evita romper "el o la docente" ya corregido
        t2 = re.sub(r'(?<!o )\bel docente\b', 'el o la docente', texto_nuevo, flags=re.IGNORECASE)
        t2 = re.sub(r'(?<!o )la docente\b',   'el o la docente', t2,          flags=re.IGNORECASE)
        if t2 != texto_nuevo:
            cambios_celda.append('"la/el docente" → "el o la docente"')
            texto_nuevo = t2

        # ── 3. Verbos 3ª plural → 2ª singular al inicio de ítems ─────────
        def _plural_a_singular(m):
            verbo = m.group(2).lower()
            if verbo in _VERBOS_PLURAL_SINGULAR:
                singular = _VERBOS_PLURAL_SINGULAR[verbo]
                singular = singular[0].upper() + singular[1:]
                cambios_celda.append(f'"{m.group(2)}" → "{singular}" (3ª plural → 2ª singular)')
                return f'{m.group(1)}. {singular}'
            return m.group(0)

        t2 = re.sub(r'(\d+)\.\s+([A-ZÁÉÍÓÚ][a-záéíóúü]+)\b', _plural_a_singular, texto_nuevo)
        if t2 != texto_nuevo:
            texto_nuevo = t2

        # ── 3a2. Reordenar "En equipos, verbo..." → "Verbo... en equipos" ──
        def _reordenar_modalidad(linea: str) -> str:
            m = _PAT_MODALIDAD_PREFIJA.match(linea)
            if not m:
                return linea
            num_part  = m.group(1) or ''   # "2. " o ""
            modalidad = m.group(2).lower() # "en equipos"
            resto     = m.group(3).strip() # "analiza un caso..."
            if not resto:
                return linea
            # Capitalizar primera letra del resto
            resto_cap = resto[0].upper() + resto[1:]
            cambios_celda.append(
                f'reordenar inicio: "{num_part}{modalidad}, {resto[:40]}…"'
                f' → "{num_part}{resto_cap}, {modalidad}"'
            )
            return f'{num_part}{resto_cap}, {modalidad}'

        lineas_orig2 = texto_nuevo.split('\n')
        lineas_new2  = [_reordenar_modalidad(l) for l in lineas_orig2]
        t2 = '\n'.join(lineas_new2)
        if t2 != texto_nuevo:
            texto_nuevo = t2

        # ── 3b. Imperativas con acento incorrecto (clasifíca → clasifica) ──
        for mal, bien in _ACENTOS_IMPERATIVO.items():
            patron_acento = re.compile(r'\b' + re.escape(mal) + r'\b', re.IGNORECASE)
            t2 = patron_acento.sub(bien, texto_nuevo)
            if t2 != texto_nuevo:
                cambios_celda.append(f'acento incorrecto: "{mal}" → "{bien}"')
                texto_nuevo = t2

        # ── 3c. Correcciones ortográficas generales ───────────────────────
        # Solo aplica correcciones con alta confianza (valor no-None en el dict)
        for mal, bien in _ORTOGRAFIA_GENERAL.items():
            if bien is None:
                continue   # contexto-dependiente, no corregir
            # Buscar palabra completa (con posibles tildes perdidas)
            pat = re.compile(r'\b' + re.escape(mal) + r'\b', re.IGNORECASE)
            t2 = pat.sub(bien, texto_nuevo)
            if t2 != texto_nuevo:
                cambios_celda.append(f'ortografía: "{mal}" → "{bien}"')
                texto_nuevo = t2

        # ── Aplicar cambios si los hubo ───────────────────────────────────
        if texto_nuevo != texto:
            cell.value = texto_nuevo
            aplicar_azul_diff(cell, texto, texto_nuevo)  # azul solo en líneas modificadas
            resumen_cambios = '; '.join(cambios_celda)
            for c in cambios_celda:
                log.append(f'    [Plan F{r}] Lenguaje actividad: {c}')
            _reg(registro, ws, cell,
                 f'Lenguaje/ortografía actividad ({resumen_cambios[:80]})',
                 texto, texto_nuevo)
            celdas_modificadas += 1

        # ── 4. Detección de problemas (numerados Y sin número) ───────────────
        # Divide el bloque en líneas y determina cuál es el título del momento
        # (primera línea no vacía solo con letras y espacios).
        # Todos los checks usan `if` independientes para reportar múltiples issues.
        lineas_bloque = [l.strip() for l in texto_nuevo.split('\n') if l.strip()]
        titulo_bloque = None
        for _l in lineas_bloque:
            if _PAT_TITULO_MOMENTO.match(_l):
                titulo_bloque = _l
                break  # solo la primera línea "limpia" es el título

        for linea in lineas_bloque:
            es_proposito_l = bool(re.match(r'Prop[oó]sito\s*:', linea, re.IGNORECASE))
            es_titulo_l    = (linea == titulo_bloque)
            es_numerado    = bool(re.match(r'^\d+\.', linea))

            # No verificar la línea de título del momento ni el Propósito
            if es_titulo_l or es_proposito_l:
                continue

            # a) Sujeto explícito es el o la docente
            if _PAT_ACCION_DOCENTE.match(linea):
                advertencias.append(
                    f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Ítem con docente como sujeto '
                    f'(reescribir dirigido al estudiante): "{linea[:90]}"'
                )

            # b) Frase nominal — empieza con sustantivo conocido, no verbo
            if _PAT_FRASE_NOMINAL.match(linea):
                advertencias.append(
                    f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Ítem nominal sin verbo imperativo '
                    f'(iniciar con verbo: Analiza, Revisa, Identifica…): "{linea[:90]}"'
                )

            # c) Separador "+"
            if '+' in linea:
                advertencias.append(
                    f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Usa "+" como separador — '
                    f'es una nota de planificación docente, NO una instrucción al estudiante. '
                    f'Reescribir como ítems numerados con verbo imperativo: "{linea[:90]}"'
                )

            # d) Cualquier ítem (numerado o no) sin verbo imperativo al inicio
            # Extraer primera palabra real después del número (si existe)
            m_primer = re.match(r'^(?:\d+\.\s+)?([A-ZÁÉÍÓÚa-záéíóúüñÑ]+)', linea)
            if m_primer and not _PAT_ACCION_DOCENTE.match(linea):
                primer = m_primer.group(1).lower()
                if (primer not in _VERBOS_IMPERATIVO
                        and not _PAT_FRASE_NOMINAL.match(linea)):
                    advertencias.append(
                        f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Ítem sin verbo imperativo '
                        f'("{m_primer.group(1)}" no es imperativo — '
                        f'usar Analiza, Revisa, Identifica…): "{linea[:90]}"'
                    )

            # e) "Propósito:" embebido dentro de un ítem (no al final del bloque)
            if re.search(r'Prop[oó]sito\s*:', linea, re.IGNORECASE):
                advertencias.append(
                    f'    [Plan F{r} {str(momento)[:12]}] ⚠️  "Propósito:" dentro de un ítem '
                    f'— debe ser línea independiente al final del bloque: "{linea[:90]}"'
                )

            # f) Recepción pasiva — estudiante solo recibe del o la docente
            if _PAT_RECEPCION_PASIVA.match(linea):
                advertencias.append(
                    f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Ítem pasivo (docente actúa, '
                    f'estudiante recibe) — reformular con verbo activo '
                    f'(Reflexiona, Registra, Compara…): "{linea[:90]}"'
                )

            # g) Nota editorial entre paréntesis para el o la docente
            if _PAT_NOTA_EDITORIAL.search(linea):
                advertencias.append(
                    f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Nota editorial entre paréntesis '
                    f'(eliminar — el texto debe dirigirse al estudiante, '
                    f'no contener instrucciones para el docente): "{linea[:90]}"'
                )

        # h) Bloque con contenido de actividad pero sin "Propósito:" al final
        # Aplica tanto si los ítems son numerados como si no lo son.
        # Fuente: Manual Diseño Instruccional UST, p. 19
        lineas_item = [l for l in lineas_bloque
                       if l != titulo_bloque
                       and not re.match(r'Prop[oó]sito\s*:', l, re.IGNORECASE)]
        tiene_prop = bool(re.search(r'Prop[oó]sito\s*:', texto_nuevo, re.IGNORECASE))
        if len(lineas_item) >= 2 and not tiene_prop:
            advertencias.append(
                f'    [Plan F{r} {str(momento)[:12]}] ⚠️  Falta "Propósito:" al final '
                f'del bloque (estructura obligatoria: título + ítems + '
                f'Propósito, Manual UST p.19)'
            )

        # i) Coherencia semántica: Propósito ↔ ítems, genéricos, breves,
        #    verbos repetidos, incoherencias por tipo de momento
        m_prop = re.search(r'Prop[oó]sito\s*:\s*(.+)', texto_nuevo, re.IGNORECASE)
        proposito_extraido = m_prop.group(1).strip() if m_prop else None
        if len(lineas_item) >= 1:
            _analizar_coherencia_bloque(
                titulo_bloque, proposito_extraido, lineas_item,
                momento, r, advertencias
            )

    if advertencias:
        log.append('\n  [Advertencias de lenguaje — requieren revisión manual]')
        log.extend(advertencias)

    return celdas_modificadas


# ══════════════════════════════════════════════════════════════════════════
#  VERIFICACIÓN DE HORAS
# ══════════════════════════════════════════════════════════════════════════

def _detectar_col_horas(ws):
    """
    Detecta la primera columna de horas pedagógicas en la hoja de planificación,
    buscando 'hora', 'hrs.' o 'horas' (case insensitive) en las primeras 4 filas.
    Excluye columnas TPE (trabajo personal del estudiante).
    Devuelve número de columna (1-based) o None si no se encuentra.
    """
    for fila in range(1, 5):
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(fila, col).value or '').lower()
            if not val:
                continue
            # Detectar encabezados de horas pedagógicas, excluyendo TPE
            if ('hora' in val or 'hrs' in val) and 'tpe' not in val:
                return col
    return None


def _leer_horas_sintesis(ws_sint):
    """
    Lee horas por unidad desde la Síntesis didáctica.
    Busca dinámicamente el encabezado ('Nombre de la unidad o módulo' / 'Horas
    pedagógicas') y lee desde la fila siguiente hasta encontrar la fila TOTAL.

    Devuelve (horas_por_unidad, total_horas):
      horas_por_unidad — dict {nombre_unidad → float}
      total_horas      — float o None
    """
    horas_por_unidad: dict[str, float] = {}
    total_horas = None

    if not ws_sint:
        return horas_por_unidad, total_horas

    # Buscar la fila de encabezado de la tabla de unidades
    fila_inicio = None
    col_nombre  = 2    # col B por defecto
    col_horas   = 7    # col G por defecto
    max_row     = ws_sint.max_row or 40

    for fila in range(1, min(max_row + 1, 35)):
        b = str(ws_sint.cell(fila, 2).value or '').strip()
        g = str(ws_sint.cell(fila, 7).value or '').strip()
        if ('nombre' in b.lower() and 'unidad' in b.lower()) \
                or 'horas pedagógicas' in g.lower():
            fila_inicio = fila + 1
            break

    # Fallback: si no encontró header, leer desde fila 17 (formato estándar UST)
    if fila_inicio is None:
        fila_inicio = 17

    for fila in range(fila_inicio, fila_inicio + 12):
        nombre = ws_sint.cell(fila, col_nombre).value
        horas  = ws_sint.cell(fila, col_horas).value

        # Fila TOTAL: buscar "TOTAL" en cualquier celda de la fila
        row_values = [str(ws_sint.cell(fila, c).value or '') for c in range(1, 10)]
        if any('TOTAL' in v.upper() for v in row_values):
            try:
                total_horas = float(str(horas).replace(',', '.'))
            except (ValueError, TypeError):
                pass
                break   # el TOTAL marca el fin de la sección de unidades

        if nombre is None and horas is None:
            continue

        if nombre and horas:
            try:
                h = float(str(horas).replace(',', '.'))
                horas_por_unidad[str(nombre).strip()] = h
            except (ValueError, TypeError):
                pass

    # Fallback: si el total era una fórmula (=SUM…), calcularlo como suma de unidades
    if total_horas is None and horas_por_unidad:
        total_horas = sum(horas_por_unidad.values())

    return horas_por_unidad, total_horas


def _horas_por_unidad_plan(ws_plan, col_horas):
    """
    Suma horas por unidad y por momento en la hoja Planificación por unidades,
    usando la columna de horas detectada.
    Devuelve:
      totales_unidad:   dict {unidad → total_horas}
      por_momento:      dict {unidad → {momento → horas}}
    """
    totales_unidad = {}
    por_momento    = {}
    unidad_actual  = None

    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        unidad  = row[COL['UNIDAD']  - 1]
        momento = row[COL['MOMENTO'] - 1]
        horas_v = row[col_horas      - 1] if col_horas else None

        if unidad:
            unidad_actual = str(unidad).strip()

        if not unidad_actual:
            continue

        # Ignorar fila de totales
        if 'total' in str(unidad_actual).lower():
            continue

        try:
            h = float(str(horas_v).replace(',', '.')) if horas_v is not None else 0
        except (ValueError, TypeError):
            h = 0

        if h <= 0:
            continue

        if unidad_actual not in totales_unidad:
            totales_unidad[unidad_actual] = 0
            por_momento[unidad_actual]    = {}

        totales_unidad[unidad_actual] += h

        if momento:
            m = str(momento).strip()
            por_momento[unidad_actual][m] = por_momento[unidad_actual].get(m, 0) + h

    return totales_unidad, por_momento


def verificar_horas(ws_plan, ws_sint, programa, log):
    """
    Verifica la distribución de horas en tres niveles:

    Nivel 0 — Total global PDF vs Excel:
      Total horas pedagógicas del programa (tabla distribución) vs total
      declarado en la Síntesis didáctica.

    Nivel 1 — Unidad vs Programa:
      Las horas declaradas por unidad en la Síntesis deben coincidir con
      el programa oficial (tolerancia ±1 hora pedagógica).

    Nivel 2 — Momentos dentro de la unidad:
      Preparación + Desarrollo + Trabajo Independiente = total de la unidad.

    Devuelve (n_ok, n_error).
    """
    errores = []
    oks     = []

    horas_sint, total_sint = _leer_horas_sintesis(ws_sint)

    # ── Nivel 0: Total PDF vs total Síntesis ──────────────────────────────
    total_pdf = (programa or {}).get('total_pedagogicas')
    if total_pdf is not None and total_sint is not None:
        diff0 = abs(total_pdf - total_sint)
        if diff0 <= 1:
            oks.append(
                f'Total global: programa={int(total_pdf)}h = Síntesis={int(total_sint)}h ✓')
        else:
            errores.append(
                f'Total global: programa={int(total_pdf)}h ≠ Síntesis={int(total_sint)}h '
                f'(diferencia {diff0:.0f}h)')

    # ── Nivel 1: Síntesis vs Programa ─────────────────────────────────────
    if programa and not programa.get('_error') and programa.get('unidades'):
        # Construir mapa romano → horas del programa (excluir EXAMEN para el match posicional)
        unidades_prog = [u for u in programa['unidades'] if u['numero'] != 'EXAMEN']
        prog_horas    = {u['numero']: u['horas'] for u in unidades_prog}

        ROMANO     = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII']
        sint_lista = [(n, h) for n, h in horas_sint.items()
                      if not re.search(r'examen|evaluación final', n, re.IGNORECASE)]

        for i, (nombre_sint, hrs_sint) in enumerate(sint_lista):
            num_romano = ROMANO[i] if i < len(ROMANO) else None

            # Intentar leer romano desde el nombre ("Unidad I", "Módulo II"…)
            m = re.search(r'\b([IVX]{1,4})\b', str(nombre_sint))
            if m:
                num_romano = m.group(1).upper()

            if num_romano and num_romano in prog_horas:
                hrs_prog = prog_horas[num_romano]
                diff = abs(hrs_sint - hrs_prog)
                if diff <= 1:
                    oks.append(
                        f'Unidad {num_romano} ({nombre_sint[:30]}): '
                        f'{int(hrs_sint)}h (programa: {hrs_prog}h) ✓')
                else:
                    errores.append(
                        f'Unidad {num_romano} ({nombre_sint[:30]}): '
                        f'Síntesis={int(hrs_sint)}h ≠ programa={hrs_prog}h '
                        f'(diferencia {diff:.0f}h)')
            elif horas_sint:
                oks.append(f'{nombre_sint[:35]}: {int(hrs_sint)}h declaradas en Síntesis')
    else:
        log.append('    ℹ️  Nivel 1 omitido: sin datos de programa oficial disponibles')

    # ── Nivel 2: Momentos dentro de la unidad ─────────────────────────────
    if ws_plan:
        col_horas = _detectar_col_horas(ws_plan)

        if col_horas:
            totales_u, por_mom = _horas_por_unidad_plan(ws_plan, col_horas)
            EXCLUIR = {'total', 'evaluación final', 'lga', 'examen'}

            for unidad, hrs_total_plan in totales_u.items():
                if any(ex in unidad.lower() for ex in EXCLUIR):
                    continue

                momentos = por_mom.get(unidad, {})
                hrs_suma = sum(momentos.values())
                diff_int = abs(hrs_suma - hrs_total_plan)
                u_corta  = unidad[:35]

                if diff_int <= 0.5:
                    oks.append(f'{u_corta}: momentos suman {hrs_suma:.0f}h = total unidad ✓')
                else:
                    detalle_mom = ', '.join(
                        f'{m[:15]}={v:.0f}h' for m, v in momentos.items()
                    )
                    errores.append(
                        f'{u_corta}: momentos suman {hrs_suma:.0f}h ≠ total {hrs_total_plan:.0f}h '
                        f'({detalle_mom})')

                mom_lower = {str(k).lower(): v for k, v in momentos.items()}
                for lbl in ('preparación', 'desarrollo', 'trabajo independiente'):
                    if not any(lbl in k for k in mom_lower):
                        errores.append(
                            f'{u_corta}: sin horas asignadas a "{lbl.capitalize()}"')

        else:
            log.append(
                '    ℹ️  Nivel 2 omitido: no se detectó columna de horas en la Planificación '
                '(verifica que el encabezado contenga "hora")')

    # ── Reporte ───────────────────────────────────────────────────────────
    log.append('\n  [Verificación de horas — Total / Síntesis vs Programa / Momentos]')
    for o in oks:
        log.append(f'    ✅ {o}')
    for e in errores:
        log.append(f'    ❌ {e}')
    if not oks and not errores:
        log.append('    ⚠️  Sin datos suficientes para verificar horas')

    log.append(f'\n    Horas: {len(oks)}✅  {len(errores)}❌')
    return len(oks), len(errores)


# ══════════════════════════════════════════════════════════════════════════
#  VERIFICACIÓN DE MOMENTOS (Manual Diseño Instruccional UST, Res. N°481/24)
# ══════════════════════════════════════════════════════════════════════════
#
# Reglas según el Manual de Diseño Instruccional en E-Learning (pp. 17-22):
#
# PREPARACIÓN (TPE — Tiempo de Preparación del Estudiante)
#   • Debe incluir al menos un recurso didáctico (col H) de tipo T1/T2/T3
#   • Las actividades deben ser de exploración/revisión previa (no sumativas)
#   • La evaluación Diagnóstica debe ubicarse en Preparación (inicio de unidad)
#   • Cada bloque debe terminar con "Propósito:" (estructura de actividad)
#
# DESARROLLO (Sincrónico / Asincrónico / Presencial)
#   • Momento central de aprendizaje — debe existir en cada unidad
#   • Contiene las evaluaciones Formativas y/o la principal actividad de la unidad
#   • Si la modalidad es sincrónica, Medio de entrega debe ser "No aplica"
#   • Si es asincrónica, puede tener Buzón de tareas
#
# TRABAJO INDEPENDIENTE (Autónomo del estudiante)
#   • Debe mostrar autonomía del estudiante (producto, entrega, evaluación)
#   • Medio de entrega debe ser "Buzón de tareas" (ya corregido automáticamente)
#   • Evaluaciones Sumativas pueden ubicarse aquí o en Desarrollo
#   • Debe tener instrucciones claras en imperativo (ya verificado)

# Tipos de recursos válidos según Manual (p. 25-27)
_RECURSOS_T1 = {'videoclase', 'video cápsula', 'video capsula', 'podcast', 'video'}
_RECURSOS_T2 = {'genially'}
_RECURSOS_T3 = {'guía de aprendizaje', 'guia de aprendizaje', 'guía de ejercicios',
                'guia de ejercicios', 'apunte', 'infografía', 'infografia'}
_RECURSOS_T4 = {'foro', 'quiz', 'tarea', 'cuestionario'}
_RECURSOS_VALIDOS = _RECURSOS_T1 | _RECURSOS_T2 | _RECURSOS_T3 | _RECURSOS_T4


def verificar_momentos(ws_plan, log):
    """
    Verifica reglas de los 3 momentos de aprendizaje según el Manual de
    Diseño Instruccional en E-Learning UST (Resolución N°481/24).

    Reglas verificadas:
    1. Preparación: debe tener recursos didácticos
    2. Preparación: Diagnóstica debe ubicarse aquí (no en Desarrollo/TI)
    3. Preparación: no debe contener evaluaciones Sumativas
    4. Desarrollo: debe existir en cada unidad
    5. Trabajo Independiente: debe existir en cada unidad
    6. Actividades: deben contener "Propósito:" (estructura obligatoria)
    7. Recursos: deben corresponder a tipos T1/T2/T3/T4 del manual

    Devuelve (n_ok, n_advertencias).
    """
    if not ws_plan:
        return 0, 0

    # ── Recolectar datos por fila ──────────────────────────────────────────
    unidad_actual  = None
    datos_unidades = {}   # unidad → {'prep': [], 'des': [], 'ti': []}

    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        unidad    = row[COL['UNIDAD']    - 1]
        momento   = row[COL['MOMENTO']  - 1]
        actividad = row[COL['ACTIVIDAD']- 1]
        recursos  = row[COL['RECURSOS'] - 1]
        tipo      = row[COL['TIPO']     - 1]

        if unidad:
            unidad_actual = str(unidad).strip()
        if not unidad_actual:
            continue

        if unidad_actual not in datos_unidades:
            datos_unidades[unidad_actual] = {'prep': [], 'des': [], 'ti': []}

        m = str(momento or '').lower()
        fila = {
            'actividad': str(actividad or ''),
            'recursos':  str(recursos  or ''),
            'tipo':      str(tipo      or '').lower(),
        }

        if 'preparación' in m or 'preparacion' in m:
            datos_unidades[unidad_actual]['prep'].append(fila)
        elif 'desarrollo' in m:
            datos_unidades[unidad_actual]['des'].append(fila)
        elif 'independiente' in m:
            datos_unidades[unidad_actual]['ti'].append(fila)

    # ── Verificaciones por unidad ──────────────────────────────────────────
    advertencias = []
    oks = []

    EXCLUIR = {'evaluación final', 'lga', 'examen', 'total'}

    for unidad, momentos_data in datos_unidades.items():
        if any(ex in unidad.lower() for ex in EXCLUIR):
            continue

        u_corta = unidad[:35]
        prep = momentos_data['prep']
        des  = momentos_data['des']
        ti   = momentos_data['ti']

        # Regla 1: Preparación debe existir
        if not prep:
            advertencias.append(
                f'    ⚠️  {u_corta}: sin momento Preparación '
                f'(Manual UST p.17: obligatorio antes de cada sesión)')
        else:
            oks.append(f'{u_corta}: Preparación presente ({len(prep)} fila(s))')

        # Regla 2: Preparación debe tener recursos
        if prep:
            prep_sin_recursos = [f for f in prep if not f['recursos'].strip()]
            if prep_sin_recursos:
                advertencias.append(
                    f'    ⚠️  {u_corta}: {len(prep_sin_recursos)} fila(s) de Preparación '
                    f'sin recursos didácticos (Manual UST p.25: T1/T2/T3 obligatorios)')
            else:
                oks.append(f'{u_corta}: Preparación tiene recursos')

        # Regla 3: Diagnóstica debe estar en Preparación (no en Desarrollo/TI)
        diag_en_prep = any('diagnóst' in f['tipo'] for f in prep)
        diag_en_des  = any('diagnóst' in f['tipo'] for f in des)
        diag_en_ti   = any('diagnóst' in f['tipo'] for f in ti)
        if (diag_en_des or diag_en_ti) and not diag_en_prep:
            advertencias.append(
                f'    ⚠️  {u_corta}: Diagnóstica en Desarrollo/TI — '
                f'Manual UST p.18: debe ubicarse en Preparación (inicio de unidad)')
        elif diag_en_prep:
            oks.append(f'{u_corta}: Diagnóstica correctamente en Preparación')

        # Regla 4: No debe haber Sumativa en Preparación
        sumativa_en_prep = any('sumativ' in f['tipo'] for f in prep)
        if sumativa_en_prep:
            advertencias.append(
                f'    ⚠️  {u_corta}: Evaluación Sumativa en momento Preparación '
                f'(Manual UST p.18: Sumativa va en Desarrollo o Trabajo Independiente)')

        # Regla 5: Desarrollo debe existir
        if not des:
            advertencias.append(
                f'    ⚠️  {u_corta}: sin momento Desarrollo '
                f'(Manual UST p.17: momento central obligatorio por unidad)')
        else:
            oks.append(f'{u_corta}: Desarrollo presente ({len(des)} fila(s))')

        # Regla 6: Trabajo Independiente debe existir
        if not ti:
            advertencias.append(
                f'    ⚠️  {u_corta}: sin Trabajo Independiente '
                f'(Manual UST p.17: autonomía del estudiante obligatoria por unidad)')
        else:
            oks.append(f'{u_corta}: Trabajo Independiente presente ({len(ti)} fila(s))')

        # Regla 7: TI debe tener al menos una evaluación o actividad con producto
        if ti:
            ti_con_eval = any(f['tipo'] for f in ti)
            ti_con_actividad = any(f['actividad'].strip() for f in ti)
            if not ti_con_eval and not ti_con_actividad:
                advertencias.append(
                    f'    ⚠️  {u_corta}: Trabajo Independiente sin actividad ni evaluación '
                    f'(Manual UST p.20: debe evidenciar producto o entrega del estudiante)')

    # ── Verificación global: estructura de actividades (Propósito) ─────────
    n_sin_proposito = 0
    n_con_proposito = 0
    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        actividad = row[COL['ACTIVIDAD'] - 1]
        momento   = row[COL['MOMENTO']  - 1]
        if not actividad or not momento:
            continue
        texto = str(actividad)
        tiene_items = bool(re.search(r'^\d+\.', texto, re.MULTILINE))
        tiene_prop  = bool(re.search(r'Prop[oó]sito\s*:', texto, re.IGNORECASE))
        if tiene_items and tiene_prop:
            n_con_proposito += 1
        elif tiene_items and not tiene_prop:
            n_sin_proposito += 1

    if n_sin_proposito > 0:
        advertencias.append(
            f'    ⚠️  {n_sin_proposito} celda(s) con ítems numerados pero sin "Propósito:" '
            f'(Manual UST p.19: estructura: título + ítems numerados + Propósito obligatorio)')
    elif n_con_proposito > 0:
        oks.append(f'{n_con_proposito} bloque(s) con Propósito correctamente declarado')

    # ── Verificación global: tipos de recursos ─────────────────────────────
    recursos_desconocidos = []
    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        recursos = str(row[COL['RECURSOS'] - 1] or '').strip().lower()
        if not recursos:
            continue
        # Verificar que al menos un recurso reconocido esté presente
        reconocido = any(r in recursos for r in _RECURSOS_VALIDOS)
        if not reconocido and len(recursos) > 3:
            recursos_desconocidos.append(recursos[:60])

    if recursos_desconocidos:
        n_desc = len(set(recursos_desconocidos))
        advertencias.append(
            f'    ⚠️  {n_desc} tipo(s) de recurso no reconocido(s) '
            f'(Manual UST p.25: T1=Video/Podcast, T2=Genially, T3=Guía/Apunte, T4=Foro/Quiz/Tarea)')

    # ── Reporte ───────────────────────────────────────────────────────────
    log.append('\n  [Verificación de momentos — Manual Diseño Instruccional UST (Res. N°481/24)]')
    for o in oks:
        log.append(f'    ✅ {o}')
    for a in advertencias:
        log.append(a)

    if not advertencias:
        log.append('    ✅ Estructura de momentos correcta en todas las unidades')

    log.append(f'\n    Momentos: {len(oks)}✅  {len(advertencias)}⚠️  advertencias')
    return len(oks), len(advertencias)


# ── Patrones para validación de consignas de Foro (T4) ──────────────────────

# Referencia a sesión sincrónica (no válido en e-learning asíncrono)
_PAT_FORO_SESION = re.compile(
    r'\b(?:trabajados?\s+en\s+(?:la\s+)?sesi[oó]n|vistos?\s+en\s+clase|'
    r'comentados?\s+en\s+(?:la\s+)?sesi[oó]n|discutidos?\s+en\s+(?:la\s+)?sesi[oó]n|'
    r'revisados?\s+en\s+(?:la\s+)?sesi[oó]n|seg[uú]n\s+(?:lo\s+)?visto\s+en\s+clase|'
    r'a\s+partir\s+de\s+(?:lo\s+)?trabajado\s+en\s+(?:la\s+)?sesi[oó]n)\b',
    re.IGNORECASE,
)

# Instrucción de participación entre pares (obligatoria en foros)
_PAT_FORO_PARTICIPACION = re.compile(
    r'\b(?:responde?\s+(?:al\s+menos|a\s+(?:un|dos|tres|\d+))\s+(?:compa[ñn]ero|participaci[oó]n)|'
    r'comenta?\s+(?:al\s+menos|la\s+publicaci[oó]n)|'
    r'interact[uú]a?\s+con|retroalimenta?|responde?\s+a\s+(?:dos|tres|\d+)|'
    r'comenta?\s+(?:dos|tres|\d+))\b',
    re.IGNORECASE,
)

# Criterio de evaluación explícito
_PAT_FORO_CRITERIO = re.compile(
    r'\b(?:criterio|r[uú]brica|se\s+evaluar[aá]|ser[aá]\s+evaluad|puntaje|'
    r'se\s+considerar[aá]|se\s+valorar[aá]|indicador)\b',
    re.IGNORECASE,
)

# Fuente citada en la consigna (apellido + año entre paréntesis)
_PAT_FORO_CITA = re.compile(r'[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+\s+\(\d{4}\)')

# Referencia APA completa al pie (línea que comienza con Apellido, I.)
_PAT_FORO_REFERENCIA_APA = re.compile(
    r'^[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+,\s+[A-Z]\.',
    re.MULTILINE,
)


def verificar_consignas_foro(ws_plan, log):
    """
    Valida que cada celda de actividad asociada a un recurso Foro cumpla:
    1. No referencie la sesión sincrónica (autonomía e-learning).
    2. Incluya instrucción de participación entre pares.
    3. Incluya criterio de evaluación explícito.
    4. Si cita fuentes (Autor, año), incluya referencia APA 7 completa.

    Devuelve (n_ok, n_advertencias).
    """
    if not ws_plan:
        return 0, 0

    adv = []
    ok  = 0

    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        recursos  = str(row[COL['RECURSOS']  - 1] or '').strip().lower()
        actividad = str(row[COL['ACTIVIDAD'] - 1] or '').strip()
        unidad    = str(row[COL['UNIDAD']    - 1] or '').strip()
        semana    = str(row[COL['SEMANA']    - 1] or '').strip()

        if 'foro' not in recursos:
            continue
        if not actividad or len(actividad) < 20:
            continue

        ref = f'Foro {("U"+unidad) if unidad else ""}{(" S"+semana) if semana else ""}'.strip()
        problemas = []

        # 1. Referencia a sesión sincrónica
        if _PAT_FORO_SESION.search(actividad):
            problemas.append(
                'hace referencia a "la sesión" — la consigna debe ser autónoma '
                '(el estudiante a distancia no asistió a clase)'
            )

        # 2. Instrucción de participación entre pares
        if not _PAT_FORO_PARTICIPACION.search(actividad):
            problemas.append(
                'falta instrucción de participación entre pares '
                '(ej: "Responde al menos a dos compañeros") — obligatorio en Foro T4'
            )

        # 3. Criterio de evaluación
        if not _PAT_FORO_CRITERIO.search(actividad):
            problemas.append(
                'falta criterio de evaluación explícito '
                '(qué hace la reflexión buena, no solo cuántas palabras)'
            )

        # 4. Cita sin referencia APA completa
        citas = _PAT_FORO_CITA.findall(actividad)
        if citas and not _PAT_FORO_REFERENCIA_APA.search(actividad):
            autores = ', '.join(set(citas[:3]))
            problemas.append(
                f'cita fuente(s) ({autores}) pero no incluye referencia APA 7 completa al pie'
            )

        if problemas:
            adv.append(f'    ⚠️  {ref} (Foro):')
            for p in problemas:
                adv.append(f'        → {p}')
        else:
            ok += 1

    log.append('\n  [Verificación de consignas de Foro (T4)]')
    if adv:
        for a in adv:
            log.append(a)
    elif ok > 0:
        log.append(f'    ✅ {ok} foro(s) con consigna completa')
    else:
        log.append('    —  No se detectaron recursos Foro en la planificación')

    n_adv = len([a for a in adv if a.strip().startswith('⚠️')])
    return ok, n_adv


# ══════════════════════════════════════════════════════════════════════════
#  REVISIÓN ORTOGRÁFICA Y GRAMATICAL — LanguageTool API (gratuita, es)
# ══════════════════════════════════════════════════════════════════════════

_LT_URL     = 'https://api.languagetool.org/v2/check'
_LT_LANG    = 'es'
# Reglas que generan demasiado ruido en textos educativos estructurados
_LT_DISABLE = ','.join([
    'WHITESPACE_RULE',
    'UPPERCASE_SENTENCE_START',
    'ES_QUESTION_MARK',             # signos de interrogación invertidos
    'COMMA_PARENTHESIS_WHITESPACE',
    'DOUBLE_PUNCTUATION',
])
_LT_MIN_CHARS   = 30    # ignorar celdas muy cortas
_LT_PAUSA_SEG   = 3.5   # pausa entre llamadas (free tier ~20 req/min)

# Flag global para activar autocorrección conservadora
_LANGUAGETOOL_AUTOCORREGIR = True   # aplica correcciones seguras (tildes, typos claros)

# Reglas de LanguageTool consideradas SEGURAS para autocorrección
# Estas son correcciones ortográficas unívocas o errores gramaticales simples
_LT_REGLAS_SEGURAS = {
    # Ortografía básica — typos y acentos
    'MORFOLOGIK_RULE_ES',          # errores ortográficos generales
    'ES_ACCENT_ERRORS',            # errores de acentuación
    'HUNSPELL_NO_SUGGEST_RULE',    # palabras no reconocidas
    'ES_SIMPLE_REPLACE',           # reemplazos simples
    # Errores comunes en español
    'ES_CONFUSION_B_V',            # confusión b/v
    'ES_CONFUSION_H',              # h muda
    'ES_CONFUSION_LL_Y',           # confusión ll/y
    'ES_CONFUSION_C_S_Z',          # confusión c/s/z (seseo)
    'ES_CONFUSION_G_J',            # confusión g/j
    'ES_CONFUSION_S_X',            # confusión s/x
    # Mayúsculas
    'UPPERCASE_SENTENCE_START',    # mayúscula inicial
}


def autocorregir_con_languagetool(texto, umbral='conservador'):
    """
    Aplica correcciones SEGURAS de LanguageTool automáticamente.

    Args:
        texto: Texto a corregir
        umbral: Nivel de confianza para aplicar correcciones
            - 'conservador': solo cuando hay 1 sugerencia única y la regla es segura
            - 'moderado': acepta hasta 2 sugerencias si la primera es muy probable
            - 'todos': aplica siempre la primera sugerencia (riesgoso)

    Returns:
        (texto_corregido, lista_cambios)
        - texto_corregido: texto con correcciones aplicadas
        - lista_cambios: lista de strings describiendo cada cambio
    """
    if not texto or len(texto.strip()) < _LT_MIN_CHARS:
        return texto, []

    errores = revisar_con_languagetool(texto)
    if not errores:
        return texto, []

    cambios = []
    texto_corr = texto

    # Ordenar errores por offset descendente para corregir de atrás hacia adelante
    # (así no se desplazan los offsets de errores anteriores)
    errores_ordenados = sorted(errores, key=lambda e: e.get('offset', 0), reverse=True)

    for error in errores_ordenados:
        offset = error.get('offset', 0)
        largo = error.get('length', 1)
        sugerencias = [s['value'] for s in error.get('replacements', [])]
        regla_id = error.get('rule', {}).get('id', '')
        mensaje = error.get('message', '')

        # Sin sugerencias → no hay nada que corregir
        if not sugerencias:
            continue

        # Extraer el fragmento original con error
        if offset + largo > len(texto_corr):
            continue
        original = texto_corr[offset:offset + largo]
        sugerencia = sugerencias[0]

        # ── Criterios de SEGURIDAD ─────────────────────────────────────────
        aplicar = False
        razon = ''

        if umbral == 'conservador':
            # Solo aplicar si:
            # 1. Hay una ÚNICA sugerencia (no ambiguo)
            # 2. La regla está en la lista de reglas seguras
            # 3. El cambio no altera significativamente la longitud
            if len(sugerencias) == 1:
                if regla_id in _LT_REGLAS_SEGURAS:
                    aplicar = True
                    razon = f'regla segura: {regla_id}'
                # También aceptar si el cambio es solo ortográfico (mismos caracteres visibles)
                elif original.lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ü','u') == \
                     sugerencia.lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ü','u'):
                    # Es solo un cambio de acentos o mayúsculas
                    if len(original) == len(sugerencia):
                        aplicar = True
                        razon = 'corrección de acento/mayúscula'

        elif umbral == 'moderado':
            # Aceptar hasta 2 sugerencias si la primera parece correcta
            if len(sugerencias) <= 2:
                aplicar = True
                razon = f'umbral moderado ({len(sugerencias)} sugerencias)'

        elif umbral == 'todos':
            # Aplicar siempre la primera sugerencia (riesgoso)
            aplicar = True
            razon = 'umbral todos'

        # ── Aplicar corrección ───────────────────────────────────────────────
        if aplicar:
            texto_corr = texto_corr[:offset] + sugerencia + texto_corr[offset + largo:]
            cambios.append(f'«{original}» → «{sugerencia}» ({razon})')

    return texto_corr, cambios


def revisar_con_languagetool(texto):
    """
    Llama a la API gratuita de LanguageTool y devuelve lista de errores.
    Cada error es un dict con: mensaje, sugerencias, offset, largo, regla_id.
    Devuelve lista vacía si falla la conexión o el texto es muy corto.
    """
    if not texto or len(texto.strip()) < _LT_MIN_CHARS:
        return []
    try:
        data = urllib.parse.urlencode({
            'text':          texto,
            'language':      _LT_LANG,
            'disabledRules': _LT_DISABLE,
        }).encode('utf-8')
        req = urllib.request.Request(
            _LT_URL,
            data=data,
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            resultado = json.loads(resp.read().decode('utf-8'))
        return resultado.get('matches', [])
    except Exception:
        return None  # señal de fallo de conexión — los llamadores buscan None


def _lt_fragmento(texto, offset, largo, max_ctx=40):
    """Extrae el fragmento con error y agrega contexto."""
    inicio = max(0, offset - 10)
    fin    = min(len(texto), offset + largo + 10)
    frag   = texto[inicio:fin].replace('\n', ' ')
    if len(frag) > max_ctx:
        frag = frag[:max_ctx] + '…'
    return frag


def verificar_lenguaje_momentos(ws_plan, log, autocorregir=False):
    """
    Recorre las celdas de ACTIVIDAD en los 3 momentos de aprendizaje y
    reporta errores ortográficos y gramaticales usando LanguageTool API.

    Args:
        ws_plan: Worksheet de planificación
        log: Lista para mensajes de log
        autocorregir: Si True, aplica correcciones SEGURAS automáticamente
                      (solo cambios unívocos de reglas seguras)

    Devuelve (n_revisadas, n_errores, n_correcciones).
    """
    if not ws_plan:
        return 0, 0, 0

    modo_str = 'autocorrección conservadora' if autocorregir else 'solo detección'
    log.append(f'\n  [Revisión ortográfica/gramatical — LanguageTool es ({modo_str})]')

    revisadas    = 0
    n_errores    = 0
    n_correcciones = 0
    sin_acceso   = False

    # Usar values_only=False para poder modificar celdas si autocorregir=True
    for row in ws_plan.iter_rows(min_row=4, values_only=False):
        cell_momento = row[COL['MOMENTO']   - 1]
        cell_act    = row[COL['ACTIVIDAD'] - 1]

        momento = cell_momento.value if cell_momento else None
        actividad = cell_act.value if cell_act else None

        if not momento or not actividad:
            continue
        m = str(momento).lower()
        if not any(k in m for k in ('preparación', 'preparacion', 'desarrollo', 'independiente')):
            continue

        texto = str(actividad).strip()
        if len(texto) < _LT_MIN_CHARS:
            continue

        if sin_acceso:
            break

        errores = revisar_con_languagetool(texto)
        if errores is None:
            sin_acceso = True
            log.append('    ⚠️  Sin acceso a LanguageTool API — revisión omitida')
            break

        revisadas += 1
        momento_label = str(momento).strip()
        errores_encontrados = 0

        # ── Modo AUTOCORRECCIÓN ─────────────────────────────────────────────
        if autocorregir and errores:
            texto_corr, cambios = autocorregir_con_languagetool(texto, umbral='conservador')

            if cambios:
                # Aplicar correcciones a la celda
                cell_act.value = texto_corr
                aplicar_azul_diff(cell_act, texto, texto_corr)  # azul solo en líneas modificadas
                n_correcciones += len(cambios)

                for cambio in cambios:
                    log.append(f'    ✅ [{momento_label}] Corregido: {cambio}')
                    errores_encontrados += 1

            # Reportar errores NO corregidos automáticamente
            for e in errores:
                offset  = e.get('offset', 0)
                largo   = e.get('length', 1)
                mensaje = e.get('message', '')
                regla   = e.get('rule', {}).get('id', '')
                sugs    = [s['value'] for s in e.get('replacements', [])[:3]]
                frag    = _lt_fragmento(texto, offset, largo)

                # Verificar si este error fue corregido
                frag_original = texto[offset:offset+largo] if offset + largo <= len(texto) else ''
                if frag_original.lower() not in [c.split('→')[0].strip('«» ') for c in cambios]:
                    sug_txt = f' → sugerencia(s): {", ".join(sugs)}' if sugs else ''
                    log.append(
                        f'    ⚠️  [{momento_label}] «{frag}» — {mensaje}{sug_txt}'
                        f'  ({regla})'
                    )
                    errores_encontrados += 1

        # ── Modo SOLO DETECCIÓN ───────────────────────────────────────────────
        else:
            for e in errores:
                offset  = e.get('offset', 0)
                largo   = e.get('length', 1)
                mensaje = e.get('message', '')
                regla   = e.get('rule', {}).get('id', '')
                sugs    = [s['value'] for s in e.get('replacements', [])[:3]]
                frag    = _lt_fragmento(texto, offset, largo)

                sug_txt = f' → sugerencia(s): {", ".join(sugs)}' if sugs else ''
                log.append(
                    f'    ⚠️  [{momento_label}] «{frag}» — {mensaje}{sug_txt}'
                    f'  ({regla})'
                )
                errores_encontrados += 1

        n_errores += errores_encontrados
        time.sleep(_LT_PAUSA_SEG)

    # ── Resumen ─────────────────────────────────────────────────────────────
    if not sin_acceso:
        if n_errores == 0 and revisadas > 0:
            log.append(f'    ✅ Sin errores detectados ({revisadas} celda(s) revisadas)')
        elif revisadas == 0:
            log.append('    ℹ️  Sin celdas de actividad suficientemente largas para revisar')

        resumen = f'LanguageTool: {revisadas} celda(s) revisadas, {n_errores} error(es)'
        if autocorregir:
            resumen += f', {n_correcciones} corrección(es) aplicada(s)'
        log.append(f'\n    {resumen}')

    return revisadas, n_errores, n_correcciones


# ══════════════════════════════════════════════════════════════════════════
#  VERIFICACIÓN A+Se (APRENDIZAJE + SERVICIO E-LEARNING)
# ══════════════════════════════════════════════════════════════════════════

HITOS_AS = [
    # (id, descripción, modo)
    ('RE1',   'Reflexión Estructurada 1 presente (diagnóstico)',               'auto'),
    ('RE2',   'Reflexión Estructurada 2 presente (diseño del servicio)',        'auto'),
    ('RE3',   'Reflexión Estructurada 3 presente (cierre/evaluación)',          'auto'),
    ('SC',    'Socio Comunitario mencionado en actividades',                    'auto'),
    ('SC_S7', 'Estudiantes conocen al SC antes de semana 7',                   'auto'),
    ('ENC',   'Encuestas A+S programadas en semanas 17-18',                    'auto'),
    ('GRUP',  'Trabajo grupal/equipo presente (duplas o grupos)',               'auto'),
    ('PROD',  'Evaluación sumativa incluye producto o actuación del estudiante','auto'),
    ('RE3_S', 'Reflexión Estructurada 3 en momento sincrónico',                'auto'),
    ('ETAP',  'Las 4 etapas A+S representadas en la planificación',            'manual'),
    ('CIER',  'Actividad de cierre con el socio comunitario presente',         'manual'),
]


def verificar_as(ws_plan, log):
    """
    Verifica los hitos obligatorios A+Se según el lineamiento UST 2025
    en la hoja Planificación por unidades.
    Devuelve (n_ok, n_error, n_manual).
    """
    # Recolectar datos por fila
    filas = []
    semana_actual = None
    for row in ws_plan.iter_rows(min_row=4, values_only=True):
        semana    = row[COL['SEMANA']    - 1]
        momento   = row[COL['MOMENTO']  - 1]
        actividad = row[COL['ACTIVIDAD']- 1]
        tipo      = row[COL['TIPO']     - 1]
        proc      = row[COL['PROC']     - 1]
        indiv     = row[COL['INDIV']    - 1]
        if semana:
            try:
                semana_actual = int(re.sub(r'[^\d]', '', str(semana)))
            except ValueError:
                pass
        if actividad or tipo:
            filas.append({
                'semana':    semana_actual,
                'momento':   str(momento   or '').lower(),
                'actividad': str(actividad or '').lower(),
                'tipo':      str(tipo      or '').lower(),
                'proc':      str(proc      or '').lower(),
                'indiv':     str(indiv     or '').lower(),
            })

    resultados = {}

    # ── RE1 / RE2 / RE3 — contar Reflexiones Estructuradas ───────────────
    def es_re(texto):
        return bool(re.search(
            r'reflexi[oó]n\s+estructurada|RE\s*[123]',
            texto, re.IGNORECASE
        ))

    filas_re = [f for f in filas if es_re(f['actividad'])]
    n_re = len(filas_re)

    resultados['RE1'] = (
        ('✅', 'Reflexión Estructurada 1 encontrada') if n_re >= 1
        else ('❌', 'No se encontró ninguna Reflexión Estructurada en actividades')
    )
    resultados['RE2'] = (
        ('✅', 'Reflexión Estructurada 2 encontrada') if n_re >= 2
        else ('❌', f'Solo {n_re} reflexión(es) encontrada(s) — se necesitan 3')
    )
    resultados['RE3'] = (
        ('✅', 'Reflexión Estructurada 3 encontrada') if n_re >= 3
        else ('❌', f'Solo {n_re} reflexión(es) encontrada(s) — se necesitan 3')
    )

    # ── SC — Socio Comunitario mencionado ─────────────────────────────────
    def menciona_sc(texto):
        return bool(re.search(r'socio\s+comunitario', texto, re.IGNORECASE))

    filas_sc = [f for f in filas if menciona_sc(f['actividad'])]
    resultados['SC'] = (
        ('✅', f'"Socio comunitario" encontrado ({len(filas_sc)} fila(s))') if filas_sc
        else ('❌', '"Socio comunitario" no aparece en ninguna actividad')
    )

    # ── SC_S7 — SC conocido antes de semana 7 ────────────────────────────
    primera_semana_sc = next(
        (f['semana'] for f in filas_sc if f['semana'] is not None), None
    )
    if primera_semana_sc is None:
        resultados['SC_S7'] = ('❌', 'Socio comunitario sin semana asignada')
    elif primera_semana_sc <= 7:
        resultados['SC_S7'] = ('✅', f'SC en semana {primera_semana_sc} (≤ 7 ✓)')
    else:
        resultados['SC_S7'] = ('⚠️',
            f'SC en semana {primera_semana_sc} — lineamiento recomienda ≤ semana 7')

    # ── ENC — Encuestas A+S en semanas 17-18 ─────────────────────────────
    filas_enc = [f for f in filas
                 if re.search(r'encuesta', f['actividad'], re.IGNORECASE)]
    semanas_enc = [f['semana'] for f in filas_enc if f['semana'] is not None]
    if not semanas_enc:
        resultados['ENC'] = ('⚠️',
            'Encuesta A+S no encontrada — debe programarse en semanas 17-18')
    elif all(17 <= s <= 18 for s in semanas_enc):
        resultados['ENC'] = ('✅', f'Encuesta A+S en semana(s) {semanas_enc}')
    else:
        resultados['ENC'] = ('⚠️',
            f'Encuesta en semana(s) {semanas_enc} — debe estar en semanas 17-18')

    # ── GRUP — Trabajo grupal presente ───────────────────────────────────
    tiene_grupal = any('grupal' in f['indiv'] for f in filas)
    resultados['GRUP'] = (
        ('✅', 'Trabajo grupal/equipo presente') if tiene_grupal
        else ('❌', 'Sin trabajo grupal — A+S requiere trabajo en duplas o equipos')
    )

    # ── PROD — Sumativa con producto o actuación ──────────────────────────
    tiene_prod = any(
        'sumativa' in f['tipo'] and
        any(k in f['proc'] for k in ('producciones', 'actuación', 'actuacion', 'informe'))
        for f in filas
    )
    resultados['PROD'] = (
        ('✅', 'Evaluación sumativa incluye producto o actuación del estudiante')
        if tiene_prod
        else ('❌',
              'Sumativa sin producto/actuación — A+S evalúa el servicio entregado')
    )

    # ── RE3_S — Reflexión 3 en momento sincrónico ─────────────────────────
    re3_sinc = any(
        es_re(f['actividad']) and 'sincr' in f['momento']
        for f in filas
    )
    resultados['RE3_S'] = (
        ('✅', 'Reflexión Estructurada 3 en momento sincrónico') if re3_sinc
        else ('⚠️', 'No se confirmó RE3 sincrónica — el lineamiento lo recomienda')
    )

    # ── Criterios manuales ────────────────────────────────────────────────
    resultados['ETAP'] = ('⚠️ manual',
        'Verificar que las 4 etapas A+S estén representadas (Diagnóstico, '
        'Planificación, Implementación, Evaluación/Cierre)')
    resultados['CIER'] = ('⚠️ manual',
        'Verificar que haya una actividad de cierre con el socio comunitario')

    # ── Reporte ───────────────────────────────────────────────────────────
    log.append('\n  [Verificación A+Se — Hitos obligatorios (Lineamiento UST 2025)]')
    n_ok = n_err = n_manual = 0
    for hito_id, desc, _ in HITOS_AS:
        estado, detalle = resultados.get(hito_id, ('⚠️', 'No verificado'))
        desc_corta = desc[:65] + '…' if len(desc) > 65 else desc
        log.append(f'    {estado} {desc_corta}')
        if estado != '✅':
            log.append(f'         → {detalle}')
        if estado == '✅':
            n_ok += 1
        elif 'manual' in str(estado):
            n_manual += 1
        else:
            n_err += 1

    log.append(
        f'\n    Resultado A+Se: {n_ok}✅  {n_err}❌  {n_manual}⚠️ manual'
        f'  (de {len(HITOS_AS)} hitos)'
    )
    return n_ok, n_err, n_manual


# ══════════════════════════════════════════════════════════════════════════
#  PROCESAMIENTO POR ASIGNATURA
# ══════════════════════════════════════════════════════════════════════════

def procesar_asignatura(carpeta_asig, dry_run=False, programa=None, es_as=False):
    nombre = os.path.basename(carpeta_asig)
    log = [f'\n{"═"*70}', f'  ASIGNATURA: {nombre}', f'{"═"*70}']

    carpeta_correcciones = os.path.join(carpeta_asig, 'Correcciones')
    carpeta_enviado      = os.path.join(carpeta_asig, 'Enviado a DEL')
    carpeta_revisado     = os.path.join(carpeta_asig, 'Revisado')

    # ── Determinar fuente ──────────────────────────────────────────────────
    plan_path   = encontrar_archivo(carpeta_correcciones, '*.xlsx')
    escala_path = None
    es_original = False

    if plan_path:
        fuente = 'Correcciones'
        log.append(f'  Fuente : Correcciones/ (revisada por DEL)')
        # intentar encontrar escala igual
        _, escala_path = encontrar_planificacion_enviada(carpeta_enviado)
    else:
        # Buscar en Enviado a DEL
        if os.path.isdir(carpeta_enviado):
            plan_path, escala_path = encontrar_planificacion_enviada(carpeta_enviado)
        if plan_path:
            fuente = 'Enviado a DEL'
            es_original = True
            log.append(f'  Fuente : Enviado a DEL/ (planificación original — sin revisión DEL)')
        else:
            log.append('  ⚠️  No se encontró ninguna planificación (.xlsx) — omitida')
            return log, False

    log.append(f'  Archivo: {os.path.basename(plan_path)}')

    # ── Observaciones específicas del archivo de escala (contexto adicional) ─
    obs_escala = leer_observaciones_escala(escala_path)
    if escala_path:
        log.append(f'  Escala : {os.path.basename(escala_path)}')
        if obs_escala:
            log.append('  [Observaciones DEL registradas en la escala — solo referencia]')
            for criterio, obs in obs_escala:
                log.append(f'    • {criterio[:65]}')
                log.append(f'      → {obs[:130]}')
    else:
        log.append('  Escala : [no disponible — se aplican los 41 criterios estándar UST]')

    # ── Cargar libro ───────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(plan_path)
    hojas = wb.sheetnames
    log.append(f'\n  Hojas detectadas: {hojas}')

    total_sintesis = 0
    total_alta = 0
    total_media_j = 0
    total_media_m = 0
    total_baja = 0
    registro = []   # Registro estructurado de cambios para la hoja final

    # ── Síntesis didáctica ─────────────────────────────────────────────────
    if _tiene_hoja(wb, 'Síntesis didáctica'):
        log.append('\n  [Síntesis didáctica]')
        n = corregir_sintesis(_hoja(wb, 'Síntesis didáctica'), log, registro)
        total_sintesis = n
        if n == 0:
            log.append('    Sin cambios necesarios.')
    else:
        log.append('\n  ⚠️  Hoja "Síntesis didáctica" no encontrada')

    # ── Planificación por unidades ─────────────────────────────────────────
    total_lenguaje = 0
    if _tiene_hoja(wb, 'Planificación por unidades'):
        log.append('\n  [Planificación por unidades]')
        a, mj, mm, b = corregir_planificacion(_hoja(wb, 'Planificación por unidades'), log, registro)
        total_alta, total_media_j, total_media_m, total_baja = a, mj, mm, b
        total_lenguaje = corregir_lenguaje_actividades(_hoja(wb, 'Planificación por unidades'), log, registro)
        if a + mj + mm + b + total_lenguaje == 0:
            log.append('    Sin cambios necesarios.')
    else:
        log.append('\n  ⚠️  Hoja "Planificación por unidades" no encontrada')

    # ── Verificación contra escala (post-correcciones) ─────────────────────
    fuente_escala = 'archivo' if escala_path else 'estandar'
    resultados_escala = verificar_escala(wb)
    log.extend(formatear_verificacion(resultados_escala, fuente_escala))

    # ── Verificación de momentos (Manual Diseño Instruccional UST) ──────────
    n_mom_ok = n_mom_adv = 0
    if _tiene_hoja(wb, 'Planificación por unidades'):
        n_mom_ok, n_mom_adv = verificar_momentos(
            _hoja(wb, 'Planificación por unidades'), log
        )
        verificar_consignas_foro(_hoja(wb, 'Planificación por unidades'), log)
        verificar_imperativo_momentos(_hoja(wb, 'Planificación por unidades'), log)

    # ── Revisión ortográfica/gramatical con LanguageTool ─────────────────
    if _tiene_hoja(wb, 'Planificación por unidades') and _LANGUAGETOOL_ACTIVO:
        verificar_lenguaje_momentos(
            _hoja(wb, 'Planificación por unidades'), log,
            autocorregir=_LANGUAGETOOL_AUTOCORREGIR
        )

    # ── Verificación de horas ─────────────────────────────────────────────
    n_hrs_ok = n_hrs_err = 0
    ws_sint_ref = _hoja(wb, 'Síntesis didáctica') if _tiene_hoja(wb, 'Síntesis didáctica') else None
    ws_plan_ref = _hoja(wb, 'Planificación por unidades') if _tiene_hoja(wb, 'Planificación por unidades') else None
    n_hrs_ok, n_hrs_err = verificar_horas(ws_plan_ref, ws_sint_ref, programa, log)

    # ── Verificación contra programa oficial (si se proveyó PDF) ───────────
    n_discrepancias = verificar_contra_programa(wb, programa, log)

    # ── Verificación Taxonomía de Bloom en RA ───────────────────────────────
    n_bloom_ok, n_bloom_prob, _ = verificar_verbos_bloom_planificacion(wb, log)

    # ── Verificación A+Se (si corresponde) ──────────────────────────────
    n_as_ok = n_as_err = n_as_manual = 0
    if es_as:
        if _tiene_hoja(wb, 'Planificación por unidades'):
            n_as_ok, n_as_err, n_as_manual = verificar_as(
                _hoja(wb, 'Planificación por unidades'), log
            )
        else:
            log.append('\n  ⚠️  A+Se: hoja "Planificación por unidades" no encontrada')

    # ── Guardar ────────────────────────────────────────────────────────────
    if not dry_run:
        os.makedirs(carpeta_revisado, exist_ok=True)
        # Si viene de original, renombrar para distinguirla
        nombre_salida = (nombre_revisado(os.path.basename(plan_path))
                         if es_original else os.path.basename(plan_path))
        dest = os.path.join(carpeta_revisado, nombre_salida)
        escribir_registro_cambios(wb, registro)
        wb.save(dest)
        log.append(f'\n  💾 Guardado en: Revisado/{nombre_salida}')
        log.append(f'  📋 Hoja "Registro de cambios": {len(registro)} cambio(s) documentado(s)')

    total = total_sintesis + total_alta + total_media_j + total_media_m + total_baja + total_lenguaje
    log.append(f'\n  RESUMEN DE CAMBIOS:')
    log.append(f'    Síntesis didáctica      : {total_sintesis} correcciones')
    log.append(f'    Alta  (proc/instr/tipo) : {total_alta} correcciones')
    log.append(f'    Media (Medio entrega J) : {total_media_j} correcciones')
    log.append(f'    Media (Indiv/Grupal  M) : {total_media_m} correcciones')
    log.append(f'    Baja  (% vacíos → 0)   : {total_baja} correcciones')
    log.append(f'    Lenguaje (actividades)  : {total_lenguaje} celdas corregidas')
    log.append(f'    ─────────────────────────────')
    log.append(f'    TOTAL                  : {total} correcciones')
    log.append(f'    🔵 Celdas modificadas marcadas con fuente azul (#0070C0)')
    log.append(f'    📖 Verificación momentos (Manual UST): {n_mom_ok}✅  {n_mom_adv}⚠️ advertencias')
    log.append(f'    ⏱️  Verificación de horas: {n_hrs_ok}✅  {n_hrs_err}❌')
    log.append(f'    🎯 Taxonomía de Bloom (RA): {n_bloom_ok}✅ verbos correctos, {n_bloom_prob}⚠️ verbos débiles')
    if programa:
        log.append(f'    📄 Verificación vs programa: {n_discrepancias} discrepancia(s) encontrada(s)')
    if es_as:
        log.append(f'    📌 Verificación A+Se: {n_as_ok}✅  {n_as_err}❌  {n_as_manual}⚠️ manual')
    if es_original:
        log.append(f'    ⚠️  Revisar manualmente las observaciones de la escala')

    return log, True


# ══════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description='Revisa y corrige planificaciones didácticas UST en ~/Desktop/DEL'
    )
    parser.add_argument(
        '--base', type=str, default=None,
        help='Carpeta raíz con las asignaturas (por defecto: ~/Desktop/DEL). '
             'Útil en Google Colab: --base /content/drive/MyDrive/DEL'
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Muestra correcciones sin guardar archivos'
    )
    parser.add_argument(
        '--asignatura', type=str, default=None,
        help='Procesa solo esta asignatura (nombre exacto de la carpeta)'
    )
    parser.add_argument(
        '--log', type=str, default=None,
        help='Guarda el reporte en este archivo (ej: reporte.txt)'
    )
    parser.add_argument(
        '--no-languagetool', action='store_true',
        help='Desactiva la revisión ortográfica/gramatical con LanguageTool API'
    )
    args = parser.parse_args()

    # Activar/desactivar LanguageTool globalmente
    global _LANGUAGETOOL_ACTIVO
    _LANGUAGETOOL_ACTIVO = not args.no_languagetool

    base = os.path.expanduser(args.base) if args.base else os.path.expanduser('~/Desktop/DEL')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')

    lineas = [
        f'REPORTE DE REVISIÓN — PLANIFICACIONES UST',
        f'Generado: {timestamp}',
        f'Carpeta base: {base}',
        f'Modo: {"DRY-RUN (sin guardar)" if args.dry_run else "CORRECCIÓN APLICADA"}',
        f'LanguageTool: {"desactivado (--no-languagetool)" if args.no_languagetool else "activo (es)"}',
    ]

    # Detectar carpetas de asignaturas
    if args.asignatura:
        carpetas = [os.path.join(base, args.asignatura)]
    else:
        carpetas = sorted([
            os.path.join(base, d)
            for d in os.listdir(base)
            if os.path.isdir(os.path.join(base, d)) and not d.startswith('.')
        ])

    if not carpetas:
        print('No se encontraron carpetas de asignaturas en', base)
        return

    procesadas = 0
    omitidas = 0

    for carpeta in carpetas:
        log, ok = procesar_asignatura(carpeta, dry_run=args.dry_run)
        lineas.extend(log)
        if ok:
            procesadas += 1
        else:
            omitidas += 1

    lineas.append(f'\n{"═"*70}')
    lineas.append(f'  TOTAL: {procesadas} asignatura(s) procesada(s), {omitidas} omitida(s)')
    lineas.append(f'{"═"*70}')

    reporte = '\n'.join(lineas)
    print(reporte)

    # Guardar log si se solicitó
    log_path = args.log or os.path.join(base, f'reporte_{datetime.now().strftime("%Y%m%d_%H%M")}.txt')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write(reporte)
    print(f'\n📄 Reporte guardado en: {log_path}')


if __name__ == '__main__':
    main()
