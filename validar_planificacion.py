"""
validar_planificacion.py — Validaciones preventivas DEL · UST
Verifica el .xlsx antes de procesarlo para detectar problemas estructurales.
"""

from __future__ import annotations
import openpyxl
from io import BytesIO

# Hojas requeridas
HOJAS_REQUERIDAS = ["Síntesis didáctica", "Planificación por unidades"]

# Cabeceras esperadas en fila 3 de "Planificación por unidades" (col 1-15)
# None = no se valida esa columna
CABECERAS_ESPERADAS = {
    1:  "resultado",          # RA o Resultado de Aprendizaje
    2:  "unidad",
    3:  "semana",
    6:  "momento",
    7:  "actividad",
    11: "tipo",               # Tipo de evaluación
    12: "procedimiento",
    14: "instrumento",
}

# Valores válidos por columna (insensible a mayúsculas/tildes)
MOMENTOS_VALIDOS = {"preparación", "preparacion", "desarrollo",
                    "trabajo independiente", "trabajo  independiente"}

TIPOS_VALIDOS = {"formativa", "sumativa", "diagnóstica", "diagnostica",
                 "formativo", "sumativo", "diagnóstico", "diagnostico", ""}


def _normalizar(val) -> str:
    """Convierte a str minúsculas sin espacios extra."""
    if val is None:
        return ""
    return str(val).strip().lower()


def validar_xlsx(fuente) -> list[dict]:
    """
    Valida un archivo .xlsx de planificación didáctica UST.

    Parameters
    ----------
    fuente : bytes | str
        Bytes del archivo o ruta en disco.

    Returns
    -------
    list[dict]  Lista de problemas detectados. Cada problema tiene:
        - nivel   : "error" | "advertencia"
        - codigo  : identificador corto
        - mensaje : descripción legible
    """
    problemas: list[dict] = []

    def problema(nivel, codigo, mensaje):
        problemas.append({"nivel": nivel, "codigo": codigo, "mensaje": mensaje})

    # ── 1. Cargar workbook ─────────────────────────────────────────────────
    try:
        if isinstance(fuente, (bytes, bytearray)):
            wb = openpyxl.load_workbook(BytesIO(fuente), read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(fuente, read_only=True, data_only=True)
    except Exception as exc:
        problema("error", "XLSX_INVALIDO",
                 f"No se pudo abrir el archivo: {exc}")
        return problemas

    hojas = wb.sheetnames

    # ── 2. Hojas requeridas ────────────────────────────────────────────────
    for hoja in HOJAS_REQUERIDAS:
        if hoja not in hojas:
            problema("error", "HOJA_FALTANTE",
                     f"Falta la hoja obligatoria: «{hoja}»")

    if any(p["codigo"] == "HOJA_FALTANTE" for p in problemas):
        return problemas  # sin hojas no tiene sentido continuar

    # ── 3. Síntesis didáctica — celdas clave no vacías ────────────────────
    ws_sint = wb["Síntesis didáctica"]
    _check_sint = [
        (7, 1, "Nombre de la asignatura"),
        (4, 4, "Código de asignatura"),
        (7, 3, "Créditos"),
        (11, 1, "Competencia / Resultado general"),
    ]
    for fila, col, etiqueta in _check_sint:
        try:
            val = ws_sint.cell(fila, col).value
        except Exception:
            val = None
        if not val or str(val).strip() == "":
            problema("advertencia", "SINT_CELDA_VACIA",
                     f"Síntesis didáctica: celda vacía esperada → {etiqueta} "
                     f"(fila {fila}, col {col})")

    # ── 4. Planificación por unidades — cabeceras ─────────────────────────
    ws_plan = wb["Planificación por unidades"]

    # Leer fila 3 como cabeceras
    fila_cabeceras = {}
    try:
        for cell in list(ws_plan.iter_rows(min_row=3, max_row=3, values_only=True))[0]:
            pass  # iter completo para read_only
        row3 = [c for c in ws_plan.iter_rows(min_row=3, max_row=3, values_only=True)][0]
        for idx, val in enumerate(row3, start=1):
            fila_cabeceras[idx] = _normalizar(val)
    except Exception:
        problema("advertencia", "CABECERAS_NO_LEIDAS",
                 "No se pudieron leer las cabeceras de «Planificación por unidades» (fila 3).")
        fila_cabeceras = {}

    for col_num, clave_esperada in CABECERAS_ESPERADAS.items():
        real = fila_cabeceras.get(col_num, "")
        if clave_esperada not in real:
            problema("advertencia", "CABECERA_INCORRECTA",
                     f"Columna {col_num}: se esperaba «{clave_esperada}», "
                     f"se encontró «{real or '(vacía)'}»")

    # ── 5. Datos mínimos en plan ──────────────────────────────────────────
    filas_datos = 0
    filas_momento_invalido = []
    filas_tipo_invalido = []

    try:
        for i, row in enumerate(
            ws_plan.iter_rows(min_row=4, max_col=15, values_only=True), start=4
        ):
            if any(c is not None for c in row):
                filas_datos += 1
                momento = _normalizar(row[5])   # col 6 → índice 5
                tipo    = _normalizar(row[10])  # col 11 → índice 10

                if momento and momento not in MOMENTOS_VALIDOS:
                    filas_momento_invalido.append((i, row[5]))
                if tipo and tipo not in TIPOS_VALIDOS:
                    filas_tipo_invalido.append((i, row[10]))
    except Exception:
        pass

    if filas_datos == 0:
        problema("error", "PLAN_VACIO",
                 "La hoja «Planificación por unidades» no tiene filas con datos.")

    for fila_n, val in filas_momento_invalido[:5]:
        problema("advertencia", "MOMENTO_INVALIDO",
                 f"Fila {fila_n}: Momento no reconocido → «{val}» "
                 "(esperado: Preparación / Desarrollo / Trabajo Independiente)")

    for fila_n, val in filas_tipo_invalido[:5]:
        problema("advertencia", "TIPO_INVALIDO",
                 f"Fila {fila_n}: Tipo de evaluación no reconocido → «{val}» "
                 "(esperado: Formativa / Sumativa / Diagnóstica)")

    # ── 6. Verificar que código asignatura tenga formato UST ──────────────
    import re
    codigo_cell = ws_sint.cell(4, 4).value
    if codigo_cell:
        codigo_str = str(codigo_cell).strip()
        if not re.match(r'^[A-Z]{3}-[A-Z]\d{4}$', codigo_str):
            problema("advertencia", "CODIGO_FORMATO",
                     f"El código «{codigo_str}» no tiene el formato esperado UST "
                     "(ej: AMI-S0112)")

    return problemas
